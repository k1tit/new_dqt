import os
import json
import pandas as pd
import logging
import traceback
import re
from datetime import datetime
import sys
import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)
try:
    from utils.symbols import Symbols
    from utils.column_matcher import ColumnMatcher
    from utils.file_manager import ErrorFileManager
    from core.memory_manager import MemoryManager
    from validators.completeness import CompletenessValidator
    from validators.conformity import ConformityValidator
    from validators.cross_column import CrossColumnEqualityValidator
    from validators.text_validators import SpecialCharactersValidator, ConsecutiveSpacesValidator, UppercaseValidator
    from validators.logical_validator import LogicalValidator
except ImportError as e:
    print(f'Ошибка импорта: {e}')
    raise

class FastDataQualityChecker:

    def __init__(self, db_path: str, rules_file: str, output_dir: str='quality_reports'):
        self.db_path = db_path
        self.rules_file = rules_file
        self.output_dir = output_dir
        self.memory_manager = MemoryManager(db_path)
        self.error_manager = ErrorFileManager(output_dir)
        self.column_matcher = ColumnMatcher()
        self.symbols = Symbols()
        self.results = []
        self.rule_errors = {}
        self.suspicious_rules = []
        self.processed_rules = 0
        self.skipped_rules = 0
        self.logger = logging.getLogger('FastDQChecker')
        self.MAX_ERRORS_TO_SAVE = 100000
        self.MASS_ERROR_THRESHOLD = 0.5
        self.colors = {'green': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'), 'red': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'), 'orange': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'), 'dark_red': PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid'), 'header': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'), 'header_font': Font(color='FFFFFF', bold=True), 'normal_font': Font(name='Calibri', size=11), 'bold_font': Font(bold=True), 'error_font': Font(color='FF0000', bold=True), 'success_font': Font(color='00B050', bold=True)}
        self.current_table = None
        self.current_rule = None
        self.start_time = None
        self.table_start_time = None
        os.makedirs(output_dir, exist_ok=True)
        os.makedirs(os.path.join(output_dir, 'errors'), exist_ok=True)
        logging.basicConfig(level=logging.INFO, format='%(levelname)s:%(name)s: %(message)s')
        self.column_mapping = self._load_column_mapping()
        self.table_handlers = self._load_table_handlers()
        self.FORCE_STANDARD_METHOD = ['DFKKBPTAXNUM1', 'DFKKBPTAXNUM2', 'DFKKBPTAXNUM3', 'DFKKBPTAXNUM4', 'DFKKBPTAXNUM5', 'DFKKBPTAXNUM6', 'DFKKBPTAXNUM', 'KNB1', 'KNB5', 'KNVP']
        self.DEBUG_TABLES = ['DFKKBPTAXNUM1', 'DFKKBPTAXNUM2', 'DFKKBPTAXNUM3', 'DFKKBPTAXNUM4', 'DFKKBPTAXNUM5', 'DFKKBPTAXNUM6', 'KNB1', 'KNB5', 'KNVP', 'BUT000']

    def _load_column_mapping(self):
        try:
            rules_dir = os.path.dirname(self.rules_file)
            mapping_path = os.path.join(rules_dir, 'map_column.json')
            if os.path.exists(mapping_path):
                with open(mapping_path, 'r', encoding='utf-8') as f:
                    mapping = json.load(f)
                print(f'   [INFO] Загружен map_column.json: {len(mapping)} таблиц')
                return mapping
            else:
                print(f'   [WARN] Файл map_column.json не найден: {mapping_path}')
                return {}
        except Exception as e:
            print(f'   [WARN] Ошибка загрузки map_column.json: {e}')
            return {}

    def _print_progress_bar(self, iteration, total, prefix='', suffix='', length=50, fill='█', print_end='\r'):
        percent = '{0:.1f}'.format(100 * (iteration / float(total)))
        filled_length = int(length * iteration // total)
        bar = fill * filled_length + '░' * (length - filled_length)
        sys.stdout.write(f'\r{prefix} |{bar}| {percent}% {suffix}')
        sys.stdout.flush()
        if iteration == total:
            sys.stdout.write('\n')

    def _print_rule_stats(self, rule_code, total_rows, error_count, exec_time, is_suspicious=False, mass_error=False):
        if total_rows > 0:
            success_rate = (total_rows - error_count) / total_rows * 100
            error_percent = error_count / total_rows * 100
        else:
            success_rate = 0
            error_percent = 0
        if error_count == 0:
            color = '\x1b[92m'
            status = '[OK] УСПЕШНО'
        elif mass_error:
            color = '\x1b[91m'
            status = '[!] МАССОВЫЕ'
        elif is_suspicious:
            color = '\x1b[93m'
            status = '[!] ПОДОЗР.'
        else:
            color = '\x1b[91m'
            status = '[!] ОШИБКИ'
        print(f'\r    {color}{status}\x1b[0m {rule_code:20} | Строк: {total_rows:8,} | Успех: {success_rate:6.1f}% | Ошибок: {error_count:8,} ({error_percent:5.1f}%) | Время: {exec_time:6.2f}с')

    def _print_table_header(self, table_name, rule_count, row_count):
        print(f'\n{'=' * 100}')
        print(f'ТАБЛИЦА: \x1b[1m{table_name}\x1b[0m')
        print(f'  Правил: {rule_count:3d} | Строк: {row_count:,} | Начало: {datetime.now().strftime('%H:%M:%S')}')
        print(f'{'-' * 100}')

    def _print_table_summary(self, table_name, elapsed_time):
        table_results = [r for r in self.results if r.get('table_name') == table_name]
        if table_results:
            success_count = len([r for r in table_results if r.get('status') == 'УСПЕШНО'])
            error_count = len([r for r in table_results if r.get('status') == 'ОШИБКИ'])
            suspicious_count = len([r for r in table_results if r.get('status') in ['ПОДОЗРИТЕЛЬНО', 'МАССОВЫЕ ОШИБКИ']])
            skipped_count = len([r for r in table_results if r.get('status') == 'ПРОПУЩЕНО'])
            failed_count = len([r for r in table_results if r.get('status') in ['ОШИБКА ВЫПОЛНЕНИЯ', 'ОШИБКА ОБРАБОТЧИКА']])
            total_rules = len(table_results)
            success_percent = success_count / total_rules * 100 if total_rules > 0 else 0
        else:
            total_rules = 0
            success_count = 0
            error_count = 0
            suspicious_count = 0
            skipped_count = 0
            failed_count = 0
            success_percent = 0
        print(f'{'-' * 100}')
        print(f'ИТОГ ТАБЛИЦЫ \x1b[1m{table_name}\x1b[0m:')
        print(f'  Всего правил: {total_rules:3d} | Время: {elapsed_time:.2f}с')
        print(f'  [OK] Успешно:    {success_count:3d} ({success_percent:.1f}%)')
        print(f'  [!] Ошибки:     {error_count:3d}')
        print(f'  [!] Подозрительные: {suspicious_count:3d}')
        if skipped_count > 0:
            print(f'  [~] Пропущено:  {skipped_count:3d}')
        if failed_count > 0:
            print(f'  [!] Ошибки выполнения: {failed_count:3d}')
        print(f'{'=' * 100}\n')

    def _load_table_handlers(self):
        handlers = {}
        try:
            import importlib
            try:
                module = importlib.import_module('table_scripts.but000_handler')
                if hasattr(module, 'BUT000Handler'):
                    handlers['BUT000'] = module.BUT000Handler
                    print(f'   [INFO] Загружен обработчик BUT000')
            except ImportError as e:
                print(f'   [WARN] Не удалось загрузить BUT000 обработчик: {e}')
            try:
                module = importlib.import_module('table_scripts.kna1_handler')
                if hasattr(module, 'KNA1Handler'):
                    handlers['KNA1'] = module.KNA1Handler
                    print(f'   [INFO] Загружен обработчик KNA1')
            except ImportError:
                pass
        except Exception as e:
            print(f'   [WARN] Ошибка загрузки обработчиков: {e}')
        print(f'   [INFO] Всего загружено обработчиков: {len(handlers)}')
        return handlers

    def load_configuration(self):
        try:
            with open(self.rules_file, 'r', encoding='utf-8') as f:
                rules = json.load(f)
            total_tables = len(rules)
            total_rules = sum((len(rules[table]) for table in rules))
            print(f'\n\x1b[1m[INFO]\x1b[0m Загружено {total_tables} таблиц, {total_rules} правил')
            return rules
        except Exception as e:
            self.logger.error(f'Ошибка загрузки конфигурации: {e}')
            return {}

    def run_quality_checks_fast(self, specific_table: str=None):
        print(f'\n' + '=' * 100)
        print(f'\x1b[1mЗАПУСК ПРОВЕРОК КАЧЕСТВА ДАННЫХ\x1b[0m')
        print(f'=' * 100)
        self.start_time = time.time()
        rules_config = self.load_configuration()
        if not rules_config:
            self.logger.error('[ERROR] Не удалось загрузить конфигурацию правил')
            return pd.DataFrame()
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.results = []
        self.rule_errors = {}
        self.suspicious_rules = []
        self.processed_rules = 0
        self.skipped_rules = 0
        if specific_table:
            print(f'[INFO] Проверяем только таблицу: {specific_table}')
        else:
            print(f'[INFO] Проверяем все таблицы')
        print(f'[INFO] Настройки: Сохраняется максимум {self.MAX_ERRORS_TO_SAVE:,} ошибок на правило')
        print(f'[INFO] Используется map_column.json: {len(self.column_mapping)} таблиц')
        print(f'[INFO] Стандартный метод для: {', '.join(self.FORCE_STANDARD_METHOD[:5])}...')
        print(f'\n[INFO] Загружаем данные из базы...')
        load_start = time.time()
        if hasattr(self.memory_manager, 'load_all_data_to_ram'):
            self.memory_manager.load_all_data_to_ram()
        available_tables = []
        if hasattr(self.memory_manager, 'data_cache'):
            available_tables = list(self.memory_manager.data_cache.keys())
        load_time = time.time() - load_start
        print(f'   [INFO] Загрузка завершена за {load_time:.2f} сек')
        print(f'   [INFO] Доступно таблиц в памяти: {len(available_tables)}')
        print(f'\n[DEBUG] ПРОВЕРКА ТАБЛИЦ ИЗ ПРАВИЛ:')
        tables_in_rules = list(rules_config.keys())
        for table in self.DEBUG_TABLES:
            if table in tables_in_rules:
                if table in available_tables:
                    df = self.memory_manager.get_table(table)
                    if df is not None:
                        print(f'  [OK] {table}: найдено {len(df):,} строк, {len(df.columns)} колонок')
                        if len(df.columns) > 0:
                            print(f'     Колонки: {list(df.columns)[:5]}{('...' if len(df.columns) > 5 else '')}')
                    else:
                        print(f'  [!] {table}: есть в кэше, но DataFrame = None')
                else:
                    print(f'  [!] {table}: НЕ НАЙДЕНА в доступных таблицах')
            else:
                print(f'  - {table}: нет в правилах')
        print(f'\n[DEBUG] ПРОВЕРКА НАЛОГОВЫХ ТАБЛИЦ:')
        tax_tables = [t for t in available_tables if 'TAX' in t.upper() or 'DFKK' in t.upper()]
        for table in tax_tables[:10]:
            df = self.memory_manager.get_table(table)
            if df is not None:
                print(f'  {table}: {len(df):,} строк, колонки: {list(df.columns)}')
        print(f'\n\x1b[1m[INFO]\x1b[0m Обрабатываем правила:')
        if specific_table:
            if specific_table in rules_config:
                self._process_table_rules(specific_table, rules_config[specific_table], available_tables, timestamp)
            else:
                print(f"\n[ERROR] В конфигурации нет правил для таблицы '{specific_table}'")
                print(f'   Доступные таблицы: {list(rules_config.keys())}')
                return pd.DataFrame()
        else:
            total_tables = len(rules_config)
            print(f'   [INFO] Всего таблиц для проверки: {total_tables}')
            for i, (table_name, table_rules) in enumerate(rules_config.items(), 1):
                print(f'\n[ПРОГРЕСС] Таблица {i}/{total_tables}...')
                self._process_table_rules(table_name, table_rules, available_tables, timestamp)
        overall_time = time.time() - self.start_time
        self._print_final_statistics()
        self._save_rule_errors()
        report_name = 'quality_check_report'
        if specific_table:
            report_name = f'quality_check_report_{specific_table}'
        self._create_correct_report(report_name)
        print(f'\n' + '=' * 100)
        print(f'\x1b[1mПРОВЕРКА ЗАВЕРШЕНА\x1b[0m')
        print(f'   Общее время: {overall_time:.2f} сек')
        print(f'   Скорость: {self.processed_rules / overall_time:.1f} правил/сек' if overall_time > 0 else '')
        print(f'=' * 100)
        results_df = pd.DataFrame(self.results)
        return results_df

    def _process_table_rules(self, table_name, table_rules, available_tables, timestamp):
        self.current_table = table_name
        self.table_start_time = time.time()
        if table_name in self.DEBUG_TABLES:
            print(f'\n[DEBUG] ОБРАБОТКА ТАБЛИЦЫ {table_name}:')
            print(f'  Доступна в памяти: {table_name in available_tables}')
            print(f'  Количество правил: {len(table_rules)}')
            if table_name in available_tables:
                df = self.memory_manager.get_table(table_name)
                if df is not None:
                    print(f'  Размер DataFrame: {len(df)} строк, {len(df.columns)} колонок')
                    print(f'  Колонки таблицы: {list(df.columns)}')
                else:
                    print(f'  ОШИБКА: DataFrame = None')
        if table_name not in available_tables:
            print(f'   \x1b[91m[ERROR]\x1b[0m Таблица {table_name} НЕ НАЙДЕНА в БД!')
            for rule in table_rules:
                self.skipped_rules += 1
                self._log_skipped_rule(rule, table_name, 'Таблица не найдена в БД', timestamp)
            elapsed_time = time.time() - self.table_start_time
            self._print_table_summary(table_name, elapsed_time)
            return
        df = self.memory_manager.get_table(table_name)
        if df is None or df.empty:
            print(f'   \x1b[93m[WARN]\x1b[0m Таблица {table_name} пуста! Пропускаем...')
            for rule in table_rules:
                self.skipped_rules += 1
                self._log_skipped_rule(rule, table_name, 'Таблица пуста', timestamp)
            elapsed_time = time.time() - self.table_start_time
            self._print_table_summary(table_name, elapsed_time)
            return
        self._print_table_header(table_name, len(table_rules), len(df))
        if table_name in self.FORCE_STANDARD_METHOD:
            print(f'   [INFO] Принудительно используем стандартный метод (обход обработчика)')
            self._process_with_standard_method_forced(table_name, df, table_rules, timestamp)
        elif table_name in self.table_handlers:
            print(f'   [INFO] Используем специальный обработчик')
            self._process_with_table_handler_real(table_name, df, table_rules, timestamp)
        else:
            print(f'   [INFO] Используем стандартный метод проверки')
            self._process_with_standard_method_forced(table_name, df, table_rules, timestamp)
        elapsed_time = time.time() - self.table_start_time
        self._print_table_summary(table_name, elapsed_time)

    def _process_with_standard_method_forced(self, table_name, df, table_rules, timestamp):
        total_rules = len(table_rules)
        if table_name in self.DEBUG_TABLES:
            print(f'\n    [DEBUG] Стандартный метод для таблицы: {table_name}')
            print(f'    [DEBUG] Колонки таблицы: {list(df.columns)}')
            print(f'    [DEBUG] Всего правил: {total_rules}')
        for i, rule in enumerate(table_rules, 1):
            self.processed_rules += 1
            rule_code = rule.get('rule_code', 'UNKNOWN')
            rule_desc = rule.get('rule_description', '')
            column_to_check = rule.get('column_name_checked', '')
            if table_name in self.DEBUG_TABLES:
                print(f'\n    [DEBUG] Правило {i}/{total_rules}: {rule_code}')
                print(f'    [DEBUG] Описание: {rule_desc}')
                print(f"    [DEBUG] Колонка из правила: '{column_to_check}'")
            matched_column = self._find_column_with_mapping_enhanced(table_name, df.columns, column_to_check, debug_flag=table_name in self.DEBUG_TABLES)
            if not matched_column:
                matched_column = self.column_matcher.find_column_match(df.columns, column_to_check)
                if table_name in self.DEBUG_TABLES and matched_column:
                    print(f'    [DEBUG] ColumnMatcher нашел: {matched_column}')
            if not matched_column:
                matched_column = self._find_most_similar_column_enhanced(df.columns, column_to_check)
                if table_name in self.DEBUG_TABLES and matched_column:
                    print(f'    [DEBUG] Поиск похожих нашел: {matched_column}')
            sys.stdout.write(f'\r    [{i:3d}/{total_rules:3d}] {rule_code:20} | ')
            if matched_column:
                sys.stdout.write(f'Колонка: {matched_column}')
            else:
                sys.stdout.write(f'\x1b[91mКолонка не найдена: {column_to_check}\x1b[0m')
            sys.stdout.flush()
            rule_start_time = time.time()
            if not matched_column:
                print(f"\n      \x1b[93m[WARN]\x1b[0m Колонка '{column_to_check}' не найдена в таблице {table_name}")
                if table_name in self.DEBUG_TABLES:
                    print(f'      [DEBUG] Доступные колонки: {list(df.columns)}')
                    print(f'      [DEBUG] Искали: {column_to_check}')
                execution_time = time.time() - rule_start_time
                self.results.append({'rule_code': rule_code, 'rule_description': rule_desc, 'quality_category': rule.get('quality_category', 'Unknown'), 'table_name': table_name, 'column_checked': column_to_check, 'matched_column': '', 'total_records': len(df), 'passed': len(df), 'failed': 0, 'success_rate_%': 100.0, 'execution_time_sec': round(execution_time, 2), 'check_date': timestamp, 'status': 'ПРОПУЩЕНО', 'status_color': 'gray', 'error_file': 'Нет', 'comments': f"Колонка '{column_to_check}' не найдена в таблице {table_name}"})
                self.skipped_rules += 1
                self._print_rule_stats(rule_code, len(df), 0, execution_time, False, False)
                continue
            try:
                if table_name in self.DEBUG_TABLES:
                    print(f'      [DEBUG] Выполняем проверку для колонки: {matched_column}')
                    print(f'      [DEBUG] Значения колонки (первые 3):')
                    if len(df) > 0:
                        for idx in range(min(3, len(df))):
                            val = df.iloc[idx].get(matched_column)
                            print(f"        Row {idx}: '{val}'")
                error_count, total_rows = self._execute_real_check(rule, table_name, df, matched_column, timestamp)
                execution_time = time.time() - rule_start_time
                is_suspicious = self._check_if_suspicious(rule_code, error_count, total_rows)
                mass_error = error_count > self.MAX_ERRORS_TO_SAVE
                self._print_rule_stats(rule_code, total_rows, error_count, execution_time, is_suspicious, mass_error)
            except Exception as e:
                print(f'\n      \x1b[91m[ERROR]\x1b[0m Ошибка при проверке правила {rule_code}: {str(e)}')
                if table_name in self.DEBUG_TABLES:
                    traceback.print_exc()
                execution_time = time.time() - rule_start_time
                self.results.append({'rule_code': rule_code, 'rule_description': rule_desc, 'quality_category': rule.get('quality_category', 'Unknown'), 'table_name': table_name, 'column_checked': column_to_check, 'matched_column': matched_column if 'matched_column' in locals() else '', 'total_records': len(df), 'passed': 0, 'failed': 0, 'success_rate_%': 0, 'execution_time_sec': round(execution_time, 2), 'check_date': timestamp, 'status': 'ОШИБКА ВЫПОЛНЕНИЯ', 'status_color': 'dark_red', 'error_file': 'Нет', 'comments': f'Ошибка выполнения: {str(e)[:100]}...'})
                self.skipped_rules += 1
                self._print_rule_stats(rule_code, len(df), 0, execution_time, False, False)

    def _process_with_table_handler_real(self, table_name, df, table_rules, timestamp):
        handler_class = self.table_handlers[table_name]
        print(f'\n    [DEBUG] Создаем обработчик для {table_name}')
        try:
            handler = handler_class(table_name, df, self.memory_manager, self)
            total_rules = len(table_rules)
            for i, rule in enumerate(table_rules, 1):
                self.processed_rules += 1
                rule_code = rule.get('rule_code', 'UNKNOWN')
                rule_desc = rule.get('rule_description', '')
                column_to_check = rule.get('column_name_checked', '')
                print(f'\n    [DEBUG] Правило {i}/{total_rules}: {rule_code}')
                print(f"    [DEBUG] Колонка из правила: '{column_to_check}'")
                sys.stdout.write(f'\r    [{i:3d}/{total_rules:3d}] {rule_code:20} | ')
                sys.stdout.write(f'Обработчик: {table_name}')
                sys.stdout.flush()
                rule_start_time = time.time()
                try:
                    print(f'    [DEBUG] Вызываем handler.validate_rule()...')
                    result = handler.validate_rule(rule)
                    execution_time = time.time() - rule_start_time
                    print(f'    [DEBUG] Обработчик вернул результат')
                    error_count = 0
                    error_df = pd.DataFrame()
                    matched_column = ''
                    total_rows = len(df)
                    if isinstance(result, dict):
                        print(f'    [DEBUG] Результат - словарь с ключами: {list(result.keys())}')
                        if 'error_count' in result:
                            error_count = result.get('error_count', 0)
                            error_df = result.get('error_df', pd.DataFrame())
                            matched_column = result.get('matched_column', '')
                            total_rows = result.get('total_rows', len(df))
                            print(f'    [DEBUG] Получили из validate_rule(): {error_count} ошибок')
                        elif 'success' in result and (not result.get('success', False)):
                            error_message = result.get('error_message', 'Неизвестная ошибка обработчика')
                            print(f'    [DEBUG] Обработчик вернул ошибку: {error_message}')
                            self._create_handler_error_result(rule, table_name, df, error_message, execution_time, timestamp)
                            continue
                    if error_count == 0 and hasattr(handler, 'get_results'):
                        print(f'    [DEBUG] Пробуем получить результаты через get_results()')
                        handler_results = handler.get_results()
                        if handler_results and len(handler_results) > 0:
                            handler_result = handler_results[0]
                            error_count = handler_result.get('failed', 0)
                            total_rows = handler_result.get('total_records', len(df))
                            matched_column = handler_result.get('matched_column', '')
                            print(f'    [DEBUG] Получили из get_results(): {error_count} ошибок')
                    if error_count == 0 and (not matched_column):
                        print(f'    [DEBUG] Не удалось получить результаты, считаем 0 ошибок')
                        matched_column = self._find_column_with_mapping_enhanced(table_name, df.columns, column_to_check)
                        if not matched_column:
                            matched_column = self.column_matcher.find_column_match(df.columns, column_to_check)
                    if hasattr(handler, 'get_errors'):
                        handler_errors = handler.get_errors()
                        if rule_code in handler_errors:
                            error_data = handler_errors[rule_code]
                            if 'error_df' in error_data and (not error_data['error_df'].empty):
                                error_df = error_data['error_df']
                    is_suspicious = self._check_if_suspicious(rule_code, error_count, total_rows)
                    mass_error = error_count > self.MAX_ERRORS_TO_SAVE
                    self._print_rule_stats(rule_code, total_rows, error_count, execution_time, is_suspicious, mass_error)
                    success_rows = total_rows - error_count
                    success_rate = success_rows / total_rows * 100 if total_rows > 0 else 0
                    if error_count == 0:
                        status = 'УСПЕШНО'
                        status_color = 'green'
                    elif is_suspicious:
                        if mass_error:
                            status = 'МАССОВЫЕ ОШИБКИ'
                        else:
                            status = 'ПОДОЗРИТЕЛЬНО'
                        status_color = 'orange'
                    else:
                        status = 'ОШИБКИ'
                        status_color = 'red'
                    if error_count > self.MAX_ERRORS_TO_SAVE:
                        error_file_status = f'Частично ({self.MAX_ERRORS_TO_SAVE:,} из {error_count:,})'
                    elif error_count > 0:
                        error_file_status = 'Есть'
                    else:
                        error_file_status = 'Нет'
                    comments = ''
                    if error_count > self.MAX_ERRORS_TO_SAVE:
                        comments = f'СОХРАНЕНО ТОЛЬКО {self.MAX_ERRORS_TO_SAVE:,} ИЗ {error_count:,} ОШИБОК!'
                    elif is_suspicious and total_rows > 0:
                        error_percent = error_count / total_rows * 100
                        comments = f'ПОДОЗРИТЕЛЬНО: {error_percent:.1f}% ДАННЫХ С ОШИБКАМИ'
                    result_dict = {'rule_code': rule_code, 'rule_description': rule_desc, 'quality_category': rule.get('quality_category', 'Unknown'), 'table_name': table_name, 'column_checked': column_to_check, 'matched_column': matched_column, 'total_records': total_rows, 'passed': success_rows, 'failed': error_count, 'success_rate_%': round(success_rate, 2), 'execution_time_sec': round(execution_time, 2), 'check_date': timestamp, 'status': status, 'status_color': status_color, 'error_file': error_file_status, 'comments': comments}
                    self.results.append(result_dict)
                    if error_df is not None and (not error_df.empty):
                        self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows)
                except Exception as e:
                    print(f'\n      \x1b[91m[ERROR]\x1b[0m Ошибка в обработчике: {str(e)}')
                    traceback.print_exc()
                    self._create_handler_error_result(rule, table_name, df, str(e), time.time() - rule_start_time, timestamp)
        except Exception as e:
            print(f'\n   \x1b[91m[ERROR]\x1b[0m Ошибка создания обработчика {table_name}: {str(e)}')
            traceback.print_exc()
            print(f'   [INFO] Используем стандартный метод вместо обработчика')
            self._process_with_standard_method_forced(table_name, df, table_rules, timestamp)

    def _execute_real_check(self, rule, table_name, df, matched_column, timestamp):
        rule_code = rule.get('rule_code', 'UNKNOWN')
        rule_description = rule.get('rule_description', 'Unknown rule')
        quality_category = rule.get('quality_category', 'Unknown')
        column_to_check = rule.get('column_name_checked', '')
        debug_mode = table_name in self.DEBUG_TABLES
        if debug_mode:
            print(f'    [DEBUG] Выполнение проверки для правила {rule_code}')
            print(f'    [DEBUG] Категория качества: {quality_category}')
            print(f'    [DEBUG] Используем колонку: {matched_column}')
            print(f'    [DEBUG] Описание правила: {rule_description}')
        rule_info = {'table_name': table_name, 'rule_code': rule_code, 'rule_description': rule_description, 'quality_category': quality_category, 'matched_column': matched_column, 'original_column': column_to_check}
        validator = self._get_validator_for_rule(rule_description, quality_category, rule_info)
        if debug_mode:
            print(f'    [DEBUG] Выбран валидатор: {type(validator).__name__}')
        params = {}
        if isinstance(validator, CrossColumnEqualityValidator):
            second_column = self._extract_second_column_from_description(rule_code, rule_description, df.columns, matched_column)
            if second_column and second_column in df.columns:
                params['second_column'] = second_column
                if debug_mode:
                    print(f'    [DEBUG] Найдена вторая колонка для сравнения: {second_column}')
            elif debug_mode:
                print(f'    [DEBUG] Вторая колонка не найдена для сравнения')
        if rule_code == 'RCCONF_13.2':
            validator = InequalityValidator(rule_info)
            if debug_mode:
                print(f'    [DEBUG] Используем InequalityValidator')
        if rule_code.startswith('RCCONF_5') or 'TAXNUM' in rule_code or 'TAX' in rule_description.upper():
            validator = TaxNumberValidator(rule_info)
            if debug_mode:
                print(f'    [DEBUG] Используем TaxNumberValidator для налогового правила')
        if rule_code == 'RCCONF_15.1' and str(table_name or '').strip().upper() == 'BUT000':
            org3_res = self._find_column_with_mapping_enhanced(table_name, df.columns, 'NAME_ORG3')
            if org3_res:
                params['org3_column_resolved'] = org3_res
        try:
            total_rows, error_count, error_df = validator.validate(df, matched_column, **params)
            if debug_mode:
                print(f'    [DEBUG] Результат проверки: {error_count} ошибок из {total_rows} строк')
            is_suspicious = self._check_if_suspicious(rule_code, error_count, total_rows)
            if error_df is not None and (not error_df.empty):
                self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows)
            success_rows = total_rows - error_count
            success_rate = success_rows / total_rows * 100 if total_rows > 0 else 0
            if error_count == 0:
                status = 'УСПЕШНО'
                status_color = 'green'
            elif is_suspicious:
                if error_count > self.MAX_ERRORS_TO_SAVE:
                    status = 'МАССОВЫЕ ОШИБКИ'
                else:
                    status = 'ПОДОЗРИТЕЛЬНО'
                status_color = 'orange'
            else:
                status = 'ОШИБКИ'
                status_color = 'red'
            if error_count > self.MAX_ERRORS_TO_SAVE:
                error_file_status = f'Частично ({self.MAX_ERRORS_TO_SAVE:,} из {error_count:,})'
            elif error_count > 0:
                error_file_status = 'Есть'
            else:
                error_file_status = 'Нет'
            comments = ''
            if error_count > self.MAX_ERRORS_TO_SAVE:
                comments = f'СОХРАНЕНО ТОЛЬКО {self.MAX_ERRORS_TO_SAVE:,} ИЗ {error_count:,} ОШИБОК!'
            elif is_suspicious and total_rows > 0:
                error_percent = error_count / total_rows * 100
                comments = f'ПОДОЗРИТЕЛЬНО: {error_percent:.1f}% ДАННЫХ С ОШИБКАМИ'
            result = {'rule_code': rule_code, 'rule_description': rule_description, 'quality_category': quality_category, 'table_name': table_name, 'column_checked': column_to_check, 'matched_column': matched_column, 'total_records': total_rows, 'passed': success_rows, 'failed': error_count, 'success_rate_%': round(success_rate, 2), 'execution_time_sec': 0, 'check_date': timestamp, 'status': status, 'status_color': status_color, 'error_file': error_file_status, 'comments': comments}
            self.results.append(result)
            return (error_count, total_rows)
        except Exception as e:
            print(f'\n      \x1b[91m[ERROR]\x1b[0m Ошибка при выполнении проверки: {str(e)}')
            if debug_mode:
                traceback.print_exc()
            raise

    def _find_column_with_mapping_enhanced(self, table_name, available_columns, target_column, debug_flag=False):
        import re
        if not target_column:
            if debug_flag:
                print(f'    [DEBUG] Целевая колонка пуста')
            return None
        if debug_flag:
            print(f"    [DEBUG] Поиск колонки: таблица='{table_name}', цель='{target_column}'")
            print(f'    [DEBUG] Доступные колонки ({len(available_columns)}): {available_columns}')
        target_upper = target_column.upper()
        for col in available_columns:
            if col.upper() == target_upper:
                if debug_flag:
                    print(f"    [DEBUG] Точное совпадение: '{col}'")
                return col
        if table_name in self.column_mapping:
            table_mapping = self.column_mapping[table_name]
            if debug_flag:
                print(f'    [DEBUG] Используем маппинг из JSON: {table_mapping}')
            if target_column in table_mapping:
                mapped_name = table_mapping[target_column]
                if mapped_name in available_columns:
                    if debug_flag:
                        print(f"    [DEBUG] Прямой маппинг: '{target_column}' -> '{mapped_name}'")
                    return mapped_name
            for mapped_key, mapped_value in table_mapping.items():
                if isinstance(mapped_value, str) and mapped_value.upper() == target_upper:
                    if mapped_key in available_columns:
                        if debug_flag:
                            print(f"    [DEBUG] Обратный маппинг: '{mapped_value}' -> '{mapped_key}'")
                        return mapped_key
        if 'TAX' in target_upper or 'DFKK' in table_name.upper():
            tax_variants = [target_upper, target_upper.replace('TAX', 'TAX_'), target_upper.replace('NUM', '_NUM'), target_upper.replace('TAXNUM', 'TAX_NUM')]
            for variant in tax_variants:
                for col in available_columns:
                    if col.upper() == variant:
                        if debug_flag:
                            print(f"    [DEBUG] Налоговый вариант: '{variant}' -> '{col}'")
                        return col
        for col in available_columns:
            col_upper = col.upper()
            target_for_match = target_upper.replace('_', '').replace('-', '').replace(' ', '')
            col_for_match = col_upper.replace('_', '').replace('-', '').replace(' ', '')
            if target_for_match in col_for_match or col_for_match in target_for_match:
                if debug_flag:
                    print(f"    [DEBUG] Частичное совпадение: '{col}'")
                return col
            target_words = set(re.findall('[A-Z][a-z]*|[A-Z]+', target_upper))
            col_words = set(re.findall('[A-Z][a-z]*|[A-Z]+', col_upper))
            if target_words & col_words:
                if debug_flag:
                    print(f"    [DEBUG] Общие слова: '{col}'")
                return col
        if debug_flag:
            print(f"    [DEBUG] Колонка '{target_column}' не найдена!")
        return None

    def _find_most_similar_column_enhanced(self, columns, target_column):
        if not target_column:
            return None
        target_upper = target_column.upper().replace('_', '').replace('-', '').replace(' ', '')
        best_match = None
        best_score = 0
        for col in columns:
            col_upper = col.upper().replace('_', '').replace('-', '').replace(' ', '')
            score = 0
            if col_upper == target_upper:
                score += 100
            if target_upper in col_upper:
                score += 50 + len(target_upper) * 2
            elif col_upper in target_upper:
                score += 50 + len(col_upper) * 2
            common = set(target_upper) & set(col_upper)
            if len(common) > 0:
                score += len(common) * 3
            target_words = set(re.findall('[A-Z][a-z]*|[A-Z]+', target_column.upper()))
            col_words = set(re.findall('[A-Z][a-z]*|[A-Z]+', col.upper()))
            common_words = target_words & col_words
            if common_words:
                score += len(common_words) * 20
            if len(target_upper) > 2 and len(col_upper) > 2:
                if target_upper[-3:] == col_upper[-3:]:
                    score += 15
                if target_upper[-2:] == col_upper[-2:]:
                    score += 10
            if score > best_score:
                best_score = score
                best_match = col
        return best_match if best_score > 15 else None

    def _save_rule_result_complete(self, rule, table_name, df, matched_column, execution_time, timestamp, error_count, error_df=None):
        if isinstance(rule, dict):
            rule_code = rule.get('rule_code', 'UNKNOWN')
            rule_desc = rule.get('rule_description', '')
            quality_category = rule.get('quality_category', 'Unknown')
            original_column = rule.get('column_name_checked', '')
        else:
            rule_code = rule
            rule_desc = ''
            quality_category = 'Unknown'
            original_column = ''
        total_rows = len(df)
        success_rows = total_rows - error_count
        success_rate = success_rows / total_rows * 100 if total_rows > 0 else 0
        is_suspicious = self._check_if_suspicious(rule_code, error_count, total_rows)
        if error_count == 0:
            status = 'УСПЕШНО'
            status_color = 'green'
        elif is_suspicious:
            if error_count > self.MAX_ERRORS_TO_SAVE:
                status = 'МАССОВЫЕ ОШИБКИ'
            else:
                status = 'ПОДОЗРИТЕЛЬНО'
            status_color = 'orange'
        else:
            status = 'ОШИБКИ'
            status_color = 'red'
        if error_count > self.MAX_ERRORS_TO_SAVE:
            error_file_status = f'Частично ({self.MAX_ERRORS_TO_SAVE:,} из {error_count:,})'
        elif error_count > 0:
            error_file_status = 'Есть'
        else:
            error_file_status = 'Нет'
        comments = ''
        if error_count > self.MAX_ERRORS_TO_SAVE:
            comments = f'СОХРАНЕНО ТОЛЬКО {self.MAX_ERRORS_TO_SAVE:,} ИЗ {error_count:,} ОШИБОК!'
        elif is_suspicious and total_rows > 0:
            error_percent = error_count / total_rows * 100
            comments = f'ПОДОЗРИТЕЛЬНО: {error_percent:.1f}% ДАННЫХ С ОШИБКАМИ'
        result = {'rule_code': rule_code, 'rule_description': rule_desc, 'quality_category': quality_category, 'table_name': table_name, 'column_checked': original_column, 'matched_column': matched_column, 'total_records': total_rows, 'passed': success_rows, 'failed': error_count, 'success_rate_%': round(success_rate, 2), 'execution_time_sec': round(execution_time, 2), 'check_date': timestamp, 'status': status, 'status_color': status_color, 'error_file': error_file_status, 'comments': comments}
        self.results.append(result)

    def _save_rule_error_with_limit(self, rule_code, table_name, error_df, error_count, is_suspicious, total_rows):
        if error_df is None or error_df.empty:
            return
        if len(error_df) > self.MAX_ERRORS_TO_SAVE:
            error_df_sampled = error_df.head(self.MAX_ERRORS_TO_SAVE).copy()
            error_df_sampled['LIMIT_NOTICE'] = f'ПОКАЗАНЫ ПЕРВЫЕ {self.MAX_ERRORS_TO_SAVE:,} ИЗ {error_count:,} ОШИБОК'
            error_df_to_save = error_df_sampled
        else:
            error_df_to_save = error_df.copy()
        key = f'{rule_code}_{table_name}'
        self.rule_errors[key] = {'rule_code': rule_code, 'table_name': table_name, 'error_df': error_df_to_save, 'error_count': error_count, 'is_suspicious': is_suspicious, 'total_rows': total_rows}

    def _check_if_suspicious(self, rule_code, error_count, total_rows):
        if rule_code in ['RCCONF_12.2', 'RCCONF_12.3']:
            return True
        if error_count > 1000000:
            return True
        if total_rows > 0 and error_count / total_rows > self.MASS_ERROR_THRESHOLD:
            return True
        if rule_code == 'RCCONF_15.1' and total_rows > 0 and (error_count / total_rows > 0.3):
            return True
        return False

    def _get_validator_for_rule(self, rule_description, quality_category, rule_info):
        rule_desc_lower = rule_description.lower()
        rule_code = rule_info.get('rule_code', '')
        if rule_code.startswith('RCCONF_5') or 'TAXNUM' in rule_code or 'TAX' in rule_description.upper():
            return TaxNumberValidator(rule_info)
        if rule_code == 'RCCONF_15.1':
            return LogicalValidator(rule_info, self.error_manager)
        if rule_code == 'RCCONF_13.2':
            return InequalityValidator(rule_info)
        elif 'равен' in rule_desc_lower or 'равны' in rule_desc_lower:
            return CrossColumnEqualityValidator(rule_info)
        elif 'специальные символы' in rule_desc_lower:
            return SpecialCharactersValidator(rule_info)
        elif 'недопустимые пробелы' in rule_desc_lower:
            return ConsecutiveSpacesValidator(rule_info)
        elif 'верхний регистр' in rule_desc_lower:
            return UppercaseValidator(rule_info)
        elif 'отсутствует' in rule_desc_lower or 'missing' in rule_desc_lower:
            return CompletenessValidator(rule_info)
        else:
            return ConformityValidator(rule_info)

    def _extract_second_column_from_description(self, rule_code, rule_description, columns, first_column):
        desc_lower = rule_description.lower()
        if 'name_org1' in desc_lower and 'name_org2' in desc_lower:
            if 'name_org1' in first_column.lower():
                return self._find_column_with_mapping_enhanced(None, columns, 'NAME_ORG2')
            else:
                return self._find_column_with_mapping_enhanced(None, columns, 'NAME_ORG1')
        if 'name_org1' in desc_lower and ('name_org3' in desc_lower or 'name_org4' in desc_lower):
            if 'name_org1' in first_column.lower():
                if 'name_org3' in desc_lower:
                    return self._find_column_with_mapping_enhanced(None, columns, 'NAME_ORG3')
                else:
                    return self._find_column_with_mapping_enhanced(None, columns, 'NAME_ORG4')
        if 'taxnum' in desc_lower:
            if '1' in first_column.lower():
                return self._find_column_with_mapping_enhanced(None, columns, 'TAXNUM2')
            elif '2' in first_column.lower():
                return self._find_column_with_mapping_enhanced(None, columns, 'TAXNUM1')
            elif '3' in first_column.lower():
                return self._find_column_with_mapping_enhanced(None, columns, 'TAXNUM4')
            elif '5' in first_column.lower():
                return self._find_column_with_mapping_enhanced(None, columns, 'TAXNUM6')
        return None

    def _log_skipped_rule(self, rule, table_name, reason, timestamp):
        rule_code = rule.get('rule_code', 'UNKNOWN')
        self.results.append({'rule_code': rule_code, 'rule_description': rule.get('rule_description', 'Unknown rule'), 'quality_category': rule.get('quality_category', 'Unknown'), 'table_name': table_name, 'column_checked': rule.get('column_name_checked', ''), 'matched_column': '', 'total_records': 0, 'passed': 0, 'failed': 0, 'success_rate_%': 0, 'execution_time_sec': 0, 'check_date': timestamp, 'status': 'ПРОПУЩЕНО', 'status_color': 'gray', 'error_file': 'Нет', 'comments': f'Пропущено: {reason}'})

    def _log_failed_rule(self, rule, table_name, error_message, timestamp):
        rule_code = rule.get('rule_code', 'UNKNOWN')
        self.results.append({'rule_code': rule_code, 'rule_description': rule.get('rule_description', 'Unknown rule'), 'quality_category': rule.get('quality_category', 'Unknown'), 'table_name': table_name, 'column_checked': rule.get('column_name_checked', ''), 'matched_column': '', 'total_records': 0, 'passed': 0, 'failed': 0, 'success_rate_%': 0, 'execution_time_sec': 0, 'check_date': timestamp, 'status': 'ОШИБКА ВЫПОЛНЕНИЯ', 'status_color': 'dark_red', 'error_file': 'Нет', 'comments': f'Ошибка: {error_message}'})

    def _create_handler_error_result(self, rule, table_name, df, error_message, execution_time, timestamp):
        rule_code = rule.get('rule_code', 'UNKNOWN')
        rule_desc = rule.get('rule_description', '')
        print(f'      \x1b[93m[WARN]\x1b[0m Ошибка обработчика: {error_message}')
        result = {'rule_code': rule_code, 'rule_description': rule_desc, 'quality_category': rule.get('quality_category', 'Unknown'), 'table_name': table_name, 'column_checked': rule.get('column_name_checked', ''), 'matched_column': '', 'total_records': len(df), 'passed': 0, 'failed': 0, 'success_rate_%': 0, 'execution_time_sec': round(execution_time, 2), 'check_date': timestamp, 'status': 'ОШИБКА ОБРАБОТЧИКА', 'status_color': 'dark_red', 'error_file': 'Нет', 'comments': f'Ошибка обработчика: {error_message[:100]}...'}
        self.results.append(result)
        self.skipped_rules += 1
        self._print_rule_stats(rule_code, len(df), 0, execution_time, False, False)

    def run(self, specific_table: str=None, table_list: list=None):
        if table_list:
            return self.run_with_selected_tables(table_list)
        elif specific_table:
            return self.run_quality_checks_fast(specific_table)
        else:
            return self.run_quality_checks_fast()

    def run_with_selected_tables(self, table_names: list):
        return self.run_quality_checks_fast()

    def _save_rule_errors(self):
        if not self.rule_errors:
            print(f'\n[INFO] Нет ошибок для сохранения')
            return
        print(f'\n[INFO] Сохранение ошибок по правилам...')
        errors_dir = os.path.join(self.output_dir, 'errors')
        os.makedirs(errors_dir, exist_ok=True)
        saved_count = 0
        for key, error_data in self.rule_errors.items():
            try:
                rule_code = error_data['rule_code']
                table_name = error_data['table_name']
                error_df = error_data['error_df']
                if error_df is None or error_df.empty:
                    continue
                filename = f'{rule_code}_{table_name}_errors.xlsx'
                filepath = os.path.join(errors_dir, filename)
                error_df.to_excel(filepath, index=False, engine='openpyxl')
                saved_count += 1
            except Exception as e:
                print(f'   [ERROR] Ошибка сохранения {key}: {e}')
        print(f'   Сохранено файлов: {saved_count}')

    def _create_correct_report(self, report_name: str='quality_check_report'):
        try:
            if not self.results:
                print(f'\n[INFO] Нет данных для отчета')
                return
            excel_path = os.path.join(self.output_dir, f'{report_name}.xlsx')
            wb = Workbook()
            ws = wb.active
            ws.title = 'Сводка проверок'
            ws['A1'] = 'СВОДКА ПРОВЕРОК КАЧЕСТВА ДАННЫХ'
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f'Настройки: Сохраняется максимум {self.MAX_ERRORS_TO_SAVE:,} ошибок на правило'
            ws['A2'].font = Font(size=10)
            check_date = self.results[0].get('check_date', '') if self.results else ''
            ws['A3'] = f'Дата проверки: {check_date}'
            ws['A3'].font = Font(size=9, italic=True)
            ws['A4'] = ''
            headers = ['Код правила', 'Описание', 'Категория', 'Таблица', 'Колонка', 'Всего записей', 'Успешно', 'Ошибок', '% успеха', 'Статус', 'Время (сек)', 'Файл ошибок', 'Комментарии']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col_num, value=header)
                cell.fill = self.colors['header']
                cell.font = self.colors['header_font']
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row_num = 7
            for result in self.results:
                values = [result.get('rule_code', ''), result.get('rule_description', ''), result.get('quality_category', ''), result.get('table_name', ''), result.get('column_checked', ''), result.get('total_records', 0), result.get('passed', 0), result.get('failed', 0), result.get('success_rate_%', 0), result.get('status', ''), result.get('execution_time_sec', 0), result.get('error_file', 'Нет'), result.get('comments', '')]
                for col_num, value in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.font = self.colors['normal_font']
                    if col_num == 10:
                        status_color = result.get('status_color', '')
                        if status_color == 'green':
                            cell.fill = self.colors['green']
                        elif status_color == 'red':
                            cell.fill = self.colors['red']
                        elif status_color == 'orange':
                            cell.fill = self.colors['orange']
                        elif status_color == 'dark_red':
                            cell.fill = self.colors['dark_red']
                            cell.font = Font(color='FFFFFF', bold=True)
                    elif col_num == 8:
                        failed_count = result.get('failed', 0)
                        if failed_count == 0:
                            cell.fill = self.colors['green']
                        elif failed_count > self.MAX_ERRORS_TO_SAVE:
                            cell.fill = self.colors['orange']
                            cell.font = Font(color='FF0000', bold=True)
                        else:
                            cell.fill = self.colors['red']
                row_num += 1
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            wb.save(excel_path)
            print(f'\n[INFO] Цветной отчет сохранен: {excel_path}')
        except Exception as e:
            print(f'\n[ERROR] Ошибка при создании отчета: {e}')
            traceback.print_exc()

    def _print_final_statistics(self):
        successful_rules = len([r for r in self.results if r.get('status') == 'УСПЕШНО'])
        failed_rules = len([r for r in self.results if r.get('status') == 'ОШИБКИ'])
        suspicious_rules = len([r for r in self.results if r.get('status') in ['ПОДОЗРИТЕЛЬНО', 'МАССОВЫЕ ОШИБКИ']])
        skipped_rules = len([r for r in self.results if r.get('status') == 'ПРОПУЩЕНО'])
        error_rules = len([r for r in self.results if r.get('status') in ['ОШИБКА ВЫПОЛНЕНИЯ', 'ОШИБКА ОБРАБОТЧИКА']])
        total_rules = successful_rules + failed_rules + suspicious_rules + skipped_rules + error_rules
        print(f'\n' + '=' * 100)
        print(f'\x1b[1mИТОГОВАЯ СТАТИСТИКА:\x1b[0m')
        print(f'=' * 100)
        print(f'Всего правил: \x1b[1m{total_rules}\x1b[0m')
        if total_rules > 0:
            print(f'  [OK] Успешно:      \x1b[92m{successful_rules:4d}\x1b[0m ({successful_rules / total_rules * 100:.1f}%)')
            print(f'  [!] Ошибки:       \x1b[91m{failed_rules:4d}\x1b[0m ({failed_rules / total_rules * 100:.1f}%)')
            print(f'  [!] Подозрительные: \x1b[93m{suspicious_rules:4d}\x1b[0m ({suspicious_rules / total_rules * 100:.1f}%)')
            print(f'  [~] Пропущено:    \x1b[90m{skipped_rules:4d}\x1b[0m ({skipped_rules / total_rules * 100:.1f}%)')
            print(f'  [!] Ошибки выполнения: \x1b[91m{error_rules:4d}\x1b[0m ({error_rules / total_rules * 100:.1f}%)')
        print(f'=' * 100)
        suspicious_list = [r for r in self.results if r.get('status') in ['ПОДОЗРИТЕЛЬНО', 'МАССОВЫЕ ОШИБКИ']]
        if suspicious_list:
            print(f'\n\x1b[93mПОДОЗРИТЕЛЬНЫЕ ПРАВИЛА (требуют проверки логики):\x1b[0m')
            for rule in suspicious_list[:10]:
                print(f'  • {rule['rule_code']:20} - {rule['table_name']:15} - {rule.get('failed', 0):,} ошибок')
        overall_time = time.time() - self.start_time
        print(f'\nВремя выполнения: {overall_time:.2f} сек')
        if overall_time > 0 and total_rules > 0:
            print(f'Скорость: {total_rules / overall_time:.1f} правил/сек')

class TaxNumberValidator:

    def __init__(self, rule_info):
        self.rule_info = rule_info

    def validate(self, df, column_name, second_column=None):
        total_rows = len(df)
        errors = []
        rule_code = self.rule_info.get('rule_code', '')
        print(f'    [TaxValidator] Проверка правила {rule_code} для колонки {column_name}')
        if column_name not in df.columns:
            print(f'    [TaxValidator] ОШИБКА: Колонка {column_name} не найдена в DataFrame')
            return (total_rows, 0, pd.DataFrame())
        if 'accuracy' in self.rule_info.get('rule_description', '').lower() or 'format' in self.rule_info.get('rule_description', '').lower():
            print(f'    [TaxValidator] Проверка формата налогового номера')
            errors = self._validate_tax_format(df, column_name)
        elif 'same value' in self.rule_info.get('rule_description', '').lower() or 'uniqueness' in self.rule_info.get('rule_description', '').lower():
            print(f'    [TaxValidator] Проверка уникальности')
            errors = self._validate_tax_uniqueness(df, column_name)
        else:
            print(f'    [TaxValidator] Общая проверка конформности')
            errors = self._validate_tax_conformity(df, column_name)
        error_df = pd.DataFrame(errors) if errors else pd.DataFrame()
        print(f'    [TaxValidator] Найдено {len(errors)} ошибок')
        return (total_rows, len(errors), error_df)

    def _validate_tax_format(self, df, column_name):
        errors = []
        for idx, row in df.iterrows():
            tax_value = row.get(column_name)
            if pd.isna(tax_value) or not str(tax_value).strip():
                continue
            tax_str = str(tax_value).strip()
            clean_tax = re.sub('[\\s\\-]', '', tax_str)
            if len(clean_tax) < 5:
                errors.append({'row_index': idx, 'row_id': idx + 1, column_name: tax_value, 'error_message': f'Налоговый номер слишком короткий: {len(clean_tax)} символов (минимум 5)', 'tax_value': tax_str, 'clean_length': len(clean_tax)})
                continue
            if not any((char.isdigit() for char in clean_tax)):
                errors.append({'row_index': idx, 'row_id': idx + 1, column_name: tax_value, 'error_message': 'Налоговый номер должен содержать цифры', 'tax_value': tax_str})
                continue
            if len(clean_tax) > 20:
                errors.append({'row_index': idx, 'row_id': idx + 1, column_name: tax_value, 'error_message': f'Налоговый номер слишком длинный: {len(clean_tax)} символов (максимум 20)', 'tax_value': tax_str, 'clean_length': len(clean_tax)})
        return errors

    def _validate_tax_uniqueness(self, df, column_name):
        errors = []
        tax_columns = []
        for col in df.columns:
            if 'TAX' in col.upper() or 'NUM' in col.upper():
                tax_columns.append(col)
        print(f'    [TaxValidator] Проверяем уникальность {column_name} среди {tax_columns}')
        for idx, row in df.iterrows():
            current_tax = row.get(column_name)
            if pd.isna(current_tax) or not str(current_tax).strip():
                continue
            current_tax_str = str(current_tax).strip()
            for other_col in tax_columns:
                if other_col == column_name:
                    continue
                other_tax = row.get(other_col)
                if pd.isna(other_tax) or not str(other_tax).strip():
                    continue
                other_tax_str = str(other_tax).strip()
                if current_tax_str == other_tax_str:
                    errors.append({'row_index': idx, 'row_id': idx + 1, column_name: current_tax, other_col: other_tax, 'error_message': f'{column_name} совпадает с {other_col}', 'values': f'"{current_tax_str}" == "{other_tax_str}"'})
                    break
        return errors

    def _validate_tax_conformity(self, df, column_name):
        errors = []
        for idx, row in df.iterrows():
            tax_value = row.get(column_name)
            if pd.notna(tax_value):
                tax_str = str(tax_value)
                forbidden_chars = ['<', '>', '&', '"', "'", '\\', '/', '|']
                found_chars = []
                for char in forbidden_chars:
                    if char in tax_str:
                        found_chars.append(char)
                if found_chars:
                    errors.append({'row_index': idx, 'row_id': idx + 1, column_name: tax_value, 'error_message': f'Найдены запрещенные символы: {', '.join(found_chars)}', 'forbidden_chars': list(found_chars)})
        return errors

class InequalityValidator:

    def __init__(self, rule_info):
        self.rule_info = rule_info

    def validate(self, df, column_name, second_column=None):
        total_rows = len(df)
        errors = []
        if not second_column:
            second_column = self._find_second_column(df.columns, column_name)
        if not second_column or second_column not in df.columns:
            for idx, row in df.iterrows():
                val1 = row.get(column_name)
                if pd.notna(val1) and str(val1).strip():
                    errors.append({'row_index': idx, 'row_id': idx + 1, column_name: val1, 'error_message': f'{column_name} должно быть пустым (проверка неравенства)', 'values': f'"{val1}" != ""'})
        else:
            for idx, row in df.iterrows():
                val1 = row.get(column_name)
                val2 = row.get(second_column)
                if pd.isna(val1) or pd.isna(val2):
                    continue
                str1 = str(val1).strip().lower()
                str2 = str(val2).strip().lower()
                if not str1 or not str2:
                    continue
                if str1 == str2:
                    errors.append({'row_index': idx, 'row_id': idx + 1, column_name: val1, second_column: val2, 'error_message': f'{column_name} не должно быть равно {second_column}', 'values': f'"{val1}" == "{val2}"'})
        error_df = pd.DataFrame(errors) if errors else pd.DataFrame()
        return (total_rows, len(errors), error_df)

    def _find_second_column(self, available_columns, first_column):
        first_lower = first_column.lower()
        if '2' in first_lower or 'org2' in first_lower or 'name2' in first_lower:
            for col in available_columns:
                col_lower = col.lower()
                if '1' in col_lower or 'org1' in col_lower or 'name1' in col_lower:
                    return col
        elif '1' in first_lower or 'org1' in first_lower or 'name1' in first_lower:
            for col in available_columns:
                col_lower = col.lower()
                if '2' in col_lower or 'org2' in col_lower or 'name2' in col_lower:
                    return col
        return None
__all__ = ['FastDataQualityChecker', 'InequalityValidator', 'TaxNumberValidator']