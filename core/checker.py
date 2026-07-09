import os
import json
import shutil
import pandas as pd
import logging
import traceback
import re
from datetime import datetime
from typing import Optional
import sys
import time
import threading
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)
try:
    from utils.empty_rows import fully_empty_rows_mask
    from utils.sqlite_safe import connect_sqlite
    from utils.symbols import Symbols
    from utils.column_matcher import ColumnMatcher
    from utils.file_manager import ErrorFileManager
    from core.memory_manager import MemoryManager
    from validators.completeness import CompletenessValidator
    from validators.conformity import ConformityValidator
    from validators.cross_column import CrossColumnEqualityValidator
    from validators.cross_column_equality import CrossColumnEqualityCheckValidator
    from validators.payment_terms_consistency import PaymentTermsConsistencyValidator
    from validators.recon_account_consistency import ReconAccountConsistencyValidator
    from validators.text_validators import SpecialCharactersValidator, ConsecutiveSpacesValidator, UppercaseValidator
    from validators.advanced_special_characters import AdvancedSpecialCharactersValidator
    from validators.logical_validator import LogicalValidator
except ImportError as e:
    print(f'Ошибка импорта: {e}')
    raise

class FastDataQualityChecker:
    CHECKER_BUILD_ID = '2026-07-09-adr2-scope-9038-01-01'
    ADRC_TABLE_ALIASES = frozenset({'ADRC', 'DM_CUSTOMER_ADDRESS', '/LOT/GC_ADR', 'LOTGC_ADR'})
    RULES_KTOKD_ONLY_9038_SCOPE = frozenset({'RCCOMP_113.1', 'RCCOMP_115.1', 'RCCOMP_142.1', 'RCCOMP_143.1'})
    RULES_FORCE_KNA1_KTOKD_JOIN = frozenset({'RCCONF_113.1', 'RCCONF_115.11', 'RCCONF_24.1', 'RCCOMP_113.1', 'RCCOMP_115.1', 'RCCOMP_142.1', 'RCCOMP_143.1', 'RCCONF_154.4', 'RCCOMP_149.1', 'RCCOMP_149.2'})
    RULES_ERROR_EXPORT_KNA1_KTOKD = frozenset({'RCCOMP_113.1', 'RCCOMP_115.1', 'RCCONF_113.1', 'RCCONF_24.1', 'RCCONF_115.11', 'RCCOMP_142.1', 'RCCOMP_143.1', 'RCCONF_154.4'})
    RULES_SAVE_ALL_ERRORS = frozenset({'RCCONF_39.5', 'RCCONF_39.5.2', 'RCCONF_18.2', 'RCCONF_63.1'})
    TABLES_SAVE_ALL_ERRORS = frozenset({'ADR2', 'BUT000'})
    ADR2_NON_BLOCKED_MOBILE_RULES = frozenset({
        'RCCOMP_375.1', 'RCCOMP_375.1.2',
        'RCCONF_39.3', 'RCCONF_39.3.2', 'RCCONF_39.5', 'RCCONF_39.5.2',
    })
    ADR2_DM_CUSTOMER_SCOPE_ACCOUNT_GROUP = '9038'
    KNA1_JOIN_BLOCKED_COLUMNS = frozenset({'CLIENT', 'CL', 'MANDT', 'MANDANT'})
    KNA1_JOIN_VIA_BUT020_TABLES = frozenset({'ADRC', 'ADR2'})

    def __init__(self, db_path: str, rules_file: str, output_dir: str='quality_reports', parallel_tables: int=0, use_async_load: bool=False, debug: bool=False, reference_datetime=None):
        self.db_path = db_path
        self.rules_file = rules_file
        self.output_dir = output_dir
        self.parallel_tables = max(0, int(parallel_tables))
        self.use_async_load = bool(use_async_load)
        self.debug = bool(debug)
        self.reference_datetime = reference_datetime
        self._parallel_lock = threading.Lock() if self.parallel_tables else None
        self.memory_manager = MemoryManager(db_path)
        print(f'[CHECKER] {self.CHECKER_BUILD_ID} | {os.path.abspath(__file__)}', flush=True)
        self.error_manager = ErrorFileManager(output_dir)
        self.column_matcher = ColumnMatcher()
        self.symbols = Symbols()
        self.results = []
        self.rule_errors = {}
        self.saved_error_files = {}
        self.last_errors_dir = None
        self.last_report_path = None
        self.last_stable_report_path = None
        self.last_file_timestamp = None
        self.suspicious_rules = []
        self.processed_rules = 0
        self.skipped_rules = 0
        self.logger = logging.getLogger('FastDQChecker')
        self.MAX_ERRORS_TO_SAVE = 100000
        self.EXCEL_MAX_ROWS = 1048576
        self.MASS_ERROR_THRESHOLD = 0.5
        self.column_map = self._load_column_map()
        self.ausp_atinn_mapping = self._load_ausp_atinn_mapping()
        self.colors = {'green': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'), 'red': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'), 'orange': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'), 'dark_red': PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid'), 'header': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'), 'header_font': Font(color='FFFFFF', bold=True), 'normal_font': Font(name='Calibri', size=11), 'bold_font': Font(bold=True), 'error_font': Font(color='FF0000', bold=True), 'success_font': Font(color='00B050', bold=True)}
        self.current_table = None
        self.current_rule = None
        self.start_time = None
        self.table_start_time = None
        os.makedirs(output_dir, exist_ok=True)
        os.makedirs(os.path.join(output_dir, 'errors'), exist_ok=True)
        logging.basicConfig(level=logging.INFO, format='%(levelname)s:%(name)s: %(message)s')
        self.table_handlers = self._load_table_handlers()

    def _saves_all_errors(self, rule_code, table_name):
        rule_u = self._normalize_rule_code(rule_code)
        tbl = str(table_name or '').strip().upper()
        if rule_u in self.RULES_SAVE_ALL_ERRORS:
            return True
        if tbl in self.TABLES_SAVE_ALL_ERRORS:
            return True
        if tbl.startswith('DFKKBPTAXNUM'):
            return True
        return False

    def _error_save_limit(self, rule_code, table_name):
        if self._saves_all_errors(rule_code, table_name):
            return self.EXCEL_MAX_ROWS
        return self.MAX_ERRORS_TO_SAVE

    def _load_column_map(self):
        column_map_path = os.path.join(parent_dir, 'json files', 'column_map.json')
        try:
            if os.path.exists(column_map_path):
                with open(column_map_path, 'r', encoding='utf-8') as f:
                    column_map = json.load(f)
                print(f'[INFO] Загружен маппинг колонок для {len(column_map)} таблиц')
                return column_map
            else:
                print(f'[WARN] Файл column_map.json не найден: {column_map_path}')
                return {}
        except Exception as e:
            print(f'[WARN] Ошибка загрузки column_map.json: {e}')
            return {}

    def _load_ausp_atinn_mapping(self):
        conf_path = os.path.join(parent_dir, 'json files', 'conf_ausp_atinn_mapping.json')
        try:
            if os.path.exists(conf_path):
                with open(conf_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                mapping = data.get('column_name_checked_to_atinn', data) if isinstance(data, dict) else {}
                self.ausp_atinn_to_temporary_name = data.get('atinn_to_temporary_column_name', {}) if isinstance(data, dict) else {}
                if mapping:
                    print(f'[INFO] Загружен маппинг AUSP ATINN для {len(mapping)} правил, временные имена колонок: {len(self.ausp_atinn_to_temporary_name)}')
                return mapping
            self.ausp_atinn_to_temporary_name = {}
            return {}
        except Exception as e:
            print(f'[WARN] Ошибка загрузки conf_ausp_atinn_mapping.json: {e}')
            self.ausp_atinn_to_temporary_name = {}
            return {}

    def _normalize_ausp_name(self, name):
        if not name:
            return ''
        s = (name or '').strip().upper().replace(' ', '_')
        if s.startswith('Z_'):
            return s[2:]
        if s.startswith('Z') and len(s) > 1 and (s[1] != '_'):
            return s[1:]
        return s

    def _find_ausp_columns(self, columns, table_name):
        cols = [c for c in columns if c is not None]
        atinn_col = None
        atwrt_col = None
        for c in cols:
            raw = str(c).strip().upper()
            cu = raw.replace(' ', '').replace('_', '')
            if cu == 'ATINN':
                atinn_col = c
            if cu == 'ATWRT':
                atwrt_col = c
        if not atinn_col:
            atinn_col = self._find_column_alternative(cols, 'ATINN', table_name)
        if not atwrt_col:
            atwrt_col = self._find_column_alternative(cols, 'ATWRT', table_name)
        if not atinn_col:
            for c in cols:
                if 'ATINN' in str(c).upper() or re.sub('[^A-Za-z0-9]', '', str(c).upper()) == 'ATINN':
                    atinn_col = c
                    break
        if not atwrt_col:
            for c in cols:
                if 'ATWRT' in str(c).upper() or re.sub('[^A-Za-z0-9]', '', str(c).upper()) == 'ATWRT':
                    atwrt_col = c
                    break
        if not atinn_col:
            for c in cols:
                if re.sub('[^A-Za-z0-9]', '', str(c).upper()) == 'ATINN':
                    atinn_col = c
                    break
        if not atwrt_col:
            for c in cols:
                if re.sub('[^A-Za-z0-9]', '', str(c).upper()) == 'ATWRT':
                    atwrt_col = c
                    break
        if (table_name or '').strip().upper() == 'AUSP' and len(cols) >= 3:
            if not atinn_col:
                atinn_col = cols[1]
            if not atwrt_col:
                atwrt_col = cols[2]
        if (table_name or '').strip().upper() == 'AUSP' and (not atinn_col or not atwrt_col):
            self._debug_ausp_columns(columns, table_name)
        return (atinn_col, atwrt_col)

    def _debug_ausp_columns(self, columns, table_name):
        cols = [c for c in columns if c is not None]
        lines = ['', '   [AUSP DEBUG] === почему не найдены колонки ATINN/ATWRT ===', f'   [AUSP DEBUG] Всего колонок: {len(cols)}']
        for i, c in enumerate(cols):
            raw = str(c).strip().upper()
            cu = raw.replace(' ', '').replace('_', '')
            clean = re.sub('[^A-Za-z0-9]', '', raw)
            repr_c = repr(c)[:80]
            lines.append(f'   [AUSP DEBUG]   {i}: repr={repr_c}')
            lines.append(f'   [AUSP DEBUG]       upper={raw!r}  replace(space,_)={cu!r}  clean(only letters)={clean!r}')
            lines.append(f"   [AUSP DEBUG]       ==ATINN? {cu == 'ATINN'}  ==ATWRT? {cu == 'ATWRT'}  'ATINN' in name? {'ATINN' in raw}  'ATWRT' in name? {'ATWRT' in raw}")
        lines.append('   [AUSP DEBUG] === конец отладки ===')
        print('\n'.join(lines))
    AUSP_ATINN_TO_COLUMN = {'143': 'CCAF', '604': 'RED_OUTLET', '148': 'ZGLOBAL_CUSTOMER', '151': 'ZTRADE_NAME'}

    def _ausp_atinn_to_column_name(self, atinn_value):
        if atinn_value is None:
            return None
        return self.AUSP_ATINN_TO_COLUMN.get(str(atinn_value).strip())

    def _normalize_atinn_for_filter(self, value):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return ''
        s = str(value).strip()
        try:
            return str(int(float(s)))
        except (ValueError, TypeError):
            return re.sub('\\.0+$', '', s) if s else ''

    def _ausp_atinn_mask(self, series, atinn_value):
        target = self._normalize_atinn_for_filter(atinn_value)
        if not target:
            return pd.Series(False, index=series.index)
        normalized = series.apply(lambda x: self._normalize_atinn_for_filter(x))
        return normalized == target

    def _resolve_ausp_atinn_value(self, rule):
        column_to_check = (rule.get('column_name_checked') or '').strip()
        biz = (rule.get('business_attribute_name') or '').strip()
        atinn_value = None
        if column_to_check:
            cu = column_to_check.upper()
            if '143' in cu or 'ATINN(143)' in cu or 'ATINN=143' in cu or ('ATINN =143' in cu):
                atinn_value = '143'
            elif '604' in cu or 'ATINN=604' in cu or 'ATINN = 604' in cu:
                atinn_value = '604'
            elif '148' in cu or 'ATINN=148' in cu:
                atinn_value = '148'
            elif '151' in cu or 'ATINN=151' in cu:
                atinn_value = '151'
        if atinn_value:
            return atinn_value
        if biz:
            bu = biz.upper()
            if 'CCAF' in bu or 'CAF' in bu:
                return '143'
            if 'RED' in bu and 'OUTLET' in bu or bu == 'OUTLET':
                return '604'
            if 'GLOBAL' in bu and 'CUSTOMER' in bu:
                return '148'
            if 'TRADE' in bu and 'NAME' in bu or ('TRADING' in bu and 'GROUP' in bu):
                return '151'
        candidates = [column_to_check] if column_to_check else []
        if biz and biz not in candidates:
            candidates.append(biz)
        if self.ausp_atinn_mapping and candidates:
            for key, val in self.ausp_atinn_mapping.items():
                key_norm = self._normalize_ausp_name(key)
                key_upper = (key or '').strip().upper()
                for cand in candidates:
                    if not cand:
                        continue
                    cand_norm = self._normalize_ausp_name(cand)
                    cand_upper = cand.strip().upper()
                    if key_upper == cand_upper or key_norm == cand_norm:
                        atinn_value = str(val).strip()
                        break
                if atinn_value:
                    break
        if not atinn_value and column_to_check:
            atinn_match = re.search('ATINN\\s*[=\\(]\\s*(\\d+)', column_to_check, re.IGNORECASE)
            if atinn_match:
                atinn_value = atinn_match.group(1).strip()
        return atinn_value

    def _build_ausp_split(self, df, table_name):
        if df is None or df.empty or (table_name or '').strip().upper() != 'AUSP':
            return None
        atinn_col, atwrt_col = self._find_ausp_columns(df.columns, table_name)
        if not atinn_col or not atwrt_col:
            return None
        out = {}
        for atinn_val, temp_name in self.AUSP_ATINN_TO_COLUMN.items():
            mask = self._ausp_atinn_mask(df[atinn_col], atinn_val)
            slice_df = df.loc[mask].copy()
            slice_df = slice_df.rename(columns={atwrt_col: temp_name})
            out[atinn_val] = (slice_df, temp_name)
        return out

    def _apply_ausp_filter(self, df, column_to_check, table_name, rule=None):
        if df is None or df.empty:
            return (None, None, None)
        t = (table_name or '').strip().upper()
        if t != 'AUSP':
            return (None, None, None)
        candidates = [(column_to_check or '').strip()]
        if rule:
            biz = (rule.get('business_attribute_name') or '').strip()
            if biz and biz not in candidates:
                candidates.append(biz)
        atinn_value = None
        if self.ausp_atinn_mapping:
            for key, val in self.ausp_atinn_mapping.items():
                key_norm = self._normalize_ausp_name(key)
                key_upper = (key or '').strip().upper()
                for cand in candidates:
                    if not cand:
                        continue
                    cand_norm = self._normalize_ausp_name(cand)
                    cand_upper = cand.strip().upper()
                    if key_upper == cand_upper or key_norm == cand_norm:
                        atinn_value = str(val).strip()
                        break
                if atinn_value:
                    break
        if not atinn_value and column_to_check:
            atinn_match = re.search('ATINN\\s*[=\\(]\\s*(\\d+)', column_to_check or '', re.IGNORECASE)
            if atinn_match:
                atinn_value = atinn_match.group(1).strip()
        if not atinn_value:
            return (None, None, None)
        atinn_col, atwrt_col = self._find_ausp_columns(df.columns, table_name)
        if not atinn_col or not atwrt_col:
            return (None, None, None)
        mask = self._ausp_atinn_mask(df[atinn_col], atinn_value)
        filtered = df.loc[mask].copy()
        temporary_name = self._ausp_atinn_to_column_name(atinn_value)
        if not temporary_name and rule:
            temporary_name = (rule.get('business_attribute_name') or '').strip()
        if not temporary_name:
            temporary_name = atwrt_col
        return (filtered, atwrt_col, temporary_name)

    def _print_progress_bar(self, iteration, total, prefix='', suffix='', length=50, fill='█', print_end='\r'):
        percent = '{0:.1f}'.format(100 * (iteration / float(total)))
        filled_length = int(length * iteration // total)
        bar = fill * filled_length + '░' * (length - filled_length)
        sys.stdout.write(f'\r{prefix} |{bar}| {percent}% {suffix}')
        sys.stdout.flush()
        if iteration == total:
            sys.stdout.write('\n')

    def _print_rule_stats(self, rule_code, total_rows, error_count, exec_time, is_suspicious=False, mass_error=False, not_evaluated=False):
        total_records = total_rows - error_count + error_count
        if total_records > 0:
            success_rate = (total_rows - error_count) / total_records * 100
            error_percent = error_count / total_records * 100
        else:
            success_rate = 0
            error_percent = 0
        if not_evaluated or (total_rows == 0 and error_count == 0):
            color = '\x1b[91m'
            status = '[!] НЕ ОЦЕНЕНО'
        elif error_count == 0:
            color = '\x1b[92m'
            status = '[OK] УСПЕШНО'
        elif mass_error:
            color = '\x1b[91m'
            status = '[!] МАССОВЫЕ'
        elif is_suspicious:
            color = '\x1b[93m'
            status = '[!] ПОДОЗР.'
        else:
            color = '\x1b[91m'
            status = '[!] ОШИБКИ'
        print(f'\r    {color}{status}\x1b[0m {rule_code:20} | Оценено: {total_rows:8,} | Успех: {success_rate:6.1f}% | Ошибок: {error_count:8,} ({error_percent:5.1f}%) | Время: {exec_time:6.2f}с')

    def _print_table_header(self, table_name, rule_count, row_count):
        print('\n' + '=' * 100)
        print(f'ТАБЛИЦА: \x1b[1m{table_name}\x1b[0m')
        print(f"  Правил: {rule_count:3d} | Строк: {row_count:,} | Начало: {datetime.now().strftime('%H:%M:%S')}")
        print('-' * 100)

    def _print_table_summary(self, table_name, elapsed_time, success_count, error_count, suspicious_count):
        total_rules = success_count + error_count + suspicious_count
        success_percent = success_count / total_rules * 100 if total_rules > 0 else 0
        print('-' * 100)
        print(f'ИТОГ ТАБЛИЦЫ \x1b[1m{table_name}\x1b[0m:')
        print(f'  Всего правил: {total_rules:3d} | Время: {elapsed_time:.2f}с')
        print(f'  [OK] Успешно:    {success_count:3d} ({success_percent:.1f}%)')
        print(f'  [!] Ошибки:     {error_count:3d}')
        print(f'  [!] Подозрительные: {suspicious_count:3d}')
        print('=' * 100 + '\n')

    def _load_table_handlers(self, silent=False):
        handlers = {}
        import importlib

        def _log(msg):
            if not silent:
                print(msg)
        try:
            module = importlib.import_module('table_scripts.but000_handler')
            if hasattr(module, 'BUT000Handler'):
                handlers['BUT000'] = module.BUT000Handler
                _log(f'   [INFO] Загружен обработчик BUT000')
        except ImportError as e:
            _log(f'   [WARN] Не удалось загрузить обработчик BUT000: {e}')
        try:
            module = importlib.import_module('table_scripts.adrc_handler')
            if hasattr(module, 'ADRCHandler'):
                handlers['ADRC'] = module.ADRCHandler
                _log(f'   [INFO] Загружен обработчик ADRC')
        except ImportError as e:
            _log(f'   [WARN] Не удалось загрузить обработчик ADRC: {e}')
        try:
            module = importlib.import_module('table_scripts.kna1_handler')
            if hasattr(module, 'KNA1Handler'):
                handlers['KNA1'] = module.KNA1Handler
                _log(f'   [INFO] Загружен обработчик KNA1')
        except ImportError as e:
            _log(f'   [WARN] Не удалось загрузить обработчик KNA1: {e}')
        try:
            module = importlib.import_module('table_scripts.taxnum_handler')
            if hasattr(module, 'TaxNumHandler'):
                handlers['DFKKBPTAXNUM'] = module.TaxNumHandler
                handlers['DFKKBPTAXNUM1'] = module.TaxNumHandler
                handlers['DFKKBPTAXNUM2'] = module.TaxNumHandler
                handlers['DFKKBPTAXNUM3'] = module.TaxNumHandler
                handlers['DFKKBPTAXNUM4'] = module.TaxNumHandler
                handlers['DFKKBPTAXNUM5'] = module.TaxNumHandler
                handlers['DFKKBPTAXNUM6'] = module.TaxNumHandler
                _log(f'   [INFO] Загружен обработчик TaxNumHandler')
        except ImportError as e:
            _log(f'   [WARN] Не удалось загрузить обработчик TaxNumHandler: {e}')
        _log(f'   [INFO] Всего загружено обработчиков: {len(handlers)}')
        return handlers

    def reload_table_handlers(self):
        import importlib
        import sys
        mod_names = ('table_scripts.but000_handler', 'table_scripts.adrc_handler', 'table_scripts.kna1_handler', 'table_scripts.taxnum_handler')
        for name in mod_names:
            if name in sys.modules:
                try:
                    importlib.reload(sys.modules[name])
                except Exception as e:
                    print(f'   [WARN] reload {name}: {e}')
        self.table_handlers = self._load_table_handlers(silent=True)
        print(f'   [INFO] Обработчики таблиц обновлены с диска ({len(self.table_handlers)} шт.)')

    def _apply_rule_time_column_map(self, df, table_name: str):
        try:
            from utils.column_map_resolver import apply_column_headers_for_rules
            return apply_column_headers_for_rules(df, table_name, self.column_map, parent_dir, log_renames=True)
        except ImportError:
            return df.copy() if df is not None else df

    def _get_table_for_rules(self, table_name: str):
        if not hasattr(self, '_rule_time_column_cache'):
            self._rule_time_column_cache = {}
        cache_key = str(table_name or '').strip().upper()
        if cache_key in self._rule_time_column_cache:
            return self._rule_time_column_cache[cache_key].copy()
        raw = self.memory_manager.get_table(table_name)
        if raw is None or raw.empty:
            return raw
        mapped = self._apply_rule_time_column_map(raw, table_name)
        self._rule_time_column_cache[cache_key] = mapped.copy()
        return self._rule_time_column_cache[cache_key].copy()

    def _get_mapped_column_name(self, table_name, column_name):
        try:
            from utils.column_map_resolver import map_logical_to_sap
            mapped = map_logical_to_sap(table_name, column_name, self.column_map, parent_dir)
            if mapped and mapped != column_name:
                print(f"      [MAP] {table_name}: '{column_name}' -> '{mapped}' (из column_map.json)")
            return mapped
        except ImportError:
            pass
        if not self.column_map:
            return column_name
        table_mapping = None
        if table_name in self.column_map:
            table_mapping = self.column_map[table_name]
        else:
            tn = str(table_name or '').strip().upper()
            for k, v in self.column_map.items():
                if str(k).strip().upper() == tn and isinstance(v, dict):
                    table_mapping = v
                    break
        if not table_mapping:
            return column_name
        for logical_name, real_name in table_mapping.items():
            if str(logical_name).startswith('_'):
                continue
            if str(real_name).upper() == str(column_name or '').upper():
                return column_name
        if column_name in table_mapping:
            return table_mapping[column_name]
        for logical_name, real_name in table_mapping.items():
            if str(logical_name).startswith('_'):
                continue
            if str(logical_name).upper() == str(column_name or '').upper():
                return real_name
        return column_name

    def _resolve_column_for_rule(self, df, column_name, table_name):
        if not column_name or df is None:
            return None
        try:
            from utils.column_map_resolver import resolve_column_in_df, map_logical_to_sap
            sap_name = map_logical_to_sap(table_name, column_name, self.column_map, parent_dir)
            for target in (sap_name, column_name):
                if not target:
                    continue
                found = resolve_column_in_df(df, target, table_name, self.column_map, parent_dir)
                if found:
                    if target != column_name or found != column_name:
                        print(f"      [MAP] Колонка по column_map: '{column_name}' -> '{found}'")
                    return found
        except ImportError:
            pass
        return self._find_column_alternative(df.columns, column_name, table_name)

    def _fix_column_name_for_taxnum(self, column_name, table_name):
        if table_name.startswith('DFKKBPTAXNUM') and len(table_name) > 12 and table_name[12:].isdigit():
            if column_name.upper().startswith('TAXNUM') and len(column_name) > 6 and column_name[6:].isdigit():
                return 'TAXNUM'
        if table_name == 'DFKKBPTAXNUM' and column_name.upper().startswith('TAXNUM'):
            return 'TAXNUM'
        return column_name

    def load_configuration(self):
        try:
            with open(self.rules_file, 'r', encoding='utf-8') as f:
                rules = json.load(f)
            total_tables = len(rules)
            total_rules = sum((len(rules[table]) for table in rules))
            print(f'\n\x1b[1m[INFO]\x1b[0m Загружено {total_tables} таблиц, {total_rules} правил')
            return rules
        except Exception as e:
            self.logger.error(f'Ошибка загрузки конфигурации: {e}')
            return {}

    def run_quality_checks_fast(self, specific_table: str=None, table_list: list=None, only_rule_codes: set=None):

        def _filter_rules(rules, codes):
            if not codes:
                return rules
            return [r for r in rules if r.get('rule_code') in codes]
        print(f'\n' + '=' * 100)
        print(f'\x1b[1mЗАПУСК ПРОВЕРОК КАЧЕСТВА ДАННЫХ\x1b[0m')
        print(f'Сборка checker: {self.CHECKER_BUILD_ID}')
        print(f'=' * 100)
        self.start_time = time.time()
        self._rule_time_column_cache = {}
        rules_config = self.load_configuration()
        if not rules_config:
            self.logger.error('[ERROR] Не удалось загрузить конфигурацию правил')
            self._save_totals_by_table()
            return pd.DataFrame()
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.results = []
        self.rule_errors = {}
        self.suspicious_rules = []
        self.processed_rules = 0
        self.skipped_rules = 0
        if table_list:
            print(f'[INFO] Проверяем выбранные таблицы: {len(table_list)} шт.')
        elif specific_table:
            print(f'[INFO] Проверяем только таблицу: {specific_table}')
        else:
            print(f'[INFO] Проверяем все таблицы')
        if only_rule_codes:
            print(f'[INFO] Изолированный режим: только правила {sorted(only_rule_codes)}')
        print(f'[INFO] Настройки: Сохраняется максимум {self.MAX_ERRORS_TO_SAVE:,} ошибок на правило')
        print(f'\n[INFO] Загружаем данные из базы...')
        load_start = time.time()
        if hasattr(self.memory_manager, 'load_all_data_to_ram'):
            if specific_table:
                tables_to_load = self._expand_ausp_for_load([specific_table])
                if self.use_async_load and hasattr(self.memory_manager, 'load_selected_tables_to_ram_async_sync'):
                    self.memory_manager.load_selected_tables_to_ram_async_sync(tables_to_load)
                elif hasattr(self.memory_manager, 'load_selected_tables_to_ram'):
                    self.memory_manager.load_selected_tables_to_ram(tables_to_load)
                else:
                    self.memory_manager.load_all_data_to_ram()
            elif table_list:
                tables_to_load = self._expand_ausp_for_load(list(table_list))
                if self.use_async_load and hasattr(self.memory_manager, 'load_selected_tables_to_ram_async_sync'):
                    self.memory_manager.load_selected_tables_to_ram_async_sync(tables_to_load)
                elif hasattr(self.memory_manager, 'load_selected_tables_to_ram'):
                    self.memory_manager.load_selected_tables_to_ram(tables_to_load)
                else:
                    self.memory_manager.load_all_data_to_ram()
            elif self.use_async_load and hasattr(self.memory_manager, 'load_selected_tables_to_ram_async_sync'):
                all_tables = self.memory_manager._get_all_table_names()
                self.memory_manager.load_selected_tables_to_ram_async_sync(all_tables, add_reference_tables=True)
            else:
                self.memory_manager.load_all_data_to_ram()
        if hasattr(self.memory_manager, '_finalize_load_postprocess'):
            self.memory_manager._finalize_load_postprocess()
        available_tables = []
        if hasattr(self.memory_manager, 'data_cache'):
            available_tables = list(self.memory_manager.data_cache.keys())
        load_time = time.time() - load_start
        print(f'   [INFO] Загрузка завершена за {load_time:.2f} сек')
        print(f'   [INFO] Доступно таблиц в памяти: {len(available_tables)}')
        if available_tables:
            table_sizes = []
            for table in available_tables[:10]:
                df = self.memory_manager.get_table(table)
                if df is not None:
                    table_sizes.append((table, len(df)))
            if table_sizes:
                print(f'\n   [INFO] Размеры таблиц (первые {len(table_sizes)}):')
                for table, size in sorted(table_sizes, key=lambda x: x[1], reverse=True)[:5]:
                    print(f'      {table:25} -> {size:10,} строк')
            if 'DFKKBPTAXNUM' in self.memory_manager.data_cache:
                print(f'\n   [INFO] Подсчёт строк по таблицам-алиасам DFKKBPTAXNUM (по каждой в отдельности):')
                aliases_to_show = getattr(self.memory_manager, 'DFKKBPTAXNUM_TABLES', ())
                for alias in aliases_to_show:
                    if alias in self.memory_manager.data_cache:
                        alias_df = self.memory_manager.get_table(alias)
                        n = len(alias_df) if alias_df is not None else 0
                        print(f'      {alias:25} -> {n:10,} строк')
        if self.parallel_tables:
            print(f'   [INFO] Параллельная обработка таблиц: {self.parallel_tables} потоков')
        print(f'\n\x1b[1m[INFO]\x1b[0m Обрабатываем правила:')
        if specific_table:
            if specific_table in rules_config:
                table_rules = _filter_rules(rules_config[specific_table], only_rule_codes)
                if not table_rules:
                    print(f"\n[WARN] Для таблицы '{specific_table}' нет правил с кодами {only_rule_codes}. Завершение.")
                    self._save_totals_by_table()
                    return pd.DataFrame()
                self._process_table_rules(specific_table, table_rules, available_tables, timestamp)
            elif specific_table == 'AUSP':
                for t in self.AUSP_TABLE_GROUP:
                    if t in rules_config and t in available_tables:
                        table_rules = _filter_rules(rules_config[t], only_rule_codes)
                        if table_rules:
                            self._process_table_rules(t, table_rules, available_tables, timestamp)
            elif specific_table == 'DFKKBPTAXNUM':
                for alias in self.DFKKBPTAXNUM_ALIASES:
                    if alias in rules_config and alias in available_tables:
                        table_rules = _filter_rules(rules_config[alias], only_rule_codes)
                        if table_rules:
                            self._process_table_rules(alias, table_rules, available_tables, timestamp)
            else:
                print(f"\n[ERROR] В конфигурации нет правил для таблицы '{specific_table}'")
                print(f'   Доступные таблицы: {list(rules_config.keys())}')
                self._save_totals_by_table()
                return pd.DataFrame()
        elif table_list:
            tables_to_process = []
            for t in table_list:
                if t == 'AUSP':
                    for a in self.AUSP_TABLE_GROUP:
                        if a in rules_config:
                            tr = _filter_rules(rules_config[a], only_rule_codes)
                            if tr:
                                tables_to_process.append((a, tr))
                elif t in rules_config:
                    tr = _filter_rules(rules_config[t], only_rule_codes)
                    if tr:
                        tables_to_process.append((t, tr))
            missing = [t for t in table_list if t != 'AUSP' and t not in rules_config]
            if missing:
                print(f'\n[WARN] Таблицы не найдены в конфигурации: {missing}')
            total_tables = len(tables_to_process)
            print(f'   [INFO] Таблиц для проверки: {total_tables}')
            self._run_tables_loop(tables_to_process, available_tables, timestamp, total_tables)
        else:
            tables_to_process = []
            for table_name, table_rules in rules_config.items():
                tr = _filter_rules(table_rules, only_rule_codes)
                if tr:
                    tables_to_process.append((table_name, tr))
            total_tables = len(tables_to_process)
            print(f'   [INFO] Всего таблиц для проверки: {total_tables}')
            self._run_tables_loop(tables_to_process, available_tables, timestamp, total_tables)
        overall_time = time.time() - self.start_time
        file_timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        self._print_final_statistics()
        self._save_rule_errors(file_timestamp)
        self._sync_results_error_files()
        try:
            self._save_totals_by_table(file_timestamp)
        except Exception as e:
            print(f'\n[ERROR] Ошибка записи в папку total: {e}')
            traceback.print_exc()
        report_name = 'quality_check_report'
        if specific_table:
            report_name = f'quality_check_report_{self._safe_filename_token(specific_table)}'
        self._create_correct_report(report_name, file_timestamp)
        print(f'\n' + '=' * 100)
        print(f'\x1b[1mПРОВЕРКА ЗАВЕРШЕНА\x1b[0m')
        print(f'   Общее время: {overall_time:.2f} сек')
        print(f'   Скорость: {self.processed_rules / overall_time:.1f} правил/сек' if overall_time > 0 else '')
        print(f'=' * 100)
        results_df = pd.DataFrame(self.results)
        return results_df

    def _run_tables_loop(self, tables_to_process, available_tables, timestamp, total_tables):

        def do_one(item):
            i, (table_name, table_rules) = item
            try:
                if self._parallel_lock:
                    with self._parallel_lock:
                        print(f'\n[ПРОГРЕСС] Таблица {i}/{total_tables}: {table_name}')
                else:
                    print(f'\n[ПРОГРЕСС] Таблица {i}/{total_tables}: {table_name}')
                self._process_table_rules(table_name, table_rules, available_tables, timestamp)
            except Exception as e:
                if self._parallel_lock:
                    with self._parallel_lock:
                        print(f'   \x1b[91m[ERROR]\x1b[0m Ошибка при обработке таблицы {table_name}: {str(e)}')
                else:
                    print(f'   \x1b[91m[ERROR]\x1b[0m Ошибка при обработке таблицы {table_name}: {str(e)}')
                traceback.print_exc()
        if self.parallel_tables and len(tables_to_process) > 1:
            enumerated = list(enumerate(tables_to_process, 1))
            with ThreadPoolExecutor(max_workers=self.parallel_tables) as executor:
                list(executor.map(do_one, enumerated))
        else:
            for i, (table_name, table_rules) in enumerate(tables_to_process, 1):
                do_one((i, (table_name, table_rules)))

    def _process_table_rules(self, table_name, table_rules, available_tables, timestamp):
        if table_name == 'AUSP' and table_rules:
            by_table = {}
            for r in table_rules:
                t = r.get('table_name_checked') or r.get('table_name') or ''
                if t not in by_table:
                    by_table[t] = []
                by_table[t].append(r)
            for t in self.AUSP_TABLE_GROUP:
                if t in by_table and by_table[t]:
                    self._process_table_rules(t, by_table[t], available_tables, timestamp)
            return
        self.current_table = table_name
        self.table_start_time = time.time()
        if str(table_name or '').strip().upper() == 'KNB1':
            setattr(self, '_kna1_ktokd_lookup_df', None)

        def _table_available(tname, avail):
            if hasattr(self.memory_manager, 'table_exists') and self.memory_manager.table_exists(tname):
                return True
            if tname in avail:
                return True
            if str(tname or '').strip().upper() == 'ADRC':
                return any((str(t).strip().upper() == 'ADRC' for t in avail))
            return False
        if not _table_available(table_name, available_tables):
            print(f"   \x1b[91m[ERROR]\x1b[0m Таблица '{table_name}' НЕ НАЙДЕНА в БД!")
            suffix = '...' if len(available_tables) > 10 else ''
            print(f'   [DEBUG] Доступные таблицы в БД: {sorted(available_tables)[:10]}{suffix}')
            print(f'   [DEBUG] Всего таблиц в БД: {len(available_tables)}')
            similar_tables = [t for t in available_tables if table_name.replace('/', '_') in t or t.replace('/', '_') == table_name.replace('/', '_')]
            if similar_tables:
                print(f'   [DEBUG] Найдены похожие таблицы: {similar_tables}')
            if self._parallel_lock:
                with self._parallel_lock:
                    for _ in table_rules:
                        self.skipped_rules += 1
                    for rule in table_rules:
                        self._log_skipped_rule(rule, table_name, 'Таблица не найдена в БД', timestamp)
            else:
                for rule in table_rules:
                    self.skipped_rules += 1
                    self._log_skipped_rule(rule, table_name, 'Таблица не найдена в БД', timestamp)
            return
        df_raw = self.memory_manager.get_table(table_name)
        if df_raw is None or df_raw.empty:
            skip_reason = 'Таблица пуста'
            if table_name in self.AUSP_TABLE_GROUP:
                atinn_val = table_name.replace('AUSP_', '')
                skip_reason = f'Таблица пуста (нет строк с ATINN={atinn_val} в AUSP)'
            print(f'   \x1b[93m[WARN]\x1b[0m Таблица {table_name} пуста! Пропускаем...')
            if self._parallel_lock:
                with self._parallel_lock:
                    for rule in table_rules:
                        self.skipped_rules += 1
                    self._log_skipped_rule(rule, table_name, skip_reason, timestamp)
            else:
                for rule in table_rules:
                    self.skipped_rules += 1
                    self._log_skipped_rule(rule, table_name, skip_reason, timestamp)
            return
        df = self._get_table_for_rules(table_name)
        if df is None or df.empty:
            print(f'   \x1b[93m[WARN]\x1b[0m Таблица {table_name} пуста после маппинга колонок. Пропускаем...')
            for rule in table_rules:
                self.skipped_rules += 1
                self._log_skipped_rule(rule, table_name, 'Таблица пуста после маппинга колонок', timestamp)
            return
        if (table_name or '').strip().upper() == 'ADRC':
            name1_col = None
            for c in df.columns:
                if str(c).strip().upper() == 'NAME1':
                    name1_col = c
                    break
            if name1_col is None:
                name1_col = self._find_column_alternative(df.columns, 'NAME1', table_name)
            if name1_col is None:
                name1_col = self._find_most_similar_column(df.columns, 'NAME1')
            if name1_col and name1_col in df.columns:
                before = len(df)
                val_str = df[name1_col].astype(str).str.strip().str.upper()
                df = df[val_str != 'RESERVED'].copy()
                dropped = before - len(df)
                if dropped > 0:
                    print(f'   [ADRC] Исключены строки с NAME1=RESERVED: {dropped:,} (осталось {len(df):,})')
                if df.empty:
                    print(f'   \x1b[93m[WARN]\x1b[0m После фильтра NAME1!=RESERVED таблица ADRC пуста. Пропускаем...')
                    if self._parallel_lock:
                        with self._parallel_lock:
                            for rule in table_rules:
                                self.skipped_rules += 1
                                self._log_skipped_rule(rule, table_name, 'Нет строк после исключения NAME1=RESERVED', timestamp)
                    else:
                        for rule in table_rules:
                            self.skipped_rules += 1
                            self._log_skipped_rule(rule, table_name, 'Нет строк после исключения NAME1=RESERVED', timestamp)
                    return
            else:
                print(f'   [WARN] В ADRC не найдена колонка NAME1, фильтр RESERVED не применён (колонки: {list(df.columns)[:15]})')
        before_empty_filter = len(df)
        empty_mask = fully_empty_rows_mask(df)
        df = df.loc[~empty_mask].copy()
        empty_rows_dropped = before_empty_filter - len(df)
        if empty_rows_dropped > 0:
            print(f'   [{table_name}] Удалено полностью пустых строк: {empty_rows_dropped:,} (осталось {len(df):,})')
        if df.empty:
            print(f'   \x1b[93m[WARN]\x1b[0m После фильтрации пустых строк таблица {table_name} пуста. Пропускаем...')
            if self._parallel_lock:
                with self._parallel_lock:
                    for rule in table_rules:
                        self.skipped_rules += 1
                        self._log_skipped_rule(rule, table_name, 'Таблица пуста после фильтрации', timestamp)
            else:
                for rule in table_rules:
                    self.skipped_rules += 1
                    self._log_skipped_rule(rule, table_name, 'Таблица пуста после фильтрации', timestamp)
            return
        display_row_count = len(df)
        if table_name in self.TABLE_UNIQUE_PARTNER:
            count = getattr(self.memory_manager, 'get_unique_partner_count', lambda t: None)(table_name)
            if count is not None:
                display_row_count = count
            else:
                partner_col = self._find_partner_column(df, table_name=table_name)
                if partner_col and partner_col in df.columns:
                    display_row_count = int(df[partner_col].nunique())
        self._print_table_header(table_name, len(table_rules), display_row_count)
        ausp_split = None
        if (table_name or '').strip().upper() == 'AUSP':
            ausp_split = self._build_ausp_split(df, table_name)
            if ausp_split:
                total_slices = sum((len(s[0]) for s in ausp_split.values()))
                print(f'   [AUSP] Таблица разбита по ATINN на {len(ausp_split)} срезов (всего строк в срезах: {total_slices:,})')
            else:
                atinn_c, atwrt_c = self._find_ausp_columns(df.columns, table_name)
                names = list(df.columns) if hasattr(df.columns, '__iter__') else []
                suffix = '...' if len(names) > 15 else ''
                print(f'   [AUSP] Не удалось разбить по имени: колонки ATINN/ATWRT не найдены. Заголовки ({len(names)}): {names[:15]}{suffix}')
                self._debug_ausp_columns(df.columns, table_name)
        if table_name in self.table_handlers:
            is_taxnum_table = str(table_name or '').strip().upper().startswith('DFKKBPTAXNUM')
            rule_codes_in_table = {str(r.get('rule_code') or '').strip() for r in table_rules or [] if r}
            same_row_here = self.TAXNUM_SAME_ROW_RULES & rule_codes_in_table
            if is_taxnum_table:
                print(f'   [DEBUG] {table_name}: кодов правил в таблице: {len(rule_codes_in_table)}, same_row: {same_row_here}')
            if is_taxnum_table and same_row_here:
                standard_rules = [r for r in table_rules if str(r.get('rule_code') or '').strip() in self.TAXNUM_SAME_ROW_RULES]
                handler_rules = [r for r in table_rules if str(r.get('rule_code') or '').strip() not in self.TAXNUM_SAME_ROW_RULES]
                success_count, error_count, suspicious_count = (0, 0, 0)
                if standard_rules:
                    print(f'   [INFO] Правила «same value as other» (в одной строке): стандартный метод ({len(standard_rules)} правил)')
                    s, e, sus = self._process_with_standard_method(table_name, df, standard_rules, timestamp, ausp_split=None)
                    success_count += s
                    error_count += e
                    suspicious_count += sus
                if handler_rules:
                    print(f'   [INFO] Остальные правила DFKKBPTAXNUM: специальный обработчик ({len(handler_rules)} правил)')
                    s, e, sus = self._process_with_table_handler(table_name, df, handler_rules, timestamp)
                    success_count += s
                    error_count += e
                    suspicious_count += sus
            else:
                print(f'   [INFO] Используем специальный обработчик')
                success_count, error_count, suspicious_count = self._process_with_table_handler(table_name, df, table_rules, timestamp)
        else:
            print(f'   [INFO] Используем стандартный метод проверки')
            success_count, error_count, suspicious_count = self._process_with_standard_method(table_name, df, table_rules, timestamp, ausp_split=ausp_split)
        elapsed_time = time.time() - self.table_start_time
        self._print_table_summary(table_name, elapsed_time, success_count, error_count, suspicious_count)
        return (success_count, error_count, suspicious_count)

    def _process_with_table_handler(self, table_name, df, table_rules, timestamp):
        handler_class = self.table_handlers[table_name]
        success_count = 0
        error_count = 0
        suspicious_count = 0
        try:
            handler = handler_class(table_name, df, self.memory_manager, self)
            total_rules = len(table_rules)
            for i, rule in enumerate(table_rules, 1):
                if self._parallel_lock:
                    with self._parallel_lock:
                        self.processed_rules += 1
                else:
                    self.processed_rules += 1
                self.current_rule = rule.get('rule_code', 'UNKNOWN')
                sys.stdout.write(f'\r    [{i:3d}/{total_rules:3d}] {self.current_rule:20} | ')
                sys.stdout.flush()
                rule_start_time = time.time()
                result = handler.validate_rule(rule)
                execution_time = time.time() - rule_start_time
                if result and isinstance(result, dict):
                    error_df_res = result.get('error_df', pd.DataFrame())
                    error_count_result = int(result.get('error_count', result.get('failed', 0)))
                    failed = int(result.get('failed', error_count_result))
                    passed = int(result.get('passed', 0))
                    if result.get('total_records') is not None:
                        total_rows = int(result['total_records'])
                    elif result.get('total_evaluated') is not None:
                        total_rows = int(result['total_evaluated'])
                    else:
                        total_rows = passed + failed
                    if passed + failed != total_rows and total_rows > 0:
                        passed = max(total_rows - failed, 0)
                    is_suspicious = self._check_if_suspicious(self.current_rule, error_count_result, total_rows)
                    mass_error = error_count_result > self.MAX_ERRORS_TO_SAVE
                    handler_status = str(result.get('status', '')).strip().upper()
                    not_evaluated = total_rows == 0 and error_count_result == 0
                    exec_failed = not_evaluated or handler_status == 'ОШИБКА ВЫПОЛНЕНИЯ'
                    if error_count_result == 0 and total_rows > 0:
                        success_count += 1
                    elif exec_failed:
                        error_count += 1
                    elif is_suspicious or mass_error:
                        suspicious_count += 1
                    else:
                        error_count += 1
                    self._print_rule_stats(self.current_rule, total_rows, error_count_result, execution_time, is_suspicious, mass_error, not_evaluated=exec_failed)
                    result['check_date'] = timestamp
                    result['passed'] = passed
                    result['failed'] = failed
                    result['total_records'] = total_rows
                    result['total_evaluated'] = total_rows
                    result['error_count'] = error_count_result
                    if self._parallel_lock:
                        with self._parallel_lock:
                            self.results.append(result)
                            if error_count_result > 0:
                                key = f'{self.current_rule}_{table_name}'
                                error_df = result.get('error_df', pd.DataFrame())
                                if error_df is not None and (not error_df.empty):
                                    if len(error_df) > error_count_result * 1.1:
                                        error_df = error_df.head(error_count_result)
                                    self._save_rule_error_with_limit(self.current_rule, table_name, error_df, error_count_result, is_suspicious, total_rows)
                                else:
                                    print(f'      [WARN] {self.current_rule} ({table_name}): найдено {error_count_result:,} ошибок, но error_df пустой — детальный файл не будет создан')
                                    self.rule_errors[key] = {'rule_code': self.current_rule, 'table_name': table_name, 'error_df': pd.DataFrame(), 'error_count': error_count_result, 'is_suspicious': is_suspicious, 'total_rows': total_rows}
                    else:
                        self.results.append(result)
                        if error_count_result > 0:
                            key = f'{self.current_rule}_{table_name}'
                            error_df = result.get('error_df', pd.DataFrame())
                            if error_df is not None and (not error_df.empty):
                                if len(error_df) > error_count_result * 1.1:
                                    error_df = error_df.head(error_count_result)
                                self._save_rule_error_with_limit(self.current_rule, table_name, error_df, error_count_result, is_suspicious, total_rows)
                            else:
                                print(f'      [WARN] {self.current_rule} ({table_name}): найдено {error_count_result:,} ошибок, но error_df пустой — детальный файл не будет создан')
                                self.rule_errors[key] = {'rule_code': self.current_rule, 'table_name': table_name, 'error_df': pd.DataFrame(), 'error_count': error_count_result, 'is_suspicious': is_suspicious, 'total_rows': total_rows}
                elif self._parallel_lock:
                    with self._parallel_lock:
                        self.skipped_rules += 1
                        self._log_skipped_rule(rule, table_name, 'Обработчик вернул неверный формат', timestamp)
                else:
                    self.skipped_rules += 1
                    self._log_skipped_rule(rule, table_name, 'Обработчик вернул неверный формат', timestamp)
        except Exception as e:
            print(f'\n   \x1b[91m[ERROR]\x1b[0m Ошибка в обработчике {table_name}: {str(e)}')
            traceback.print_exc()
            if self._parallel_lock:
                with self._parallel_lock:
                    for rule in table_rules:
                        self.skipped_rules += 1
                        self._log_skipped_rule(rule, table_name, f'Ошибка обработчика: {str(e)}', timestamp)
            else:
                for rule in table_rules:
                    self.skipped_rules += 1
                    self._log_skipped_rule(rule, table_name, f'Ошибка обработчика: {str(e)}', timestamp)
        return (success_count, error_count, suspicious_count)

    def _process_with_standard_method(self, table_name, df, table_rules, timestamp, ausp_split=None):
        success_count = 0
        error_count = 0
        suspicious_count = 0
        total_rules = len(table_rules)
        for i, rule in enumerate(table_rules, 1):
            if self._parallel_lock:
                with self._parallel_lock:
                    self.processed_rules += 1
            else:
                self.processed_rules += 1
            self.current_rule = rule.get('rule_code', 'UNKNOWN')
            sys.stdout.write(f'\r    [{i:3d}/{total_rules:3d}] {self.current_rule:20} | ')
            sys.stdout.flush()
            rule_start_time = time.time()
            rule_df = df.copy() if str(table_name or '').strip().upper() == 'KNVV' else df
            error_count_result, total_rows = self._process_single_rule(rule, table_name, rule_df, timestamp, ausp_split=ausp_split)
            execution_time = time.time() - rule_start_time
            is_suspicious = self._check_if_suspicious(self.current_rule, error_count_result, total_rows)
            mass_error = error_count_result > self.MAX_ERRORS_TO_SAVE
            not_evaluated = total_rows == 0 and error_count_result == 0
            if error_count_result == 0 and total_rows > 0:
                success_count += 1
            elif not_evaluated:
                error_count += 1
            elif is_suspicious or mass_error:
                suspicious_count += 1
            else:
                error_count += 1
            self._print_rule_stats(self.current_rule, total_rows, error_count_result, execution_time, is_suspicious, mass_error, not_evaluated=not_evaluated)
        return (success_count, error_count, suspicious_count)

    def _process_single_rule_without_save(self, rule, table_name, df, timestamp):
        return self._process_single_rule(rule, table_name, df, timestamp, save_result=False)

    def _process_single_rule(self, rule, table_name, df, timestamp, save_result=True, ausp_split=None):
        import re
        self._last_rule_error = None
        self._last_rule_skip_reason = None
        self._current_save_result = save_result
        rule_code_raw = str(rule.get('rule_code', 'UNKNOWN'))
        rule_code = re.sub('[^A-Za-z0-9._-]', '', rule_code_raw).upper()
        rule_description = rule.get('rule_description', 'Unknown rule')
        rule_desc_lower = str(rule_description).lower()
        quality_category = rule.get('quality_category', 'Unknown')
        column_to_check = rule.get('column_name_checked', '')
        value_checked = rule.get('value_checked', '')
        tn = str(table_name or '').strip().upper()
        if rule_code in self.RCCOMP_149_RULES and tn == 'KNVP':
            validator = self._get_validator_for_rule(rule_description, quality_category, {'rule_code': rule_code, 'rule_description': rule_description, 'quality_category': quality_category, 'table_name': table_name, 'matched_column': column_to_check, 'original_column': column_to_check})
            return self._process_rcccomp_149_knvp(rule_code, df, table_name, rule, validator, rule.get('column_name_checked', 'PARVW'), column_to_check, save_result, timestamp)
        if tn.startswith('DFKKBPTAXNUM'):
            handler_cls = self.table_handlers.get(table_name) or self.table_handlers.get(tn)
            if rule_code in self.TAXNUM_FORMAT_RULES and handler_cls is not None:
                try:
                    h = handler_cls(table_name, df, self.memory_manager, self)
                    result = h.validate_rule(rule)
                    if result and isinstance(result, dict):
                        err = int(result.get('failed', result.get('error_count', 0)))
                        tot = int(result.get('total_records', 0))
                        if save_result:
                            result['check_date'] = timestamp
                            result['table_name'] = table_name
                            self.results.append(result)
                            if err > 0 and result.get('error_df') is not None:
                                self._save_rule_error_with_limit(rule_code, table_name, result['error_df'], err, self._check_if_suspicious(rule_code, err, tot), tot)
                        return (err, tot)
                except Exception as e:
                    print(f'      [WARN] {rule_code} TaxNumHandler: {e}')
        matched_column = None
        actual_column_to_check = column_to_check
        if (table_name or '').strip().upper() == 'AUSP' and ausp_split:
            atinn_value = self._resolve_ausp_atinn_value(rule)
            if atinn_value and atinn_value in ausp_split:
                df_slice, col_name = ausp_split[atinn_value]
                df = df_slice
                matched_column = col_name
                actual_column_to_check = col_name
                if df.empty:
                    self._log_skipped_rule(rule, table_name, f'Нет строк с ATINN={atinn_value} ({column_to_check})', timestamp)
                    return (0, 0)
                print(f"      [AUSP] Правило → срез ATINN={atinn_value}, колонка '{col_name}', строк: {len(df):,}")
        if not matched_column:
            actual_column_to_check = None
            table_name_norm = str(table_name or '').strip().upper()
            if table_name_norm == 'KNB5' and column_to_check == 'MAHNA':
                actual_column_to_check = column_to_check
                print(f"      [MAP] KNB5: используем column_name_checked '{column_to_check}' напрямую (альтернативы через _find_column_alternative)")
            elif table_name_norm == 'KNB1' and column_to_check in ('AKONT', 'FDGRV', 'ZTERM'):
                actual_column_to_check = column_to_check
                print(f"      [MAP] KNB1: используем column_name_checked '{column_to_check}' (AKONT/FDGRV/ZTERM), альтернативы через _find_column_alternative)")
            elif value_checked and self.column_map:
                table_mapping = None
                tn = str(table_name or '').strip().upper()
                if table_name in self.column_map:
                    table_mapping = self.column_map[table_name]
                else:
                    for k, v in self.column_map.items():
                        if str(k).strip().upper() == tn and isinstance(v, dict):
                            table_mapping = v
                            break
                if table_mapping:
                    logical_keys = sorted((k for k in table_mapping.keys() if not str(k).startswith('_')), key=lambda x: len(str(x)), reverse=True)
                    for logical_name in logical_keys:
                        pattern = f'(?:^|[\\.\\s\\+\\-_])({re.escape(logical_name)})(?:[\\s\\+\\-_\\.]|$)'
                        if re.search(pattern, value_checked, re.IGNORECASE):
                            actual_column_to_check = table_mapping[logical_name]
                            print(f"      [MAP] Найдено логическое имя в value_checked: '{logical_name}' -> '{actual_column_to_check}'")
                            break
            if not actual_column_to_check:
                actual_column_to_check = self._get_mapped_column_name(table_name, column_to_check)
            actual_column_to_check = self._fix_column_name_for_taxnum(actual_column_to_check, table_name)
            print(f"      [COL] Ищем колонку: '{column_to_check}' (value_checked: '{value_checked}') -> '{actual_column_to_check}'")
            matched_column = None
            ausp_filtered, ausp_atwrt_col, ausp_temporary_name = self._apply_ausp_filter(df, column_to_check, table_name, rule=rule)
            if ausp_filtered is not None and ausp_atwrt_col is not None and ausp_temporary_name:
                if ausp_filtered.empty:
                    atinn_val = None
                    if self.ausp_atinn_mapping:
                        for k, v in self.ausp_atinn_mapping.items():
                            if (k or '').strip().upper() == (column_to_check or '').strip().upper():
                                atinn_val = v
                                break
                    if atinn_val is None and column_to_check:
                        m = re.search('ATINN\\s*[=\\(]\\s*(\\d+)', column_to_check or '', re.IGNORECASE)
                        if m:
                            atinn_val = m.group(1)
                    self._log_skipped_rule(rule, table_name, f'Нет строк с ATINN={atinn_val} ({column_to_check})', timestamp)
                    return (0, 0)
                df = ausp_filtered.copy()
                if ausp_atwrt_col in df.columns and ausp_temporary_name != ausp_atwrt_col:
                    df = df.rename(columns={ausp_atwrt_col: ausp_temporary_name})
                matched_column = ausp_temporary_name
                print(f"      [AUSP] ATINN отфильтрован, колонка ATWRT временно переименована в '{ausp_temporary_name}', строк: {len(df):,}")
            if (table_name or '').strip().upper() == 'AUSP' and (not matched_column):
                atinn_col, atwrt_col = self._find_ausp_columns(df.columns, table_name)
                if not atinn_col or not atwrt_col:
                    self._log_skipped_rule(rule, table_name, 'В таблице AUSP не найдены колонки ATINN или ATWRT', timestamp)
                    return (0, 0)
                self._log_skipped_rule(rule, table_name, f"Для правила не определено значение ATINN (column_name_checked='{column_to_check}'). Добавьте запись в conf_ausp_atinn_mapping.json", timestamp)
                return (0, 0)
            if not matched_column:
                matched_column = self._resolve_column_for_rule(df, actual_column_to_check, table_name)
            if not matched_column and actual_column_to_check != column_to_check:
                matched_column = self._resolve_column_for_rule(df, column_to_check, table_name)
            if not matched_column:
                matched_column = self._find_most_similar_column(df.columns, actual_column_to_check)
                if not matched_column:
                    self._log_skipped_rule(rule, table_name, f"Колонка '{actual_column_to_check}' не найдена", timestamp)
                    return (0, 0)
        rule_info = {'table_name': table_name, 'rule_code': rule_code, 'rule_description': rule_description, 'quality_category': quality_category, 'matched_column': matched_column, 'original_column': column_to_check, 'actual_column_searched': actual_column_to_check}
        validator = self._get_validator_for_rule(rule_description, quality_category, rule_info)
        is_recon_1131 = self._normalize_rule_code(rule_code) == 'RCCONF_113.1' or ('recon' in rule_desc_lower and 'account group' in rule_desc_lower)
        if rule_code == 'RCCONF_24.1' and (not isinstance(validator, ConformityValidator)):
            print(f'      [WARN] RCCONF_24.1: выбран {validator.__class__.__name__}, принудительно ConformityValidator')
            validator = ConformityValidator(rule_info)
        if rule_code == 'RCCONF_119.2' and (not isinstance(validator, PaymentTermsConsistencyValidator)):
            print(f'      [WARN] RCCONF_119.2: выбран {validator.__class__.__name__}, принудительно переключаем на PaymentTermsConsistencyValidator')
            validator = PaymentTermsConsistencyValidator(rule_info)
        if is_recon_1131 and (not isinstance(validator, ReconAccountConsistencyValidator)):
            print(f'      [WARN] RCCONF_113.1: выбран {validator.__class__.__name__}, принудительно переключаем на ReconAccountConsistencyValidator')
            validator = ReconAccountConsistencyValidator(rule_info)
        if rule_code == 'RCCONF_119.2':
            print(f'      [DEBUG] RCCONF_119.2: validator={validator.__class__.__name__}')
        if is_recon_1131:
            print(f'      [DEBUG] RCCONF_113.1: validator={validator.__class__.__name__}')
        params = {}
        need_second_column = False
        if rule_code == 'RCCONF_119.2' or is_recon_1131:
            need_second_column = False
        elif isinstance(validator, CrossColumnEqualityValidator) or isinstance(validator, CrossColumnEqualityCheckValidator):
            need_second_column = True
        if rule_code == 'RCCONF_119.2':
            print(f'      [DEBUG] RCCONF_119.2: need_second_column={need_second_column}')
        if is_recon_1131:
            print(f'      [DEBUG] RCCONF_113.1: need_second_column={need_second_column}')
        if need_second_column:
            technical_def = rule.get('technical_definition_RU', '')
            if isinstance(technical_def, list):
                technical_def = ' '.join((str(x) for x in technical_def))
            value_checked = rule.get('value_checked', '')
            second_column_candidate = None
            if table_name == 'BUT000' and rule_code in ('RCCONF_15.2.1', 'RCCONF_15.2.2', 'RCCONF_15.2.3', 'RCCONF_14.1.1', 'RCCONF_14.1.2', 'RCCONF_13.2'):
                second_by_rule = {'RCCONF_15.2.1': 'NAME_ORG1', 'RCCONF_15.2.2': 'NAME_ORG2', 'RCCONF_15.2.3': 'NAME_ORG3', 'RCCONF_14.1.1': 'NAME_ORG1', 'RCCONF_14.1.2': 'NAME_ORG2', 'RCCONF_13.2': 'NAME_ORG1'}
                wanted = second_by_rule.get(rule_code)
                if wanted:
                    found = self._resolve_column_for_rule(df, wanted, table_name) or (wanted if wanted in df.columns else None)
                    if not found:
                        for c in df.columns:
                            if c.upper() == wanted.upper():
                                found = c
                                break
                    if found:
                        params['second_column'] = found
                        print(f"      [COL] Вторая колонка по коду правила {rule_code}: '{found}'")
            if not params.get('second_column') and str(table_name or '').strip().upper().startswith('DFKKBPTAXNUM'):
                is_tax_same_as_other = rule_code in self.TAXNUM_SAME_ROW_RULES or (technical_def and ('tax_0_value OR' in technical_def or '= tax_0_value OR' in technical_def or '= tax_1_value OR' in technical_def))
                if is_tax_same_as_other and matched_column:
                    tax_like = [c for c in df.columns if c != matched_column and (re.match('^tax_\\d+_value$', str(c), re.I) or re.match('^tax_\\d+$', str(c), re.I) or (str(c).upper().startswith('TAXNUM') and str(c).upper() != matched_column.upper()))]
                    if not tax_like:
                        col_upper = {str(c).strip().upper(): c for c in df.columns}
                        for i in range(0, 7):
                            for cand in (f'tax_{i}_value', f'tax_{i}', f'TAXNUM{i}', f'TAXNUM_{i}'):
                                if cand.upper() in col_upper:
                                    c = col_upper[cand.upper()]
                                    if c != matched_column and c not in tax_like:
                                        tax_like.append(c)
                                    break
                            else:
                                for c in df.columns:
                                    cu = str(c).upper().replace(' ', '').replace('_', '')
                                    if (f'TAX{i}VALUE' in cu or f'TAXNUM{i}' in cu) and c != matched_column and (c not in tax_like):
                                        tax_like.append(c)
                                        break
                    if tax_like:
                        params['other_columns'] = tax_like
                        print(f"      [COL] TAXNUM «same as other»: проверяем '{matched_column}' против {len(tax_like)} колонок: {[str(c) for c in tax_like[:5]]}{('...' if len(tax_like) > 5 else '')}")
            first_num = None
            if value_checked:
                first_match = re.search('organization_(\\d+)_name', value_checked)
                if first_match:
                    first_num = first_match.group(1)
            if not params.get('second_column') and technical_def:
                name_org_pattern = 'NAME_ORG(\\d+)'
                name_org_matches = re.findall(name_org_pattern, technical_def, re.IGNORECASE)
                if name_org_matches:
                    name_org_comparison = re.search('NAME_ORG(\\d+)\\s*=\\s*NAME_ORG(\\d+)', technical_def, re.IGNORECASE)
                    if name_org_comparison:
                        col1_num = name_org_comparison.group(1)
                        col2_num = name_org_comparison.group(2)
                        matched_num = None
                        if matched_column and 'ORG' in matched_column.upper():
                            matched_num_match = re.search('ORG(\\d+)', matched_column.upper())
                            if matched_num_match:
                                matched_num = matched_num_match.group(1)
                        if matched_num == col1_num:
                            second_col_name = f'NAME_ORG{col2_num}'
                        elif matched_num == col2_num:
                            second_col_name = f'NAME_ORG{col1_num}'
                        elif matched_column and matched_column.upper() == f'NAME_ORG{col1_num}':
                            second_col_name = f'NAME_ORG{col2_num}'
                        else:
                            second_col_name = f'NAME_ORG{col1_num}'
                        if second_col_name in df.columns:
                            params['second_column'] = second_col_name
                            print(f"      [COL] Вторая колонка из technical_definition (прямое имя): '{second_col_name}'")
                    if not params.get('second_column'):
                        matched_num = None
                        if matched_column and 'ORG' in matched_column.upper():
                            matched_num_match = re.search('ORG(\\d+)', matched_column.upper())
                            if matched_num_match:
                                matched_num = matched_num_match.group(1)
                        for num in name_org_matches:
                            if matched_num and num != matched_num:
                                candidate = f'NAME_ORG{num}'
                                if candidate in df.columns:
                                    params['second_column'] = candidate
                                    print(f"      [COL] Вторая колонка из technical_definition (прямое имя): '{candidate}'")
                                    break
                if not params.get('second_column'):
                    pattern = 'organization_(\\d+)_name'
                    matches = re.findall(pattern, technical_def)
                    if matches:
                        comparison_pattern = 'organization_(\\d+)_name\\s*=\\s*organization_(\\d+)_name'
                        comparison_match = re.search(comparison_pattern, technical_def)
                        if comparison_match:
                            col1_num = comparison_match.group(1)
                            col2_num = comparison_match.group(2)
                            if col1_num == first_num:
                                second_num = col2_num
                            elif col2_num == first_num:
                                second_num = col1_num
                            else:
                                second_num = col2_num
                            second_logical = f'organization_{second_num}_name'
                            second_column_candidate = self._get_mapped_column_name(table_name, second_logical)
                            resolved = self._resolve_column_for_rule(df, second_column_candidate, table_name)
                            if not resolved and second_column_candidate:
                                for c in df.columns:
                                    if c.upper() == second_column_candidate.upper():
                                        resolved = c
                                        break
                            if resolved:
                                params['second_column'] = resolved
                                print(f"      [COL] Вторая колонка из сравнения в technical_definition: '{second_logical}' -> '{resolved}'")
                        if not params.get('second_column'):
                            for match_num in matches:
                                if match_num != first_num:
                                    second_logical = f'organization_{match_num}_name'
                                    second_column_candidate = self._get_mapped_column_name(table_name, second_logical)
                                    resolved = self._resolve_column_for_rule(df, second_column_candidate, table_name)
                                    if not resolved and second_column_candidate:
                                        for c in df.columns:
                                            if c.upper() == second_column_candidate.upper():
                                                resolved = c
                                                break
                                    if resolved:
                                        params['second_column'] = resolved
                                        print(f"      [COL] Вторая колонка из technical_definition: '{second_logical}' -> '{resolved}'")
                                        break
                    if not params.get('second_column'):
                        desc_lower = rule_description.lower()
                        second_logical = None
                        if 'cannot be the same as name 2' in desc_lower or 'cannot be the same as name2' in desc_lower:
                            second_logical = 'organization_2_name'
                        elif 'cannot be the same as name 3' in desc_lower or 'cannot be the same as name3' in desc_lower:
                            second_logical = 'organization_3_name'
                        elif 'cannot be the same as name 4' in desc_lower or 'cannot be the same as name4' in desc_lower:
                            second_logical = 'organization_4_name'
                        elif 'cannot be the same as name 1' in desc_lower or 'cannot be the same as name1' in desc_lower:
                            second_logical = 'organization_1_name'
                        elif 'equals name 2' in desc_lower or 'equals name2' in desc_lower:
                            second_logical = 'organization_2_name'
                        elif 'equals name 3' in desc_lower or 'equals name3' in desc_lower:
                            second_logical = 'organization_3_name'
                        elif 'equals name 4' in desc_lower or 'equals name4' in desc_lower:
                            second_logical = 'organization_4_name'
                        elif 'equals name 1' in desc_lower or 'equals name1' in desc_lower:
                            second_logical = 'organization_1_name'
                        elif ('name 2' in desc_lower or 'name2' in desc_lower) and first_num != '2':
                            second_logical = 'organization_2_name'
                        elif ('name 3' in desc_lower or 'name3' in desc_lower) and first_num != '3':
                            second_logical = 'organization_3_name'
                        elif ('name 4' in desc_lower or 'name4' in desc_lower) and first_num != '4':
                            second_logical = 'organization_4_name'
                        elif ('name 1' in desc_lower or 'name1' in desc_lower) and first_num != '1':
                            second_logical = 'organization_1_name'
                        if second_logical:
                            second_column_candidate = self._get_mapped_column_name(table_name, second_logical)
                            resolved = self._resolve_column_for_rule(df, second_column_candidate, table_name)
                            if not resolved and second_column_candidate:
                                for c in df.columns:
                                    if c.upper() == second_column_candidate.upper():
                                        resolved = c
                                        break
                            if resolved:
                                params['second_column'] = resolved
                                print(f"      [COL] Вторая колонка из описания правила: '{second_logical}' -> '{resolved}'")
            if not params.get('second_column'):
                second_column = self._extract_second_column_from_description(rule_code, rule_description, df.columns, matched_column, table_name)
                if second_column and second_column in df.columns:
                    params['second_column'] = second_column
                    print(f"      [COL] Вторая колонка из описания: '{second_column}'")
            if not params.get('second_column') and (not params.get('other_columns')):
                self._log_failed_rule(rule, table_name, f"Не найдена вторая колонка для сравнения в правиле '{rule_description}'", timestamp)
                return (0, 0)
        df_to_validate = df
        if str(table_name or '').strip().upper() == 'ADRC':
            mapped_adrc = self._get_table_for_rules('ADRC')
            if mapped_adrc is not None and (not mapped_adrc.empty):
                df_to_validate = mapped_adrc
                df = mapped_adrc
        taxnum_baseline_total = None
        if str(table_name or '').strip().upper().startswith('DFKKBPTAXNUM'):
            typ = self._get_dfkkbptaxnum_type_from_table(table_name)
            if typ is not None:
                df_to_validate, taxnum_baseline_total = self._scope_dfkkbptaxnum_by_taxtype(df_to_validate, table_name)
                if df_to_validate.empty:
                    self._log_skipped_rule(rule, table_name, f'Нет строк с Tax_Number_Category=RU{typ}', timestamp)
                    return (0, 0)
        if str(table_name or '').strip().upper() == 'ADRC':
            name1_col = None
            for c in df_to_validate.columns:
                if str(c).strip().upper() == 'NAME1':
                    name1_col = c
                    break
            if name1_col is None:
                name1_col = self._find_column_alternative(df_to_validate.columns, 'NAME1', table_name)
            if name1_col and name1_col in df_to_validate.columns:
                reserved_mask = df_to_validate[name1_col].astype(str).str.strip().str.upper() == 'RESERVED'
                if reserved_mask.any():
                    before_adrc = len(df_to_validate)
                    df_to_validate = df_to_validate[~reserved_mask].copy()
                    print(f'      [FILTER] ADRC: скип строк с NAME1=RESERVED (как пустые) — {reserved_mask.sum():,} исключено, к оценке {len(df_to_validate):,} из {before_adrc:,}')
                if df_to_validate.empty:
                    self._log_skipped_rule(rule, table_name, 'Нет данных после исключения NAME1=RESERVED', timestamp)
                    return (0, 0)
        technical_def = rule.get('technical_definition_RU', '')
        if isinstance(technical_def, list):
            technical_def = ' '.join((str(x) for x in technical_def))
        if technical_def and ('contact_medium_type' in technical_def or 'source' in technical_def or rule_code in ['RCCONF_38.3', 'RCCONF_38.5', 'RCCONF_39.3', 'RCCONF_39.3.2', 'RCCONF_39.5', 'RCCONF_39.5.2', 'RCCOMP_369.1', 'RCCONF_369.1', 'RCCOMP_375.1', 'RCCOMP_375.1.2']):
            df_to_validate = self._apply_conditional_filter(df_to_validate, technical_def, rule_code, table_name)
            if df_to_validate.empty:
                self._log_skipped_rule(rule, table_name, 'Нет данных, соответствующих условиям правила', timestamp)
                return (0, 0)
        if str(table_name or '').strip().upper() == 'ADR2':
            before_adr2_scope = len(df_to_validate)
            df_to_validate = self._filter_adr2_dm_customer_scope(df_to_validate, rule_code, table_name)
            df = df_to_validate
            if before_adr2_scope > 0 and df_to_validate.empty:
                self._log_skipped_rule(rule, table_name, f'Нет строк ADR2 в scope (KNA1 Group={self.ADR2_DM_CUSTOMER_SCOPE_ACCOUNT_GROUP}, KNVV 01-01)', timestamp)
                return (0, 0)
        if rule_code == 'RCCOMP_180.1' and str(table_name or '').strip().upper() == 'BUT0BK':
            before_cnt = len(df_to_validate)
            df_to_validate = self._scope_but0bk_to_kna1_partners(df_to_validate, table_name, rule_code)
            print(f'      [FILTER] {rule_code}: scope партнёры из KNA1 -> {len(df_to_validate):,} из {before_cnt:,}')
            if df_to_validate.empty:
                self._log_skipped_rule(rule, table_name, 'Нет строк BUT0BK для партнёров из KNA1', timestamp)
                return (0, 0)
            if matched_column in df_to_validate.columns:
                banks_norm = df_to_validate[matched_column].astype(str).str.strip().str.upper()
                ru_mask = banks_norm == 'RU'
                ru_cnt = int(ru_mask.sum())
                before_ru = len(df_to_validate)
                df_to_validate = df_to_validate[ru_mask].copy()
                print(f"      [FILTER] {rule_code}: учитываем только BANKS='RU' -> {ru_cnt:,} из {before_ru:,}")
                if df_to_validate.empty:
                    print(f"      [FILTER] {rule_code}: нет строк с BANKS='RU' -> ошибок = 0")
                    return (0, 0)
            params['allowed_values'] = ['RU']
        needs_account_group_code = False
        if technical_def and table_name != 'KNA1':
            technical_def_lower = technical_def.lower()
            if 'account_group_code' in technical_def_lower or 'accountgroupcode' in technical_def_lower.replace(' ', '').replace('_', '') or 'ktokd' in technical_def_lower:
                needs_account_group_code = True
        if needs_account_group_code:
            if rule_code == 'RCCONF_24.1' and self._is_adrc_table(table_name):
                print(f'      [JOIN] [{self.CHECKER_BUILD_ID}] RCCONF_24.1 — принудительный путь ADRC→BUT020→KNA1 (CLIENT запрещён)')
                df_to_validate = self._join_kna1_ktokd_rconf_24_1_adrc(df_to_validate, table_name, rule_code)
            else:
                df_to_validate = self._add_account_group_code_from_kna1(df_to_validate, table_name, rule_code)
            if rule_code == 'RCCONF_24.1' and self._is_adrc_table(table_name):
                ag_col = self._find_account_group_column(df_to_validate)
                if not ag_col:
                    self._log_skipped_rule(rule, table_name, 'RCCONF_24.1: account_group_code (KTOKD) не получен после JOIN ADRC->BUT020->KNA1', timestamp)
                    return (0, 0)
                from utils.sap_account_keys import norm_sap_account_group
                if int((df_to_validate[ag_col].apply(norm_sap_account_group) != '').sum()) == 0:
                    self._log_skipped_rule(rule, table_name, 'RCCONF_24.1: KTOKD пустой у всех строк после JOIN ADRC->BUT020->KNA1', timestamp)
                    return (0, 0)
            if str(rule_code).strip().upper() in self.RULES_KTOKD_ONLY_9038_SCOPE:
                before_scope = len(df_to_validate)
                df_to_validate = self._filter_rows_only_ktokd_9038(df_to_validate, rule_code)
                skipped_non_9038 = before_scope - len(df_to_validate)
                print(f"      [FILTER] {rule_code}: только account_group_code='9038' -> {len(df_to_validate):,} из {before_scope:,} (пропущено не 9038: {skipped_non_9038:,})")
                if df_to_validate.empty:
                    st = getattr(self, '_last_kna1_join_stats', {}) or {}
                    self._log_skipped_rule(rule, table_name, f"Нет строк KNB1 с KTOKD=9038 из KNA1. После JOIN: {st.get('rows_after_join', before_scope):,} строк; с заполненным KTOKD: {st.get('filled_ktokd', '?')}; с KTOKD=9038: {st.get('n9038', 0):,}. Проверьте JOIN KNB1.Customer=KNA1.Customer и поле Group_1.", timestamp)
                    return (0, 0)
                df_to_validate = self._attach_kna1_ktokd_export_columns(df_to_validate, rule_code)
            elif self._normalize_rule_code(rule_code) in self.RULES_ERROR_EXPORT_KNA1_KTOKD:
                df_to_validate = self._attach_kna1_ktokd_export_columns(df_to_validate, rule_code)
            if is_recon_1131:
                has_account_group = any((str(c).strip().lower() in ('account_group_code', 'b.account_group_code', 'ktokd') for c in df_to_validate.columns))
                if not has_account_group:
                    print('      [SKIP] RCCONF_113.1: account_group_code не доступен после JOIN (memory/SQLite), правило пропущено')
                    self._log_skipped_rule(rule, table_name, 'RCCONF_113.1 skipped: account_group_code (KTOKD) not available after KNA1 join', timestamp)
                    return (0, 0)
        if str(table_name or '').strip().upper() == 'KNVV':
            scope_source = df_to_validate
            if rule_code == 'RCCONF_153.4':
                full_knvv = self._get_table_for_rules('KNVV')
                if full_knvv is not None and (not full_knvv.empty):
                    scope_source = full_knvv
            before_knvv_scope = len(scope_source)
            df_to_validate = self._apply_knvv_dm_sales_org_scope(scope_source, rule_code, table_name)
            if before_knvv_scope > 0 and df_to_validate.empty:
                if rule_code == 'RCCONF_153.4':
                    skip_msg = 'В выгрузке KNVV нет VTWEG=04 и SPART=02 (дамп только 01-01) — RCCONF_153.4 вне scope'
                else:
                    skip_msg = 'Нет строк KNVV в scope dm_customer_sales_org (VTWEG=01, SPART=01; exclude order_block S,NH,S3,S4,SY,U,R,PR; exclude KDGRP=ZIN)'
                self._log_skipped_rule(rule, table_name, skip_msg, timestamp)
                return (0, 0)
        if rule_code == 'RCCONF_371.2' and str(table_name or '').strip().upper() == 'BUT051':
            df_to_validate = self._add_central_order_block_code_from_kna1(df_to_validate, table_name, rule_code)
            block_col = next((c for c in df_to_validate.columns if str(c).strip().lower() == 'central_order_block_code'), None)
            if not block_col:
                self._log_skipped_rule(rule, table_name, 'Для RCCONF_371.2 не удалось получить central_order_block_code (KNA1.AUFSD)', timestamp)
                return (0, 0)
            pafkt_norm = df_to_validate[matched_column].astype(str).str.strip()
            block_norm = df_to_validate[block_col].astype(str).str.strip().str.upper()
            eval_mask = (pafkt_norm != '') & (block_norm == 'M')
            before_count = len(df_to_validate)
            df_to_validate = df_to_validate[eval_mask].copy()
            print(f"      [FILTER] RCCONF_371.2 scope: PAFKT not null AND central_order_block_code='M' -> {len(df_to_validate):,} из {before_count:,}")
            if df_to_validate.empty:
                self._log_skipped_rule(rule, table_name, "Нет строк для оценки RCCONF_371.2 после фильтра PAFKT + central_order_block_code='M'", timestamp)
                return (0, 0)
            params['allowed_values'] = ['0001', '0002', '0003', '0004', '0005', '0006', '0007', '0008', '0009', '0010', 'DF', 'F1', 'F2', 'F3', 'F4', 'F5', 'FX', 'Z1', 'Z2', 'Z3', 'Z4', 'Z5', 'Z6', 'Z7']
        if rule_code in ['RCCONF_39.5', 'RCCONF_39.5.2'] and matched_column and (matched_column in df_to_validate.columns):
            before_filter_count = len(df_to_validate)
            col_series = df_to_validate[matched_column]
            null_or_empty = col_series.isna() | (col_series.astype(str).str.strip() == '') | col_series.astype(str).str.strip().str.lower().isin(['none', 'null', 'nan', 'na'])
            after_filter_count = (~null_or_empty).sum()
            dropped = before_filter_count - after_filter_count
            if self.debug and before_filter_count > 0:
                null_count = col_series.isna().sum()
                empty_str = (col_series.astype(str).str.strip() == '').sum()
                null_like = col_series.astype(str).str.strip().str.lower().isin(['null', 'nan', 'none', '']).sum()
                print(f'      [DEBUG] {rule_code}: до фильтра — всего строк={before_filter_count}, isna()={null_count}, empty_str={empty_str}, null_like={null_like}')
                if null_or_empty.any():
                    sample_null = col_series[null_or_empty].head(3).tolist()
                    print(f'      [DEBUG] {rule_code}: примеры отброшенных (NULL/пустых): {sample_null}')
            df_to_validate = df_to_validate[~null_or_empty].copy()
            print(f'      [FILTER] {rule_code}: только непустые {matched_column} — {len(df_to_validate):,} из {before_filter_count:,} (исключено {dropped:,})')
            if self.debug and len(df_to_validate) > 0:
                sample_filled = df_to_validate[matched_column].head(5).tolist()
                print(f'      [DEBUG] {rule_code}: в валидатор передано строк: {len(df_to_validate):,}, первые 5 TEL_NUMBER: {sample_filled}')
            if df_to_validate.empty:
                self._log_skipped_rule(rule, table_name, 'Нет записей с заполненным TEL_NUMBER для проверки формата', timestamp)
                return (0, 0)
        try:
            if rule_code == 'RCCOMP_180.1' and str(table_name or '').strip().upper() == 'BUT0BK':
                total_rows = len(df_to_validate)
                result = {'rule_code': rule_code, 'rule_description': rule_description, 'quality_category': quality_category, 'table_name': table_name, 'column_checked': column_to_check, 'matched_column': matched_column, 'total_records': total_rows, 'passed': total_rows, 'failed': 0, 'success_rate_%': 100.0 if total_rows > 0 else 0, 'execution_time_sec': 0, 'check_date': timestamp, 'status': 'УСПЕШНО', 'status_color': 'green', 'error_file': 'Нет', 'comments': ''}
                self.results.append(result)
                return (0, total_rows)
            if rule_code == 'RCCONF_143.7':
                ref_table_name = self._get_reference_table_for_rule(rule_code, 'RCCONF_143.7_reference_table') or 'TVBVK'
                print(f"      [REF] Правило {rule_code}: данные в KNVV, справочник допустимых пар VKGRP+VKBUR: '{ref_table_name}' (или conf_sales_group_office.json)")

                def _norm_code(v):
                    s = self._norm_lookup_value(v).upper()
                    if not s:
                        return ''
                    if re.fullmatch('\\d+', s):
                        return s.lstrip('0').zfill(len(s)) if len(s) > 1 else (s.lstrip('0') or '0')
                    return s

                def _norm_num(v):
                    s = self._norm_lookup_value(v).upper()
                    if not s:
                        return ''
                    if re.fullmatch('\\d+', s):
                        return s.lstrip('0') or '0'
                    return s

                allowed_pairs, ref_source = self._load_allowed_vkgrp_vkbur_pairs(ref_table_name, _norm_num)
                if not allowed_pairs:
                    self._log_skipped_rule(rule, table_name, f'Справочник {ref_table_name} не найден или не дал пар VKGRP+VKBUR (и allowed_combinations в conf не задан).', timestamp)
                    return (0, 0)
                print(f"      [REF] Загружено {len(allowed_pairs)} пар VKGRP+VKBUR из '{ref_source}'")

                vkgrp_col_knvv = self._resolve_column_for_rule(df_to_validate, 'VKGRP', 'KNVV') or matched_column
                vkbur_col_knvv = self._resolve_column_for_rule(df_to_validate, 'VKBUR', 'KNVV')
                if not vkgrp_col_knvv or vkgrp_col_knvv not in df_to_validate.columns:
                    self._log_skipped_rule(rule, table_name, 'Колонка VKGRP не найдена в KNVV', timestamp)
                    return (0, 0)
                if not vkbur_col_knvv or vkbur_col_knvv not in df_to_validate.columns:
                    self._log_skipped_rule(rule, table_name, 'Колонка VKBUR не найдена в KNVV', timestamp)
                    return (0, 0)

                eval_mask = df_to_validate[vkgrp_col_knvv].notna() & (df_to_validate[vkgrp_col_knvv].astype(str).str.strip() != '')
                if not eval_mask.any():
                    self._log_skipped_rule(rule, table_name, 'Нет строк для оценки RCCONF_143.7 после фильтров/NULL VKGRP', timestamp)
                    return (0, 0)

                vkgrp_norm = df_to_validate.loc[eval_mask, vkgrp_col_knvv].apply(_norm_num)
                vkbur_norm = df_to_validate.loc[eval_mask, vkbur_col_knvv].apply(_norm_num)
                pairs = list(zip(vkgrp_norm.tolist(), vkbur_norm.tolist()))
                bad = [i for i, p in enumerate(pairs) if (not p[0]) or (not p[1]) or (p not in allowed_pairs)]
                error_mask = pd.Series(False, index=df_to_validate.index)
                if bad:
                    idx_err = df_to_validate.loc[eval_mask].iloc[bad].index
                    error_mask.loc[idx_err] = True
                error_count = int(error_mask.sum())
                total_rows = int(eval_mask.sum())
                error_description = (
                    f"Consistency Between Sales Group and Sales Office: pair (VKGRP, VKBUR) must exist in {ref_source} "
                    f"(is_sales_group_office_allowed='1'). Scope: VTWEG=01, SPART=01; exclude AUFSD in (S,NH,S3,S4,SY,U,R,PR); exclude KDGRP=ZIN."
                )
                error_df = validator._prepare_error_dataframe(df_to_validate, error_mask, 'CONFORMITY', error_description) if error_count > 0 else None
                is_suspicious = self._check_if_suspicious(rule_code, error_count, total_rows)
                if save_result:
                    rule_info = {'rule_code': rule_code, 'rule_description': rule.get('rule_description', 'Unknown rule'), 'quality_category': rule.get('quality_category', 'Unknown'), 'table_name': table_name, 'original_column': column_to_check, 'matched_column': matched_column}
                    total_rows, error_count = self._apply_unique_partner_counts_if_needed(table_name, df_to_validate, error_df, total_rows, error_count)
                    if self._parallel_lock:
                        with self._parallel_lock:
                            self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                    else:
                        self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                        self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                elif self._parallel_lock:
                    with self._parallel_lock:
                        self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                else:
                    self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                return (error_count, total_rows)
            if rule_code == 'RCCONF_153.4' and str(table_name or '').strip().upper() == 'KNVV':
                vsbed_col = matched_column if matched_column in df_to_validate.columns else self._resolve_column_for_rule(df_to_validate, 'VSBED', 'KNVV')
                if not vsbed_col or vsbed_col not in df_to_validate.columns:
                    self._log_skipped_rule(rule, table_name, 'RCCONF_153.4: колонка VSBED не найдена', timestamp)
                    return (0, 0)
                s = df_to_validate[vsbed_col].astype(str).str.strip()
                eval_mask = df_to_validate[vsbed_col].notna() & (s != '') & ~s.str.lower().isin(['none', 'null', 'nan', 'na'])
                if not eval_mask.any():
                    self._log_skipped_rule(rule, table_name, 'RCCONF_153.4: нет заполненных VSBED в scope VTWEG=04/SPART=02', timestamp)
                    return (0, 0)
                vsbed_norm = s.str.upper()
                error_mask = eval_mask & (vsbed_norm != '05')
                error_count = int(error_mask.sum())
                total_rows = int(eval_mask.sum())
                error_description = "VSBED must be '05' when VTWEG='04' and SPART='02'."
                error_df = validator._prepare_error_dataframe(df_to_validate, error_mask, 'CONFORMITY', error_description) if error_count > 0 else None
                is_suspicious = self._check_if_suspicious(rule_code, error_count, total_rows)
                if save_result:
                    rule_info = {'rule_code': rule_code, 'rule_description': rule.get('rule_description', 'Unknown rule'), 'quality_category': rule.get('quality_category', 'Unknown'), 'table_name': table_name, 'original_column': column_to_check, 'matched_column': vsbed_col}
                    if self._parallel_lock:
                        with self._parallel_lock:
                            self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                    else:
                        self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                        self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                elif error_df is not None and (not error_df.empty):
                    if self._parallel_lock:
                        with self._parallel_lock:
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                    else:
                        self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                return (error_count, total_rows)
            if rule_code == 'RCCONF_372.1':
                print(f'      [REF] {rule_code}: ADR2.COUNTRY (ISO 3166) vs справочник T005 (SAP LAND1 = 2-симв. код, в выгрузке часто C/R)...')
                t005_df = self._get_table_for_rules('T005')
                if t005_df is None or t005_df.empty:
                    print(f'      [WARN] Справочник T005 не найден или пуст для правила {rule_code}')
                    self._log_skipped_rule(rule, table_name, 'RCCONF_372.1: справочник T005 не найден (нужен для ADR2.COUNTRY)', timestamp)
                    return (0, 0)
                land1_col = None
                for ref_candidate in ('LAND1', 'C/R', 'ISO_Code'):
                    land1_col = self._resolve_column_for_rule(t005_df, ref_candidate, 'T005')
                    if land1_col:
                        break
                if land1_col:
                    print(f'      [REF] T005 — колонка кодов стран (LAND1): {land1_col}')
                else:
                    print(f'      [WARN] В T005 не найдена колонка кодов стран (LAND1 / C/R / ISO_Code). Колонки: {list(t005_df.columns)[:12]}...')
                    self._log_skipped_rule(rule, table_name, 'RCCONF_372.1: в справочнике T005 нет колонки кодов стран (LAND1, C/R, ISO_Code) для сравнения с ADR2.COUNTRY', timestamp)
                    return (0, 0)
                valid_countries = set()
                for val in t005_df[land1_col].dropna():
                    val_str = str(val).strip().upper()
                    if val_str:
                        valid_countries.add(val_str)
                print(f'      [REF] Загружено {len(valid_countries)} валидных кодов стран из T005')
                country_col = matched_column
                if country_col not in df_to_validate.columns:
                    for col in df_to_validate.columns:
                        col_lower = col.lower()
                        if col_lower in ['country', 'country_code', 'co', 'cntry'] or 'country' in col_lower:
                            country_col = col
                            print(f'      [REF] Найдена альтернативная колонка COUNTRY: {col}')
                            break
                if country_col not in df_to_validate.columns:
                    print(f'      [WARN] Колонка COUNTRY не найдена в данных для правила {rule_code}')
                    print(f'      [WARN] Искали: {matched_column}, доступные колонки: {list(df_to_validate.columns)[:10]}...')
                    self._log_skipped_rule(rule, table_name, f'Колонка COUNTRY не найдена', timestamp)
                    return (0, 0)
                null_mask = df_to_validate[country_col].isna() | (df_to_validate[country_col].astype(str).str.strip() == '')
                non_null_mask = ~null_mask
                if non_null_mask.any():
                    country_values_upper = df_to_validate.loc[non_null_mask, country_col].astype(str).str.strip().str.upper()
                    error_mask = non_null_mask & ~country_values_upper.isin(valid_countries)
                else:
                    error_mask = pd.Series([False] * len(df_to_validate), index=df_to_validate.index)
                error_count = error_mask.sum()
                total_rows = int(non_null_mask.sum())
                error_description = f'Country code in {country_col} not found in reference table T005 (ISO 3166 standard). Valid codes must exist in T005.LAND1.'
                if error_count > 0:
                    error_df = validator._prepare_error_dataframe(df_to_validate, error_mask, 'CONFORMITY', error_description)
                else:
                    error_df = None
                is_suspicious = self._check_if_suspicious(rule_code, error_count, total_rows)
                if save_result:
                    rule_info = {'rule_code': rule_code, 'rule_description': rule.get('rule_description', 'Unknown rule'), 'quality_category': rule.get('quality_category', 'Unknown'), 'table_name': table_name, 'original_column': column_to_check, 'matched_column': matched_column}
                    total_rows, error_count = self._apply_unique_partner_counts_if_needed(table_name, df_to_validate, error_df, total_rows, error_count)
                    if self._parallel_lock:
                        with self._parallel_lock:
                            if error_df is not None and (not error_df.empty):
                                self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                            self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                    else:
                        if error_df is not None and (not error_df.empty):
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                        self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                elif error_df is not None and (not error_df.empty):
                    if self._parallel_lock:
                        with self._parallel_lock:
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                    else:
                        self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                return (error_count, total_rows)
            if rule_code in ('RCCONF_383.1', 'RCCONF_384.1'):
                coord_col = matched_column if matched_column in df_to_validate.columns else None
                if coord_col is None:
                    is_long = rule_code == 'RCCONF_383.1'
                    candidates = ['/LOT/GC_LONGITUD', '_LOT_GC_LONGITUD', 'LONGITUDE', 'longitude'] if is_long else ['/LOT/GC_LATITUDE', '_LOT_GC_LATITUDE', 'LATITUDE', 'latitude']
                    for cand in candidates:
                        c = self._find_column_alternative(df_to_validate.columns, cand, table_name)
                        if c and c in df_to_validate.columns:
                            coord_col = c
                            break
                if coord_col is None:
                    self._log_skipped_rule(rule, table_name, f'Для {rule_code} не найдена колонка координаты', timestamp)
                    return (0, 0)
                account_group_col = next((c for c in df_to_validate.columns if str(c).strip().lower() in ('account_group_code', 'b.account_group_code', 'ktokd', 'b.ktokd', 'kna.ktokd', 'kna.ktokd')), None)
                account_group_skip = pd.Series(False, index=df_to_validate.index)
                if account_group_col is not None:
                    account_group_skip = df_to_validate[account_group_col].astype(str).str.strip().str.startswith('7')
                s = df_to_validate[coord_col].astype(str).str.strip()
                null_like = df_to_validate[coord_col].isna() | (s == '') | s.str.lower().isin(['none', 'null', 'nan', 'na'])
                integer_part = s.str.replace(',', '.', regex=False).str.extract('^\\s*([+-]?\\d+)', expand=False).fillna('')
                integer_zero = integer_part.str.lstrip('+-').str.lstrip('0').eq('')
                zero_skip = integer_zero & ~null_like
                skip_mask = null_like | zero_skip | account_group_skip
                evaluated_mask = ~skip_mask
                total_rows = int(evaluated_mask.sum())
                print(f'      [DEBUG] {rule_code}: total={len(df_to_validate):,}, null_like={int(null_like.sum()):,}, to_integer_zero={int(zero_skip.sum()):,}, account_group_7xx={int(account_group_skip.sum()):,}, evaluated={total_rows:,}')
                if total_rows == 0:
                    return (0, 0)
                fmt_ok = s.str.match('^-?\\d{1,3}\\.\\d{6}$', na=False)
                error_mask = evaluated_mask & ~fmt_ok
                error_count = int(error_mask.sum())
                print(f'      [DEBUG] {rule_code}: coord_col={coord_col}, evaluated={total_rows:,}, errors={error_count:,}')
                error_description = f'Invalid coordinate format in {coord_col}. Expected (-)x.xxxxxx / (-)xx.xxxxxx / (-)xxx.xxxxxx with dot as decimal separator.'
                error_df = validator._prepare_error_dataframe(df_to_validate, error_mask, 'CONFORMITY', error_description) if error_count > 0 else None
                is_suspicious = self._check_if_suspicious(rule_code, error_count, total_rows)
                if save_result:
                    rule_info = {'rule_code': rule_code, 'rule_description': rule.get('rule_description', 'Unknown rule'), 'quality_category': rule.get('quality_category', 'Unknown'), 'table_name': table_name, 'original_column': column_to_check, 'matched_column': coord_col}
                    total_rows, error_count = self._apply_unique_partner_counts_if_needed(table_name, df_to_validate, error_df, total_rows, error_count)
                    if self._parallel_lock:
                        with self._parallel_lock:
                            if error_df is not None and (not error_df.empty):
                                self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                            self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                    else:
                        if error_df is not None and (not error_df.empty):
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                        self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                elif error_df is not None and (not error_df.empty):
                    if self._parallel_lock:
                        with self._parallel_lock:
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                    else:
                        self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                return (error_count, total_rows)
            if rule_code == 'RCCONF_372.2':
                tel_col = None
                for c in df_to_validate.columns:
                    if str(c).strip().upper() == 'TEL_NUMBER':
                        tel_col = c
                        break
                if tel_col is None:
                    tel_col = self._find_column_alternative(df_to_validate.columns, 'TEL_NUMBER', table_name)
                country_col = None
                for c in df_to_validate.columns:
                    if str(c).strip().upper() in ('COUNTRY', 'COUNTRY_CODE'):
                        country_col = c
                        break
                if country_col is None:
                    country_col = matched_column if matched_column in df_to_validate.columns else None
                if country_col is None:
                    for c in df_to_validate.columns:
                        if 'COUNTRY' in str(c).strip().upper():
                            country_col = c
                            break
                if not tel_col or tel_col not in df_to_validate.columns:
                    self._log_skipped_rule(rule, table_name, 'Для RCCONF_372.2 не найдена колонка TEL_NUMBER', timestamp)
                    return (0, 0)
                if not country_col or country_col not in df_to_validate.columns:
                    self._log_skipped_rule(rule, table_name, 'Для RCCONF_372.2 не найдена колонка COUNTRY', timestamp)
                    return (0, 0)
                tel_s = df_to_validate[tel_col].astype(str).str.strip()
                tel_filled = df_to_validate[tel_col].notna() & (tel_s != '') & ~tel_s.str.lower().isin(['none', 'null', 'nan', 'na'])
                if not tel_filled.any():
                    return (0, 0)
                country_s = df_to_validate[country_col].astype(str).str.strip()
                country_missing = df_to_validate[country_col].isna() | (country_s == '') | country_s.str.lower().isin(['none', 'null', 'nan', 'na'])
                error_mask = tel_filled & country_missing
                error_count = int(error_mask.sum())
                total_rows = int(tel_filled.sum())
                error_description = f'{country_col} is required when {tel_col} has a value (ADR2 scope).'
                error_df = validator._prepare_error_dataframe(df_to_validate, error_mask, 'CONFORMITY', error_description) if error_count > 0 else None
                is_suspicious = self._check_if_suspicious(rule_code, error_count, total_rows)
                if save_result:
                    rule_info = {'rule_code': rule_code, 'rule_description': rule.get('rule_description', 'Unknown rule'), 'quality_category': rule.get('quality_category', 'Unknown'), 'table_name': table_name, 'original_column': column_to_check, 'matched_column': country_col}
                    total_rows, error_count = self._apply_unique_partner_counts_if_needed(table_name, df_to_validate, error_df, total_rows, error_count)
                    if self._parallel_lock:
                        with self._parallel_lock:
                            if error_df is not None and (not error_df.empty):
                                self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                            self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                    else:
                        if error_df is not None and (not error_df.empty):
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                        self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                elif error_df is not None and (not error_df.empty):
                    if self._parallel_lock:
                        with self._parallel_lock:
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                    else:
                        self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                return (error_count, total_rows)
            if rule_code == 'RCCONF_170.7':
                print(f'      [FILTER] Обработка правила {rule_code}: пропускаем записи с пустым Customer Group 4...')
                customer_group_4_col = None
                for col in df_to_validate.columns:
                    col_lower = col.lower()
                    if col_lower == 'kvgr4' or col_lower == 'customer_group_4_code' or 'customer_group_4' in col_lower:
                        customer_group_4_col = col
                        print(f'      [FILTER] Найдена колонка Customer Group 4: {col}')
                        break
                if customer_group_4_col:
                    before_filter = len(df_to_validate)
                    non_null_mask = df_to_validate[customer_group_4_col].notna() & (df_to_validate[customer_group_4_col].astype(str).str.strip() != '') & (df_to_validate[customer_group_4_col].astype(str).str.strip() != '-1') & (df_to_validate[customer_group_4_col].astype(str).str.strip().str.lower() != 'null')
                    df_to_validate = df_to_validate[non_null_mask].copy()
                    filtered_count = before_filter - len(df_to_validate)
                    if filtered_count > 0:
                        print(f'      [FILTER] Пропущено записей с пустым Customer Group 4: {filtered_count} (осталось {len(df_to_validate)} для проверки)')
                    if df_to_validate.empty:
                        print(f'      [INFO] После фильтрации все записи пропущены (Customer Group 4 пустое)')
                        self._log_skipped_rule(rule, table_name, 'Все записи имеют пустое Customer Group 4', timestamp)
                        return (0, 0)
                else:
                    print(f'      [WARN] Колонка Customer Group 4 (KVGR4) не найдена для правила {rule_code}')
            if rule_code in ['RCCONF_38.3', 'RCCONF_38.5', 'RCCONF_39.3', 'RCCONF_39.3.2', 'RCCONF_39.5', 'RCCONF_39.5.2']:
                params['technical_definition'] = technical_def
                params['rule_code'] = rule_code
            elif rule_code in ['RCCONF_18.2', 'RCCONF_22.2']:
                params['rule_code'] = rule_code
                params['technical_definition'] = technical_def
            else:
                params['rule_code'] = rule_code
            if table_name == 'BUT000' and isinstance(validator, (CrossColumnEqualityValidator, CrossColumnEqualityCheckValidator)):
                second_col = params.get('second_column', 'НЕ НАЙДЕНА')
                print(f"      [DEBUG] Правило {rule_code}: сравниваем '{matched_column}' с '{second_col}'")
                if second_col in df_to_validate.columns:
                    sample_vals1 = df_to_validate[matched_column].head(3).tolist() if matched_column in df_to_validate.columns else []
                    sample_vals2 = df_to_validate[second_col].head(3).tolist()
                    print(f'      [DEBUG] Примеры значений: {matched_column}={sample_vals1}, {second_col}={sample_vals2}')
            if self.debug and rule_code in ['RCCONF_39.5', 'RCCONF_39.5.2']:
                print(f'      [DEBUG] Вызов validator.validate() для {rule_code}: строк={len(df_to_validate):,}, колонка={matched_column}')
            if rule_code == 'RCCONF_119.2' and isinstance(validator, PaymentTermsConsistencyValidator):
                knb1_kunnr_col = self._find_kunnr_column(df_to_validate) or next((c for c in df_to_validate.columns if str(c).strip().upper() == 'KUNNR'), None)
                knvv_df = self.memory_manager.get_table('KNVV')
                knvv_kunnr_col = None
                knvv_zterm_col = None
                if knvv_df is not None and (not knvv_df.empty):
                    knvv_kunnr_col = next((c for c in knvv_df.columns if str(c).strip().upper() == 'KUNNR'), None) or self._find_kunnr_column(knvv_df)
                    knvv_zterm_col = next((c for c in knvv_df.columns if str(c).strip().upper() == 'ZTERM'), None) or self._find_column_alternative(knvv_df.columns, 'ZTERM', 'KNVV')
                params.update({'knb1_kunnr_col': knb1_kunnr_col, 'db_path': getattr(self, 'db_path', None), 'knvv_df': knvv_df, 'knvv_kunnr_col': knvv_kunnr_col, 'knvv_zterm_col': knvv_zterm_col})
            if is_recon_1131 and isinstance(validator, ReconAccountConsistencyValidator):
                account_group_col = None
                for c in df_to_validate.columns:
                    cu = str(c).strip().lower()
                    if cu in ('account_group_code', 'b.account_group_code', 'ktokd', 'b.ktokd'):
                        account_group_col = c
                        break
                if not account_group_col:
                    account_group_col = self._find_column_alternative(df_to_validate.columns, 'account_group_code', table_name)
                recon_ref_path = os.path.join(parent_dir, 'json files', 'conf_recon_accounts.json')
                if not os.path.isfile(recon_ref_path):
                    self._log_skipped_rule(rule, table_name, f'RCCONF_113.1: не найден {recon_ref_path}', timestamp)
                    return (0, 0)
                params.update({'account_group_col': account_group_col, 'reference_path': recon_ref_path})
                print(f'      [DEBUG] RCCONF_113.1: recon_col={matched_column}, account_group_col={account_group_col}, ref={recon_ref_path}')
            if rule_code == 'RCCONF_115.11':
                matrix_rules, matrix_path = self._load_planning_group_matrix()
                if not matrix_rules:
                    self._log_skipped_rule(rule, table_name, 'conf_planning_group_matrix.json не найден или пуст', timestamp)
                    return (0, 0)
                account_group_col = next((c for c in df_to_validate.columns if str(c).strip() in ('kna.KTOKD', 'kna.ktokd')), None) or next((c for c in df_to_validate.columns if str(c).strip().lower() in ('account_group_code', 'b.account_group_code', 'ktokd', 'b.ktokd')), None) or self._find_column_alternative(df_to_validate.columns, 'account_group_code', table_name)
                zterm_col = next((c for c in df_to_validate.columns if str(c).strip().upper() == 'ZTERM'), None) or self._find_column_alternative(df_to_validate.columns, 'ZTERM', table_name)
                if not account_group_col:
                    self._log_skipped_rule(rule, table_name, 'Для RCCONF_115.11 не найден account_group_code (KTOKD из KNA1 после JOIN)', timestamp)
                    return (0, 0)
                if not zterm_col:
                    self._log_skipped_rule(rule, table_name, 'Для RCCONF_115.11 не найдена колонка ZTERM', timestamp)
                    return (0, 0)
                print(f'      [DEBUG] RCCONF_115.11: fdgrv_col={matched_column}, account_group_col={account_group_col} (KNA1.KTOKD), zterm_col={zterm_col}, matrix={matrix_path}')
                print("      [DEBUG] RCCONF_115.11 logic: compare ZTERM + FDGRV only; '*' in ZTERM is wildcard")
                term_to_planning = {}
                for item in matrix_rules:
                    if not isinstance(item, dict):
                        continue
                    t = self._norm_lookup_value(item.get('terms_of_payment_code')).upper()
                    p = self._norm_lookup_value(item.get('planning_group_code')).upper()
                    if not p:
                        continue
                    if not t:
                        t = '*'
                    term_to_planning.setdefault(t, set()).add(p)
                if not term_to_planning:
                    self._log_skipped_rule(rule, table_name, 'RCCONF_115.11: conf_planning_group_matrix не содержит валидных строк rules[]', timestamp)
                    return (0, 0)
                fdgrv_norm = df_to_validate[matched_column].apply(self._norm_lookup_value)
                account_group_norm = df_to_validate[account_group_col].apply(self._norm_lookup_value)
                zterm_norm = df_to_validate[zterm_col].apply(self._norm_lookup_value)
                skip_mask = (fdgrv_norm == '') | account_group_norm.str.startswith('7') | account_group_norm.isin({'9096', '9022', '9023', '9095'})
                evaluated_mask = ~skip_mask
                total_rows = int(evaluated_mask.sum())
                if total_rows == 0:
                    self._log_skipped_rule(rule, table_name, 'Нет записей для оценки RCCONF_115.11 после skip-условий', timestamp)
                    return (0, 0)
                fdgrv_upper = fdgrv_norm.str.upper()
                zterm_upper = zterm_norm.str.upper()
                allowed_series = pd.Series([set()] * len(df_to_validate), index=df_to_validate.index, dtype=object)
                for idx in df_to_validate.index[evaluated_mask]:
                    t = zterm_upper.loc[idx] if zterm_upper.loc[idx] else '*'
                    allowed = set()
                    allowed |= term_to_planning.get(t, set())
                    allowed |= term_to_planning.get('*', set())
                    allowed_series.loc[idx] = allowed
                error_mask = evaluated_mask & ~df_to_validate.index.to_series().apply(lambda i: fdgrv_upper.loc[i] in (allowed_series.loc[i] or set()))
                error_count = int(error_mask.sum())
                error_description = "Invalid value in column FDGRV: planning_group_code not found in conf_planning_group_matrix for terms_of_payment_code (ZTERM) with wildcard '*'. Skip if FDGRV empty or KNA1 account_group_code (KTOKD) starts with 7 or IN (9096,9022,9023,9095)."
                error_df = validator._prepare_error_dataframe(df_to_validate, error_mask, 'CONFORMITY', error_description) if error_count > 0 else None
                if error_df is not None and (not error_df.empty):
                    error_df['FDGRV_ACTUAL'] = fdgrv_norm.loc[error_df.index].values
                    error_df['LOOKUP_ACCOUNT_GROUP_KTOKD'] = account_group_norm.loc[error_df.index].values
                    error_df['LOOKUP_TERMS_OF_PAYMENT'] = zterm_norm.loc[error_df.index].values
                    error_df['ALLOWED_PLANNING_GROUPS'] = error_df.index.to_series().apply(lambda i: ', '.join(sorted(list(allowed_series.loc[i] or set())))[:2000]).values
                    error_df['DQ_RULE_CHECK_COLUMNS'] = f"KNB1 planning_group_code [{matched_column}] (FDGRV) сверяется с conf_planning_group_matrix по terms_of_payment_code [{zterm_col}] (ZTERM) с wildcard '*'; фильтр по KNA1 account_group_code [{account_group_col}] (не «7…», не 9096/9022/9023/9095); пустой FDGRV — не оценивается"
                    error_df['DQ_ERROR_DESCRIPTION'] = error_df.apply(lambda row: f"FDGRV='{row['FDGRV_ACTUAL']}' не найден в матрице для ZTERM='{row['LOOKUP_TERMS_OF_PAYMENT']}' (допустимо: {row['ALLOWED_PLANNING_GROUPS']}). KTOKD='{row['LOOKUP_ACCOUNT_GROUP_KTOKD']}'", axis=1)
                    empty_fdgrv_mask = error_df['FDGRV_ACTUAL'].apply(lambda v: self._norm_lookup_value(v) == '')
                    if empty_fdgrv_mask.any():
                        before_cnt = len(error_df)
                        error_df = error_df[~empty_fdgrv_mask].copy()
                        error_count = len(error_df)
                        print(f'      [FILTER] RCCONF_115.11: исключены строки с пустым FDGRV из error_df: {before_cnt} -> {error_count}')
                    if error_df.empty:
                        error_df = None
                        error_count = 0
                is_suspicious = self._check_if_suspicious(rule_code, error_count, total_rows)
                if save_result:
                    rule_info = {'rule_code': rule_code, 'rule_description': rule.get('rule_description', 'Unknown rule'), 'quality_category': rule.get('quality_category', 'Unknown'), 'table_name': table_name, 'original_column': column_to_check, 'matched_column': matched_column}
                    total_rows, error_count = self._apply_unique_partner_counts_if_needed(table_name, df_to_validate, error_df, total_rows, error_count)
                    if self._parallel_lock:
                        with self._parallel_lock:
                            self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                            if error_df is not None and (not error_df.empty):
                                self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                    else:
                        self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                        if error_df is not None and (not error_df.empty):
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                elif error_df is not None and (not error_df.empty):
                    if self._parallel_lock:
                        with self._parallel_lock:
                            self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                    else:
                        self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
                return (error_count, total_rows)
            if rule_code == 'RCCONF_15.1' and str(table_name or '').strip().upper() == 'BUT000':
                org3_res = self._find_column_alternative(df_to_validate.columns, 'NAME_ORG3', table_name)
                if org3_res:
                    params['org3_column_resolved'] = org3_res
            if rule_code == 'RCCONF_12.4' and str(table_name or '').strip().upper() == 'BUT000':
                country_col = self._resolve_column_for_rule(df_to_validate, 'C/R', table_name)
                if not country_col:
                    for c in df_to_validate.columns:
                        if str(c).strip().upper() in ('C/R', 'CR', 'COUNTRY', 'LAND1', 'COUNTRY_CODE'):
                            country_col = c
                            break
                if country_col and country_col in df_to_validate.columns:
                    params['country_column'] = country_col
                    params['excluded_countries'] = ['AM', 'BY', 'CZ', 'SK']
                    print(f"      [FILTER] RCCONF_12.4: исключены страны AM/BY/CZ/SK по колонке '{country_col}'")
            if rule_code == 'RCCONF_22.3' and str(table_name or '').strip().upper() == 'ADRC':
                country_col = self._resolve_column_for_rule(df_to_validate, 'COUNTRY', table_name)
                if not country_col:
                    country_col = self._resolve_column_for_rule(df_to_validate, 'C/R', table_name)
                if not country_col:
                    for c in df_to_validate.columns:
                        if str(c).strip().upper() in ('C/R', 'CR', 'COUNTRY', 'LAND1', 'COUNTRY_CODE'):
                            country_col = c
                            break
                if country_col and country_col in df_to_validate.columns:
                    params['country_column'] = country_col
                    params['excluded_countries'] = ['AM', 'BY', 'CZ', 'SK']
                    print(f"      [FILTER] RCCONF_22.3: исключены страны AM/BY/CZ/SK по колонке '{country_col}'")
            if rule_code == 'RCCONF_21.1' and str(table_name or '').strip().upper() == 'ADRC':
                country_col = self._resolve_column_for_rule(df_to_validate, 'COUNTRY', table_name)
                if not country_col:
                    for cand in ('C/R', 'C_R', 'PO_C/R', 'LAND1'):
                        country_col = self._resolve_column_for_rule(df_to_validate, cand, table_name)
                        if country_col:
                            break
                if not country_col:
                    for c in df_to_validate.columns:
                        if str(c).strip().upper() in ('C/R', 'C_R', 'PO_C/R', 'COUNTRY', 'LAND1'):
                            country_col = c
                            break
                if country_col and country_col in df_to_validate.columns:
                    params['country_column'] = country_col
                t005_df = self._get_table_for_rules('T005')
                t005_by_country = {}
                if t005_df is not None and (not t005_df.empty):
                    land_col = next((c for c in t005_df.columns if str(c).strip().upper() in ('LAND1', 'C/R', 'COUNTRY')), None)
                    if land_col:
                        for _, row in t005_df.iterrows():
                            cc = str(row.get(land_col, '')).strip().upper()
                            if cc:
                                t005_by_country[cc] = row.to_dict()
                params['t005_by_country'] = t005_by_country
                print(f'      [REF] RCCONF_21.1: T005 country rules loaded: {len(t005_by_country)}')
            if rule_code == 'RCCOMP_375.1.2':
                adr6_df = self._get_adr6_df()
                if adr6_df is not None and (not adr6_df.empty):
                    params['adr6_df'] = adr6_df
                    print(f'      [ADR6] RCCOMP_375.1.2: подключена таблица ADR6 ({len(adr6_df):,} строк) для проверки e-mail')
                else:
                    print('      [WARN] RCCOMP_375.1.2: таблица ADR6 не найдена — проверяется только TEL_NUMBER')
            total_rows, error_count, error_df = validator.validate(df_to_validate, matched_column, **params)
            if taxnum_baseline_total is not None:
                total_rows = taxnum_baseline_total
            if is_recon_1131 and total_rows == 0 and (error_count == 0):
                st = getattr(self, '_last_kna1_join_stats', {}) or {}
                ag_col = params.get('account_group_col') or self._find_account_group_column(df_to_validate)
                skip_reason = f'RCCONF_113.1: нет строк для оценки (нужны заполненные AKONT и KTOKD из KNA1). Строк KNB1 в срезе: {len(df_to_validate):,}'
                if ag_col and ag_col in df_to_validate.columns:
                    from utils.sap_account_keys import norm_sap_account_group, norm_sap_recon_account
                    has_k = df_to_validate[ag_col].apply(norm_sap_account_group) != ''
                    has_a = df_to_validate[matched_column].apply(norm_sap_recon_account) != ''
                    skip_reason += f'; с KTOKD: {int(has_k.sum()):,}; с AKONT: {int(has_a.sum()):,}; с обоими: {int((has_k & has_a).sum()):,}'
                self._log_skipped_rule(rule, table_name, skip_reason, timestamp)
                return (0, 0)
            if rule_code == 'RCCONF_38.3' and error_df is not None and (not error_df.empty):
                r3_user_col = None
                for col in error_df.columns:
                    col_lower = col.lower()
                    if col_lower == 'r3_user' or col_lower == 'r3user' or 'r3_user' in col_lower or ('r3user' in col_lower):
                        r3_user_col = col
                        break
                if r3_user_col:
                    before_count = len(error_df)
                    error_df = error_df[error_df[r3_user_col].astype(str).str.strip() == '1'].copy()
                    after_count = len(error_df)
                    if before_count != after_count:
                        print(f"      [FILTER] Дополнительная фильтрация error_df: {before_count} → {after_count} записей (только R3_USER='1')")
                        error_count = after_count
                else:
                    print(f'      [WARN] Колонка R3_USER не найдена в error_df для правила {rule_code}')
                    for col in df.columns:
                        col_lower = col.lower()
                        if col_lower == 'r3_user' or col_lower == 'r3user' or 'r3_user' in col_lower or ('r3user' in col_lower):
                            if error_df.index.isin(df.index).any():
                                original_mask = df[col].astype(str).str.strip() == '1'
                                error_df = error_df[error_df.index.isin(df[original_mask].index)].copy()
                                error_count = len(error_df)
                                print(f'      [FILTER] Отфильтровано error_df по индексам исходного DataFrame: {len(error_df)} записей')
                            break
            if error_df is not None and (not error_df.empty):
                if len(error_df) > total_rows:
                    print(f'      [ERROR] Валидатор вернул error_df с {len(error_df)} строками, но total_rows={total_rows}')
                    error_df = pd.DataFrame()
                    error_count = 0
                elif len(error_df) > error_count * 1.1:
                    print(f'      [WARN] error_df содержит {len(error_df)} строк, но error_count={error_count}')
                    error_df = error_df.head(error_count)
            total_rows, error_count = self._apply_unique_partner_counts_if_needed(table_name, df_to_validate, error_df, total_rows, error_count)
            is_suspicious = self._check_if_suspicious(rule_code, error_count, total_rows)
            filtered_adr2_addr_partner_df = None
            if self._parallel_lock:
                with self._parallel_lock:
                    if error_df is not None and (not error_df.empty):
                        self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows)
                    if save_result:
                        self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious, filtered_adr2_addr_partner_df=filtered_adr2_addr_partner_df)
            else:
                if error_df is not None and (not error_df.empty):
                    self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows)
                if save_result:
                    self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious, filtered_adr2_addr_partner_df=filtered_adr2_addr_partner_df)
            return (error_count, total_rows)
        except Exception as e:
            self._last_rule_error = str(e)
            if save_result:
                self._log_failed_rule(rule, table_name, str(e), timestamp)
            return (0, 0)

    def _save_rule_error_with_limit(self, rule_code, table_name, error_df, error_count, is_suspicious, total_rows):
        if error_df is None or error_df.empty:
            return
        if self._normalize_rule_code(rule_code) == 'RCCOMP_113.1' and 'AKONT' not in error_df.columns:
            src_col = None
            if 'DQ_COLUMN_CHECKED' in error_df.columns:
                try:
                    cn = str(error_df['DQ_COLUMN_CHECKED'].iloc[0]).strip()
                    if cn and cn in error_df.columns:
                        src_col = cn
                except Exception:
                    src_col = None
            if src_col is None:
                src_col = self._find_column_alternative(error_df.columns, 'AKONT', table_name)
            if src_col and src_col in error_df.columns:
                error_df = error_df.copy()
                error_df['AKONT'] = error_df[src_col]
            else:
                error_df = error_df.copy()
                error_df['AKONT'] = ''
        if self._normalize_rule_code(rule_code) in self.RULES_ERROR_EXPORT_KNA1_KTOKD:
            error_df = self._enrich_error_df_kna1_ktokd(error_df, table_name, rule_code)
        if str(table_name or '').strip().upper() == 'ADRC':
            name1_col = None
            for c in error_df.columns:
                if str(c).strip().upper() == 'NAME1':
                    name1_col = c
                    break
            if name1_col is None:
                name1_col = self._find_column_alternative(error_df.columns, 'NAME1', table_name)
            if name1_col is None:
                best_col, best_count = (None, 0)
                for c in error_df.columns:
                    try:
                        cnt = (error_df[c].astype(str).str.strip().str.upper() == 'RESERVED').sum()
                        if cnt > best_count:
                            best_count, best_col = (cnt, c)
                    except Exception:
                        pass
                if best_col and best_count > 0:
                    name1_col = best_col
            if name1_col and name1_col in error_df.columns:
                val_str = error_df[name1_col].astype(str).str.strip().str.upper()
                error_df = error_df[val_str != 'RESERVED'].copy()
                error_count = len(error_df)
                if error_df.empty:
                    return
        if str(table_name or '').strip().upper() == 'ADR2':
            error_df = self._filter_adr2_dm_customer_scope(error_df, rule_code, table_name)
            if error_df.empty:
                return
            error_df = self._dedupe_adr2_by_partner(error_df, log_prefix=f'      [{rule_code}] ')
            error_count = len(error_df)
        if rule_code == 'RCCONF_39.5':

            def _norm(s):
                return str(s).strip().upper().replace('_', '').replace(' ', '')
            persnumber_col = next((c for c in error_df.columns if _norm(c) in ('PERSNUMBER', 'PERSONNUMBER') or 'PERSNUMBER' in _norm(c)), None)
            if persnumber_col is not None:
                s = error_df[persnumber_col].astype(str).str.strip().str.upper()
                empty_mask = error_df[persnumber_col].isna() | (s == '') | s.isin(['NONE', 'NAN', 'NULL', '-', '.'])
                error_df = error_df.loc[empty_mask].copy()
                error_count = len(error_df)
                if error_df.empty:
                    return
        if rule_code in ['RCCONF_39.5', 'RCCONF_39.5.2']:
            from utils.ru_tel_format import is_valid_rccconf_39_5_value
            tel_col = None
            if 'DQ_COLUMN_CHECKED' in error_df.columns:
                try:
                    cn = error_df['DQ_COLUMN_CHECKED'].iloc[0]
                    if cn and str(cn).strip() and (str(cn).strip() in error_df.columns):
                        tel_col = str(cn).strip()
                except Exception:
                    pass
            if tel_col is None:
                for c in error_df.columns:
                    if c in ('DQ_ERROR_TYPE', 'DQ_RULE_CODE', 'DQ_COLUMN_CHECKED', 'DQ_ERROR_DESCRIPTION', 'DQ_TIMESTAMP', 'DQ_RULE_DESCRIPTION'):
                        continue
                    cn = str(c).upper().replace('_', '').replace(' ', '')
                    if 'TELNUMBER' in cn or 'TELNR' in cn or ('TEL' in cn and 'NUM' in cn):
                        tel_col = c
                        break
            _fmt_ok = lambda v: is_valid_rccconf_39_5_value(v, rule_code)
            if tel_col is not None:
                drop_mask = error_df[tel_col].apply(_fmt_ok)
            else:
                drop_mask = pd.Series(False, index=error_df.index)
                for c in error_df.columns:
                    if c in ('DQ_ERROR_TYPE', 'DQ_RULE_CODE', 'DQ_COLUMN_CHECKED', 'DQ_ERROR_DESCRIPTION', 'DQ_TIMESTAMP', 'DQ_RULE_DESCRIPTION'):
                        continue
                    drop_mask = drop_mask | error_df[c].apply(_fmt_ok)
            if drop_mask.any():
                before = len(error_df)
                error_df = error_df.loc[~drop_mask].copy()
                error_count = len(error_df)
                if before > error_count:
                    print(f'      [{rule_code}] Убраны из ошибок номера с валидным форматом: {before - error_count} строк, осталось {error_count}')
            if error_df.empty:
                return
        key = f'{rule_code}_{table_name}'
        limit_errors = self._error_save_limit(rule_code, table_name)
        total_error_count = int(error_count)
        if len(error_df) > error_count * 1.1:
            print(f'      [WARN] Для {rule_code} ({table_name}): error_df содержит {len(error_df)} строк, но error_count={error_count}')
            if error_count > 0:
                error_df = error_df.head(error_count)
            else:
                print(f'      [ERROR] error_count=0, но error_df не пустой. Пропускаем сохранение.')
                return
        error_df_to_save = error_df.copy()
        is_truncated = False
        if key in self.rule_errors:
            existing_count = self.rule_errors[key].get('error_count', 0)
            existing_df = self.rule_errors[key]['error_df']
            if existing_count >= limit_errors:
                print(f'      [WARN] {rule_code} ({table_name}): уже накоплено {existing_count:,} ошибок (лимит {limit_errors:,}), новые ошибки не добавляются')
                return
            combined_df = pd.concat([existing_df, error_df_to_save], ignore_index=True)
            total_combined = len(combined_df)
            if rule_code in ['RCCONF_39.5', 'RCCONF_39.5.2']:
                from utils.ru_tel_format import is_valid_rccconf_39_5_value
                tel_col = None
                if 'DQ_COLUMN_CHECKED' in combined_df.columns:
                    try:
                        cn = combined_df['DQ_COLUMN_CHECKED'].iloc[0]
                        if cn and str(cn).strip() in combined_df.columns:
                            tel_col = str(cn).strip()
                    except Exception:
                        pass
                if tel_col is None:
                    tel_col = next((c for c in combined_df.columns if 'TEL' in str(c).upper() and ('NUMBER' in str(c).upper() or 'NR' in str(c).upper() or 'NUM' in str(c).upper())), None)
                if tel_col is not None:
                    drop = combined_df[tel_col].apply(lambda v: is_valid_rccconf_39_5_value(v, rule_code))
                else:
                    drop = pd.Series(False, index=combined_df.index)
                    for c in combined_df.columns:
                        if 'DQ_' in str(c):
                            continue
                        drop = drop | combined_df[c].apply(lambda v: is_valid_rccconf_39_5_value(v, rule_code))
                if drop.any():
                    combined_df = combined_df.loc[~drop].copy()
                    total_combined = len(combined_df)
            if total_combined > limit_errors:
                combined_df = combined_df.head(limit_errors)
                is_truncated = True
                print(f'      [WARN] {rule_code} ({table_name}): ошибок {total_combined:,}, сохранено только {limit_errors:,} (первые {limit_errors:,})')
            saved_rows = len(combined_df)
            is_truncated = is_truncated or saved_rows < total_error_count or self.rule_errors[key].get('is_truncated', False)
            if is_truncated and saved_rows >= limit_errors:
                print(f'      [WARN] {rule_code} ({table_name}): ошибок {total_error_count:,}, сохранено только {saved_rows:,} (первые {limit_errors:,})')
            self.rule_errors[key] = {'rule_code': rule_code, 'table_name': table_name, 'error_df': combined_df, 'error_count': total_error_count, 'total_error_count': total_error_count, 'saved_error_count': saved_rows, 'is_suspicious': is_suspicious, 'total_rows': total_rows, 'is_truncated': is_truncated}
        else:
            if len(error_df_to_save) > limit_errors:
                error_df_to_save = error_df_to_save.head(limit_errors)
                is_truncated = True
                print(f'      [WARN] {rule_code} ({table_name}): ошибок {total_error_count:,}, сохранено только {limit_errors:,} (первые {limit_errors:,})')
            saved_rows = len(error_df_to_save)
            is_truncated = is_truncated or saved_rows < total_error_count
            self.rule_errors[key] = {'rule_code': rule_code, 'table_name': table_name, 'error_df': error_df_to_save, 'error_count': total_error_count, 'total_error_count': total_error_count, 'saved_error_count': saved_rows, 'is_suspicious': is_suspicious, 'total_rows': total_rows, 'is_truncated': is_truncated}

    def _check_if_suspicious(self, rule_code, error_count, total_rows):
        if rule_code in ['RCCONF_12.2', 'RCCONF_12.3']:
            return True
        if error_count > 1000000:
            return True
        if total_rows > 0 and error_count / total_rows > self.MASS_ERROR_THRESHOLD:
            return True
        if rule_code == 'RCCONF_15.1' and total_rows > 0 and (error_count / total_rows > 0.3):
            return True
        return False

    def _save_rule_result(self, rule_info, total_rows, error_count, execution_time, timestamp, is_suspicious, filtered_adr2_addr_partner_df=None):
        passed_count = total_rows - error_count
        total_records_evaluated = passed_count + error_count
        success_rate = passed_count / total_records_evaluated * 100 if total_records_evaluated > 0 else 0
        if total_records_evaluated == 0:
            status = 'ОШИБКА ВЫПОЛНЕНИЯ'
            status_color = 'red'
        elif error_count == 0:
            status = 'УСПЕШНО'
            status_color = 'green'
        elif is_suspicious:
            if error_count > self.MAX_ERRORS_TO_SAVE:
                status = 'МАССОВЫЕ ОШИБКИ'
            else:
                status = 'ПОДОЗРИТЕЛЬНО'
            status_color = 'orange'
        else:
            status = 'ОШИБКИ'
            status_color = 'red'
        rule_code = rule_info['rule_code']
        table_name = rule_info['table_name']
        key = f'{rule_code}_{table_name}'
        stored = self.rule_errors.get(key, {})
        has_errors_saved = stored.get('error_df') is not None and (not stored.get('error_df').empty)
        saved_rows = int(stored.get('saved_error_count') or (len(stored['error_df']) if has_errors_saved else 0))
        is_truncated = bool(stored.get('is_truncated', False))
        if error_count > 0 and is_truncated:
            error_file_status = f'[!] Частично ({saved_rows:,} из {error_count:,})' if has_errors_saved else 'Нет (не сохранено)'
        elif error_count > 0:
            error_file_status = 'Есть' if has_errors_saved else 'Нет (не сохранено)'
        else:
            error_file_status = 'Нет'
        comments = ''
        if total_records_evaluated == 0:
            comments = 'Правило не смогло оценить ни одной строки (total_rows=0, errors=0). Проверьте техническое условие и входные данные.'
        if error_count > 0 and is_truncated:
            comments = f'[!] ВНИМАНИЕ: Всего ошибок {error_count:,}, сохранено только первые {saved_rows:,}! Обратите внимание!'
        elif is_suspicious and total_records_evaluated > 0:
            error_percent = error_count / total_records_evaluated * 100
            comments = f'ПОДОЗРИТЕЛЬНО: {error_percent:.1f}% ДАННЫХ С ОШИБКАМИ - ПРОВЕРИТЬ ЛОГИКУ ПРАВИЛА'
        result = {'rule_code': rule_info['rule_code'], 'rule_description': rule_info['rule_description'], 'quality_category': rule_info['quality_category'], 'table_name': rule_info['table_name'], 'column_checked': rule_info.get('original_column', ''), 'matched_column': rule_info['matched_column'], 'total_records': total_records_evaluated, 'passed': passed_count, 'failed': error_count, 'total_evaluated': total_records_evaluated, 'success_rate_%': round(success_rate, 2), 'execution_time_sec': round(execution_time, 2), 'check_date': timestamp, 'status': status, 'status_color': status_color, 'error_file': error_file_status, 'comments': comments}
        rule_code = str(rule_info.get('rule_code', '')).strip()
        if rule_code == 'RCCOMP_375.1.2':
            result['filtered_adr2_count'] = total_rows
            result['filtered_adr2_file'] = ''
        elif filtered_adr2_addr_partner_df is not None and (not filtered_adr2_addr_partner_df.empty):
            result['filtered_adr2_count'] = len(filtered_adr2_addr_partner_df)
            result['filtered_adr2_file'] = ''
        self.results.append(result)
    ADR2_RULE_PARTNERS_TABLE = 'adr2_rule_partners'
    ADR2_RULE_ERRORS_TABLE = 'adr2_rule_errors'

    def _insert_adr2_partners_batch(self, rows: list, rule_code: str=None, run_ts: str=None):
        if not rows or not getattr(self, 'db_path', None):
            return
        try:
            import sqlite3
            r0 = rows[0]
            rule_code = rule_code or (r0[0] if len(r0) >= 4 else None)
            run_ts = run_ts or (r0[1] if len(r0) >= 4 else None)
            if not rule_code or not run_ts:
                return
            has_aufsd = len(r0) >= 5
            key = (rule_code, run_ts)
            cleared = getattr(self, '_adr2_partners_cleared', None)
            if cleared is None:
                self._adr2_partners_cleared = set()
                cleared = self._adr2_partners_cleared
            conn = connect_sqlite(self.db_path)
            if key not in cleared:
                conn.execute('\n                    CREATE TABLE IF NOT EXISTS ' + self.ADR2_RULE_PARTNERS_TABLE + ' (\n                        rule_code TEXT NOT NULL,\n                        run_ts TEXT NOT NULL,\n                        ADDRNUMBER TEXT,\n                        PARTNER TEXT,\n                        AUFSD TEXT\n                    )\n                ')
                try:
                    info = conn.execute('PRAGMA table_info(' + self.ADR2_RULE_PARTNERS_TABLE + ')').fetchall()
                    col_names = [c[1] for c in info]
                    if 'AUFSD' not in col_names:
                        conn.execute('ALTER TABLE ' + self.ADR2_RULE_PARTNERS_TABLE + ' ADD COLUMN AUFSD TEXT')
                except Exception:
                    pass
                conn.execute('DELETE FROM ' + self.ADR2_RULE_PARTNERS_TABLE + ' WHERE rule_code = ? AND run_ts = ?', (rule_code, run_ts))
                cleared.add(key)
            if has_aufsd:
                conn.executemany('INSERT INTO ' + self.ADR2_RULE_PARTNERS_TABLE + ' (rule_code, run_ts, ADDRNUMBER, PARTNER, AUFSD) VALUES (?, ?, ?, ?, ?)', rows)
            else:
                conn.executemany('INSERT INTO ' + self.ADR2_RULE_PARTNERS_TABLE + ' (rule_code, run_ts, ADDRNUMBER, PARTNER, AUFSD) VALUES (?, ?, ?, ?, ?)', [(r[0], r[1], r[2], r[3], '') for r in rows])
            conn.commit()
            conn.close()
        except Exception as e:
            print(f'      [WARN] Вставка пачки в {self.ADR2_RULE_PARTNERS_TABLE}: {e}')

    def _save_adr2_partners_to_db(self, rule_code: str, run_ts: str, addr_partner_df: pd.DataFrame):
        if addr_partner_df is None or addr_partner_df.empty or (not rule_code):
            return
        try:
            import sqlite3
            conn = connect_sqlite(self.db_path)
            conn.execute('\n                CREATE TABLE IF NOT EXISTS ' + self.ADR2_RULE_PARTNERS_TABLE + ' (\n                    rule_code TEXT NOT NULL,\n                    run_ts TEXT NOT NULL,\n                    ADDRNUMBER TEXT,\n                    PARTNER TEXT\n                )\n            ')
            conn.execute('DELETE FROM ' + self.ADR2_RULE_PARTNERS_TABLE + ' WHERE rule_code = ? AND run_ts = ?', (rule_code, run_ts))
            rows = [(rule_code, run_ts, str(row.get('ADDRNUMBER', '') or '').strip(), str(row.get('PARTNER', '') or '').strip()) for _, row in addr_partner_df.iterrows()]
            conn.executemany('INSERT INTO ' + self.ADR2_RULE_PARTNERS_TABLE + ' (rule_code, run_ts, ADDRNUMBER, PARTNER) VALUES (?, ?, ?, ?)', rows)
            conn.commit()
            conn.close()
            print(f'      [DB] Записано {len(addr_partner_df):,} партнёров в таблицу {self.ADR2_RULE_PARTNERS_TABLE} для правила {rule_code}')
        except Exception as e:
            print(f'      [WARN] Запись в таблицу {self.ADR2_RULE_PARTNERS_TABLE}: {e}')

    def _save_adr2_rule_errors_to_db(self, error_df: pd.DataFrame, rule_code: str, run_ts: str):
        if error_df is None or error_df.empty:
            return
        if not getattr(self, 'db_path', None):
            return
        if not run_ts or not rule_code:
            return
        try:
            import sqlite3
            conn = connect_sqlite(self.db_path)
            adr2_info = conn.execute('PRAGMA table_info("ADR2")').fetchall()
            if not adr2_info:
                conn.close()
                print(f'      [WARN] ADR2: таблица не найдена в БД, запись ошибок пропущена')
                return
            adr2_cols = [r[1] for r in adr2_info]
            adr2_col_types = {r[1]: r[2] or 'TEXT' for r in adr2_info}
            rename_map = {}
            for df_col in error_df.columns:
                for schema_col in adr2_cols:
                    if str(df_col).upper() == str(schema_col).upper() and df_col != schema_col:
                        rename_map[df_col] = schema_col
                        break
            df = error_df.rename(columns=rename_map).copy()
            extra_cols = [c for c in df.columns if c not in adr2_cols]
            data_cols = adr2_cols + extra_cols
            table = self.ADR2_RULE_ERRORS_TABLE
            exists = conn.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone()
            desired_cols = ['run_ts', 'rule_code'] + data_cols
            if not exists:
                col_defs = ['"run_ts" TEXT NOT NULL', '"rule_code" TEXT NOT NULL']
                for c in data_cols:
                    c_type = adr2_col_types.get(c, 'TEXT') or 'TEXT'
                    col_defs.append(f'"{c}" {c_type}')
                col_defs_sql = ', '.join(col_defs)
                conn.execute(f'CREATE TABLE IF NOT EXISTS "{table}" ({col_defs_sql})')
            existing_cols = {r[1] for r in conn.execute(f'PRAGMA table_info("{table}")').fetchall()}
            for c in desired_cols:
                if c not in existing_cols:
                    conn.execute(f'ALTER TABLE "{table}" ADD COLUMN "{c}" TEXT')
            conn.execute(f'DELETE FROM "{table}" WHERE run_ts = ? AND rule_code = ?', (run_ts, rule_code))
            df_insert = df.copy()
            df_insert['run_ts'] = run_ts
            df_insert['rule_code'] = rule_code
            for c in desired_cols:
                if c not in df_insert.columns:
                    df_insert[c] = None
            df_insert = df_insert[desired_cols]
            df_insert = df_insert.where(pd.notnull(df_insert), None)
            cols_sql = ', '.join([f'"{c}"' for c in desired_cols])
            placeholders = ', '.join(['?'] * len(desired_cols))
            insert_sql = f'INSERT INTO "{table}" ({cols_sql}) VALUES ({placeholders})'
            values = df_insert.to_numpy(dtype=object)
            chunk_size = 5000
            for start in range(0, len(values), chunk_size):
                chunk = values[start:start + chunk_size]
                conn.executemany(insert_sql, [tuple(r) for r in chunk])
            conn.commit()
            conn.close()
            print(f'   [DB] ADR2 errors saved: table={table}, rule={rule_code}, run_ts={run_ts}, rows={len(df_insert):,}')
        except Exception as e:
            print(f'   [WARN] Не удалось сохранить ADR2 ошибки в БД: {e}')

    def run(self, specific_table: str=None, table_list: list=None, only_rule_codes: set=None):
        print(f'\n' + '=' * 100)
        print(f'\x1b[1mЗАПУСК СИСТЕМЫ ПРОВЕРКИ КАЧЕСТВА\x1b[0m')
        print(f'=' * 100)
        if table_list:
            print(f'[INFO] Проверяем выбранные таблицы: {len(table_list)} шт.')
        elif specific_table:
            print(f'[INFO] Проверяем только таблицу: {specific_table}')
        else:
            print(f'[INFO] Проверяем все таблицы')
        if only_rule_codes:
            print(f'[INFO] Изолированный режим: только правила {sorted(only_rule_codes)}')
        print(f'[INFO] Настройки: Сохраняется максимум {self.MAX_ERRORS_TO_SAVE:,} ошибок на правило')
        print(f'=' * 100)
        results_df = self.run_quality_checks_fast(specific_table=specific_table, table_list=table_list, only_rule_codes=only_rule_codes)
        if not results_df.empty:
            print(f'\n' + '=' * 100)
            print(f'\x1b[1mПРОВЕРКА ЗАВЕРШЕНА УСПЕШНО!\x1b[0m')
            print(f'=' * 100)
            if specific_table:
                report_name = f'quality_check_report_{self._safe_filename_token(specific_table)}.xlsx'
            elif table_list and len(table_list) == 1:
                report_name = f'quality_check_report_{self._safe_filename_token(table_list[0])}.xlsx'
            else:
                report_name = 'quality_check_report.xlsx'
            report_path = getattr(self, 'last_stable_report_path', None) or getattr(self, 'last_report_path', None) or os.path.join(self.output_dir, report_name)
            print(f'   Отчет: {report_path}')
            if getattr(self, 'last_errors_dir', None) and os.path.isdir(self.last_errors_dir):
                saved_n = len(getattr(self, 'saved_error_files', {}) or {})
                print(f'   Ошибки: {self.last_errors_dir} ({saved_n} файлов)')
            elif self.rule_errors:
                ts = getattr(self, 'last_file_timestamp', '') or ''
                print(f'   Ошибки: {os.path.join(self.output_dir, "errors_" + ts)} ({len(self.rule_errors)} правил с ошибками в памяти)')
            adr2_with_data = [r for r in self.results or [] if str(r.get('table_name', '')).strip().upper() == 'ADR2' and r.get('filtered_adr2_count')]
            if adr2_with_data:
                print(f'   ADR2 — список партнёров по правилам: таблица «{self.ADR2_RULE_PARTNERS_TABLE}» в БД (rule_code, run_ts, ADDRNUMBER, PARTNER)')
        else:
            print(f'\n[INFO] Нет результатов проверки')

    def list_available_tables(self):
        rules_config = self.load_configuration()
        if not rules_config:
            print(f'\n[ERROR] Не удалось загрузить конфигурацию правил')
            return []
        tables = list(rules_config.keys())
        print(f'\n[INFO] Доступные таблицы для проверки:')
        print(f'=' * 50)
        for i, table in enumerate(tables, 1):
            rule_count = len(rules_config[table])
            print(f'{i:3}. {table:25} - {rule_count:3} правил')
        print(f'=' * 50)
        print(f'[INFO] Всего таблиц: {len(tables)}')
        return tables
    DFKKBPTAXNUM_ALIASES = ('DFKKBPTAXNUM1', 'DFKKBPTAXNUM2', 'DFKKBPTAXNUM3', 'DFKKBPTAXNUM5')
    DFKKBPTAXNUM_SHARED_RULE_CODES = frozenset({'RCCONF_52.4', 'RCCONF_52.3', 'RCCONF_52.2', 'RCCOMP_52.2'})
    TAXNUM_SAME_ROW_RULES = frozenset({'RCCONF_50.11', 'RCCONF_52.11', 'RCCONF_54.9', 'RCCONF_63.7'})
    TAXNUM_FORMAT_RULES = frozenset({'RCCONF_50.1', 'RCCONF_52.1', 'RCCONF_54.1', 'RCCONF_63.1'})
    RCCOMP_149_RULES = frozenset({'RCCOMP_149.1', 'RCCOMP_149.2'})
    RCCOMP_149_ORDER_BLOCK_SKIP = frozenset({'S', 'SP', 'E', 'G', 'S2', 'S3', 'S4', 'S5', 'S9', 'R', 'U', 'S1', 'SY', 'IA', 'IB', 'RN'})
    RCCOMP_149_ACCOUNT_GROUP_SKIP_PATTERN = '90%'
    RCCOMP_149_1_REQUIRED_PF = frozenset({'BP', 'PY', 'ZY', 'SP', 'SH', 'YR'})
    KNVV_ORDER_BLOCK_BLOCKED = frozenset({'S', 'NH', 'S3', 'S4', 'SY', 'U', 'R', 'PR'})
    KNVV_DM_SALES_ORG_SCOPE_RULES = frozenset({
        'RCCOMP_142.1', 'RCCOMP_143.1', 'RCCOMP_144.1', 'RCCOMP_153.1', 'RCCOMP_154.1',
        'RCCOMP_163.1', 'RCCOMP_164.1', 'RCCOMP_170.1', 'RCCOMP_148.1',
        'RCCONF_143.7', 'RCCONF_154.4', 'RCCONF_154.1', 'RCCONF_170.7', 'RCCONF_164.1',
    })
    TABLE_UNIQUE_PARTNER = ('ZBUT0000P3VVI9', 'ZBUT0000P', 'ZBUT0000P3VV19')
    AUSP_TABLE_GROUP = ('AUSP_143', 'AUSP_604', 'AUSP_148', 'AUSP_151')

    def _get_dfkkbptaxnum_type_from_table(self, table_name):
        m = re.match('DFKKBPTAXNUM(\\d)$', str(table_name or '').strip(), re.I)
        return int(m.group(1)) if m else None

    def _find_dfkkbptaxnum_taxtype_column(self, df, table_name=None):
        if df is None or df.empty:
            return None
        cols = list(df.columns)
        root = parent_dir
        for rel in ('json files/conf_dfkkbptaxnum.json', 'config/conf_dfkkbptaxnum.json'):
            path = os.path.join(root, rel)
            if not os.path.isfile(path):
                continue
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
                for key in ('taxtype_column', 'taxtype_column_name'):
                    cand = cfg.get(key)
                    if not cand:
                        continue
                    if cand in cols:
                        return cand
                    for c in cols:
                        if str(c).strip().upper() == str(cand).strip().upper():
                            return c
                for alt in cfg.get('taxtype_column_alternatives') or cfg.get('taxtype_columns') or []:
                    if alt in cols:
                        return alt
                    for c in cols:
                        if str(c).strip().upper() == str(alt).strip().upper():
                            return c
            except Exception:
                pass
        resolved = self._resolve_column_for_rule(df, 'TAXTYPE', table_name) if table_name else None
        if resolved and resolved in cols:
            return resolved
        for col in cols:
            cu = str(col).upper().replace(' ', '').replace('_', '')
            if cu in ('TAXNUMBERCATEGORY', 'TAXTYPE', 'TAXTYP', 'TAXNUMTYPE', 'TYPE') or 'TAXTYPE' in cu:
                return col
        return None

    def _scope_dfkkbptaxnum_by_taxtype(self, df, table_name):
        typ = self._get_dfkkbptaxnum_type_from_table(table_name)
        if typ is None or df is None or df.empty:
            return df, None
        type_col = self._find_dfkkbptaxnum_taxtype_column(df, table_name)
        if not type_col or type_col not in df.columns:
            print(f'      [WARN] [{table_name}] Tax_Number_Category не найдена — «всего записей» по всей таблице ({len(df):,})')
            return df, len(df)
        type_ser = df[type_col].astype(str).str.strip()
        numeric_type = pd.to_numeric(type_ser, errors='coerce')
        want_ru = f'RU{typ}'
        mask = (numeric_type == typ) | (type_ser == str(typ)) | (type_ser.str.upper() == want_ru)
        scoped = df.loc[mask].copy()
        baseline = len(scoped)
        before = len(df)
        if before != baseline:
            print(f"      [FILTER] {table_name}: «всего записей» только Tax_Number_Category={want_ru} -> {baseline:,} из {before:,}")
        else:
            print(f"      [FILTER] {table_name}: «всего записей» по Tax_Number_Category={want_ru}: {baseline:,}")
        return scoped, baseline

    def _norm_knvv_so_code_val(self, value):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return ''
        s = str(value).strip()
        if not s or s.lower() in ('none', 'null', 'nan', 'na'):
            return ''
        try:
            if isinstance(value, (int, float)) and (not pd.isna(value)) and float(value) == int(float(value)):
                return str(int(float(value))).zfill(2)
        except (ValueError, TypeError, OverflowError):
            pass
        s_num = s.replace(',', '.')
        try:
            num = float(s_num)
            if num == int(num):
                return str(int(num)).zfill(2)
        except ValueError:
            pass
        if s.isdigit() and len(s) < 2:
            return s.zfill(2)
        return s

    def _norm_knvv_so_code_series(self, series):
        if series is None:
            return pd.Series(dtype=str)
        return series.map(self._norm_knvv_so_code_val)

    def _apply_knvv_dm_sales_org_scope(self, df, rule_code, table_name):
        """dm_customer_sales_org scope: VTWEG=01 & SPART=01, exclude blocked OrBlk and KDGRP=ZIN.
        RCCONF_153.4 uses VTWEG=04 & SPART=02 per technical_definition (outside 01-01 dump)."""
        if df is None or df.empty:
            return df
        rule_code = str(rule_code or '').strip().upper()
        if str(table_name or '').strip().upper() != 'KNVV':
            return df
        before = len(df)
        if rule_code == 'RCCONF_153.4':
            vtweg_col = self._resolve_column_for_rule(df, 'VTWEG', 'KNVV')
            spart_col = self._resolve_column_for_rule(df, 'SPART', 'KNVV')
            if not vtweg_col or not spart_col or vtweg_col not in df.columns or spart_col not in df.columns:
                print(f'      [WARN] {rule_code}: VTWEG/SPART не найдены в KNVV')
                return df.iloc[0:0].copy()
            vt = self._norm_knvv_so_code_series(df[vtweg_col])
            sp = self._norm_knvv_so_code_series(df[spart_col])
            scoped = df.loc[(vt == '04') & (sp == '02')].copy()
            print(f"      [FILTER] {rule_code} scope: VTWEG='04' AND SPART='02' -> {len(scoped):,} из {before:,}")
            if scoped.empty:
                print(f"      [INFO] {rule_code}: в выгрузке нет VTWEG=04/SPART=02 (дамп только 01-01) — правило вне scope")
            return scoped
        if rule_code not in self.KNVV_DM_SALES_ORG_SCOPE_RULES:
            return df
        mask = pd.Series(True, index=df.index)
        vtweg_col = self._resolve_column_for_rule(df, 'VTWEG', 'KNVV')
        spart_col = self._resolve_column_for_rule(df, 'SPART', 'KNVV')
        if vtweg_col and spart_col and vtweg_col in df.columns and (spart_col in df.columns):
            vt = self._norm_knvv_so_code_series(df[vtweg_col])
            sp = self._norm_knvv_so_code_series(df[spart_col])
            mask = mask & (vt == '01') & (sp == '01')
            after = int(mask.sum())
            print(f"      [FILTER] {rule_code} scope: VTWEG='01' AND SPART='01' -> {after:,} из {before:,}")
            before = after
        else:
            print(f'      [WARN] {rule_code}: VTWEG/SPART не найдены в KNVV — scope 01-01 не применён')
        kdgrp_col = self._resolve_column_for_rule(df, 'KDGRP', 'KNVV')
        if not kdgrp_col or kdgrp_col not in df.columns:
            kdgrp_col = next((c for c in df.columns if str(c).strip().upper() in ('CGRP', 'KDGRP')), None)
        if kdgrp_col and kdgrp_col in df.columns:
            before_kd = int(mask.sum())
            kd = df[kdgrp_col].astype(str).str.strip().str.upper()
            mask = mask & (kd != 'ZIN')
            after = int(mask.sum())
            if before_kd != after:
                print(f"      [FILTER] {rule_code} exclude KDGRP='ZIN' -> {after:,} из {before_kd:,}")
        aufsd_col = self._resolve_column_for_rule(df, 'AUFSD', 'KNVV')
        if not aufsd_col or aufsd_col not in df.columns:
            aufsd_col = next((c for c in df.columns if str(c).strip().upper() in ('ORBLK', 'AUFSD')), None)
        if aufsd_col and aufsd_col in df.columns:
            before_au = int(mask.sum())
            au = df[aufsd_col].astype(str).str.strip().str.upper()
            mask = mask & (~au.isin(self.KNVV_ORDER_BLOCK_BLOCKED))
            after = int(mask.sum())
            if before_au != after:
                blocked = sorted(self.KNVV_ORDER_BLOCK_BLOCKED)
                print(f"      [FILTER] {rule_code} exclude order_block in {blocked} -> {after:,} из {before_au:,}")
        return df.loc[mask].copy()

    def _partner_code_filled_mask(self, series):
        if series is None:
            return pd.Series(dtype=bool)
        s = series.astype(str).str.strip()
        s_for_zero = s.str.replace(',', '.', regex=False)
        zeroish = s_for_zero.str.match('^-?0+(?:[.][0]+)?$', na=False)
        return series.notna() & (s != '') & ~s.str.lower().isin(['none', 'null', 'nan', 'na']) & ~zeroish

    def _find_order_block_column(self, df):
        if df is None or df.empty:
            return None
        for c in df.columns:
            cl = str(c).strip().lower()
            if cl in ('order_block_code', 'central_order_block_code', 'aufsd', 'orblk'):
                return c
        return None

    def _add_order_block_code_from_kna1_customer(self, df, table_name, rule_code, join_col=None):
        try:
            if df is None or df.empty:
                return df
            if self._find_order_block_column(df):
                return df
            if not join_col:
                join_col = self._pick_best_kunnr_column(df, table_name)
            if not join_col:
                print(f'      [WARN] {rule_code}: не найден ключ клиента для JOIN KNA1.AUFSD')
                return df
            kna1_df = self._get_table_for_rules('KNA1')
            if kna1_df is None or kna1_df.empty:
                try:
                    self.memory_manager.load_selected_tables_to_ram(['KNA1'], add_reference_tables=False)
                    kna1_df = self._get_table_for_rules('KNA1')
                except Exception:
                    kna1_df = None
            if kna1_df is None or kna1_df.empty:
                try:
                    conn = connect_sqlite(self.db_path)
                    kna1_df = pd.read_sql_query('SELECT * FROM "KNA1"', conn)
                    conn.close()
                except Exception:
                    kna1_df = None
            if kna1_df is None or kna1_df.empty:
                print(f'      [WARN] {rule_code}: KNA1 пуста, order_block_code не добавлен')
                return df
            kna1_df = self._apply_rule_time_column_map(kna1_df.copy(), 'KNA1')
            kunnr_col = self._pick_best_kunnr_column(kna1_df, 'KNA1')
            aufsd_col = self._resolve_column_for_rule(kna1_df, 'AUFSD', 'KNA1') or self._resolve_column_for_rule(kna1_df, 'central_order_block_code', 'KNA1')
            if not aufsd_col:
                aufsd_col = next((c for c in kna1_df.columns if str(c).strip().upper() in ('AUFSD', 'ORBLK')), None)
            if not kunnr_col or not aufsd_col:
                print(f'      [WARN] {rule_code}: в KNA1 не найдены KUNNR/AUFSD для order_block_code')
                return df
            out = df.copy()
            out['_join_key'] = out[join_col].apply(self._norm_customer_partner_key)
            kna1_join = kna1_df[[kunnr_col, aufsd_col]].copy()
            kna1_join['_join_key'] = kna1_join[kunnr_col].apply(self._norm_customer_partner_key)
            kna1_join = kna1_join[['_join_key', aufsd_col]].drop_duplicates(subset=['_join_key'], keep='first')
            kna1_join = kna1_join.rename(columns={aufsd_col: 'order_block_code'})
            out = out.merge(kna1_join, on='_join_key', how='left')
            out['central_order_block_code'] = out['order_block_code']
            out = out.drop(columns=['_join_key'], errors='ignore')
            filled = int(out['order_block_code'].astype(str).str.strip().ne('').sum())
            print(f'      [JOIN] {rule_code}: order_block_code из KNA1.{aufsd_col} по {table_name}.{join_col} (заполнено {filled:,}/{len(out):,})')
            return out
        except Exception as e:
            print(f'      [WARN] Ошибка добавления order_block_code из KNA1 для {rule_code}: {e}')
            return df

    def _knvp_sales_org_scope_mask(self, df, vtweg_col, spart_col):
        if df is None or df.empty or not vtweg_col or not spart_col:
            return pd.Series(False, index=df.index if df is not None else [])
        if vtweg_col not in df.columns or spart_col not in df.columns:
            return pd.Series(False, index=df.index)
        vt = self._norm_knvv_so_code_series(df[vtweg_col])
        sp = self._norm_knvv_so_code_series(df[spart_col])
        return ((vt == '01') & (sp == '01')) | ((vt == '04') & (sp == '02'))

    def _get_knvv_indirect_customer_keys(self, rule_code):
        knvv_df = self._get_table_for_rules('KNVV')
        if knvv_df is None or knvv_df.empty:
            try:
                self.memory_manager.load_selected_tables_to_ram(['KNVV'], add_reference_tables=False)
                knvv_df = self._get_table_for_rules('KNVV')
            except Exception:
                knvv_df = None
        if knvv_df is None or knvv_df.empty:
            print(f'      [WARN] {rule_code}: KNVV пуста — indirect (KVGR4=IN) клиенты не определены')
            return set()
        vtweg_col = self._resolve_column_for_rule(knvv_df, 'VTWEG', 'KNVV')
        spart_col = self._resolve_column_for_rule(knvv_df, 'SPART', 'KNVV')
        kvgr4_col = self._resolve_column_for_rule(knvv_df, 'KVGR4', 'KNVV') or next((c for c in knvv_df.columns if str(c).strip().upper() in ('GRP4', 'KVGR4')), None)
        kunnr_col = self._pick_best_kunnr_column(knvv_df, 'KNVV')
        if not kunnr_col or not kvgr4_col:
            print(f'      [WARN] {rule_code}: в KNVV не найдены Customer/KVGR4')
            return set()
        so_mask = self._knvp_sales_org_scope_mask(knvv_df, vtweg_col, spart_col)
        kv4 = knvv_df[kvgr4_col].astype(str).str.strip().str.upper()
        keys = knvv_df.loc[so_mask & (kv4 == 'IN'), kunnr_col].apply(self._norm_customer_partner_key)
        return set(keys[keys != ''])

    def _process_rcccomp_149_knvp(self, rule_code, df, table_name, rule, validator, matched_column, column_to_check, save_result, timestamp):
        rule_code = str(rule_code or '').strip().upper()
        kunnr_col = self._pick_best_kunnr_column(df, table_name)
        if not kunnr_col:
            self._log_skipped_rule(rule, table_name, f'{rule_code}: не найден ключ клиента (KUNNR/Customer)', timestamp)
            return (0, 0)
        parvw_col = self._resolve_column_for_rule(df, 'PARVW', 'KNVP') or matched_column
        parc_col = self._resolve_column_for_rule(df, 'ParC', 'KNVP')
        if not parc_col:
            parc_col = self._find_column_alternative(df.columns, 'ParC', 'KNVP')
        if not parc_col:
            parc_col = self._find_column_alternative(df.columns, 'partner_code', 'KNVP')
        if not parvw_col or parvw_col not in df.columns or not parc_col or parc_col not in df.columns:
            self._log_skipped_rule(rule, table_name, f'{rule_code}: не найдены колонки partner function ({parvw_col}) / partner code ({parc_col})', timestamp)
            return (0, 0)
        df_work = df.copy()
        before_rows = len(df_work)
        vtweg_col = self._resolve_column_for_rule(df_work, 'VTWEG', 'KNVP')
        spart_col = self._resolve_column_for_rule(df_work, 'SPART', 'KNVP')
        if vtweg_col and spart_col and vtweg_col in df_work.columns and (spart_col in df_work.columns):
            so_mask = self._knvp_sales_org_scope_mask(df_work, vtweg_col, spart_col)
            df_work = df_work.loc[so_mask].copy()
            n_cust = df_work[kunnr_col].apply(self._norm_customer_partner_key).nunique()
            print(f"      [FILTER] {rule_code} KNVP SO scope (01-01 or 04-02) -> {len(df_work):,} строк, {n_cust:,} клиентов из {before_rows:,} строк")
        if df_work.empty:
            self._log_skipped_rule(rule, table_name, f'{rule_code}: нет строк KNVP в SO 01-01 / 04-02', timestamp)
            return (0, 0)
        df_work = self._add_account_group_code_from_kna1(df_work, table_name, rule_code)
        ag_col = self._find_account_group_column(df_work)
        df_work = self._add_order_block_code_from_kna1_customer(df_work, table_name, rule_code, kunnr_col)
        ob_col = self._find_order_block_column(df_work)
        if not ag_col or not ob_col:
            self._log_skipped_rule(rule, table_name, f'{rule_code}: не удалось получить account_group_code / order_block_code из KNA1', timestamp)
            return (0, 0)
        from utils.sap_account_keys import norm_sap_account_group, sap_account_group_like
        df_work['_cust_key'] = df_work[kunnr_col].apply(self._norm_customer_partner_key)
        cust_df = df_work.drop_duplicates(subset=['_cust_key'], keep='first').set_index('_cust_key', drop=False)
        ag_norm = cust_df[ag_col].apply(norm_sap_account_group)
        ob_norm = cust_df[ob_col].astype(str).str.strip().str.upper()
        ag_skip_pat = self.RCCOMP_149_ACCOUNT_GROUP_SKIP_PATTERN
        like_90 = ag_norm.apply(lambda v: sap_account_group_like(v, ag_skip_pat))
        not_blocked = ~ob_norm.isin(self.RCCOMP_149_ORDER_BLOCK_SKIP)
        skip_like_90 = like_90
        eval_scope = ~skip_like_90 & not_blocked
        total_scope = eval_scope
        n_skip_like_90 = int(skip_like_90.sum())
        n_skip_blocked = int((~skip_like_90 & ~not_blocked).sum())
        print(f"      [FILTER] {rule_code}: account_group LIKE '{ag_skip_pat}' -> пропущено {n_skip_like_90:,} клиентов (scope-колонка: {ag_col})")
        if n_skip_blocked:
            print(f"      [FILTER] {rule_code}: blocked order_block -> ещё пропущено {n_skip_blocked:,} клиентов")
        scope_desc = f"уникальные клиенты KNVP в SO 01-01/04-02, account_group NOT LIKE '{ag_skip_pat}', без blocked order_block"
        if rule_code == 'RCCOMP_149.2':
            in_keys = self._get_knvv_indirect_customer_keys(rule_code)
            if not in_keys:
                self._log_skipped_rule(rule, table_name, f'{rule_code}: нет indirect-клиентов (KVGR4=IN) в KNVV для SO 01-01/04-02', timestamp)
                return (0, 0)
            total_scope = total_scope & cust_df['_cust_key'].isin(in_keys)
            eval_scope = eval_scope & cust_df['_cust_key'].isin(in_keys)
            scope_desc = f"indirect-клиенты (KVGR4='IN'), SO 01-01/04-02, account_group NOT LIKE '{ag_skip_pat}'"
        in_scope_keys = set(cust_df.index[total_scope])
        eval_keys = set(cust_df.index[eval_scope])
        total_rows = len(in_scope_keys)
        print(f'      [FILTER] {rule_code} «Всего записей» = {total_rows:,} клиентов ({scope_desc})')
        print(f'      [FILTER] {rule_code} к оценке ошибок (без blocked order_block): {len(eval_keys):,} клиентов')
        if total_rows == 0:
            self._log_skipped_rule(rule, table_name, f'{rule_code}: нет клиентов в scope после фильтров', timestamp)
            return (0, 0)
        df_scoped = df_work[df_work['_cust_key'].isin(eval_keys)].copy()
        filled = self._partner_code_filled_mask(df_scoped[parc_col])
        df_scoped['_parvw_u'] = df_scoped[parvw_col].astype(str).str.strip().str.upper()
        rep_idx = df_scoped.groupby('_cust_key', sort=False).head(1).index
        if rule_code == 'RCCOMP_149.1':
            required = sorted(self.RCCOMP_149_1_REQUIRED_PF)
            df_pf = df_scoped.loc[filled & df_scoped['_parvw_u'].isin(self.RCCOMP_149_1_REQUIRED_PF)]
            present = df_pf.groupby(['_cust_key', '_parvw_u']).size().unstack(fill_value=0)
            for pf in required:
                if pf not in present.columns:
                    present[pf] = 0
            present = present.reindex(list(eval_keys), fill_value=0)
            bad = (present[required] == 0).any(axis=1)
            error_keys = set(present.index[bad])
            error_count = int(bad.sum())
            error_description = f'Missing Partner function: required partner_function_code in {", ".join(required)} with assigned partner_code (SO 01-01/04-02).'
        else:
            zw_ok = set(df_scoped.loc[filled & (df_scoped['_parvw_u'] == 'ZW'), '_cust_key'])
            error_keys = eval_keys - zw_ok
            error_count = len(error_keys)
            error_description = "Indirect customer (KVGR4='IN') must have partner_function_code='ZW' with assigned partner_code (SO 01-01/04-02)."
        error_mask = pd.Series(False, index=df_scoped.index)
        if error_keys:
            rep_df = df_scoped.loc[rep_idx]
            error_mask.loc[rep_df[rep_df['_cust_key'].isin(error_keys)].index] = True
        error_df = validator._prepare_error_dataframe(df_scoped, error_mask, 'COMPLETENESS', error_description) if error_count > 0 else None
        is_suspicious = self._check_if_suspicious(rule_code, error_count, total_rows)
        if save_result:
            rule_info = {'rule_code': rule_code, 'rule_description': rule.get('rule_description', 'Unknown rule'), 'quality_category': rule.get('quality_category', 'Unknown'), 'table_name': table_name, 'original_column': column_to_check, 'matched_column': matched_column}
            if self._parallel_lock:
                with self._parallel_lock:
                    self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                    self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
            else:
                self._save_rule_result(rule_info, total_rows, error_count, 0, timestamp, is_suspicious)
                self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
        elif error_df is not None and (not error_df.empty):
            if self._parallel_lock:
                with self._parallel_lock:
                    self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
            else:
                self._save_rule_error_with_limit(rule_code, table_name, error_df, error_count, is_suspicious, total_rows=total_rows)
        return (error_count, total_rows)

    def _expand_ausp_for_load(self, table_names):
        if not table_names:
            return table_names
        ausp_derived = set(self.AUSP_TABLE_GROUP)
        out = []
        needs_ausp = False
        for t in table_names:
            tu = str(t or '').strip().upper()
            if tu == 'AUSP' or tu in ausp_derived:
                needs_ausp = True
                if tu in ausp_derived:
                    out.append(t)
            else:
                out.append(t)
        if needs_ausp and 'AUSP' not in out:
            out.append('AUSP')
        kna1_dependent = {'BUT0BK', 'BUT051', 'KNB1', 'KNVV', 'KNVP', 'KNVH', 'ADR2', 'ADRC', 'BUT050'}
        if any((str(t).strip().upper() in kna1_dependent for t in out)) and 'KNA1' not in out:
            out.append('KNA1')
        if 'KNA1' in out and 'ZW2_CMDEMAND' not in out:
            out.append('ZW2_CMDEMAND')
        seen = set()
        deduped = []
        for t in out:
            key = str(t).strip().upper()
            if key not in seen:
                seen.add(key)
                deduped.append(t)
        return deduped

    def get_table_rules(self, table_name: str):
        rules_config = self.load_configuration()
        if table_name in rules_config:
            rules = list(rules_config[table_name])
            if table_name in self.DFKKBPTAXNUM_ALIASES and table_name != 'DFKKBPTAXNUM1':
                have = {str(r.get('rule_code') or '').strip() for r in rules}
                for shared in rules_config.get('DFKKBPTAXNUM1', []):
                    code = str(shared.get('rule_code') or '').strip()
                    if code in self.DFKKBPTAXNUM_SHARED_RULE_CODES and code not in have:
                        rules.append(shared)
                        have.add(code)
            return rules
        if table_name == 'DFKKBPTAXNUM':
            combined = []
            for alias in self.DFKKBPTAXNUM_ALIASES:
                combined.extend(rules_config.get(alias, []))
            return combined
        if table_name == 'AUSP':
            combined = []
            for t in self.AUSP_TABLE_GROUP:
                combined.extend(rules_config.get(t, []))
            return combined if combined else []
        print(f"\n[ERROR] Таблица '{table_name}' не найдена в конфигурации")
        return []

    def _get_validator_for_rule(self, rule_description, quality_category, rule_info):
        rule_desc_lower = rule_description.lower()
        rule_code_raw = str(rule_info.get('rule_code', ''))
        rule_code = re.sub('[^A-Za-z0-9._-]', '', rule_code_raw).upper()
        if rule_code == 'RCCONF_24.1':
            return ConformityValidator(rule_info)
        if rule_code == 'RCCONF_119.2' or ('payment terms' in rule_desc_lower and 'knb1' in rule_desc_lower and ('knvv' in rule_desc_lower)):
            return PaymentTermsConsistencyValidator(rule_info)
        if rule_code == 'RCCONF_113.1' or ('recon' in rule_desc_lower and 'account group' in rule_desc_lower):
            return ReconAccountConsistencyValidator(rule_info)
        if rule_code == 'RCCONF_63.1':
            return ConformityValidator(rule_info)
        if rule_code == 'RCCONF_15.1':
            return LogicalValidator(rule_info, self.error_manager)
        if rule_code in ['RCCOMP_375.1', 'RCCOMP_375.1.2']:
            return CompletenessValidator(rule_info)
        if rule_code in ['RCCONF_18.2', 'RCCONF_22.2']:
            return AdvancedSpecialCharactersValidator(rule_info)
        if rule_code == 'RCCONF_22.3':
            return UppercaseValidator(rule_info)
        if 'check if' in rule_desc_lower and 'equals' in rule_desc_lower:
            return CrossColumnEqualityCheckValidator(rule_info)
        elif 'cannot be the same' in rule_desc_lower or 'cannot be a the same' in rule_desc_lower:
            return CrossColumnEqualityValidator(rule_info)
        elif 'равен' in rule_desc_lower or 'равны' in rule_desc_lower:
            return CrossColumnEqualityValidator(rule_info)
        elif 'недопустимые пробелы' in rule_desc_lower or 'consecutive space' in rule_desc_lower or 'two or more consecutive' in rule_desc_lower:
            return ConsecutiveSpacesValidator(rule_info)
        elif 'специальные символы' in rule_desc_lower or 'special character' in rule_desc_lower:
            return SpecialCharactersValidator(rule_info)
        elif 'верхний регистр' in rule_desc_lower or 'uppercase' in rule_desc_lower or 'capital letters' in rule_desc_lower:
            return UppercaseValidator(rule_info)
        elif 'отсутствует' in rule_desc_lower or 'missing' in rule_desc_lower:
            return CompletenessValidator(rule_info)
        else:
            return ConformityValidator(rule_info)

    def _find_kunnr_column(self, df):
        if df is None or df.empty or (not hasattr(df, 'columns')):
            return None
        col_upper_map = {str(col).strip().upper(): col for col in df.columns}
        if 'KUNNR' in col_upper_map:
            return col_upper_map['KUNNR']
        for cu, col in col_upper_map.items():
            if 'KUNNR' in cu:
                return col
        for candidate in ('CUSTOMER', 'CUSTOMER_CODE', 'MC_CUSTOMER', 'KUNNR_KNA1', 'KUNN'):
            if candidate in col_upper_map:
                return col_upper_map[candidate]
        return None

    def _norm_customer_partner_key(self, v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ''
        s = str(v).replace('\ufeff', '').replace('\xa0', ' ').strip().strip("'").strip('"').strip()
        if s.lower() in {'', 'none', 'null', 'nan', '<na>', 'nat', '-', '.', 'n/a', 'na'}:
            return ''
        if re.fullmatch('\\d+\\.0+', s):
            s = s.split('.')[0]
        digits = re.sub('\\D', '', s)
        if not digits:
            return ''
        return digits.zfill(10)

    def _resolve_but0bk_partner_column(self, df, table_name='BUT0BK'):
        if df is None or df.empty:
            return None
        try:
            from utils.column_map_resolver import resolve_column_in_df
            for sap in ('PARTNER',):
                col = resolve_column_in_df(df, sap, table_name, self.column_map, parent_dir)
                if col:
                    return col
        except ImportError:
            pass
        for name in ('PARTNER', 'Business_Partner', 'BUSINESS_PARTNER'):
            for c in df.columns:
                if str(c).strip().upper() == name:
                    return c
        return self._find_partner_column(df, table_name=table_name)

    def _find_account_group_column(self, df):
        if df is None or df.empty:
            return None
        from utils.sap_account_keys import norm_sap_account_group
        priority = ('account_group_code', 'b.account_group_code', 'kna.ktokd', 'ktokd', 'b.ktokd', 'KTOKD', 'group_1', 'Group_1')
        col_by_lower = {str(c).strip().lower(): c for c in df.columns}
        best_col = None
        best_filled = -1
        for name in priority:
            col = col_by_lower.get(str(name).strip().lower())
            if not col:
                continue
            filled = int((df[col].apply(norm_sap_account_group) != '').sum())
            if filled > best_filled:
                best_filled = filled
                best_col = col
        return best_col

    def _filter_rows_only_ktokd_9038(self, df, rule_code):
        from utils.sap_account_keys import norm_sap_account_group
        if df is None or df.empty:
            return df
        ag_col = self._find_account_group_column(df)
        if not ag_col:
            print(f'      [WARN] {rule_code}: account_group_code/KTOKD не найден — scope «только 9038» невозможен')
            return df.iloc[0:0].copy()
        ag_norm = df[ag_col].apply(norm_sap_account_group)
        mask = ag_norm == '9038'
        matched = int(mask.sum())
        filled_ktokd = int((ag_norm != '').sum())
        self._last_kna1_join_stats = {'rows_after_join': len(df), 'filled_ktokd': filled_ktokd, 'n9038': matched}
        print(f'      [FILTER] {rule_code}: KTOKD=9038 (KNA1) -> {matched:,} из {len(df):,} (с заполненным KTOKD: {filled_ktokd:,})')
        if matched == 0 and len(df) > 0:
            top = ag_norm[ag_norm != ''].value_counts().head(8)
            if not top.empty:
                print(f'      [FILTER] топ KTOKD после JOIN KNA1: {top.to_dict()}')
            else:
                print(f'      [WARN] {rule_code}: после JOIN KNA1 колонка {ag_col} пуста у всех {len(df):,} строк — проверьте ключ Customer/KUNNR')
        return df[mask].copy()

    def _resolve_kna1_kunnr_column(self, df):
        if df is None or df.empty:
            return None
        try:
            from utils.column_map_resolver import resolve_column_in_df
            col = resolve_column_in_df(df, 'KUNNR', 'KNA1', self.column_map, parent_dir)
            if col:
                return col
        except ImportError:
            pass
        return self._find_kunnr_column(df) or next((c for c in df.columns if str(c).strip().upper() in ('KUNNR', 'CUSTOMER', 'KUNN')), None)

    def _find_partner_column(self, df, table_name=None):
        if df is None or df.empty or (not hasattr(df, 'columns')):
            return None
        cols = list(df.columns)
        if table_name and getattr(self, 'column_map', None) and (table_name in self.column_map):
            for key in ('partner', 'partner_column', 'PARTNER', 'partners'):
                if key in self.column_map[table_name]:
                    phys = self.column_map[table_name][key]
                    if phys and str(phys).strip() in cols:
                        return str(phys).strip()
                    for c in cols:
                        if str(c).strip().upper() == str(phys).strip().upper():
                            return c
        if table_name and table_name in getattr(self, 'TABLE_UNIQUE_PARTNER', ()):
            config_col = self._load_partner_column_config(table_name)
            if config_col:
                config_upper = config_col.strip().upper()
                for c in cols:
                    if str(c).strip().upper() == config_upper:
                        return c
                if config_col in cols:
                    return config_col
        col_upper = {str(c).strip().upper(): c for c in cols}
        tn = str(table_name or '').strip().upper()
        if tn == 'KNVP':
            try:
                from utils.column_map_resolver import resolve_column_in_df
                for sap in ('KUNNR', 'Customer'):
                    col = resolve_column_in_df(df, sap, table_name, self.column_map, parent_dir)
                    if col:
                        return col
            except ImportError:
                pass
            for name in ('KUNNR', 'CUSTOMER', 'CUSTOMER_1'):
                if name in col_upper:
                    return col_upper[name]
        if tn == 'BUT0BK':
            for name in ('PARTNER', 'BUSINESS_PARTNER', 'BUSINESS PARTNER'):
                if name in col_upper:
                    return col_upper[name]
        for name in ('PARTNER', 'PARTNERS', 'PARTNER_ID', 'PARTNER_NUM', 'BP', 'CUSTOMER', 'KUNNR', 'PARTNER_CODE', 'CUSTOMER_ID', 'BP_NUMBER'):
            if name in col_upper:
                return col_upper[name]
        skip_client = tn in self.KNA1_JOIN_VIA_BUT020_TABLES
        if tn != 'BUT0BK' and (not skip_client):
            for name in ('CLIENT',):
                if name in col_upper:
                    return col_upper[name]
        for cu, col in col_upper.items():
            if skip_client and self._is_blocked_kna1_join_column(col):
                continue
            if 'PARTNER' in cu or cu.startswith('PARTNER') or ('CLIENT' in cu and not skip_client) or ('KUNNR' in cu) or (cu == 'BP'):
                return col
        for col in cols:
            cu = str(col).upper()
            if skip_client and self._is_blocked_kna1_join_column(col):
                continue
            if 'PARTNER' in cu or 'CUSTOMER' in cu or 'KUNNR' in cu or ('BP' in cu) or ('CUST' in cu) or ('CLIENT' in cu and not skip_client):
                return col
        if table_name and table_name in getattr(self, 'TABLE_UNIQUE_PARTNER', ()) and (len(cols) > 0):
            try:
                best_col = None
                best_nunique = 0
                for c in cols:
                    n = df[c].nunique()
                    if n > best_nunique and n <= len(df) and (n > 1):
                        best_nunique = n
                        best_col = c
                if best_col:
                    return best_col
            except Exception:
                pass
        return None

    def _scope_but0bk_to_kna1_partners(self, df, table_name, rule_code):
        try:
            if df is None or df.empty:
                return df
            partner_col = self._resolve_but0bk_partner_column(df, table_name=table_name)
            if not partner_col:
                print(f'      [WARN] {rule_code}: в {table_name} не найдена колонка партнёра (PARTNER/Business_Partner)')
                return df.iloc[0:0].copy()
            try:
                kna1_df = self._get_table_for_rules('KNA1')
            except Exception:
                kna1_df = None
            if kna1_df is None or kna1_df.empty:
                try:
                    if hasattr(self.memory_manager, 'load_selected_tables_to_ram'):
                        self.memory_manager.load_selected_tables_to_ram(['KNA1'], add_reference_tables=False)
                    kna1_df = self._get_table_for_rules('KNA1')
                except Exception:
                    kna1_df = None
            if kna1_df is None or kna1_df.empty:
                try:
                    conn = connect_sqlite(self.db_path)
                    try:
                        kna1_df = pd.read_sql_query('SELECT "Customer" AS "KUNNR" FROM "KNA1"', conn)
                    except Exception:
                        try:
                            kna1_df = pd.read_sql_query('SELECT "KUNNR" FROM "KNA1"', conn)
                        except Exception:
                            kna1_df = pd.read_sql_query('SELECT * FROM "KNA1"', conn)
                    conn.close()
                    kna1_df = self._apply_rule_time_column_map(kna1_df, 'KNA1')
                except Exception as e:
                    print(f'      [WARN] {rule_code}: не удалось загрузить KNA1 для scope: {e}')
                    return df.iloc[0:0].copy()
            if kna1_df is None or kna1_df.empty:
                print(f'      [WARN] {rule_code}: KNA1 пуста, scope невозможен')
                return df.iloc[0:0].copy()
            kna1_kunnr_col = self._resolve_kna1_kunnr_column(kna1_df)
            if not kna1_kunnr_col:
                print(f'      [WARN] {rule_code}: в KNA1 не найдена колонка KUNNR/Customer для scope')
                return df.iloc[0:0].copy()
            left = df[partner_col].apply(self._norm_customer_partner_key)
            right = kna1_df[kna1_kunnr_col].apply(self._norm_customer_partner_key)
            kna1_keys = set(right[right != ''].unique().tolist())
            if not kna1_keys:
                return df.iloc[0:0].copy()
            mask = left.isin(kna1_keys)
            matched = int(mask.sum())
            print(f'      [JOIN] {rule_code}: {table_name}.{partner_col} -> KNA1.{kna1_kunnr_col}, совпало строк: {matched:,}')
            return df[mask].copy()
        except Exception as e:
            print(f'      [WARN] {rule_code}: ошибка scope BUT0BK по KNA1 партнёрам: {e}')
            return df.iloc[0:0].copy()

    def _get_reference_table_for_rule(self, rule_code, config_key):
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) if __file__ else os.getcwd()
        for path in [os.path.join(root, 'json files', 'conf_sales_group_office.json'), os.path.join(os.getcwd(), 'json files', 'conf_sales_group_office.json'), os.path.join(root, 'config', 'conf_sales_group_office.json')]:
            path = os.path.abspath(path)
            if os.path.isfile(path):
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        cfg = json.load(f)
                    name = cfg.get(config_key) or cfg.get('reference_table')
                    if name and isinstance(name, str) and name.strip():
                        return name.strip()
                except Exception:
                    pass
        return None

    def _load_allowed_vkgrp_vkbur_from_json(self):
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) if __file__ else os.getcwd()
        for path in [os.path.join(root, 'json files', 'conf_sales_group_office.json'), os.path.join(os.getcwd(), 'json files', 'conf_sales_group_office.json'), os.path.join(root, 'config', 'conf_sales_group_office.json')]:
            path = os.path.abspath(path)
            if os.path.isfile(path):
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        cfg = json.load(f)
                    arr = cfg.get('allowed_combinations') or cfg.get('combinations')
                    if not arr or not isinstance(arr, list):
                        return None
                    result = set()
                    for item in arr:
                        if isinstance(item, dict):
                            vg = item.get('VKGRP') or item.get('vkgrp') or item.get('sales_group_code') or item.get('sales_group')
                            vb = item.get('VKBUR') or item.get('vkbur') or item.get('sales_office_code') or item.get('sales_office')
                            if vg is not None and vb is not None:
                                result.add((str(vg).strip(), str(vb).strip()))
                        elif isinstance(item, (list, tuple)) and len(item) >= 2:
                            result.add((str(item[0]).strip(), str(item[1]).strip()))
                    return result if result else None
                except Exception:
                    pass
        return None

    def _resolve_vkgrp_vkbur_columns(self, df, table_name):
        if df is None or getattr(df, 'empty', True):
            return (None, None)
        for tbl in (table_name, 'KNVV', 'TVBVK'):
            vkgrp_col = self._resolve_column_for_rule(df, 'VKGRP', tbl)
            vkbur_col = self._resolve_column_for_rule(df, 'VKBUR', tbl)
            if vkgrp_col and vkbur_col and vkgrp_col in df.columns and vkbur_col in df.columns:
                return (vkgrp_col, vkbur_col)
        cols = {str(c).strip(): c for c in df.columns}
        vkgrp_col = cols.get('SGrp') or cols.get('VKGRP')
        vkbur_col = cols.get('SOff_') or cols.get('VKBUR')
        return (vkgrp_col, vkbur_col)

    def _load_allowed_vkgrp_vkbur_pairs(self, ref_table_name, norm_fn):
        ref_table_name = (ref_table_name or 'TVBVK').strip()
        ref_df = self.memory_manager.get_table(ref_table_name)
        if (ref_df is None or ref_df.empty) and getattr(self, 'db_path', None):
            try:
                self.memory_manager.load_selected_tables_to_ram([ref_table_name], add_reference_tables=False)
                ref_df = self.memory_manager.get_table(ref_table_name)
            except Exception:
                ref_df = self.memory_manager.get_table(ref_table_name)
        if ref_df is not None and not ref_df.empty:
            ref_filtered = ref_df.copy()
            spras_col = next((c for c in ref_filtered.columns if str(c).strip().upper() == 'SPRAS'), None)
            if spras_col:
                ref_filtered = ref_filtered[ref_filtered[spras_col].astype(str).str.strip().str.upper() == 'E'].copy()
                if self.debug:
                    print(f"      [REF] {ref_table_name}: фильтр SPRAS='E' -> {len(ref_filtered):,} строк")
            vkgrp_col, vkbur_col = self._resolve_vkgrp_vkbur_columns(ref_filtered, ref_table_name)
            if vkgrp_col and vkbur_col:
                allowed_pairs = set()
                for _, row in ref_filtered[[vkgrp_col, vkbur_col]].dropna().iterrows():
                    vg = norm_fn(row[vkgrp_col])
                    vb = norm_fn(row[vkbur_col])
                    if vg and vb:
                        allowed_pairs.add((vg, vb))
                if allowed_pairs:
                    return (allowed_pairs, ref_table_name)
        json_pairs = self._load_allowed_vkgrp_vkbur_from_json()
        if json_pairs:
            normalized = set()
            for a, b in json_pairs:
                vg = norm_fn(a)
                vb = norm_fn(b)
                if vg and vb:
                    normalized.add((vg, vb))
            if normalized:
                return (normalized, 'conf_sales_group_office.json')
        return (None, None)

    def _load_vkorg_cluster_scope_from_json(self):
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) if __file__ else os.getcwd()
        for path in [os.path.join(root, 'json files', 'conf_sales_group_office.json'), os.path.join(os.getcwd(), 'json files', 'conf_sales_group_office.json'), os.path.join(root, 'config', 'conf_sales_group_office.json')]:
            path = os.path.abspath(path)
            if os.path.isfile(path):
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        cfg = json.load(f)
                    arr = cfg.get('vkorg_cluster_scope') or cfg.get('allowed_vkorg_cluster') or cfg.get('vkorg_cluster_combinations')
                    if not arr or not isinstance(arr, list):
                        return None
                    out = set()
                    for item in arr:
                        if not isinstance(item, dict):
                            continue
                        vg = item.get('VKORG') or item.get('vkorg') or item.get('sales_org')
                        cl = item.get('CLUSTER') or item.get('cluster') or item.get('KVGR4') or item.get('customer_group_4_code')
                        vg = self._norm_lookup_value(vg).upper() if vg is not None else ''
                        cl = self._norm_lookup_value(cl).upper() if cl is not None else ''
                        if vg and cl:
                            out.add((vg, cl))
                    return out if out else None
                except Exception:
                    pass
        return None

    def _norm_lookup_value(self, value):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return ''
        s = str(value).replace('\xa0', ' ').replace('\ufeff', '').strip().strip("'").strip('"').strip()
        if s.lower() in {'', 'none', 'null', 'nan', '<na>', 'nat', '-', '.', 'n/a', 'na'}:
            return ''
        if s.endswith('.0'):
            s = re.sub('\\.0+$', '', s)
        return s

    def _load_planning_group_matrix(self):
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) if __file__ else os.getcwd()
        for path in [os.path.join(root, 'json files', 'conf_planning_group_matrix.json'), os.path.join(os.getcwd(), 'json files', 'conf_planning_group_matrix.json'), os.path.join(root, 'config', 'conf_planning_group_matrix.json')]:
            path = os.path.abspath(path)
            if os.path.isfile(path):
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        cfg = json.load(f)
                    rules = cfg.get('rules') if isinstance(cfg, dict) else None
                    if isinstance(rules, list) and rules:
                        return (rules, path)
                except Exception:
                    pass
        return (None, None)

    def _add_customer_group_4_from_knvv(self, df, table_name, rule_code):
        if df is None or df.empty:
            return df
        try:
            existing = next((c for c in df.columns if str(c).strip().lower() in ('customer_group_4_code', 'kvgr4')), None)
            if existing:
                return df
            kunnr_col = self._find_kunnr_column(df) or next((c for c in df.columns if str(c).strip().upper() == 'KUNNR'), None)
            if not kunnr_col:
                print(f'      [WARN] Колонка KUNNR не найдена в {table_name} для правила {rule_code}')
                return df
            knvv_df = self.memory_manager.get_table('KNVV')
            if (knvv_df is None or knvv_df.empty) and getattr(self, 'db_path', None):
                try:
                    self.memory_manager.load_selected_tables_to_ram(['KNVV'], add_reference_tables=False)
                    knvv_df = self.memory_manager.get_table('KNVV')
                except Exception:
                    knvv_df = self.memory_manager.get_table('KNVV')
            if knvv_df is None or knvv_df.empty:
                print(f'      [WARN] Таблица KNVV не найдена или пуста для правила {rule_code}')
                return df
            knvv_kunnr_col = next((c for c in knvv_df.columns if str(c).strip().upper() == 'KUNNR'), None) or self._find_kunnr_column(knvv_df)
            kvgr4_col = next((c for c in knvv_df.columns if str(c).strip().upper() in ('KVGR4', 'CUSTOMER_GROUP_4_CODE')), None) or self._find_column_alternative(knvv_df.columns, 'KVGR4', 'KNVV')
            if not knvv_kunnr_col or not kvgr4_col:
                print(f'      [WARN] В KNVV не найдены KUNNR/KVGR4 для правила {rule_code}')
                return df

            def _norm_kunnr(series):
                s = series.astype(str).str.strip()
                s = s.str.replace('\\.0$', '', regex=True)
                s = s.str.replace('\\D+', '', regex=True)
                return s.str.zfill(10)
            knvv_join = knvv_df[[knvv_kunnr_col, kvgr4_col]].copy()
            kvgr4_norm = knvv_join[kvgr4_col].apply(self._norm_lookup_value)
            knvv_join = knvv_join[kvgr4_norm != ''].copy()
            if knvv_join.empty:
                print(f'      [WARN] В KNVV нет заполненных KVGR4 для правила {rule_code}')
                return df
            knvv_join['_kunnr_key'] = _norm_kunnr(knvv_join[knvv_kunnr_col])
            knvv_join['customer_group_4_code'] = knvv_join[kvgr4_col].apply(self._norm_lookup_value)
            knvv_join = knvv_join.drop_duplicates(subset=['_kunnr_key'], keep='first')
            df_joined = df.copy()
            df_joined['_kunnr_key'] = _norm_kunnr(df_joined[kunnr_col])
            df_joined = df_joined.merge(knvv_join[['_kunnr_key', 'customer_group_4_code']], on='_kunnr_key', how='left').drop(columns=['_kunnr_key'], errors='ignore')
            filled = int(df_joined['customer_group_4_code'].apply(lambda v: self._norm_lookup_value(v) != '').sum())
            print(f'      [JOIN] Добавлен customer_group_4_code из KNVV (KVGR4) для правила {rule_code}. Строк: {len(df_joined)} (было {len(df)}), заполнено customer_group_4_code: {filled} ({filled / max(len(df_joined), 1) * 100:.1f}%)')
            return df_joined
        except Exception as e:
            print(f'      [WARN] Ошибка добавления customer_group_4_code из KNVV для {rule_code}: {e}')
            return df

    def _load_partner_column_config(self, table_name):
        if table_name not in getattr(self, 'TABLE_UNIQUE_PARTNER', ()):
            return None
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) if __file__ else os.getcwd()
        for path in [os.path.join(root, 'json files', 'conf_zbut0000p_partner.json'), os.path.join(os.getcwd(), 'json files', 'conf_zbut0000p_partner.json'), os.path.join(root, 'config', 'conf_zbut0000p_partner.json')]:
            path = os.path.abspath(path)
            if os.path.isfile(path):
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        cfg = json.load(f)
                    col = cfg.get('partner_column') or cfg.get('partner_column_name')
                    if col and isinstance(col, str) and col.strip():
                        return col.strip()
                except Exception:
                    pass
        return None

    def _apply_unique_partner_counts_if_needed(self, table_name, df, error_df, total_rows, error_count):
        tbl = str(table_name or '').strip().upper()
        if tbl not in self.TABLE_UNIQUE_PARTNER and tbl != 'ADR2':
            return (total_rows, error_count)
        partner_col = self._find_partner_column(df, table_name=table_name)
        if not partner_col and tbl == 'ADR2':
            df_with_partner = self._ensure_adr2_has_partner(df, '')
            partner_col = self._find_partner_column(df_with_partner, table_name='ADR2')
        if not partner_col:
            if not hasattr(self, '_partner_warned_tables'):
                self._partner_warned_tables = set()
            if table_name not in self._partner_warned_tables:
                self._partner_warned_tables.add(table_name)
                print(f'      [WARN] [{table_name}] Колонка PARTNER не найдена — подсчёт без удаления дублей (будут миллионы строк). Укажите имя колонки в conf_zbut0000p_partner.json (partner_column). Колонки: {list(df.columns)[:30]}')
            return (total_rows, error_count)
        total_rows = int(df[partner_col].nunique())
        if error_df is None or error_df.empty:
            error_count = 0
        elif partner_col in error_df.columns:
            error_count = int(error_df[partner_col].nunique())
        else:
            try:
                common = error_df.index.intersection(df.index)
                if len(common) >= len(error_df) * 0.99:
                    error_count = int(df.loc[error_df.index, partner_col].nunique())
                else:
                    error_count = int(error_count)
            except Exception:
                error_count = int(error_count)
        print(f'      [{table_name}] Подсчёт по уникальным {partner_col} (дубли исключены): всего {total_rows:,} клиентов, с ошибками: {error_count:,}')
        return (total_rows, error_count)

    def _find_column_alternative(self, columns, column_name, table_name):
        try:
            from utils.column_map_resolver import resolve_column, map_logical_to_sap
            sap = map_logical_to_sap(table_name, column_name, self.column_map, parent_dir)
            for target in (sap, column_name):
                found = resolve_column(columns, target, table_name, self.column_map, parent_dir)
                if found:
                    return found
        except ImportError:
            pass
        return None

    def _find_most_similar_column(self, columns, target_column):
        target_upper = target_column.upper().replace('_', '')
        best_match = None
        best_score = 0
        for col in columns:
            col_upper = col.upper().replace('_', '')
            if len(col_upper) < 4:
                continue
            score = 0
            if col_upper == target_upper:
                score += 100
            if len(target_upper) >= 4 and (target_upper in col_upper or col_upper in target_upper):
                score += 50
            common = set(target_upper) & set(col_upper)
            if len(common) > 0:
                score += len(common) * 2
            if score > best_score:
                best_score = score
                best_match = col
        return best_match if best_score > 10 else None

    def _extract_second_column_from_description(self, rule_code, rule_description, columns, first_column, table_name=None):
        desc_lower = rule_description.lower()
        patterns = {'name 1': ['name_org1', 'name1', 'name_1', 'organization_1_name'], 'name 2': ['name_org2', 'name2', 'name_2', 'organization_2_name'], 'name 3': ['name_org3', 'name3', 'name_3', 'organization_3_name'], 'name 4': ['name_org4', 'name4', 'name_4', 'organization_4_name'], 'tax 1': ['taxnum1', 'taxnum', 'tax_1_value'], 'tax 2': ['taxnum2', 'taxnum', 'tax_2_value'], 'tax 3': ['taxnum3', 'taxnum', 'tax_3_value'], 'tax 4': ['taxnum4', 'taxnum', 'tax_4_value'], 'tax 5': ['taxnum5', 'taxnum', 'tax_5_value'], 'tax 6': ['taxnum6', 'taxnum', 'tax_6_value']}
        second_column_candidate = None
        for key, variations in patterns.items():
            for variation in variations:
                if variation in desc_lower:
                    if variation not in first_column.lower():
                        second_column_candidate = variation
                        break
            if second_column_candidate:
                break
        if not second_column_candidate:
            if 'Tax Number 1 и Tax Number 2' in rule_description:
                if 'TAXNUM' in first_column.upper():
                    second_column_candidate = 'TAXNUM1' if 'TAXNUM2' in first_column.upper() else 'TAXNUM2'
            elif 'Name 3 и Name 4' in rule_description:
                if 'NAME' in first_column.upper():
                    second_column_candidate = 'NAME4' if 'NAME3' in first_column.upper() else 'NAME3'
            elif 'Street и House Number' in rule_description:
                if 'STRAS' in first_column.upper():
                    second_column_candidate = 'ORT01'
                elif 'ORT01' in first_column.upper():
                    second_column_candidate = 'STRAS'
        if not second_column_candidate:
            return None
        if table_name:
            mapped_second = self._get_mapped_column_name(table_name, second_column_candidate)
            second_column_candidate = mapped_second
        for col in columns:
            if col.upper() == second_column_candidate.upper():
                return col
        matched = self._find_column_alternative(columns, second_column_candidate, table_name)
        if matched:
            return matched
        return None

    def _apply_conditional_filter(self, df, technical_def, rule_code, table_name=None):
        try:
            print(f'      [FILTER] Анализ условий для {rule_code}...')
            if rule_code in ['RCCONF_39.5.2', 'RCCONF_39.3.2']:
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or ('persnumber' in col_lower):
                        person_col = col
                        print(f'      [FILTER] Найдена колонка PERSNUMBER: {col}')
                        break
                if person_col:
                    mask = df[person_col].notna() & (df[person_col].astype(str).str.strip() != '') & (df[person_col].astype(str).str.strip().str.lower() != 'none') & (df[person_col].astype(str).str.strip().str.lower() != 'null')
                    filtered_df = df[mask].copy()
                    print(f'      [FILTER] Применен фильтр PERSNUMBER IS NOT NULL (заполнено) для {rule_code}: {len(filtered_df)} из {len(df)} строк')
                    if len(filtered_df) == 0:
                        print(f'      [WARN] После фильтрации (PERSNUMBER заполнено) данных нет!')
                    return filtered_df
                else:
                    print(f'      [WARN] Колонка PERSNUMBER не найдена для правила {rule_code}')
            if rule_code == 'RCCONF_39.5' and table_name and (str(table_name).strip().upper() == 'ADR2'):
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or ('persnumber' in col_lower):
                        person_col = col
                        break
                if person_col:
                    mask = df[person_col].isna() | (df[person_col].astype(str).str.strip() == '') | (df[person_col].astype(str).str.strip().str.lower() == 'none')
                    df = df[mask].copy()
                    print(f'      [FILTER] Применен фильтр PERSNUMBER IS NULL для {rule_code}: {len(df)} строк')
                if not df.empty:
                    return df
                return df
            if rule_code == 'RCCONF_39.3':
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower in ('persnumber', 'pers_number') or col_lower == 'person' or 'persnumber' in col_lower:
                        person_col = col
                        print(f'      [FILTER] Найдена колонка PERSNUMBER: {col}')
                        break
                if person_col:
                    pers_str = df[person_col].astype(str).str.strip()
                    pers_low = pers_str.str.lower()
                    pers_str_for_zero = pers_str.str.replace(',', '.', regex=False)
                    pers_is_zeroish = pers_str_for_zero.str.match('^-?0+(?:[.][0]+)?$', na=False)
                    mask = df[person_col].isna() | (pers_str == '') | pers_low.isin(['none', 'null', 'nan']) | pers_is_zeroish
                    filtered_df = df[mask].copy()
                    print(f'      [FILTER] RCCONF_39.3: PERSNUMBER IS NULL — {len(filtered_df)} из {len(df)} строк')
                    return filtered_df
                print(f'      [WARN] RCCONF_39.3: колонка PERSNUMBER не найдена')
                return df
            if rule_code == 'RCCONF_38.3':
                r3_user_col = None
                person_col = None
                contact_medium_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'r3_user' or col_lower == 'r3user' or 'r3_user' in col_lower or ('r3user' in col_lower):
                        r3_user_col = col
                        print(f'      [FILTER] Найдена колонка R3_USER: {col}')
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or ('persnumber' in col_lower):
                        person_col = col
                        print(f'      [FILTER] Найдена колонка PERSNUMBER: {col}')
                    if 'contact_medium' in col_lower or 'medium_type' in col_lower or 'contactmedium' in col_lower:
                        contact_medium_col = col
                        print(f'      [FILTER] Найдена колонка contact_medium_type: {col}')
                if r3_user_col and person_col:
                    mask = (df[r3_user_col].astype(str).str.strip() == '1') & (df[person_col].isna() | (df[person_col].astype(str).str.strip() == '') | (df[person_col].astype(str).str.strip().str.lower() == 'none'))
                    filtered_df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр (R3_USER='1' AND PERSNUMBER IS NULL) для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                    if len(filtered_df) == 0:
                        print(f'      [WARN] После фильтрации данных нет!')
                    return filtered_df
                elif r3_user_col:
                    mask = df[r3_user_col].astype(str).str.strip() == '1'
                    filtered_df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр (только R3_USER='1') для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                    return filtered_df
                elif contact_medium_col and person_col:
                    mask = (df[contact_medium_col].astype(str).str.strip().str.lower() == 'fixed_tel_number') & (df[person_col].isna() | (df[person_col].astype(str).str.strip() == '') | (df[person_col].astype(str).str.strip().str.lower() == 'none'))
                    filtered_df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр (contact_medium_type='fixed_tel_number' AND PERSNUMBER IS NULL) для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                    if len(filtered_df) == 0:
                        print(f'      [WARN] После фильтрации данных нет!')
                    return filtered_df
                elif contact_medium_col:
                    mask = df[contact_medium_col].astype(str).str.strip().str.lower() == 'fixed_tel_number'
                    filtered_df = df[mask].copy()
                    print(f"      [FILTER] Применен фильтр (только contact_medium_type='fixed_tel_number') для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                    return filtered_df
                elif person_col:
                    mask = df[person_col].isna() | (df[person_col].astype(str).str.strip() == '') | df[person_col].astype(str).str.strip().str.lower().isin(('none', 'null', 'nan'))
                    filtered_df = df[mask].copy()
                    print(f'      [FILTER] RCCONF_38.3: PERSNUMBER IS NULL (fallback без R3_USER) — {len(filtered_df)} из {len(df)} строк')
                    return filtered_df
                else:
                    print(f'      [WARN] RCCONF_38.3: колонки R3_USER / PERSNUMBER / contact_medium_type не найдены')
            elif rule_code == 'RCCONF_38.5':
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or ('persnumber' in col_lower):
                        person_col = col
                        print(f'      [FILTER] Найдена колонка PERSNUMBER: {col}')
                        break
                if person_col:
                    mask = df[person_col].isna() | (df[person_col].astype(str).str.strip() == '') | (df[person_col].astype(str).str.strip().str.lower() == 'none')
                    filtered_df = df[mask].copy()
                    print(f'      [FILTER] Применен фильтр PERSNUMBER IS NULL для {rule_code}: {len(filtered_df)} из {len(df)} строк')
                    if len(filtered_df) == 0:
                        print(f'      [WARN] После фильтрации (PERSNUMBER IS NULL) данных нет!')
                    return filtered_df
                else:
                    print(f'      [WARN] Колонка PERSNUMBER не найдена для правила {rule_code}')
            if rule_code in ['RCCOMP_369.1', 'RCCONF_369.1']:
                source_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'source' or col_lower == 'source_file' or 'source' in col_lower:
                        source_col = col
                        print(f'      [FILTER] Найдена колонка source: {col}')
                        break
                if rule_code == 'RCCOMP_369.1':
                    if source_col:
                        mask = df[source_col].astype(str).str.strip().str.lower() == 's4'
                        filtered_df = df[mask].copy()
                        print(f"      [FILTER] Применен фильтр (source='s4') для {rule_code}: {len(filtered_df)} из {len(df)} строк")
                        if len(filtered_df) == 0:
                            print(f"      [WARN] После фильтрации (source='s4') данных нет!")
                        return filtered_df
                    else:
                        print(f'      [WARN] Колонка source не найдена для правила {rule_code}')
                        return df
                elif rule_code == 'RCCONF_369.1':
                    print(f'      [FILTER] Правило {rule_code} не требует фильтрации по source, работаем со всеми данными')
                    return df
            elif rule_code == 'RCCOMP_375.1':
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or ('persnumber' in col_lower):
                        person_col = col
                        break
                if person_col:
                    pers_str = df[person_col].astype(str).str.strip()
                    pers_low = pers_str.str.lower()
                    pers_str_for_zero = pers_str.str.replace(',', '.', regex=False)
                    pers_is_zeroish = pers_str_for_zero.str.match('^-?0+(?:[.][0]+)?$', na=False)
                    mask = df[person_col].isna() | (pers_str == '') | pers_low.isin(['none', 'null']) | pers_is_zeroish
                    df = df[mask].copy()
                    print(f'      [FILTER] Применен фильтр PERSNUMBER IS NULL для {rule_code}: {len(df)} строк')
                    if len(df) == 0:
                        return df
                return self._filter_adr2_rccomp_375_1_scope_by_kna1_aufsd(df, rule_code, table_name)
            elif rule_code == 'RCCOMP_375.1.2':
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or ('persnumber' in col_lower):
                        person_col = col
                        break
                if person_col:
                    pers_str = df[person_col].astype(str).str.strip()
                    pers_low = pers_str.str.lower()
                    pers_is_zeroish = pers_str.str.match('^0+(\\.0+)?$', na=False)
                    mask = df[person_col].notna() & (pers_str != '') & ~pers_low.isin(['none', 'null']) & ~pers_is_zeroish
                    df = df[mask].copy()
                    print(f'      [FILTER] Применен фильтр PERSNUMBER IS NOT NULL для {rule_code}: {len(df)} строк')
                    if len(df) == 0:
                        return df
                return df
            elif rule_code == 'RCCONF_39.5':
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or ('persnumber' in col_lower):
                        person_col = col
                        break
                if person_col:
                    mask = df[person_col].isna() | (df[person_col].astype(str).str.strip() == '') | (df[person_col].astype(str).str.strip().str.lower() == 'none')
                    df = df[mask].copy()
                    print(f'      [FILTER] Применен фильтр PERSNUMBER IS NULL для {rule_code}: {len(df)} строк')
                    if len(df) == 0:
                        return df
                return df
            elif rule_code in ['RCCONF_39.5.2', 'RCCONF_39.3.2']:
                person_col = None
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'persnumber' or col_lower == 'pers_number' or col_lower == 'person' or ('persnumber' in col_lower):
                        person_col = col
                        break
                if person_col:
                    mask = df[person_col].notna() & (df[person_col].astype(str).str.strip() != '') & (df[person_col].astype(str).str.strip().str.lower() != 'none') & (df[person_col].astype(str).str.strip().str.lower() != 'null')
                    df = df[mask].copy()
                    print(f'      [FILTER] Применен фильтр PERSNUMBER IS NOT NULL для {rule_code}: {len(df)} строк')
                    if len(df) == 0:
                        return df
                return df
            return df
        except Exception as e:
            print(f'      [WARN] Ошибка в _apply_conditional_filter для {rule_code}: {e}')
            import traceback
            traceback.print_exc()
            return df

    def _get_adr6_df(self):
        try:
            adr6 = self.memory_manager.get_table('ADR6')
            if (adr6 is None or adr6.empty) and getattr(self, 'db_path', None):
                conn = connect_sqlite(self.db_path)
                tables = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", conn)
                adr6_name = next((r[0] for r in tables.values if str(r[0]).strip().upper() == 'ADR6'), None)
                if adr6_name:
                    adr6 = pd.read_sql_query(f'SELECT * FROM "{adr6_name}"', conn)
                conn.close()
            if adr6 is None or adr6.empty:
                return None
            return self._apply_rule_time_column_map(adr6.copy(), 'ADR6')
        except Exception as e:
            print(f'      [WARN] _get_adr6_df: {e}')
            return None

    def _get_but020_table_for_join(self):
        but020 = self.memory_manager.get_table('BUT020')
        if (but020 is None or but020.empty) and getattr(self, 'db_path', None):
            try:
                conn = connect_sqlite(self.db_path)
                tables = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", conn)
                but_name = next((r[0] for r in tables.values if str(r[0]).strip().upper() == 'BUT020'), None)
                if but_name:
                    but020 = pd.read_sql_query(f'SELECT * FROM "{but_name}"', conn)
                conn.close()
            except Exception:
                but020 = None
        if but020 is None or but020.empty:
            try:
                self.memory_manager.load_selected_tables_to_ram(['BUT020'], add_reference_tables=False)
                but020 = self.memory_manager.get_table('BUT020')
            except Exception:
                but020 = None
        if but020 is None or but020.empty:
            return None
        return self._apply_rule_time_column_map(but020.copy(), 'BUT020')

    def _resolve_addrnumber_column(self, df, table_name='ADRC'):
        if df is None or df.empty:
            return None
        col = self._resolve_column_for_rule(df, 'ADDRNUMBER', table_name)
        if col:
            return col
        for c in df.columns:
            cu = str(c).strip().upper().replace(' ', '').replace('_', '')
            if cu in ('ADDRNUMBER', 'ADRNR', 'ADDRNRADRC', 'ADDRNO') or ('ADDR' in cu and 'NO' in cu):
                return c
        return next((c for c in df.columns if 'ADDRNUMBER' in str(c).upper()), None)

    def _resolve_but020_join_columns(self, but020_df):
        if but020_df is None or but020_df.empty:
            return (None, None)
        addr_col = self._resolve_column_for_rule(but020_df, 'ADDRNUMBER', 'BUT020')
        if not addr_col:
            addr_col = next((c for c in but020_df.columns if 'ADDRNUMBER' in str(c).upper() or str(c).upper() == 'ADRNR'), None)
        if not addr_col:
            addr_col = next((c for c in but020_df.columns if 'ADDR' in str(c).upper() and ('NO' in str(c).upper() or 'NR' in str(c).upper())), None)
        partner_col = self._resolve_column_for_rule(but020_df, 'PARTNER', 'BUT020')
        if not partner_col:
            partner_col = next((c for c in but020_df.columns if str(c).strip().upper() == 'PARTNER'), None)
        if not partner_col:
            partner_col = next((c for c in but020_df.columns if 'BUSINESS' in str(c).upper() and 'PARTNER' in str(c).upper()), None)
        return (addr_col, partner_col)

    def _merge_adrc_partner_from_but020(self, df, table_name, rule_code=None):
        """RCCONF_24.1: ADRC Addr. No. = BUT020 Addr. No. -> Business Partner (KUNNR/Customer)."""
        if df is None or df.empty:
            return (df, None)
        out = self._apply_rule_time_column_map(df.copy(), table_name or 'ADRC')
        addr_col = self._resolve_addrnumber_column(out, table_name or 'ADRC')
        if not addr_col:
            print(f'      [WARN] {rule_code or table_name}: колонка Addr. No. / ADDRNUMBER не найдена в ADRC (колонки: {[c for c in out.columns if "ADDR" in str(c).upper()][:8]})')
            return (out, None)
        but020_df = self._get_but020_table_for_join()
        if but020_df is None or but020_df.empty:
            print(f'      [WARN] {rule_code or table_name}: таблица BUT020 не найдена или пуста')
            return (out, None)
        addr_but, partner_but = self._resolve_but020_join_columns(but020_df)
        if not addr_but or not partner_but:
            print(f'      [WARN] {rule_code or table_name}: в BUT020 не найдены Addr. No. и Business Partner (колонки: {list(but020_df.columns)[:10]})')
            return (out, None)
        but_join = but020_df[[addr_but, partner_but]].copy()

        def _norm_addr_key(series: pd.Series) -> pd.Series:
            s = series.astype(str).str.strip()
            s = s.str.replace('\\.0$', '', regex=True)
            s = s.str.replace('\\D+', '', regex=True)
            return s.str.zfill(10)
        out['_addr_key_norm'] = _norm_addr_key(out[addr_col])
        but_join['_addr_key_norm'] = _norm_addr_key(but_join[addr_but])
        but_join = but_join.drop_duplicates(subset=['_addr_key_norm'], keep='first')
        out = out.merge(but_join[['_addr_key_norm', partner_but]], on='_addr_key_norm', how='left')
        out = out.drop(columns=['_addr_key_norm'], errors='ignore')
        if addr_but in out.columns and addr_but != addr_col:
            out = out.drop(columns=[addr_but], errors='ignore')
        if partner_but in out.columns and partner_but != 'PARTNER':
            out = out.rename(columns={partner_but: 'PARTNER'})
        join_col = 'PARTNER' if 'PARTNER' in out.columns else partner_but
        if join_col and join_col in out.columns:
            pf = int(out[join_col].astype(str).str.strip().ne('').sum())
            print(f'      [JOIN] {rule_code or table_name}: ADRC.[{addr_col}] = BUT020.[{addr_but}] -> Business Partner [{partner_but}]: заполнено {pf:,}/{len(out):,}')
        return (out, join_col)

    def _attach_partner_from_but020_by_addr(self, df, addr_col, log_prefix=''):
        """Подтянуть PARTNER из BUT020 по ADDRNUMBER (с маппингом Addr__No_/Business_Partner)."""
        if df is None or df.empty or not addr_col or addr_col not in df.columns:
            return df
        but020_df = self._get_but020_table_for_join()
        if but020_df is None or but020_df.empty:
            return df
        addr_but, partner_but = self._resolve_but020_join_columns(but020_df)
        if not addr_but or not partner_but:
            return df
        j = lambda x: str(x).strip().lstrip('0') or '0'
        out = df.copy()
        out['_ak'] = out[addr_col].apply(j)
        but_join = but020_df[[addr_but, partner_but]].copy()
        but_join['_ak'] = but_join[addr_but].apply(j)
        but_join = but_join.drop_duplicates('_ak', keep='first')
        if partner_but != 'PARTNER':
            but_join = but_join.rename(columns={partner_but: 'PARTNER'})
        out = out.merge(but_join[['_ak', 'PARTNER']], on='_ak', how='left')
        out = out.drop(columns=['_ak'], errors='ignore')
        if 'PARTNER_y' in out.columns:
            out['PARTNER'] = out['PARTNER_y']
            out = out.drop(columns=['PARTNER_x', 'PARTNER_y'], errors='ignore')
        if log_prefix and 'PARTNER' in out.columns:
            print(f'{log_prefix} PARTNER из BUT020: {out["PARTNER"].notna().sum():,} из {len(out):,}')
        return out

    def _ensure_adr2_has_partner(self, df, rule_code):
        if df is None or df.empty:
            return df
        addr_col = self._resolve_addrnumber_column(df, 'ADR2') or next((c for c in df.columns if 'ADDRNUMBER' in str(c).upper()), None)
        partner_col = next((c for c in df.columns if str(c).upper() == 'PARTNER'), None)
        if not addr_col or partner_col:
            return df
        try:
            but020 = self._get_but020_table_for_join()
            if but020 is None or but020.empty:
                return df
            addr_but, partner_but = self._resolve_but020_join_columns(but020)
            if not addr_but or not partner_but:
                return df
            j = lambda x: str(x).strip().lstrip('0') or '0'
            df = df.copy()
            df['_ak'] = df[addr_col].apply(j)
            but_join = but020[[addr_but, partner_but]].copy()
            but_join['_ak'] = but_join[addr_but].apply(j)
            but_join = but_join.drop_duplicates(subset=['_ak'])
            df = df.merge(but_join[['_ak', partner_but]], on='_ak', how='left')
            df = df.drop(columns=['_ak'], errors='ignore')
            if partner_but not in df.columns:
                return df
            if partner_but != 'PARTNER':
                df = df.rename(columns={partner_but: 'PARTNER'})
            return df
        except Exception as e:
            print(f'      [WARN] _ensure_adr2_has_partner для {rule_code}: {e}')
            return df

    def _build_adr2_rccomp_37512_export_df(self, df_adr2, rule_code='RCCOMP_375.1.2'):
        if df_adr2 is None or df_adr2.empty:
            return None
        try:
            person_col = None
            for col in df_adr2.columns:
                c = str(col).lower()
                if c == 'persnumber' or c == 'pers_number' or 'persnumber' in c:
                    person_col = col
                    break
            if person_col is not None:
                mask = df_adr2[person_col].notna() & (df_adr2[person_col].astype(str).str.strip() != '') & (df_adr2[person_col].astype(str).str.strip().str.lower() != 'none') & (df_adr2[person_col].astype(str).str.strip().str.lower() != 'null')
                df_adr2 = df_adr2[mask].copy()
                if df_adr2.empty:
                    return None
            df = self._filter_adr2_non_blocked_customers(df_adr2, rule_code, 'ADR2')
            if df is None or df.empty:
                return None
            addr_col = next((c for c in df.columns if 'ADDRNUMBER' in str(c).upper()), None)
            partner_col = next((c for c in df.columns if str(c).upper() == 'PARTNER'), None)
            if not addr_col:
                partner_col = next((c for c in df.columns if 'PARTNER' in str(c).upper()), None)
            aufsd_col = next((c for c in df.columns if 'AUFSD' in str(c).upper()), None)
            if not addr_col or not partner_col:
                return None
            cols = [addr_col, partner_col]
            if aufsd_col:
                cols.append(aufsd_col)
            out = df[cols].copy()
            out.columns = ['ADDRNUMBER', 'PARTNER', 'AUFSD'] if aufsd_col else ['ADDRNUMBER', 'PARTNER']
            if not aufsd_col:
                out['AUFSD'] = ''
            return out
        except Exception as e:
            print(f'      [WARN] _build_adr2_rccomp_37512_export_df: {e}')
            return None

    def _filter_adr2_non_blocked_customers(self, df, rule_code, table_name=None):
        """ADR2 rules: dm_customer scope (9038 + KNVV 01-01) and optional KNA1 order-block filter."""
        return self._filter_adr2_dm_customer_scope(df, rule_code, table_name)

    def _dedupe_adr2_by_partner(self, df, partner_col=None, log_prefix=''):
        if df is None or df.empty:
            return df
        if not partner_col:
            partner_col = self._find_partner_column(df, table_name='ADR2')
        if not partner_col or partner_col not in df.columns:
            addr_col = self._resolve_addrnumber_column(df, 'ADR2') or next((c for c in df.columns if 'ADDRNUMBER' in str(c).upper()), None)
            if addr_col:
                df = self._attach_partner_from_but020_by_addr(df, addr_col)
                partner_col = self._find_partner_column(df, table_name='ADR2')
        if not partner_col or partner_col not in df.columns:
            return df
        out = df.copy()
        out['_partner_key'] = out[partner_col].apply(self._norm_customer_partner_key)
        out = out[out['_partner_key'] != ''].copy()
        before = len(out)
        out = out.drop_duplicates(subset=['_partner_key'], keep='first')
        out = out.drop(columns=['_partner_key'], errors='ignore')
        dropped = before - len(out)
        if dropped > 0 and log_prefix:
            print(f'{log_prefix}Убраны дубли по PARTNER: {dropped:,} (осталось {len(out):,})')
        return out

    def _load_knvv_for_adr2_join(self):
        knvv = self._get_table_for_rules('KNVV')
        if (knvv is None or knvv.empty) and getattr(self, 'db_path', None):
            try:
                self.memory_manager.load_selected_tables_to_ram(['KNVV'], add_reference_tables=False)
                knvv = self._get_table_for_rules('KNVV')
            except Exception:
                knvv = self._get_table_for_rules('KNVV')
        if (knvv is None or knvv.empty) and getattr(self, 'db_path', None):
            try:
                conn = connect_sqlite(self.db_path)
                tables = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", conn)
                knvv_name = next((r[0] for r in tables.values if str(r[0]).strip().upper() == 'KNVV'), None)
                if knvv_name:
                    knvv = pd.read_sql_query(f'SELECT * FROM "{knvv_name}"', conn)
                conn.close()
            except Exception:
                knvv = None
        if knvv is not None and (not knvv.empty):
            return self._apply_rule_time_column_map(knvv.copy(), 'KNVV')
        return knvv

    def _filter_adr2_dm_customer_scope(self, df, rule_code, table_name=None):
        """ADR2 scope: PARTNER via BUT020, KNA1 Group=9038, KNVV DChl/DV=01-01, dedupe by PARTNER."""
        if table_name and str(table_name).strip().upper() != 'ADR2':
            return df
        if df is None or df.empty:
            return df
        try:
            from utils.sap_account_keys import norm_sap_account_group

            def _norm_key(series: pd.Series) -> pd.Series:
                return series.apply(self._norm_customer_partner_key)
            addr_col = self._resolve_addrnumber_column(df, 'ADR2')
            if not addr_col:
                print(f'      [FILTER] Колонка ADDRNUMBER не найдена в ADR2 для {rule_code}')
                return df
            but020 = self._get_but020_table_for_join()
            if but020 is None or but020.empty:
                print(f'      [WARN] Таблица BUT020 не найдена или пуста для {rule_code}')
                return df.iloc[0:0].copy()
            addr_but, partner_but = self._resolve_but020_join_columns(but020)
            if not addr_but or not partner_but:
                print(f'      [WARN] В BUT020 не найдены Addr. No. и Business Partner для {rule_code}')
                return df.iloc[0:0].copy()
            out = df.copy()
            out['_ak'] = _norm_key(out[addr_col])
            but_join = but020[[addr_but, partner_but]].copy()
            but_join['_ak'] = _norm_key(but_join[addr_but])
            but_join = but_join.drop_duplicates(subset=['_ak'], keep='first')
            out = out.merge(but_join[['_ak', partner_but]], on='_ak', how='left').drop(columns=['_ak'], errors='ignore')
            if partner_but != 'PARTNER' and partner_but in out.columns:
                out = out.rename(columns={partner_but: 'PARTNER'})
            partner_col = 'PARTNER'
            if partner_col not in out.columns:
                print(f'      [WARN] PARTNER не добавлен из BUT020 для {rule_code}')
                return df.iloc[0:0].copy()
            out['_partner_key'] = _norm_key(out[partner_col])
            before_partner = len(out)
            out = out[out['_partner_key'] != ''].copy()
            if len(out) < before_partner:
                print(f'      [FILTER] {rule_code}: без PARTNER из BUT020 исключено {before_partner - len(out):,} строк')
            kna1 = self._get_table_for_rules('KNA1')
            if kna1 is None or kna1.empty:
                print(f'      [WARN] Таблица KNA1 не найдена или пуста для {rule_code}')
                return df.iloc[0:0].copy()
            kna1 = self._apply_rule_time_column_map(kna1.copy(), 'KNA1')
            kunnr_col = self._pick_best_kunnr_column(kna1, 'KNA1')
            ktokd_col = self._resolve_column_for_rule(kna1, 'KTOKD', 'KNA1')
            if not ktokd_col:
                for c in kna1.columns:
                    cl = str(c).strip().lower()
                    if cl in ('group_1', 'group', 'account_group_code', 'ktokd'):
                        if ktokd_col is None or self._non_empty_key_count(kna1[c]) > self._non_empty_key_count(kna1[ktokd_col]):
                            ktokd_col = c
            if not kunnr_col or not ktokd_col:
                print(f'      [WARN] В KNA1 не найдены Customer/KUNNR или Group/KTOKD для {rule_code}')
                return df.iloc[0:0].copy()
            kna1_keys = kna1[[kunnr_col, ktokd_col]].copy()
            kna1_keys['_partner_key'] = _norm_key(kna1_keys[kunnr_col])
            kna1_keys['_ktokd'] = kna1_keys[ktokd_col].apply(norm_sap_account_group)
            kna1_keys = kna1_keys[kna1_keys['_partner_key'] != ''].drop_duplicates(subset=['_partner_key'], keep='first')
            allowed_group = self.ADR2_DM_CUSTOMER_SCOPE_ACCOUNT_GROUP
            keys_9038 = set(kna1_keys.loc[kna1_keys['_ktokd'] == allowed_group, '_partner_key'])
            before_9038 = len(out)
            out = out[out['_partner_key'].isin(keys_9038)].copy()
            print(f"      [FILTER] {rule_code}: KNA1 Group={allowed_group} -> {len(out):,} из {before_9038:,}")
            if out.empty:
                return out
            knvv = self._load_knvv_for_adr2_join()
            if knvv is None or knvv.empty:
                print(f'      [WARN] Таблица KNVV не найдена или пуста для {rule_code}')
                return df.iloc[0:0].copy()
            kunnr_knvv = self._pick_best_kunnr_column(knvv, 'KNVV')
            vtweg_col = self._resolve_column_for_rule(knvv, 'VTWEG', 'KNVV')
            spart_col = self._resolve_column_for_rule(knvv, 'SPART', 'KNVV')
            if not kunnr_knvv or not vtweg_col or not spart_col:
                print(f'      [WARN] В KNVV не найдены Customer, DChl/VTWEG или DV/SPART для {rule_code}')
                return df.iloc[0:0].copy()
            vt = self._norm_knvv_so_code_series(knvv[vtweg_col])
            sp = self._norm_knvv_so_code_series(knvv[spart_col])
            knvv_scoped = knvv.loc[(vt == '01') & (sp == '01')].copy()
            knvv_scoped['_partner_key'] = _norm_key(knvv_scoped[kunnr_knvv])
            keys_so = set(knvv_scoped.loc[knvv_scoped['_partner_key'] != '', '_partner_key'])
            before_so = len(out)
            out = out[out['_partner_key'].isin(keys_so)].copy()
            print(f"      [FILTER] {rule_code}: KNVV DChl/DV=01-01 -> {len(out):,} из {before_so:,}")
            if out.empty:
                return out.drop(columns=['_partner_key'], errors='ignore')
            if str(rule_code or '').strip().upper() in self.ADR2_NON_BLOCKED_MOBILE_RULES:
                aufsd_col = next((c for c in kna1.columns if str(c).upper() == 'AUFSD'), None)
                if not aufsd_col:
                    aufsd_col = next((c for c in kna1.columns if str(c).upper() in ('ORBLK', 'CENTRAL_ORDER_BLOCK_CODE')), None)
                if not aufsd_col:
                    aufsd_col = self._find_column_alternative(kna1.columns, 'AUFSD', 'KNA1')
                if aufsd_col:
                    kna1_aufsd = kna1[[kunnr_col, aufsd_col]].copy()
                    kna1_aufsd['_partner_key'] = _norm_key(kna1_aufsd[kunnr_col])
                    kna1_aufsd = kna1_aufsd.drop_duplicates(subset=['_partner_key'], keep='first')
                    out = out.drop(columns=['AUFSD'], errors='ignore')
                    out = out.merge(kna1_aufsd[['_partner_key', aufsd_col]].rename(columns={aufsd_col: 'AUFSD'}), on='_partner_key', how='left')
                    blocked_codes = {'E', 'G', 'SP', 'R', 'U'}
                    aufsd_norm = out['AUFSD'].astype(str).str.strip().str.upper()
                    before_blk = len(out)
                    out = out.loc[~aufsd_norm.isin(blocked_codes)].copy()
                    if before_blk > len(out):
                        print(f"      [FILTER] {rule_code}: исключены KNA1.AUFSD in {sorted(blocked_codes)}: {before_blk - len(out):,} (осталось {len(out):,})")
            before_dedup = len(out)
            out = out.drop_duplicates(subset=['_partner_key'], keep='first')
            out = out.drop(columns=['_partner_key'], errors='ignore')
            if before_dedup > len(out):
                print(f'      [FILTER] {rule_code}: дедупликация по PARTNER — убрано {before_dedup - len(out):,} (осталось {len(out):,})')
            return out
        except Exception as e:
            print(f'      [WARN] Ошибка _filter_adr2_dm_customer_scope для {rule_code}: {e}')
            import traceback
            traceback.print_exc()
            return df

    def _filter_adr2_rccomp_375_1_scope_by_kna1_aufsd(self, df, rule_code, table_name=None):
        return self._filter_adr2_dm_customer_scope(df, rule_code, table_name)

    def _filter_adr2_by_knvv_aufsd_fm(self, df, rule_code, table_name=None):
        if table_name and str(table_name).strip().upper() != 'ADR2':
            return df
        try:

            def _norm_key(series: pd.Series) -> pd.Series:
                s = series.astype(str).str.strip()
                s = s.str.replace('\\.0$', '', regex=True)
                s = s.str.replace('\\D+', '', regex=True)
                return s.str.zfill(10)
            addr_col = self._resolve_addrnumber_column(df, 'ADR2')
            if not addr_col:
                print(f'      [FILTER] Колонка ADDRNUMBER не найдена в ADR2 для {rule_code}')
                return df
            but020 = self._get_but020_table_for_join()
            if but020 is None or but020.empty:
                print(f'      [WARN] Таблица BUT020 не найдена или пуста для {rule_code}')
                return df
            addr_but, partner_but = self._resolve_but020_join_columns(but020)
            if not addr_but or not partner_but:
                print(f'      [WARN] В BUT020 не найдены Addr. No. и Business Partner для {rule_code} (колонки: {list(but020.columns)[:10]})')
                return df
            df = df.copy()
            df['_ak'] = _norm_key(df[addr_col])
            but_join = but020[[addr_but, partner_but]].copy()
            but_join['_ak'] = _norm_key(but_join[addr_but])
            but_join = but_join.drop_duplicates(subset=['_ak'], keep='first')
            df = df.merge(but_join[['_ak', partner_but]], on='_ak', how='left')
            df = df.drop(columns=['_ak'], errors='ignore')
            partner_col = 'PARTNER' if partner_but != 'PARTNER' and partner_but in df.columns else partner_but
            if partner_but != 'PARTNER' and partner_but in df.columns:
                df = df.rename(columns={partner_but: 'PARTNER'})
                partner_col = 'PARTNER'
            if partner_col not in df.columns:
                print(f'      [WARN] PARTNER не добавлен из BUT020 для {rule_code}')
                return df
            knvv = self.memory_manager.get_table('KNVV')
            if knvv is None or knvv.empty:
                if getattr(self, 'db_path', None):
                    import sqlite3
                    conn = connect_sqlite(self.db_path)
                    tables = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", conn)
                    knvv_name = next((r[0] for r in tables.values if str(r[0]).strip().upper() == 'KNVV'), None)
                    if knvv_name:
                        knvv = pd.read_sql_query(f'SELECT * FROM "{knvv_name}"', conn)
                    conn.close()
            if knvv is None or knvv.empty:
                print(f'      [WARN] Таблица KNVV не найдена или пуста для {rule_code}')
                return df
            kunnr_knvv = next((c for c in knvv.columns if str(c).upper() in ('KUNNR', 'KUNNR_KNVV')), None)
            aufsd_knvv = next((c for c in knvv.columns if 'AUFSD' in str(c).upper()), None)
            if not kunnr_knvv or not aufsd_knvv:
                print(f'      [WARN] В KNVV не найдены KUNNR или AUFSD для {rule_code}')
                return df
            knvv_fm = knvv[knvv[aufsd_knvv].astype(str).str.strip().str.upper().isin(('F', 'M'))][[kunnr_knvv, aufsd_knvv]].drop_duplicates(subset=[kunnr_knvv], keep='first')
            knvv_fm = knvv_fm.rename(columns={kunnr_knvv: '_partner', aufsd_knvv: 'AUFSD'})
            df['_partner'] = _norm_key(df[partner_col])
            knvv_fm['_partner'] = _norm_key(knvv_fm['_partner'])
            before = len(df)
            df = df.merge(knvv_fm, on='_partner', how='inner')
            df = df.drop(columns=['_partner'], errors='ignore')
            print(f'      [FILTER] Фильтр ADR2 по KNVV AUFSD in (F,M): {len(df):,} из {before:,} строк для {rule_code}')
            if df.empty:
                print(f'      [WARN] После фильтра по блокам F/M данных нет для {rule_code}')
            return df
        except Exception as e:
            print(f'      [WARN] Ошибка _filter_adr2_by_knvv_aufsd_fm для {rule_code}: {e}')
            import traceback
            traceback.print_exc()
            return df

    def _non_empty_key_count(self, series) -> int:
        if series is None:
            return 0
        filled, _distinct = self._kunnr_join_key_stats(series)
        return filled

    def _kunnr_join_key_stats(self, series) -> tuple:
        if series is None:
            return (0, 0)
        try:
            norms = series.apply(self._norm_customer_partner_key)
        except Exception:
            return (0, 0)
        mask = norms.ne('')
        filled = int(mask.sum())
        distinct = int(norms[mask].nunique()) if filled else 0
        return (filled, distinct)

    def _is_adrc_table(self, table_name) -> bool:
        tn = str(table_name or '').strip().upper()
        if tn in self.ADRC_TABLE_ALIASES:
            return True
        norm = tn.replace('/', '').replace(' ', '').replace('_', '')
        return norm in ('LOTGCADR', 'DMCUSTOMERADDRESS')

    def _is_blocked_kna1_join_column(self, col_name) -> bool:
        if col_name is None:
            return True
        cu = str(col_name).strip().upper().replace(' ', '').replace('_', '')
        if cu in self.KNA1_JOIN_BLOCKED_COLUMNS:
            return True
        if cu in ('CLIENT', 'MANDT', 'MANDANT') or cu.endswith('CLIENT'):
            return True
        return False

    def _pick_best_kunnr_column(self, df, table_name: str='KNB1'):
        if df is None or df.empty:
            return None
        tn = str(table_name or '').strip().upper()
        col_upper = {str(c).strip().upper(): c for c in df.columns}
        if tn == 'KNB1':
            name_order = ('CUSTOMER', 'KUNNR', 'CUSTOMER_CODE', 'KUNNR_KNB1')
        elif tn == 'KNA1':
            name_order = ('KUNNR', 'CUSTOMER', 'CUSTOMER_CODE', 'KUNNR_KNB1')
        elif tn in self.KNA1_JOIN_VIA_BUT020_TABLES:
            name_order = ('PARTNER', 'CUSTOMER', 'KUNNR', 'CUSTOMER_CODE')
        else:
            name_order = ('CUSTOMER', 'KUNNR', 'CUSTOMER_CODE', 'PARTNER')
        priority = {name: i for i, name in enumerate(name_order)}
        candidates = []
        for name in name_order:
            if name in col_upper:
                candidates.append(col_upper[name])
        if not candidates:
            try:
                from utils.column_map_resolver import resolve_column_in_df
                for sap in ('KUNNR', 'Customer', 'PARTNER'):
                    c = resolve_column_in_df(df, sap, table_name, self.column_map, parent_dir)
                    if c and c not in candidates:
                        candidates.append(c)
            except ImportError:
                pass
        if not candidates:
            found = self._find_kunnr_column(df)
            if found and not self._is_blocked_kna1_join_column(found):
                return found
            return None
        candidates = [c for c in candidates if not self._is_blocked_kna1_join_column(c)]
        if not candidates:
            if tn in self.KNA1_JOIN_VIA_BUT020_TABLES:
                print(f'      [WARN] {table_name}: нет ключа клиента (PARTNER/KUNNR); CLIENT/MANDT не используется для JOIN с KNA1')
            return None

        def _score(col):
            filled, distinct = self._kunnr_join_key_stats(df[col])
            prio = priority.get(str(col).strip().upper(), len(name_order))
            return (filled, distinct, -prio)
        best = max(candidates, key=_score)
        filled_b, distinct_b = self._kunnr_join_key_stats(df[best])
        if filled_b == 0:
            print(f'      [WARN] {table_name}: все кандидаты ключа клиента пусты ({candidates})')
        elif len(candidates) > 1:
            alt = [c for c in candidates if c != best]
            if alt:
                c0 = alt[0]
                f0, d0 = self._kunnr_join_key_stats(df[c0])
                if f0 == filled_b and d0 != distinct_b and (d0 < max(100, distinct_b // 100)):
                    print(f'      [JOIN] {table_name}: ключ KUNNR/Customer [{best}] (заполнено {filled_b:,}, уникальных {distinct_b:,}); отклонён [{c0}] (уникальных {d0:,})')
        return best

    def _resolve_knb1_kna1_join_column(self, df, table_name: str):
        return self._pick_best_kunnr_column(df, table_name)

    def _join_kna1_ktokd_rconf_24_1_adrc(self, df, table_name='ADRC', rule_code='RCCONF_24.1'):
        """RCCONF_24.1: только ADRC.AddrNo -> BUT020.Business Partner -> KNA1.KTOKD; CLIENT/mandant запрещён."""
        print(f'      [JOIN] [{self.CHECKER_BUILD_ID}] RCCONF_24.1: ADRC.[Addr. No.] -> BUT020.[Business Partner] -> KNA1.[KTOKD/Group_1]')
        if df is None or df.empty:
            return df
        out = self._drop_kna1_account_group_columns(df)
        out, join_col = self._merge_adrc_partner_from_but020(out, table_name or 'ADRC', rule_code)
        if not join_col or join_col not in out.columns:
            print('      [WARN] RCCONF_24.1: не удалось получить Business Partner через BUT020')
            return out
        if self._is_blocked_kna1_join_column(join_col):
            print(f'      [ERROR] RCCONF_24.1: колонка {join_col} не может использоваться для JOIN с KNA1')
            return out
        partner_filled = self._non_empty_key_count(out[join_col])
        if partner_filled == 0:
            print(f'      [WARN] RCCONF_24.1: PARTNER пустой после ADRC->BUT020 (колонка {join_col})')
            return out
        if self._get_table_for_rules('KNA1') is None or self._get_table_for_rules('KNA1').empty:
            print('      [INFO] KNA1 отсутствует в RAM для RCCONF_24.1 — загружаем...')
            try:
                self.memory_manager.load_selected_tables_to_ram(['KNA1'], add_reference_tables=False)
                setattr(self, '_kna1_ktokd_lookup_df', None)
            except Exception as e:
                print(f'      [WARN] Не удалось загрузить KNA1: {e}')
        return self._merge_kna1_account_group_from_lookup(out, table_name or 'ADRC', rule_code, join_col)

    def _add_account_group_code_from_kna1(self, df, table_name, rule_code):
        try:
            rule_code_u = str(rule_code).strip().upper()
            table_u = str(table_name or '').strip().upper()
            if rule_code_u == 'RCCONF_24.1' and self._is_adrc_table(table_name):
                return self._join_kna1_ktokd_rconf_24_1_adrc(df, table_name, rule_code)
            force_rebuild_for_rule = table_u == 'KNB1' or rule_code_u in getattr(self, 'RULES_FORCE_KNA1_KTOKD_JOIN', ('RCCONF_24.1', 'RCCONF_115.11'))
            if str(rule_code).strip().upper() == 'RCCONF_113.1':
                print('      [DEBUG] RCCONF_113.1 JOIN PATH v2 (memory->sqlite fallback)')

            def _norm_join_key(v):
                return self._norm_customer_partner_key(v)
            if not force_rebuild_for_rule and ('account_group_code' in df.columns or 'KTOKD' in df.columns):
                print(f'      [JOIN] account_group_code уже присутствует в таблице {table_name}')
                return df
            if force_rebuild_for_rule:
                df = self._drop_kna1_account_group_columns(df)
            print(f'      [JOIN] Добавление account_group_code из KNA1 для правила {rule_code} в таблице {table_name}...')
            join_col = None
            is_rule_24_1 = rule_code_u == 'RCCONF_24.1'
            if table_u == 'KNB1':
                join_col = self._resolve_knb1_kna1_join_column(df, table_name)
                if join_col:
                    print(f'      [JOIN] KNB1: ключ {table_name}.{join_col} (Customer/KUNNR) -> KNA1.Customer/KUNNR -> KNA1.KTOKD (Group_1) как account_group_code')
            if is_rule_24_1 and self._is_adrc_table(table_name):
                print(f'      [JOIN] [{self.CHECKER_BUILD_ID}] RCCONF_24.1 fallback: ADRC→BUT020→KNA1')
                df, join_col = self._merge_adrc_partner_from_but020(df, table_name, rule_code)
                if join_col:
                    print(f'      [JOIN] RCCONF_24.1: шаг 2 — Business Partner из BUT020 -> KNA1.Customer (колонка в ADRC: {join_col})')
                else:
                    print(f'      [WARN] RCCONF_24.1: не удалось получить PARTNER через ADRC->BUT020; JOIN с KNA1 пропущен')
                    return df
            elif is_rule_24_1:
                partner_direct = next((c for c in df.columns if str(c).strip().upper() == 'PARTNER'), None)
                if partner_direct and self._non_empty_key_count(df[partner_direct]) > 0:
                    join_col = partner_direct
                    print(f'      [JOIN] RCCONF_24.1: принудительный JOIN по PARTNER -> KNA1.KUNNR (колонка: {join_col})')
            if not join_col and not (is_rule_24_1 and self._is_adrc_table(table_name)):
                join_col = self._resolve_knb1_kna1_join_column(df, table_name)
                if join_col and self._is_blocked_kna1_join_column(join_col):
                    print(f'      [WARN] {table_name}: колонка {join_col} не подходит для JOIN с KNA1 (mandant/CLIENT)')
                    join_col = None
                elif join_col:
                    if is_rule_24_1 and self._is_blocked_kna1_join_column(join_col):
                        print(f'      [ERROR] [{self.CHECKER_BUILD_ID}] RCCONF_24.1: отклонён JOIN по [{join_col}] — нужен ADRC→BUT020→KNA1')
                        join_col = None
                    else:
                        print(f'      [JOIN] Найдена колонка для JOIN в {table_name}: {join_col}')
            if not join_col and not (is_rule_24_1 and self._is_adrc_table(table_name)):
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower in ['kunnr', 'customer_code', 'customer', 'kunnr_knvv', 'kunnr_knb1']:
                        if not self._is_blocked_kna1_join_column(col):
                            join_col = col
                            print(f'      [JOIN] Найдена колонка для JOIN в {table_name}: {col}')
                            break
            join_mode = 'kunnr'
            if not join_col and str(table_name or '').strip().upper() == 'ADRC':
                df, join_col = self._merge_adrc_partner_from_but020(df, table_name, rule_code)
                if join_col and is_rule_24_1:
                    print(f'      [JOIN] RCCONF_24.1: путь JOIN по PARTNER (ADRC -> BUT020 -> KNA1.KUNNR)')
                elif join_col:
                    print(f'      [JOIN] ADRC: путь JOIN по PARTNER (ADRC -> BUT020 -> KNA1.KUNNR)')
            if not join_col:
                print(f'      [WARN] Колонка для JOIN (KUNNR/customer_code или ADRC address) не найдена в {table_name} для правила {rule_code}')
                return df
            if self._get_table_for_rules('KNA1') is None or self._get_table_for_rules('KNA1').empty:
                print(f'      [INFO] KNA1 отсутствует в RAM для {rule_code} — загружаем...')
                try:
                    self.memory_manager.load_selected_tables_to_ram(['KNA1'], add_reference_tables=False)
                    setattr(self, '_kna1_ktokd_lookup_df', None)
                except Exception as e:
                    print(f'      [WARN] Не удалось загрузить KNA1: {e}')
            return self._merge_kna1_account_group_from_lookup(df, table_name, rule_code, join_col)
        except Exception as e:
            print(f'      [WARN] Ошибка при добавлении account_group_code из KNA1 для правила {rule_code}: {e}')
            import traceback
            traceback.print_exc()
            return df

    def _find_kna1_ktokd_column_in_df(self, df):
        if df is None or df.empty:
            return None
        col_lower = {str(c).strip().lower(): c for c in df.columns}
        for name in ('ktokd', 'kna.ktokd', 'account_group_code', 'b.account_group_code', 'b.ktokd', 'group_1', 'lookup_account_group_ktokd'):
            if name in col_lower:
                return col_lower[name]
        return None

    def _place_column_after(self, df, col_name, after_col_names):
        if df is None or col_name not in df.columns:
            return df
        after_col = None
        upper_map = {str(c).strip().upper(): c for c in df.columns}
        for name in after_col_names:
            if name.upper() in upper_map:
                after_col = upper_map[name.upper()]
                break
        if not after_col:
            return df
        cols = [c for c in df.columns if c != col_name]
        if after_col not in cols:
            return df
        ix = cols.index(after_col) + 1
        return df[cols[:ix] + [col_name] + cols[ix:]]

    def _normalize_rule_code(self, rule_code: str) -> str:
        return re.sub('[^A-Za-z0-9._-]', '', str(rule_code or '')).strip().upper()

    def _format_ktokd_for_export(self, series):
        if series is None:
            return pd.Series(dtype=object)
        s = series.astype(str).str.strip().str.replace('\\.0+$', '', regex=True)
        return s.replace({'nan': '', 'None': '', 'null': '', 'NaN': ''})

    def _find_customer_column_for_kna1_join(self, df, table_name: str='KNB1'):
        return self._pick_best_kunnr_column(df, table_name or 'KNB1')

    def _build_kna1_ktokd_lookup(self, force_reload: bool=False):
        cache_key = '_kna1_ktokd_lookup_df'
        if not force_reload and getattr(self, cache_key, None) is not None:
            return getattr(self, cache_key)
        kna1_df = None
        if hasattr(self, 'memory_manager'):
            kna1_df = self.memory_manager.get_table('KNA1')
        if kna1_df is None or kna1_df.empty:
            if getattr(self, 'db_path', None):
                try:
                    conn = connect_sqlite(self.db_path)
                    try:
                        kna1_df = pd.read_sql_query('SELECT "Customer", "Group_1" FROM "KNA1"', conn)
                    except Exception:
                        kna1_df = pd.read_sql_query('SELECT * FROM "KNA1"', conn)
                    conn.close()
                except Exception:
                    kna1_df = None
        lookup = pd.DataFrame(columns=['_join_key', 'KTOKD'])
        if kna1_df is None or kna1_df.empty:
            setattr(self, cache_key, lookup)
            return lookup
        kna1_mapped = self._apply_rule_time_column_map(kna1_df.copy(), 'KNA1')
        kunnr_col = self._pick_best_kunnr_column(kna1_mapped, 'KNA1')
        ktokd_col = None
        for c in kna1_mapped.columns:
            cl = str(c).strip().lower()
            if cl in ('group_1', 'ktokd', 'account_group_code'):
                if ktokd_col is None or self._non_empty_key_count(kna1_mapped[c]) > self._non_empty_key_count(kna1_mapped[ktokd_col]):
                    ktokd_col = c
        kna1_df = kna1_mapped
        if kunnr_col and ktokd_col:
            lookup = kna1_df[[kunnr_col, ktokd_col]].copy()
            lookup['_join_key'] = lookup[kunnr_col].apply(self._norm_customer_partner_key)
            lookup['KTOKD'] = self._format_ktokd_for_export(lookup[ktokd_col])
            lookup = lookup.drop_duplicates(subset=['_join_key'], keep='first')[['_join_key', 'KTOKD']]
        if kunnr_col and ktokd_col:
            _kf, _kd = self._kunnr_join_key_stats(kna1_df[kunnr_col])
            print(f'      [JOIN] справочник KNA1: ключ [{kunnr_col}] ({_kf:,} ключей, уникальных {_kd:,}), KTOKD из [{ktokd_col}]')
        setattr(self, cache_key, lookup)
        return lookup

    def _drop_kna1_account_group_columns(self, df):
        if df is None or df.empty:
            return df
        drop = []
        for c in df.columns:
            cl = str(c).strip().lower()
            cu = str(c).strip().upper()
            if cl in ('account_group_code', 'b.account_group_code', 'ktokd', 'b.ktokd', 'kna.ktokd', 'group_1') or cu == 'KTOKD':
                drop.append(c)
        if drop:
            return df.drop(columns=drop, errors='ignore')
        return df

    def _merge_kna1_account_group_from_lookup(self, df, table_name, rule_code, join_col):
        if df is None or df.empty or (not join_col):
            return df
        if self._is_blocked_kna1_join_column(join_col):
            rule_u = str(rule_code or '').strip().upper()
            if rule_u == 'RCCONF_24.1' and self._is_adrc_table(table_name):
                print(f'      [RECOVER] [{self.CHECKER_BUILD_ID}] RCCONF_24.1: CLIENT/mandant заблокирован — переключаемся на ADRC→BUT020→KNA1')
                return self._join_kna1_ktokd_rconf_24_1_adrc(df, table_name, rule_code)
            print(f'      [ERROR] {rule_code}: JOIN с KNA1 по [{join_col}] запрещён (mandant/CLIENT). Сборка: {self.CHECKER_BUILD_ID}. Для ADRC: ADRC→BUT020→Business Partner.')
            return df
        lookup = self._build_kna1_ktokd_lookup(force_reload=False)
        if lookup is None or lookup.empty:
            print(f'      [WARN] {rule_code}: справочник KNA1 (Customer->Group_1) пуст')
            return df
        out = self._drop_kna1_account_group_columns(df.copy())
        out['_join_key'] = out[join_col].apply(self._norm_customer_partner_key)
        matched_keys = out['_join_key'].isin(lookup['_join_key'])
        out = out.merge(lookup, on='_join_key', how='left')
        out = out.drop(columns=['_join_key'], errors='ignore')
        if 'KTOKD' in out.columns:
            out['account_group_code'] = out['KTOKD']
            out['b.account_group_code'] = out['KTOKD']
            out['ktokd'] = out['KTOKD']
            out['b.ktokd'] = out['KTOKD']
            out['kna.ktokd'] = out['KTOKD']
        from utils.sap_account_keys import norm_sap_account_group
        ktokd_norm = out['KTOKD'].apply(norm_sap_account_group) if 'KTOKD' in out.columns else pd.Series(dtype=str)
        filled = int((ktokd_norm != '').sum())
        n9038 = int((ktokd_norm == '9038').sum())
        self._last_kna1_join_stats = {'rows_after_join': len(out), 'filled_ktokd': filled, 'n9038': n9038, 'join_col': join_col, 'key_matched': int(matched_keys.sum())}
        rule_u = str(rule_code or '').strip().upper()
        if rule_u == 'RCCONF_24.1' and self._is_adrc_table(table_name):
            print(f'      [JOIN] [{self.CHECKER_BUILD_ID}] {rule_code}: Business Partner [{join_col}] = KNA1.Customer -> KTOKD: ключ совпал {int(matched_keys.sum()):,}/{len(out):,}, KTOKD заполнен: {filled:,}, из них 9038: {n9038:,}')
        else:
            print(f'      [JOIN] {rule_code}: KNA1.Group_1 -> {table_name} по [{join_col}]: ключ совпал {int(matched_keys.sum()):,}/{len(out):,}, KTOKD из KNA1: {filled:,}, из них 9038: {n9038:,}')
        if int(matched_keys.sum()) == 0 and filled > 0:
            print(f'      [WARN] {rule_code}: KTOKD не пустой, но ключей KNA1=0 — проверьте колонку JOIN (нужна Customer, не пустая KUNNR)')
        return out

    def _attach_kna1_ktokd_export_columns(self, df, rule_code=None):
        if df is None or df.empty:
            return df
        ag_col = self._find_account_group_column(df)
        out = df.copy()
        if ag_col is not None:
            out['KTOKD'] = self._format_ktokd_for_export(out[ag_col])
        else:
            customer_col = self._find_customer_column_for_kna1_join(out, 'KNB1')
            if customer_col:
                lookup = self._build_kna1_ktokd_lookup()
                if not lookup.empty:
                    out['_jk'] = out[customer_col].apply(self._norm_customer_partner_key)
                    out = out.merge(lookup, left_on='_jk', right_on='_join_key', how='left')
                    out = out.drop(columns=['_jk', '_join_key'], errors='ignore')
        if 'KTOKD' not in out.columns:
            out['KTOKD'] = ''
        out['KTOKD_SOURCE'] = 'KNA1'
        rule_u = self._normalize_rule_code(rule_code)
        if rule_u in self.RULES_KTOKD_ONLY_9038_SCOPE:
            out['RULE_SCOPE'] = 'only KNA1.KTOKD=9038'
        return self._place_column_after(out, 'KTOKD', ('Customer', 'KUNNR', 'CUSTOMER', 'Cl_', 'CLIENT'))

    def _enrich_error_df_kna1_ktokd(self, error_df, table_name, rule_code=None):
        if error_df is None or error_df.empty:
            return error_df
        rule_u = self._normalize_rule_code(rule_code)
        if 'KTOKD' in error_df.columns:
            out = error_df.copy()
            if 'KTOKD_SOURCE' not in out.columns:
                out['KTOKD_SOURCE'] = 'KNA1'
            if rule_u in self.RULES_KTOKD_ONLY_9038_SCOPE and 'RULE_SCOPE' not in out.columns:
                out['RULE_SCOPE'] = 'only KNA1.KTOKD=9038'
            return self._place_column_after(out, 'KTOKD', ('Customer', 'KUNNR', 'CUSTOMER', 'Cl_', 'CLIENT'))
        customer_col = self._find_customer_column_for_kna1_join(error_df, table_name or 'KNB1')
        if not customer_col:
            print(f'      [WARN] {rule_code}: в error_df нет Customer/KUNNR — колонки: {list(error_df.columns)[:12]}...')
            out = error_df.copy()
            out['KTOKD'] = ''
            out['KTOKD_SOURCE'] = 'KNA1 (join key not found)'
            return out
        try:
            lookup = self._build_kna1_ktokd_lookup()
            if lookup.empty:
                print(f'      [WARN] {rule_code}: справочник KNA1.KTOKD пуст — KTOKD не добавлен')
                out = error_df.copy()
                out['KTOKD'] = ''
                out['KTOKD_SOURCE'] = 'KNA1'
                return out
            out = error_df.copy()
            out['_join_key'] = out[customer_col].apply(self._norm_customer_partner_key)
            out = out.merge(lookup, on='_join_key', how='left')
            out = out.drop(columns=['_join_key'], errors='ignore')
            if 'KTOKD' not in out.columns:
                out['KTOKD'] = ''
            out['KTOKD_SOURCE'] = 'KNA1'
            filled = int((out['KTOKD'].astype(str).str.strip() != '').sum())
            print(f'      [JOIN] {rule_code}: KTOKD из KNA1 (Group_1) в error_df — заполнено {filled:,} из {len(out):,} (ключ: {customer_col})')
            if rule_u in self.RULES_KTOKD_ONLY_9038_SCOPE:
                out['RULE_SCOPE'] = 'only KNA1.KTOKD=9038'
                ktokd_chk = out['KTOKD'].astype(str).str.strip()
                not_9038 = ~ktokd_chk.isin({'9038', ''}) & ktokd_chk.notna()
                if not_9038.any():
                    print(f'      [WARN] {rule_code}: {int(not_9038.sum()):,} строк в error_df с KTOKD != 9038')
            out = self._place_column_after(out, 'KTOKD', ('Customer', 'KUNNR', 'CUSTOMER', 'Cl_', 'CLIENT'))
            if 'KTOKD_SOURCE' in out.columns:
                out = self._place_column_after(out, 'KTOKD_SOURCE', ('KTOKD',))
            if 'RULE_SCOPE' in out.columns:
                out = self._place_column_after(out, 'RULE_SCOPE', ('KTOKD', 'KTOKD_SOURCE'))
            return out
        except Exception as e:
            print(f'      [WARN] {rule_code}: не удалось подтянуть KTOKD из KNA1 в error_df: {e}')
            traceback.print_exc()
            out = error_df.copy()
            out['KTOKD'] = ''
            out['KTOKD_SOURCE'] = 'KNA1 (error)'
            return out

    def _add_central_order_block_code_from_kna1(self, df, table_name, rule_code):
        try:
            if df is None or df.empty:
                return df
            if 'central_order_block_code' in df.columns:
                return df

            def _norm_partner_key(v):
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    return ''
                s = str(v).replace('\ufeff', '').replace('\xa0', '').strip()
                if re.fullmatch('\\d+\\.0+', s):
                    s = s.split('.')[0]
                digits = re.sub('\\D', '', s)
                if digits:
                    s = digits
                return s.lstrip('0') or '0'
            partner_col = next((c for c in df.columns if str(c).strip().upper() == 'PARTNER1'), None)
            if not partner_col:
                partner_col = next((c for c in df.columns if 'PARTNER1' in str(c).strip().upper()), None)
            if not partner_col:
                partner_col = next((c for c in df.columns if str(c).strip().upper() == 'PARTNER2'), None)
            if not partner_col:
                partner_col = next((c for c in df.columns if 'PARTNER2' in str(c).strip().upper()), None)
            if not partner_col:
                print(f'      [WARN] {rule_code}: в {table_name} не найдены колонки PARTNER1/PARTNER2 для JOIN с KNA1')
                return df
            kna1_df = self._get_table_for_rules('KNA1')
            if kna1_df is None or kna1_df.empty:
                try:
                    self.memory_manager.load_selected_tables_to_ram(['KNA1'], add_reference_tables=False)
                    kna1_df = self._get_table_for_rules('KNA1')
                except Exception:
                    kna1_df = None
            if kna1_df is None or kna1_df.empty:
                try:
                    conn = connect_sqlite(self.db_path)
                    try:
                        kna1_df = pd.read_sql_query('SELECT "KUNNR","AUFSD" FROM "KNA1"', conn)
                    except Exception:
                        kna1_df = pd.read_sql_query('SELECT * FROM "KNA1"', conn)
                    conn.close()
                except Exception as e:
                    print(f'      [WARN] {rule_code}: не удалось загрузить KNA1 для central_order_block_code: {e}')
                    return df
            if kna1_df is None or kna1_df.empty:
                print(f'      [WARN] {rule_code}: KNA1 пуста, central_order_block_code не добавлен')
                return df
            kna1_df = self._apply_rule_time_column_map(kna1_df, 'KNA1')
            try:
                from utils.column_map_resolver import resolve_column_in_df
                kna1_kunnr_col = resolve_column_in_df(kna1_df, 'KUNNR', 'KNA1', self.column_map, parent_dir)
                aufsd_col = resolve_column_in_df(kna1_df, 'AUFSD', 'KNA1', self.column_map, parent_dir) or resolve_column_in_df(kna1_df, 'central_order_block_code', 'KNA1', self.column_map, parent_dir)
            except ImportError:
                kna1_kunnr_col = None
                aufsd_col = None
            if not kna1_kunnr_col:
                kna1_kunnr_col = next((c for c in kna1_df.columns if str(c).strip().upper() == 'KUNNR'), None)
            if not aufsd_col:
                aufsd_col = next((c for c in kna1_df.columns if str(c).strip().upper() == 'AUFSD'), None)
            if not kna1_kunnr_col or not aufsd_col:
                print(f'      [WARN] {rule_code}: в KNA1 не найдены KUNNR/AUFSD для JOIN')
                return df
            out = df.copy()
            out['_partner2_key'] = out[partner_col].apply(_norm_partner_key)
            kna1_join = kna1_df[[kna1_kunnr_col, aufsd_col]].copy()
            kna1_join['_partner2_key'] = kna1_join[kna1_kunnr_col].apply(_norm_partner_key)
            kna1_join = kna1_join[['_partner2_key', aufsd_col]].drop_duplicates(subset=['_partner2_key'], keep='first').rename(columns={aufsd_col: 'central_order_block_code'})
            out = out.merge(kna1_join, on='_partner2_key', how='left')
            out = out.drop(columns=['_partner2_key'], errors='ignore')
            print(f'      [JOIN] {rule_code}: добавлен central_order_block_code из KNA1.AUFSD по {table_name}.{partner_col} -> KNA1.{kna1_kunnr_col}')
            return out
        except Exception as e:
            print(f'      [WARN] Ошибка добавления central_order_block_code из KNA1 для {rule_code}: {e}')
            return df
    def _log_skipped_rule(self, rule, table_name, reason, timestamp):
        if not getattr(self, '_current_save_result', True):
            self._last_rule_skip_reason = reason
            return
        rule_code = rule.get('rule_code', 'UNKNOWN')
        self.results.append({'rule_code': rule_code, 'rule_description': rule.get('rule_description', 'Unknown rule'), 'quality_category': rule.get('quality_category', 'Unknown'), 'table_name': table_name, 'column_checked': rule.get('column_name_checked', ''), 'matched_column': '', 'total_records': 0, 'passed': 0, 'failed': 0, 'total_evaluated': 0, 'success_rate_%': 0, 'execution_time_sec': 0, 'check_date': timestamp, 'status': 'ПРОПУЩЕНО', 'status_color': 'gray', 'error_file': 'Нет', 'comments': f'Пропущено: {reason}'})

    def _log_failed_rule(self, rule, table_name, error_message, timestamp):
        rule_code = rule.get('rule_code', 'UNKNOWN')
        self.results.append({'rule_code': rule_code, 'rule_description': rule.get('rule_description', 'Unknown rule'), 'quality_category': rule.get('quality_category', 'Unknown'), 'table_name': table_name, 'column_checked': rule.get('column_name_checked', ''), 'matched_column': '', 'total_records': 0, 'passed': 0, 'failed': 0, 'total_evaluated': 0, 'success_rate_%': 0, 'execution_time_sec': 0, 'check_date': timestamp, 'status': 'ОШИБКА ВЫПОЛНЕНИЯ', 'status_color': 'dark_red', 'error_file': 'Нет', 'comments': f'Ошибка: {error_message}'})

    def _sync_results_error_files(self):
        saved = getattr(self, 'saved_error_files', {}) or {}
        for result in self.results:
            rule_code = str(result.get('rule_code', '') or '').strip()
            table_name = str(result.get('table_name', '') or '').strip()
            if not rule_code:
                continue
            key = f'{rule_code}_{table_name}'
            failed = int(result.get('failed') or result.get('error_count') or 0)
            path = saved.get(key, '')
            stored = self.rule_errors.get(key, {})
            is_truncated = bool(stored.get('is_truncated', False))
            saved_rows = int(stored.get('saved_error_count') or 0)
            if path and os.path.isfile(path):
                result['error_file'] = 'Есть'
                result['error_file_path'] = path
                if is_truncated and saved_rows > 0:
                    result['comments'] = f'[!] ВНИМАНИЕ: Всего ошибок {failed:,}, сохранено только первые {saved_rows:,}! Обратите внимание!'
                elif self._saves_all_errors(rule_code, table_name) and failed > self.MAX_ERRORS_TO_SAVE:
                    result['comments'] = str(result.get('comments') or '').replace(f'сохранено только первые {self.MAX_ERRORS_TO_SAVE:,}', '').strip()
                    if result['comments'].startswith('[!] ВНИМАНИЕ:') and 'сохранено только' in result['comments']:
                        result['comments'] = ''
            elif failed > 0:
                stored = self.rule_errors.get(key, {})
                has_df = stored.get('error_df') is not None and (not stored.get('error_df').empty)
                result['error_file'] = 'Нет (не сохранено)' if not has_df else 'Есть'
                result['error_file_path'] = ''
            else:
                result['error_file'] = 'Нет'
                result['error_file_path'] = ''

    def _save_rule_errors(self, timestamp=None):
        if not self.rule_errors:
            print(f'\n[INFO] Нет ошибок для сохранения')
            return
        print(f'\n[INFO] Сохранение ошибок по правилам...')
        print(f'   Всего правил с ошибками в памяти: {len(self.rule_errors)}')
        errors_by_table = {}
        for key, error_data in self.rule_errors.items():
            table_name = error_data['table_name']
            if table_name not in errors_by_table:
                errors_by_table[table_name] = 0
            errors_by_table[table_name] += 1
        print(f'   Ошибок по таблицам: {errors_by_table}')
        if timestamp is None:
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        self.last_file_timestamp = timestamp
        self.saved_error_files = {}
        errors_dir = os.path.join(self.output_dir, f'errors_{timestamp}')
        self.last_errors_dir = errors_dir
        os.makedirs(errors_dir, exist_ok=True)
        print(f'   [INFO] Создана папка для ошибок: {errors_dir}')
        key_39_5 = next((k for k in self.rule_errors if self.rule_errors[k].get('rule_code') == 'RCCONF_39.5' and str(self.rule_errors[k].get('table_name', '')).strip().upper() == 'ADR2'), None)
        if key_39_5 is not None:
            ed = self.rule_errors[key_39_5]
            raw = ed.get('error_df')
            if raw is not None and (not raw.empty):
                out = raw.copy()
                pcol = next((c for c in out.columns if 'PERSNUMBER' in str(c).upper() or ('PERSON' in str(c).upper() and 'NUMBER' in str(c).upper())), None)
                if pcol is not None:
                    v = out[pcol].astype(str).str.strip().str.upper()
                    empty = out[pcol].isna() | (v == '') | v.isin(['NONE', 'NAN', 'NULL', '-', '.'])
                    out = out.loc[empty].copy()
                    print(f'   [RCCONF_39.5] Оставлены только строки с пустым PERSNUMBER: {len(out):,} из {len(raw):,}')
                else:
                    print(f'   [RCCONF_39.5] Колонка PERSNUMBER не найдена. Имена колонок: {list(out.columns)}')

                from utils.ru_tel_format import is_valid_rccconf_39_5_value
                tcol = None
                if 'DQ_COLUMN_CHECKED' in out.columns:
                    try:
                        cn = out['DQ_COLUMN_CHECKED'].iloc[0]
                        if cn and str(cn).strip() in out.columns:
                            tcol = str(cn).strip()
                    except Exception:
                        pass
                if tcol is None:
                    tcol = next((c for c in out.columns if 'TEL' in str(c).upper() and ('NUMBER' in str(c).upper() or 'NR' in str(c).upper() or 'NUM' in str(c).upper())), None)
                if tcol is not None and (not out.empty):
                    drop = out[tcol].apply(lambda v: is_valid_rccconf_39_5_value(v, 'RCCONF_39.5'))
                else:
                    drop = pd.Series(False, index=out.index)
                    for c in out.columns:
                        if 'DQ_' in str(c):
                            continue
                        drop = drop | out[c].apply(lambda v: is_valid_rccconf_39_5_value(v, 'RCCONF_39.5'))
                if drop.any():
                    n_before = len(out)
                    out = out.loc[~drop].copy()
                    print(f'   [RCCONF_39.5] Убраны из выгрузки номера с валидным форматом: {n_before - len(out):,} строк')
                acol = next((c for c in out.columns if 'ADDRNUMBER' in str(c).upper()), None)
                if acol is not None and (not out.empty):
                    try:
                        out = self._attach_partner_from_but020_by_addr(out, acol, log_prefix='   [RCCONF_39.5] Добавлена колонка')
                    except Exception as e:
                        print(f'   [RCCONF_39.5] Ошибка при добавлении PARTNER: {e}')
                ed['error_df'] = out
                ed['saved_error_count'] = len(out)
                ed['total_error_count'] = ed.get('total_error_count') or ed.get('error_count') or len(out)
                ed['error_count'] = ed['total_error_count']
                ed['is_truncated'] = len(out) < int(ed['total_error_count'])
        saved_count = 0
        for key, error_data in self.rule_errors.items():
            try:
                rule_code = error_data['rule_code']
                table_name = error_data['table_name']
                error_df = error_data['error_df']
                if error_df is None or error_df.empty:
                    ec = int(error_data.get('error_count') or 0)
                    if ec > 0:
                        print(f'   [WARN] Пропускаем {key}: error_df пустой, но в отчёте {ec:,} ошибок — файл не создан')
                    else:
                        print(f'   [WARN] Пропускаем {key}: error_df пустой')
                    continue
                if str(table_name or '').strip().upper() == 'ADR2':
                    error_df = self._filter_adr2_dm_customer_scope(error_df, rule_code, table_name)
                    if error_df is None or error_df.empty:
                        print(f'   [WARN] Пропускаем {key}: нет ошибок ADR2 в scope (KNA1 Group={self.ADR2_DM_CUSTOMER_SCOPE_ACCOUNT_GROUP}, KNVV 01-01)')
                        continue
                    error_df = self._dedupe_adr2_by_partner(error_df, log_prefix=f'   [{rule_code}] ')
                    error_data['error_df'] = error_df
                if str(table_name or '').strip().upper() == 'ADRC':
                    name1_col = None
                    for c in error_df.columns:
                        if str(c).strip().upper() == 'NAME1':
                            name1_col = c
                            break
                    if name1_col is None:
                        name1_col = self._find_column_alternative(error_df.columns, 'NAME1', table_name)
                    if name1_col is None:
                        best_col, best_count = (None, 0)
                        for c in error_df.columns:
                            try:
                                cnt = (error_df[c].astype(str).str.strip().str.upper() == 'RESERVED').sum()
                                if cnt > best_count:
                                    best_count, best_col = (cnt, c)
                            except Exception:
                                pass
                        if best_col and best_count > 0:
                            name1_col = best_col
                    if name1_col and name1_col in error_df.columns:
                        val_str = error_df[name1_col].astype(str).str.strip().str.upper()
                        error_df = error_df[val_str != 'RESERVED'].copy()
                        if error_df.empty:
                            print(f'   [ADRC] {key}: после исключения NAME1=RESERVED записей не осталось, файл не создаётся')
                            continue
                is_adr2 = str(table_name or '').strip().upper() == 'ADR2'
                acol = next((c for c in error_df.columns if 'ADDRNUMBER' in str(c).upper()), None)
                if rule_code in ['RCCONF_38.5', 'RCCONF_39.3', 'RCCONF_39.3.2', 'RCCONF_39.5', 'RCCONF_39.5.2'] and is_adr2 and (acol is not None):
                    try:
                        error_df = self._attach_partner_from_but020_by_addr(error_df, acol, log_prefix=f'   [INFO] {rule_code}: добавлена колонка')
                    except Exception as e:
                        print(f'   [ERROR] {rule_code}: ошибка при добавлении PARTNER: {e}')
                if is_adr2:
                    need_aufsd = 'AUFSD' not in error_df.columns or error_df['AUFSD'].isna().all()
                    if need_aufsd:
                        partner_col = next((c for c in error_df.columns if str(c).upper() == 'PARTNER'), None)
                        if partner_col is None and acol is not None:
                            try:
                                error_df = self._attach_partner_from_but020_by_addr(error_df, acol)
                                partner_col = next((c for c in error_df.columns if str(c).upper() == 'PARTNER'), None)
                            except Exception as e:
                                print(f'   [WARN] {rule_code}: не удалось подтянуть PARTNER для AUFSD: {e}')
                        if partner_col is None:
                            error_df['AUFSD'] = None
                            print(f'   [WARN] {rule_code}: колонка PARTNER не найдена, AUFSD пустая')
                        else:
                            try:
                                knvv_df = self.memory_manager.get_table('KNVV')
                                if (knvv_df is None or knvv_df.empty) and getattr(self, 'db_path', None):
                                    import sqlite3
                                    conn = connect_sqlite(self.db_path)
                                    tables = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", conn)
                                    knvv_name = next((r[0] for r in tables.values if str(r[0]).strip().upper() == 'KNVV'), None)
                                    if knvv_name:
                                        knvv_df = pd.read_sql_query(f'SELECT * FROM "{knvv_name}"', conn)
                                    conn.close()
                                if knvv_df is not None and (not knvv_df.empty):
                                    kunnr_col = next((c for c in knvv_df.columns if str(c).upper() in ('KUNNR', 'KUNNR_KNVV')), None)
                                    aufsd_knvv = next((c for c in knvv_df.columns if 'AUFSD' in str(c).upper()), None)
                                    if kunnr_col and aufsd_knvv:
                                        knvv_aufsd = knvv_df[[kunnr_col, aufsd_knvv]].drop_duplicates(subset=[kunnr_col], keep='first')
                                        knvv_aufsd = knvv_aufsd.rename(columns={kunnr_col: '_partner', aufsd_knvv: 'AUFSD'})
                                        knvv_aufsd['_partner'] = knvv_aufsd['_partner'].astype(str).str.strip()
                                        error_df = error_df.drop(columns=['AUFSD'], errors='ignore')
                                        error_df['_partner'] = error_df[partner_col].astype(str).str.strip()
                                        error_df = error_df.merge(knvv_aufsd[['_partner', 'AUFSD']], on='_partner', how='left')
                                        error_df = error_df.drop(columns=['_partner'], errors='ignore')
                                        print(f'   [INFO] {rule_code}: добавлена колонка AUFSD из KNVV для ADR2')
                                    else:
                                        error_df['AUFSD'] = None
                                        print(f'   [WARN] {rule_code}: в KNVV не найдены колонки KUNNR или AUFSD. Колонки: {list(knvv_df.columns)[:20]}')
                                else:
                                    error_df['AUFSD'] = None
                                    if getattr(self, 'db_path', None):
                                        import sqlite3
                                        conn = connect_sqlite(self.db_path)
                                        tables = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", conn)
                                        knvv_in_db = any((str(r[0]).strip().upper() == 'KNVV' for r in tables.values))
                                        conn.close()
                                        if not knvv_in_db:
                                            print(f'   [WARN] {rule_code}: таблица KNVV не найдена в БД (проверьте имя таблицы), колонка AUFSD пустая')
                                        else:
                                            print(f'   [WARN] {rule_code}: таблица KNVV в БД есть, но пуста (0 строк) или не загрузилась, колонка AUFSD пустая')
                                    else:
                                        print(f'   [WARN] {rule_code}: таблица KNVV не найдена, колонка AUFSD пустая')
                            except Exception as e:
                                print(f'   [WARN] {rule_code}: ошибка при добавлении AUFSD: {e}')
                                error_df['AUFSD'] = None
                    else:
                        pass
                is_ausp_table = str(table_name or '').strip().upper() in ('AUSP_143', 'AUSP_604', 'AUSP_148', 'AUSP_151')
                if is_ausp_table:
                    need_aufsd_ausp = 'AUFSD' not in error_df.columns or error_df['AUFSD'].isna().all()
                    if need_aufsd_ausp:
                        partner_guid_col = next((c for c in error_df.columns if 'PARTNER_GUID' in str(c).upper() or str(c).upper() == 'PARTNERGUID'), None)
                        if partner_guid_col is None:
                            error_df['AUFSD'] = None
                            print(f'   [WARN] {rule_code} ({table_name}): колонка PARTNER_GUID не найдена в ошибках, AUFSD пустая')
                        else:
                            try:
                                import sqlite3
                                conn = connect_sqlite(self.db_path)
                                tables_list = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", conn)
                                but000_name = None
                                for _n in tables_list['name']:
                                    if str(_n).strip().upper() == 'BUT000':
                                        but000_name = _n
                                        break
                                if but000_name is None:
                                    conn.close()
                                    error_df['AUFSD'] = None
                                    print(f'   [WARN] {rule_code}: таблица BUT000 не найдена, AUFSD пустая')
                                else:
                                    but000_cols = pd.read_sql_query(f'PRAGMA table_info("{but000_name}")', conn)
                                    col_list = [r[1] for r in but000_cols.values]
                                    pg_but = next((c for c in col_list if 'PARTNER_GUID' in str(c).upper() or str(c).upper() == 'PARTNERGUID'), None)
                                    p_but = next((c for c in col_list if str(c).upper() == 'PARTNER'), None)
                                    if not pg_but or not p_but:
                                        conn.close()
                                        error_df['AUFSD'] = None
                                        print(f'   [WARN] {rule_code}: в BUT000 не найдены PARTNER_GUID или PARTNER')
                                    else:
                                        but000_df = pd.read_sql_query(f'SELECT "{pg_but}", "{p_but}" FROM "{but000_name}"', conn)
                                        but000_df = but000_df.rename(columns={pg_but: '_pg', p_but: '_partner'})
                                        but000_df = but000_df.drop_duplicates(subset=['_pg'], keep='first')
                                        conn.close()
                                        but000_df['_pg'] = but000_df['_pg'].astype(str).str.strip()
                                        error_df = error_df.merge(but000_df[['_pg', '_partner']], left_on=partner_guid_col, right_on='_pg', how='left')
                                        error_df = error_df.drop(columns=['_pg'], errors='ignore')
                                        if '_partner' not in error_df.columns:
                                            error_df['AUFSD'] = None
                                            print(f'   [WARN] {rule_code}: не удалось подтянуть PARTNER из BUT000')
                                        else:
                                            kna1_df = self._get_table_for_rules('KNA1')
                                            if kna1_df is None or kna1_df.empty:
                                                conn = connect_sqlite(self.db_path)
                                                kna1_tables = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", conn)
                                                kna1_name = next((r[0] for r in kna1_tables.values if 'KNA1' in str(r[0]).upper()), None)
                                                if kna1_name:
                                                    kna1_df = pd.read_sql_query(f'SELECT * FROM "{kna1_name}"', conn)
                                                conn.close()
                                            if kna1_df is not None and (not kna1_df.empty):
                                                kunnr_col = next((c for c in kna1_df.columns if str(c).upper() == 'KUNNR'), None)
                                                aufsd_col = next((c for c in kna1_df.columns if 'AUFSD' in str(c).upper()), None)
                                                if kunnr_col and aufsd_col:
                                                    kna1_aufsd = kna1_df[[kunnr_col, aufsd_col]].drop_duplicates(subset=[kunnr_col], keep='first')
                                                    kna1_aufsd = kna1_aufsd.rename(columns={kunnr_col: '_kunnr', aufsd_col: 'AUFSD'})
                                                    kna1_aufsd['_kunnr'] = kna1_aufsd['_kunnr'].astype(str).str.strip()
                                                    error_df['_partner'] = error_df['_partner'].astype(str).str.strip()
                                                    error_df = error_df.drop(columns=['AUFSD'], errors='ignore')
                                                    error_df = error_df.merge(kna1_aufsd[['_kunnr', 'AUFSD']], left_on='_partner', right_on='_kunnr', how='left')
                                                    error_df = error_df.drop(columns=['_partner', '_kunnr'], errors='ignore')
                                                    print(f'   [INFO] {rule_code} ({table_name}): добавлена колонка AUFSD из KNA1 (AUSP.PARTNER_GUID -> BUT000 -> KNA1)')
                                                else:
                                                    error_df = error_df.drop(columns=['_partner'], errors='ignore')
                                                    error_df['AUFSD'] = None
                                                    print(f'   [WARN] {rule_code}: в KNA1 не найдены KUNNR или AUFSD')
                                            else:
                                                error_df = error_df.drop(columns=['_partner'], errors='ignore')
                                                error_df['AUFSD'] = None
                                                print(f'   [WARN] {rule_code}: таблица KNA1 не найдена, AUFSD пустая')
                            except Exception as e:
                                error_df['AUFSD'] = None
                                if '_partner' in error_df.columns:
                                    error_df = error_df.drop(columns=['_partner'], errors='ignore')
                                print(f'   [WARN] {rule_code}: ошибка при добавлении AUFSD для AUSP: {e}')
                    error_data['error_df'] = error_df
                if self._normalize_rule_code(rule_code) in self.RULES_ERROR_EXPORT_KNA1_KTOKD:
                    error_df = self._enrich_error_df_kna1_ktokd(error_df, table_name, rule_code)
                    error_data['error_df'] = error_df
                if not is_adr2 and (not is_ausp_table):
                    need_aufsd = 'AUFSD' not in error_df.columns or error_df['AUFSD'].isna().all()
                    if need_aufsd:
                        partner_col = next((c for c in error_df.columns if str(c).upper() == 'KUNNR'), None)
                        if partner_col is None:
                            partner_col = next((c for c in error_df.columns if str(c).upper() == 'PARTNER'), None)
                        acol_gen = next((c for c in error_df.columns if 'ADDRNUMBER' in str(c).upper()), None)
                        if partner_col is None and acol_gen is not None:
                            try:
                                error_df = self._attach_partner_from_but020_by_addr(error_df, acol_gen, log_prefix=f'   [INFO] {rule_code}: добавлена колонка')
                                partner_col = next((c for c in error_df.columns if str(c).upper() == 'PARTNER'), None)
                            except Exception as e:
                                print(f'   [WARN] {rule_code}: не удалось подтянуть PARTNER для AUFSD: {e}')
                        if partner_col is not None:
                            aufsd_added = False
                            try:
                                knvv_df = self.memory_manager.get_table('KNVV')
                                if (knvv_df is None or knvv_df.empty) and getattr(self, 'db_path', None):
                                    import sqlite3
                                    conn = connect_sqlite(self.db_path)
                                    tables = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", conn)
                                    knvv_name = next((r[0] for r in tables.values if str(r[0]).strip().upper() == 'KNVV'), None)
                                    if knvv_name:
                                        knvv_df = pd.read_sql_query(f'SELECT * FROM "{knvv_name}"', conn)
                                    conn.close()
                                if knvv_df is not None and (not knvv_df.empty):
                                    kunnr_knvv = next((c for c in knvv_df.columns if str(c).upper() in ('KUNNR', 'KUNNR_KNVV')), None)
                                    aufsd_knvv = next((c for c in knvv_df.columns if 'AUFSD' in str(c).upper()), None)
                                    if kunnr_knvv and aufsd_knvv:
                                        knvv_aufsd = knvv_df[[kunnr_knvv, aufsd_knvv]].drop_duplicates(subset=[kunnr_knvv], keep='first')
                                        knvv_aufsd = knvv_aufsd.rename(columns={kunnr_knvv: '_partner', aufsd_knvv: 'AUFSD'})
                                        knvv_aufsd['_partner'] = knvv_aufsd['_partner'].astype(str).str.strip()
                                        error_df = error_df.drop(columns=['AUFSD'], errors='ignore')
                                        error_df['_partner'] = error_df[partner_col].astype(str).str.strip()
                                        error_df = error_df.merge(knvv_aufsd[['_partner', 'AUFSD']], on='_partner', how='left')
                                        error_df = error_df.drop(columns=['_partner'], errors='ignore')
                                        print(f'   [INFO] {rule_code} ({table_name}): добавлена колонка AUFSD из KNVV')
                                        aufsd_added = True
                                if not aufsd_added:
                                    kna1_df = self._get_table_for_rules('KNA1')
                                    if (kna1_df is None or kna1_df.empty) and getattr(self, 'db_path', None):
                                        import sqlite3
                                        conn = connect_sqlite(self.db_path)
                                        kna1_tables = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", conn)
                                        kna1_name = next((r[0] for r in kna1_tables.values if 'KNA1' in str(r[0]).upper()), None)
                                        if kna1_name:
                                            kna1_df = pd.read_sql_query(f'SELECT * FROM "{kna1_name}"', conn)
                                        conn.close()
                                    if kna1_df is not None and (not kna1_df.empty):
                                        kunnr_col = next((c for c in kna1_df.columns if str(c).upper() == 'KUNNR'), None)
                                        aufsd_col = next((c for c in kna1_df.columns if 'AUFSD' in str(c).upper()), None)
                                        if kunnr_col and aufsd_col:
                                            kna1_aufsd = kna1_df[[kunnr_col, aufsd_col]].drop_duplicates(subset=[kunnr_col], keep='first')
                                            kna1_aufsd = kna1_aufsd.rename(columns={kunnr_col: '_partner', aufsd_col: 'AUFSD'})
                                            kna1_aufsd['_partner'] = kna1_aufsd['_partner'].astype(str).str.strip()
                                            error_df = error_df.drop(columns=['AUFSD'], errors='ignore')
                                            error_df['_partner'] = error_df[partner_col].astype(str).str.strip()
                                            error_df = error_df.merge(kna1_aufsd[['_partner', 'AUFSD']], on='_partner', how='left')
                                            error_df = error_df.drop(columns=['_partner'], errors='ignore')
                                            print(f'   [INFO] {rule_code} ({table_name}): добавлена колонка AUFSD из KNA1')
                                            aufsd_added = True
                                    if not aufsd_added:
                                        error_df['AUFSD'] = None
                            except Exception as e:
                                error_df['AUFSD'] = None
                                print(f'   [WARN] {rule_code} ({table_name}): ошибка при добавлении AUFSD: {e}')
                        else:
                            error_df['AUFSD'] = None
                total_errors = len(error_df)
                is_truncated = error_data.get('is_truncated', False)
                original_error_count = error_data.get('total_error_count') or error_data.get('error_count', total_errors)
                no_100k_limit = self._saves_all_errors(rule_code, table_name)
                limit_save = self._error_save_limit(rule_code, table_name)
                if total_errors > limit_save:
                    error_df = error_df.head(limit_save)
                    is_truncated = True
                    print(f'   [WARN] {rule_code} ({table_name}): ошибок {original_error_count:,}, в файл сохранено только {limit_save:,} (первые {limit_save:,})')
                if is_adr2:
                    self._save_adr2_rule_errors_to_db(error_df, rule_code=rule_code, run_ts=timestamp)
                if is_truncated and len(error_df) > 0:
                    warning_row = pd.DataFrame([{col: f'[!] ВНИМАНИЕ: Всего ошибок {original_error_count:,}, показано только первые {limit_save:,}' if col == error_df.columns[0] else '' for col in error_df.columns}])
                    error_df = pd.concat([warning_row, error_df], ignore_index=True)
                use_csv = no_100k_limit or original_error_count > self.EXCEL_MAX_ROWS or total_errors > limit_save or (len(error_df) > limit_save)
                safe_table_name = self._safe_filename_token(table_name)
                if str(table_name or '').strip().upper() == 'KNB1' and self._normalize_rule_code(rule_code) in self.RULES_ERROR_EXPORT_KNA1_KTOKD:
                    error_df = self._enrich_error_df_kna1_ktokd(error_df, table_name, rule_code)
                    if 'KTOKD' not in error_df.columns:
                        error_df.insert(0, 'KTOKD', '')
                        error_df.insert(1, 'KTOKD_SOURCE', 'KNA1')
                if use_csv:
                    filename = f'{rule_code}_{safe_table_name}_errors_{timestamp}.csv'
                    filepath = os.path.join(errors_dir, filename)
                    error_df.to_csv(filepath, index=False, encoding='utf-8-sig', sep=';')
                    status_msg = f' ({original_error_count:,} всего, сохранено {len(error_df):,})' if is_truncated else f' ({len(error_df):,} строк)'
                    print(f'   [INFO] Сохранены ошибки в CSV: {filename}{status_msg}')
                else:
                    filename = f'{rule_code}_{safe_table_name}_errors_{timestamp}.xlsx'
                    filepath = os.path.join(errors_dir, filename)
                    error_df.to_excel(filepath, index=False, engine='openpyxl')
                    status_msg = f' ({original_error_count:,} всего, сохранено {len(error_df):,})' if is_truncated else f' ({len(error_df):,} строк)'
                    print(f'   [INFO] Сохранены ошибки в Excel: {filename}{status_msg}')
                self.saved_error_files[key] = filepath
                saved_count += 1
            except Exception as e:
                print(f'   [ERROR] Ошибка сохранения {key}: {e}')
                traceback.print_exc()
        print(f'   Сохранено файлов: {saved_count}')

    def _aggregate_result_score_by_category(self, results):
        """Суммы passed/total по Completeness и Conformity из self.results."""
        buckets = {'Completeness': {'passed': 0, 'total': 0}, 'Conformity': {'passed': 0, 'total': 0}}

        def _cat_key(raw):
            c = str(raw or '').strip().lower()
            if c in ('completeness', 'полнота', 'complete'):
                return 'Completeness'
            if c in ('conformity', 'конформность', 'соответствие'):
                return 'Conformity'
            return None

        for row in results or []:
            key = _cat_key(row.get('quality_category'))
            if not key:
                continue
            passed = int(row.get('passed', 0) or 0)
            total = int(row.get('total_records', row.get('total_evaluated', 0)) or 0)
            if total <= 0 and passed > 0:
                total = passed + int(row.get('failed', 0) or 0)
            buckets[key]['passed'] += passed
            buckets[key]['total'] += total
        return buckets

    def _create_result_score_sheet(self, wb, summary_sheet_title: str='Сводка проверок', summary_last_row: Optional[int]=None, results=None):
        """Лист result: взвешенный score по Completeness / Conformity (формулы Excel)."""
        sheet_name = 'result'
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        insert_at = 1 if len(wb.sheetnames) >= 1 else 0
        ws = wb.create_sheet(sheet_name, insert_at)
        summary_ref = summary_sheet_title.replace("'", "''")
        data_last_row = summary_last_row if summary_last_row and summary_last_row >= 7 else None
        if data_last_row:
            cat_rng = f"'{summary_ref}'!$C$7:$C${data_last_row}"
            total_rng = f"'{summary_ref}'!$G$7:$G${data_last_row}"
            passed_rng = f"'{summary_ref}'!$H$7:$H${data_last_row}"
        else:
            cat_rng = f"'{summary_ref}'!$C:$C"
            total_rng = f"'{summary_ref}'!$G:$G"
            passed_rng = f"'{summary_ref}'!$H:$H"
        header_fill = self.colors.get('header', PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'))
        header_font_white = self.colors.get('header_font', Font(bold=True, color='FFFFFF'))
        ws.cell(row=1, column=1, value='Category')
        for col, title in ((2, 'Weight'), (3, 'Successful'), (4, 'Failed'), (5, 'Total'), (6, '%'), (8, 'Successful'), (9, 'Total'), (10, 'Score')):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        categories = (('Completeness', 0.4), ('Conformity', 0.2))
        last_row = 1 + len(categories)
        for row_idx, (category, weight) in enumerate(categories, start=2):
            ws.cell(row=row_idx, column=1, value=category)
            ws.cell(row=row_idx, column=2, value=weight)
            ws.cell(row=row_idx, column=3, value=f'=SUMIF({cat_rng},$A{row_idx},{passed_rng})')
            ws.cell(row=row_idx, column=5, value=f'=SUMIF({cat_rng},$A{row_idx},{total_rng})')
            ws.cell(row=row_idx, column=4, value=f'=E{row_idx}-C{row_idx}')
            ws.cell(row=row_idx, column=6, value=f'=IF(E{row_idx}=0,0,C{row_idx}/E{row_idx})')
        ws.cell(row=2, column=8, value=f'=SUMPRODUCT($B$2:$B${last_row},$C$2:$C${last_row})')
        ws.cell(row=2, column=9, value=f'=SUMPRODUCT($B$2:$B${last_row},$E$2:$E${last_row})')
        ws.cell(row=2, column=10, value='=IF(I2=0,0,H2/I2)')
        for col_letter, width in (('A', 16), ('B', 10), ('C', 14), ('D', 12), ('E', 14), ('F', 10), ('G', 3), ('H', 14), ('I', 14), ('J', 10)):
            ws.column_dimensions[col_letter].width = width
        for row_idx in range(2, last_row + 1):
            ws.cell(row=row_idx, column=2).number_format = '0.0'
            ws.cell(row=row_idx, column=6).number_format = '0.00%'
        ws.cell(row=2, column=10).number_format = '0.00%'
        return ws

    def _create_result_score_sheet_values_only(self, wb, results=None):
        """Лист result только с числами (fallback, если формулы Excel не записались)."""
        sheet_name = 'result'
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        insert_at = 1 if len(wb.sheetnames) >= 1 else 0
        ws = wb.create_sheet(sheet_name, insert_at)
        header_fill = self.colors.get('header', PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'))
        header_font_white = self.colors.get('header_font', Font(bold=True, color='FFFFFF'))
        for col, title in ((1, 'Category'), (2, 'Weight'), (3, 'Successful'), (4, 'Failed'), (5, 'Total'), (6, '%'), (8, 'Successful'), (9, 'Total'), (10, 'Score')):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        categories = (('Completeness', 0.4), ('Conformity', 0.2))
        buckets = self._aggregate_result_score_by_category(results or getattr(self, 'results', []))
        weighted_pass = 0.0
        weighted_total = 0.0
        for row_idx, (category, weight) in enumerate(categories, start=2):
            passed_val = int(buckets.get(category, {}).get('passed', 0))
            total_val = int(buckets.get(category, {}).get('total', 0))
            failed_val = max(total_val - passed_val, 0)
            pct_val = (passed_val / total_val) if total_val > 0 else 0.0
            weighted_pass += weight * passed_val
            weighted_total += weight * total_val
            ws.cell(row=row_idx, column=1, value=category)
            ws.cell(row=row_idx, column=2, value=weight)
            ws.cell(row=row_idx, column=3, value=passed_val)
            ws.cell(row=row_idx, column=4, value=failed_val)
            ws.cell(row=row_idx, column=5, value=total_val)
            ws.cell(row=row_idx, column=6, value=pct_val)
            ws.cell(row=row_idx, column=2).number_format = '0.0'
            ws.cell(row=row_idx, column=6).number_format = '0.00%'
        score_val = (weighted_pass / weighted_total) if weighted_total > 0 else 0.0
        ws.cell(row=2, column=8, value=weighted_pass)
        ws.cell(row=2, column=9, value=weighted_total)
        ws.cell(row=2, column=10, value=score_val)
        ws.cell(row=2, column=10).number_format = '0.00%'
        return ws

    def _save_totals_by_table(self, timestamp=None):
        return

    def _create_correct_report(self, report_name: str='quality_check_report', timestamp=None):
        try:
            if not self.results:
                print(f'\n[INFO] Нет данных для отчета')
                return
            if timestamp is None:
                timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            excel_path = os.path.join(self.output_dir, f'{self._safe_filename_token(report_name)}_{timestamp}.xlsx')
            wb = Workbook()
            ws = wb.active
            ws.title = 'Сводка проверок'
            ws['A1'] = 'СВОДКА ПРОВЕРОК КАЧЕСТВА ДАННЫХ'
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f'Настройки: Сохраняется максимум {self.MAX_ERRORS_TO_SAVE:,} ошибок на правило'
            ws['A2'].font = Font(size=10)
            ws['A5'] = 'Для таблицы ADR2: «Всего записей» считается по внутренним фильтрам (ADDRNUMBER+PARTNER) для правила, отдельный список записей сейчас не формируется автоматически.'
            ws['A5'].font = Font(size=9, italic=True)
            check_date = self.results[0].get('check_date', '') if self.results else ''
            ws['A3'] = f'Дата проверки: {check_date}'
            ws['A3'].font = Font(size=9, italic=True)
            ws['A4'] = ''
            headers = ['Код правила', 'Описание', 'Категория', 'Таблица', 'Тип TAXNUM', 'Колонка', 'Всего записей', 'Успешно', 'Ошибок', '% успеха', 'Статус', 'Время (сек)', 'Комментарии', 'Список записей (файл)', 'Файл ошибок']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col_num, value=header)
                cell.fill = self.colors['header']
                cell.font = self.colors['header_font']
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row_num = 7
            for result in self.results:
                table_name = result.get('table_name', '')
                if table_name in self.DFKKBPTAXNUM_ALIASES:
                    table_display = 'DFKKBPTAXNUM'
                    suffix = table_name.replace('DFKKBPTAXNUM', '')
                    taxnum_type = 'ALL' if suffix == '_ALL' else suffix
                else:
                    table_display = table_name
                    taxnum_type = ''
                is_adr2 = str(table_name or '').strip().upper() == 'ADR2'
                if is_adr2 and result.get('filtered_adr2_count') is not None:
                    total_rec_display = result.get('filtered_adr2_count')
                    list_file_display = result.get('filtered_adr2_file', '') or ''
                else:
                    total_rec_display = result.get('total_records', 0)
                    list_file_display = ''
                error_file_display = result.get('error_file_path', '') or result.get('error_file', 'Нет')
                values = [result.get('rule_code', ''), result.get('rule_description', ''), result.get('quality_category', ''), table_display, taxnum_type, result.get('column_checked', ''), total_rec_display, result.get('passed', 0), result.get('failed', 0), result.get('success_rate_%', 0), result.get('status', ''), result.get('execution_time_sec', 0), result.get('comments', ''), list_file_display, error_file_display]
                for col_num, value in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.font = self.colors['normal_font']
                    if col_num in (14, 15) and value and isinstance(value, str) and os.path.isfile(value):
                        try:
                            path_uri = 'file:///' + value.replace('\\', '/').lstrip('/')
                            cell.hyperlink = path_uri
                            cell.font = Font(color='0563C1', underline='single')
                        except Exception:
                            pass
                    if col_num == 11:
                        status_color = result.get('status_color', '')
                        if status_color == 'green':
                            cell.fill = self.colors['green']
                            cell.font = Font(color='FFFFFF', bold=True)
                        elif status_color == 'red':
                            cell.fill = self.colors['red']
                            cell.font = Font(color='FFFFFF', bold=True)
                        elif status_color == 'orange':
                            cell.fill = self.colors['orange']
                            cell.font = Font(color='FFFFFF', bold=True)
                        elif status_color == 'dark_red':
                            cell.fill = self.colors['dark_red']
                            cell.font = Font(color='FFFFFF', bold=True)
                    elif col_num == 9:
                        failed_count = result.get('failed', 0)
                        if failed_count == 0:
                            cell.fill = self.colors['green']
                            cell.font = Font(color='000000', bold=True)
                        elif failed_count > self.MAX_ERRORS_TO_SAVE:
                            cell.fill = self.colors['orange']
                            cell.font = Font(color='000000', bold=True)
                        else:
                            cell.fill = self.colors['red']
                            cell.font = Font(color='000000', bold=True)
                row_num += 1
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            result_sheet_mode = 'failed'
            try:
                self._create_result_score_sheet(wb, summary_sheet_title=ws.title, summary_last_row=row_num - 1, results=self.results)
                result_sheet_mode = 'formulas'
            except Exception as result_err:
                print(f'\n[WARN] Лист result (формулы): {result_err} — записываем значения без формул')
                traceback.print_exc()
                try:
                    self._create_result_score_sheet_values_only(wb, results=self.results)
                    result_sheet_mode = 'values'
                except Exception as values_err:
                    print(f'\n[ERROR] Лист result не создан: {values_err}')
                    traceback.print_exc()
            wb.save(excel_path)
            if result_sheet_mode != 'failed' and 'result' not in wb.sheetnames:
                print('[WARN] Лист result отсутствует в книге перед сохранением')
            print(f'[INFO] Лист result (Score, {result_sheet_mode}) добавлен в отчёт: {excel_path}')
            self.last_report_path = excel_path
            stable_path = os.path.join(self.output_dir, f'{self._safe_filename_token(report_name)}.xlsx')
            try:
                shutil.copy2(excel_path, stable_path)
                self.last_stable_report_path = stable_path
            except Exception as copy_err:
                print(f'\n[WARN] Не удалось обновить копию отчёта {stable_path}: {copy_err}')
            print(f'\n[INFO] Цветной отчет сохранен: {excel_path}')
            if getattr(self, 'last_stable_report_path', None):
                print(f'[INFO] Актуальная копия отчёта: {self.last_stable_report_path}')
        except Exception as e:
            print(f'\n[ERROR] Ошибка при создании отчета: {e}')
            traceback.print_exc()

    def _print_final_statistics(self):
        total_rules = self.processed_rules + self.skipped_rules
        successful_rules = len([r for r in self.results if r.get('status') == 'УСПЕШНО'])
        failed_rules = len([r for r in self.results if r.get('status') == 'ОШИБКИ'])
        suspicious_rules = len([r for r in self.results if r.get('status') == 'ПОДОЗРИТЕЛЬНО'])
        mass_rules = len([r for r in self.results if r.get('status') == 'МАССОВЫЕ ОШИБКИ'])
        print(f'\n' + '=' * 100)
        print(f'\x1b[1mИТОГОВАЯ СТАТИСТИКА:\x1b[0m')
        print(f'=' * 100)
        print(f'Всего правил: \x1b[1m{total_rules}\x1b[0m')
        print(f'  [OK] Успешно:      \x1b[92m{successful_rules:4d}\x1b[0m ({successful_rules / total_rules * 100:.1f}%)')
        print(f'  [!] Ошибки:       \x1b[91m{failed_rules:4d}\x1b[0m ({failed_rules / total_rules * 100:.1f}%)')
        print(f'  [!] Подозрительные: \x1b[93m{suspicious_rules:4d}\x1b[0m ({suspicious_rules / total_rules * 100:.1f}%)')
        print(f'  ⚡ Массовые:      \x1b[91m{mass_rules:4d}\x1b[0m ({mass_rules / total_rules * 100:.1f}%)')
        print(f'  ↻ Пропущено:    \x1b[90m{self.skipped_rules:4d}\x1b[0m ({self.skipped_rules / total_rules * 100:.1f}%)')
        print(f'=' * 100)
        if suspicious_rules > 0:
            print(f'\n\x1b[93mПОДОЗРИТЕЛЬНЫЕ ПРАВИЛА (требуют проверки логики):\x1b[0m')
            suspicious_list = [r for r in self.results if r.get('status') in ['ПОДОЗРИТЕЛЬНО', 'МАССОВЫЕ ОШИБКИ']]
            for rule in suspicious_list[:10]:
                rc = rule.get('rule_code', '')
                tn = rule.get('table_name', '')
                failed = rule.get('failed', 0)
                rate = rule.get('success_rate_%', 0)
                print(f'  • {rc:20} - {tn:15} - {failed:,} ошибок ({rate:.1f}% успеха)')
        overall_time = time.time() - self.start_time
        print(f'\nВремя выполнения: {overall_time:.2f} сек')
        if overall_time > 0:
            print(f'Скорость: {self.processed_rules / overall_time:.1f} правил/сек')

    def _safe_filename_token(self, value):
        s = str(value or '').strip()
        if not s:
            return 'unknown'
        s = re.sub('[\\\\/:*?"<>|]+', '_', s)
        s = re.sub('\\s+', '_', s)
        s = re.sub('_+', '_', s).strip('._')
        return s or 'unknown'

class InequalityValidator:

    def __init__(self, rule_info):
        self.rule_info = rule_info

    def validate(self, df, column_name, second_column=None, **kwargs):
        total_rows = len(df)
        if not second_column:
            second_column = self._find_second_column(df.columns, column_name)
        if not second_column or second_column not in df.columns:
            print(f'      [WARN] Вторая колонка не найдена для сравнения')
            return (total_rows, 0, pd.DataFrame())
        v1 = df[column_name].astype(str).str.strip().str.lower().fillna('')
        v2 = df[second_column].astype(str).str.strip().str.lower().fillna('')
        both_filled = (v1 != '') & (v2 != '')
        equal = v1 == v2
        error_mask = both_filled & equal
        error_indices = df.index[error_mask].tolist()
        error_count = len(error_indices)
        if error_count > 0:
            error_df = df.loc[error_indices].copy()
            error_df['error_type'] = 'DUPLICATE_VALUES'
            error_df['error_message'] = f'{column_name} не должно быть равно {second_column}'
        else:
            error_df = pd.DataFrame()
        return (total_rows, error_count, error_df)

    def _find_second_column(self, available_columns, first_column):
        first_lower = first_column.lower()
        if '2' in first_lower or 'org2' in first_lower or 'name2' in first_lower:
            for col in available_columns:
                col_lower = col.lower()
                if '1' in col_lower or 'org1' in col_lower or 'name1' in col_lower:
                    return col
        return None
__all__ = ['FastDataQualityChecker', 'InequalityValidator']