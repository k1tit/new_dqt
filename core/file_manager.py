import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from utils.symbols import TERMINAL_SYMBOLS, ts

class ErrorFileManager:

    def __init__(self, output_dir):
        self.output_dir = output_dir
        self.all_errors_file = None
        self.all_errors = []
        self.workbook = None

    def save_errors(self, table_name, rule_code, quality_category, error_df, rule_description=''):
        if error_df is None or len(error_df) == 0:
            return None
        error_count = len(error_df)
        print(f'   {ts("INFO")} Найдено {error_count:,} ошибок, сохраняем...')
        if self.all_errors_file is None:
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            errors_dir = os.path.join(self.output_dir, f'detailed_errors_{timestamp}')
            os.makedirs(errors_dir, exist_ok=True)
            self.all_errors_file = os.path.join(errors_dir, 'ALL_ERRORS.xlsx')
            self.workbook = Workbook()
            self.workbook.remove(self.workbook.active)
            print(f'   {ts("FILE_FOLDER")} Создана папка для детальных ошибок: {errors_dir}')
        error_df_enhanced = error_df.copy()
        for col in error_df_enhanced.columns:
            if col.startswith('DQ_'):
                error_df_enhanced = error_df_enhanced.drop(columns=[col])
        error_df_enhanced.insert(0, 'DQ_TABLE_NAME', table_name)
        error_df_enhanced.insert(1, 'DQ_RULE_CODE', rule_code)
        error_df_enhanced.insert(2, 'DQ_RULE_DESCRIPTION', rule_description)
        error_df_enhanced.insert(3, 'DQ_QUALITY_CATEGORY', quality_category)
        error_df_enhanced.insert(4, 'DQ_TIMESTAMP', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        error_df_enhanced.insert(5, 'DQ_ERROR_TYPE', self._get_error_type(quality_category, rule_description))
        self.all_errors.append({'table_name': table_name, 'rule_code': rule_code, 'error_df': error_df_enhanced, 'error_count': error_count})
        if error_count > 1000:
            csv_filename = f'{rule_code}_{table_name}.csv'
            csv_path = os.path.join(os.path.dirname(self.all_errors_file), csv_filename)
            error_df_enhanced.to_csv(csv_path, index=False, encoding='utf-8-sig')
            print(f'   {ts("SAVE")} Много ошибок ({error_count:,}) - сохранен отдельный CSV: {csv_filename}')
            return csv_path
        return f'{rule_code} (будет в ALL_ERRORS.xlsx)'

    def _get_error_type(self, quality_category, rule_description):
        desc_lower = rule_description.lower()
        if quality_category == 'Completeness':
            return 'Пропущенные значения'
        elif quality_category == 'Conformity':
            if 'special character' in desc_lower:
                return 'Специальные символы'
            elif 'consecutive space' in desc_lower:
                return 'Множественные пробелы'
            elif 'capital' in desc_lower or 'upper' in desc_lower:
                return 'Регистр текста'
            elif 'cannot be the same' in desc_lower or 'equals' in desc_lower:
                return 'Совпадение значений'
            else:
                return 'Несоответствие справочнику'
        else:
            return 'Другая ошибка'

    def save_all_errors_to_excel(self):
        if not self.all_errors or not self.workbook:
            return None
        print(f'\n{ts("PALETTE")} Создаем красивый файл со всеми ошибками...')
        self._create_summary_sheet()
        self._create_error_type_sheets()
        self._create_rule_sheets()
        self.workbook.save(self.all_errors_file)
        print(f'{ts("CELEBRATION")} Файл со всеми ошибками создан: {self.all_errors_file}')
        return self.all_errors_file

    def _create_summary_sheet(self):
        ws = self.workbook.create_sheet('Сводка ошибок')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        bold_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws.merge_cells('A1:F1')
        ws['A1'] = 'СВОДКА ОШИБОК ПРОВЕРКИ КАЧЕСТВА ДАННЫХ'
        ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        total_errors = sum((item['error_count'] for item in self.all_errors))
        unique_rules = len(set((item['rule_code'] for item in self.all_errors)))
        unique_tables = len(set((item['table_name'] for item in self.all_errors)))
        ws['A3'] = 'Общая статистика:'
        ws['A3'].font = bold_font
        ws['A4'] = f'Всего ошибок: {total_errors:,}'
        ws['A5'] = f'Уникальных правил: {unique_rules}'
        ws['A6'] = f'Проверенных таблиц: {unique_tables}'
        ws['A8'] = 'Детализация по правилам:'
        ws['A8'].font = bold_font
        headers = ['№', 'Код правила', 'Таблица', 'Кол-во ошибок', 'Тип ошибки', 'Описание']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        for idx, item in enumerate(self.all_errors, 1):
            error_type = 'Не определен'
            if not item['error_df'].empty and 'DQ_ERROR_TYPE' in item['error_df'].columns:
                error_type = item['error_df']['DQ_ERROR_TYPE'].iloc[0] if len(item['error_df']) > 0 else 'Не определен'
            row_data = [idx, item['rule_code'], item['table_name'], item['error_count'], error_type, item['error_df']['DQ_RULE_DESCRIPTION'].iloc[0] if len(item['error_df']) > 0 else '']
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=9 + idx, column=col, value=value)
                cell.border = thin_border
                if col == 4 and value > 0:
                    cell.font = Font(bold=True, color='FF0000' if value > 100 else '000000')
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_error_type_sheets(self):
        errors_by_type = {}
        for item in self.all_errors:
            if not item['error_df'].empty and 'DQ_ERROR_TYPE' in item['error_df'].columns:
                error_type = item['error_df']['DQ_ERROR_TYPE'].iloc[0]
                if error_type not in errors_by_type:
                    errors_by_type[error_type] = []
                errors_by_type[error_type].append(item)
        for error_type, items in errors_by_type.items():
            sheet_name = error_type[:31] if len(error_type) > 31 else error_type
            if not sheet_name or sheet_name == 'Не определен':
                continue
            ws = self.workbook.create_sheet(sheet_name)
            all_errors_of_type = []
            for item in items:
                all_errors_of_type.append(item['error_df'])
            if all_errors_of_type:
                combined_df = pd.concat(all_errors_of_type, ignore_index=True)
                self._write_dataframe_to_sheet(ws, combined_df, f'Ошибки типа: {error_type}')

    def _create_rule_sheets(self):
        for item in self.all_errors:
            if item['error_count'] <= 500 and (not item['error_df'].empty):
                sheet_name = f"{item['rule_code']}_{item['table_name']}"
                sheet_name = sheet_name[:31]
                ws = self.workbook.create_sheet(sheet_name)
                self._write_dataframe_to_sheet(ws, item['error_df'], f'Ошибки правила: {item['rule_code']}')

    def _write_dataframe_to_sheet(self, ws, df, title=''):
        if df.empty:
            ws['A1'] = 'Нет данных'
            return
        if title:
            ws.merge_cells('A1:Z1')
            ws['A1'] = title
            ws['A1'].font = Font(bold=True, size=12, color='1F4E78')
            ws['A1'].alignment = Alignment(horizontal='center')
            start_row = 3
        else:
            start_row = 1
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
            cell.border = Border(bottom=Side(style='medium'))
        for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if headers[col_idx - 1] == 'DQ_ERROR_TYPE':
                    cell.font = Font(bold=True, color='FF0000')
                elif headers[col_idx - 1] in ['DQ_RULE_CODE', 'DQ_TABLE_NAME']:
                    cell.font = Font(bold=True)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        ws.freeze_panes = ws[f'A{start_row + 1}']

    def get_errors_directory(self):
        if self.all_errors_file:
            return os.path.dirname(self.all_errors_file)
        return None