import os
import re
import sys
import pandas as pd
import sqlite3
from glob import glob
import gc
import time

def _bootstrap_dq_project() -> str:
    here = os.path.dirname(os.path.abspath(__file__))
    parent = os.path.dirname(here)
    for candidate in (parent, here):
        if os.path.isfile(os.path.join(candidate, 'utils', 'sqlite_safe.py')):
            if candidate not in sys.path:
                sys.path.insert(0, candidate)
            return candidate
    if parent not in sys.path:
        sys.path.insert(0, parent)
    return parent
_PROJECT_ROOT = _bootstrap_dq_project()
from utils.sqlite_safe import connect_sqlite, probe_db_writable, is_lock_error
try:
    from utils.sqlite_safe import resolve_database_path
except ImportError:
    from utils.sqlite_safe import DEFAULT_DB_FILENAME

    def resolve_database_path(project_root, cli_path=None, must_exist=False):
        if cli_path and str(cli_path).strip():
            p = cli_path if os.path.isabs(cli_path) else os.path.join(project_root, cli_path)
        else:
            cfg = os.path.join(project_root, 'config', 'database.json')
            name = DEFAULT_DB_FILENAME
            if os.path.isfile(cfg):
                try:
                    import json
                    with open(cfg, encoding='utf-8') as f:
                        data = json.load(f)
                    if isinstance(data, dict) and data.get('database'):
                        name = str(data['database']).strip()
                except Exception:
                    pass
            p = os.path.join(project_root, name)
        if must_exist and (not os.path.isfile(p)):
            raise FileNotFoundError(f'Файл базы данных не найден: {p}')
        return (p, 'config/database.json или DEFAULT_DB_FILENAME')

def _resolve_db_path(db_path=None):
    if db_path is not None and str(db_path).strip():
        path, _ = resolve_database_path(_PROJECT_ROOT, db_path)
        return path
    path, _ = resolve_database_path(_PROJECT_ROOT)
    return path

def _resolve_data_path(data_folder=None):
    folder = data_folder if data_folder is not None else DEFAULT_DATA_FOLDER_REL
    if os.path.isabs(folder):
        return folder
    return os.path.join(_PROJECT_ROOT, folder)
DEFAULT_TABLE_NAME = 'T052U'
DEFAULT_DATA_FOLDER_REL = 'db'
AUSP_TABLE_NAME = 'AUSP'
AUSP_EQUIPMENT_TABLE_NAME = 'AUSP_EQUIPMENT'
AUSP_KNOWN_ATINN = ('143', '604', '148', '151')
AUSP_EQUIPMENT_ATINN = ('24', '27', '30', '52')
DEFAULT_DB_PATH = _resolve_db_path()
DEFAULT_DATA_FOLDER = _resolve_data_path()

def resolve_ausp_data_folder(project_root=None):
    root = project_root or _PROJECT_ROOT
    for rel in (os.path.join('db', 'AUSP'), 'AUSP'):
        path = os.path.join(root, rel)
        if os.path.isdir(path):
            return path
    return None


def resolve_ausp_equipment_data_folder(project_root=None):
    root = project_root or _PROJECT_ROOT
    for rel in (os.path.join('db', 'AUSP_EQUIPMENT'), 'AUSP_EQUIPMENT'):
        path = os.path.join(root, rel)
        if os.path.isdir(path):
            return path
    return None


def _ausp_target_table_for_atinn(atinn: str | None) -> str:
    a = str(atinn or '').strip()
    try:
        a = str(int(float(a))) if a else ''
    except (TypeError, ValueError):
        a = re.sub(r'\.0+$', '', a)
    if a in AUSP_EQUIPMENT_ATINN:
        return AUSP_EQUIPMENT_TABLE_NAME
    return AUSP_TABLE_NAME

def _parse_atinn_from_folder_name(name):
    n = str(name or '').strip()
    if not n:
        return None
    if n.isdigit():
        return str(int(n))
    m = re.match('^(?:ATINN[_\\s-]*)?(\\d+)$', n, re.I)
    if m:
        return str(int(m.group(1)))
    return None

def _find_atinn_column(columns):
    for c in columns:
        if re.sub('[^A-Za-z0-9]', '', str(c).upper()) == 'ATINN':
            return c
    return None

def _ausp_has_nested_atinn_folders(ausp_folder):
    if not os.path.isdir(ausp_folder) or _list_data_files(ausp_folder):
        return False
    for name in os.listdir(ausp_folder):
        sub = os.path.join(ausp_folder, name)
        if not os.path.isdir(sub) or name.startswith('.'):
            continue
        if _parse_atinn_from_folder_name(name) and _list_data_files(sub):
            return True
    return False

def _collect_ausp_atinn_file_groups(ausp_folder):
    if not os.path.isdir(ausp_folder):
        return None
    direct_files = _list_data_files(ausp_folder)
    if direct_files:
        return None
    groups = {}
    for name in sorted(os.listdir(ausp_folder)):
        sub = os.path.join(ausp_folder, name)
        if not os.path.isdir(sub) or name.startswith('.'):
            continue
        atinn = _parse_atinn_from_folder_name(name)
        if not atinn:
            print(f'  [AUSP] Пропуск подпапки (не ATINN): {name}')
            continue
        files = _list_data_files(sub)
        if not files:
            print(f'  [AUSP] Подпапка ATINN={atinn}: нет .xlsx/.xls/.csv')
            continue
        groups[atinn] = files
        print(f'  [AUSP] ATINN={atinn}: {len(files)} файл(ов) в {sub}')
    return groups or None

def _inject_atinn_column(df, atinn_value):
    if df is None or df.empty:
        return df
    out = df.copy()
    atinn_col = _find_atinn_column(out.columns)
    if atinn_col is None:
        out.insert(1 if len(out.columns) else 0, 'ATINN', str(atinn_value))
        return out
    s = out[atinn_col].astype(str).str.strip()
    empty = out[atinn_col].isna() | (s == '') | s.str.lower().isin(['none', 'nan', 'null'])
    if empty.all():
        out[atinn_col] = str(atinn_value)
    return out
sys.stdout.flush()
sys.stderr.flush()
print('=' * 80, file=sys.stderr)
print('МОДУЛЬ add_table_to_DB.py ЗАГРУЖЕН', file=sys.stderr)
print('=' * 80, file=sys.stderr)
sys.stderr.flush()

def _dedup_table_in_db(conn, table_name):
    escaped = f'"{table_name}"'
    cursor = conn.cursor()
    cursor.execute(f'SELECT COUNT(*) FROM {escaped}')
    before = cursor.fetchone()[0]
    if before == 0:
        return (before, 0, 0)
    tmp = f'"{table_name}_dedup_tmp"'
    cursor.execute(f'CREATE TABLE {tmp} AS SELECT DISTINCT * FROM {escaped}')
    cursor.execute(f'SELECT COUNT(*) FROM {tmp}')
    after = cursor.fetchone()[0]
    removed = before - after
    if removed == 0:
        cursor.execute(f'DROP TABLE {tmp}')
        return (before, after, 0)
    cursor.execute(f'DELETE FROM {escaped}')
    cursor.execute(f'INSERT INTO {escaped} SELECT * FROM {tmp}')
    cursor.execute(f'DROP TABLE {tmp}')
    return (before, after, removed)

def _sanitize_header_cell(value) -> str:
    s = str(value).strip() if value is not None else ''
    if not s or s.lower() == 'nan':
        return ''
    try:
        import unicodedata
        s = unicodedata.normalize('NFKC', s)
    except Exception:
        pass
    s = re.sub('[\\s\\u00a0\\u200b\\u200c\\u200d\\ufeff]+', '_', s)
    s = re.sub('[.\\-/\\\\:]+', '_', s)
    s = re.sub('_+', '_', s).strip('_')
    return s

def _unique_sqlite_column_names(raw_names) -> tuple[list[str], list[tuple[str, str]]]:
    result: list[str] = []
    renamed: list[tuple[str, str]] = []
    seen_count: dict[str, int] = {}
    for i, raw in enumerate(raw_names):
        base = _sanitize_header_cell(raw)
        if not base:
            base = f'col_{i}'
        key = base.upper()
        n = seen_count.get(key, 0)
        seen_count[key] = n + 1
        if n == 0:
            result.append(base)
        else:
            new_name = f'{base}_{n}'
            result.append(new_name)
            renamed.append((str(raw).strip() if raw is not None else base, new_name))
    return (result, renamed)

def _normalize_df_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    raw_cols = list(df.columns)
    new_cols, renamed = _unique_sqlite_column_names(raw_cols)
    if renamed:
        print(f'  [WARN] В выгрузке {len(renamed)} дубль(ей) заголовка — переименованы для SQLite:')
        for old, new in renamed[:8]:
            print(f'         {old!r} -> {new!r}')
        if len(renamed) > 8:
            print(f'         ... и ещё {len(renamed) - 8}')
    need_assign = raw_cols != new_cols or len(new_cols) != len(set(new_cols)) or (hasattr(df.columns, 'duplicated') and bool(df.columns.duplicated().any()))
    if need_assign:
        out = df.copy()
        out.columns = new_cols
        return out
    return df

class ProgressBar:

    @staticmethod
    def show(current, total, prefix='', suffix='', length=50):
        percent = 100 * (current / float(total))
        filled_length = int(length * current // total)
        bar = '█' * filled_length + '░' * (length - filled_length)
        percent_display = f'{percent:.1f}%'
        counter_display = f'{current:,}/{total:,}'
        sys.stdout.write(f'\r{prefix} |{bar}| {percent_display} ({counter_display}) {suffix}')
        sys.stdout.flush()
        if current == total:
            print()

def print_step(step_num, total_steps, message):
    print(f'\n[ШАГ {step_num}/{total_steps}] {message}')
    print('-' * 60)

def _list_data_files(data_folder: str) -> list[str]:
    files: list[str] = []
    for pattern in ('*.xlsx', '*.xls', '*.csv'):
        files.extend(glob(os.path.join(data_folder, pattern)))
    files = [f for f in files if not os.path.basename(f).startswith('~$')]
    return sorted(files)

def _normalize_atinn_value_simple(value) -> str:
    if value is None:
        return ''
    s = str(value).strip()
    if not s or s.lower() in ('none', 'nan', 'null'):
        return ''
    try:
        return str(int(float(s)))
    except (TypeError, ValueError):
        return re.sub(r'\.0+$', '', s)


def _peek_file_atinn_values(file_path, max_rows=8000) -> set[str]:
    """Уникальные ATINN из файла (для классификации единого дампа)."""
    try:
        df = _read_data_file(file_path, header=0, max_rows=max_rows)
    except Exception as e:
        print(f'  [WARN] не прочитать ATINN из {os.path.basename(file_path)}: {e}')
        return set()
    if df is None or df.empty:
        return set()
    col = _find_atinn_column(df.columns)
    if not col:
        return set()
    vals = set()
    for v in df[col].tolist():
        n = _normalize_atinn_value_simple(v)
        if n:
            vals.add(n)
    return vals


def _classify_flat_ausp_target(files: list[str]) -> str:
    """Единый дамп без папок ATINN: куда грузить — AUSP или AUSP_EQUIPMENT."""
    found: set[str] = set()
    for p in files[:5]:
        found |= _peek_file_atinn_values(p)
    if not found:
        return AUSP_TABLE_NAME
    eq = found & set(AUSP_EQUIPMENT_ATINN)
    cust = found & set(AUSP_KNOWN_ATINN)
    if eq and not cust:
        return AUSP_EQUIPMENT_TABLE_NAME
    if cust and not eq:
        return AUSP_TABLE_NAME
    if eq and cust:
        print(f'  [WARN] в дампе смешаны customer {sorted(cust)} и equipment {sorted(eq)} ATINN — грузим в AUSP_EQUIPMENT только если папка AUSP_EQUIPMENT; иначе AUSP')
    # по умолчанию: если есть equipment-only majority
    if len(eq) >= len(cust):
        return AUSP_EQUIPMENT_TABLE_NAME
    return AUSP_TABLE_NAME


def merge_and_load_ausp_equipment_flat(db_path=None, data_folder=None, data_files=None, skip_final_dedup=False, method='fast'):
    """Equipment AUSP: один дамп (без папок 24/27/30/52) → таблица AUSP_EQUIPMENT как есть."""
    if db_path is None:
        db_path = _resolve_db_path()
    else:
        db_path = _resolve_db_path(db_path)
    if data_files is None:
        folder = data_folder or resolve_ausp_equipment_data_folder() or os.path.join(_resolve_data_path(), AUSP_EQUIPMENT_TABLE_NAME)
        data_files = _list_data_files(folder) if folder and os.path.isdir(folder) else []
    data_files = [f for f in (data_files or []) if os.path.isfile(f)]
    if not data_files:
        print(f'ОШИБКА: нет файлов для {AUSP_EQUIPMENT_TABLE_NAME} (ожидается db/AUSP_EQUIPMENT/*.xlsx или AUSP_EQUIPMENT.xlsx)')
        return None
    print('\n' + '=' * 80)
    print(f'ЗАГРУЗКА {AUSP_EQUIPMENT_TABLE_NAME}: единый дамп (ATINN 24/27/30/52 внутри файла)')
    print('=' * 80)
    print(f'Файлов: {len(data_files)}')
    for f in data_files[:10]:
        print(f'  - {os.path.basename(f)}')
    if method == 'ultra_fast':
        return merge_and_load_xlsx_files_ultra_fast(
            db_path=db_path,
            data_folder=os.path.dirname(data_files[0]),
            target_table=AUSP_EQUIPMENT_TABLE_NAME,
            skip_final_dedup=skip_final_dedup,
            data_files=data_files,
        )
    return merge_and_load_xlsx_files_fast(
        db_path=db_path,
        data_folder=os.path.dirname(data_files[0]),
        target_table=AUSP_EQUIPMENT_TABLE_NAME,
        skip_final_dedup=skip_final_dedup,
        data_files=data_files,
    )


def merge_and_load_ausp_from_atinn_folders(db_path=None, ausp_folder=None, skip_header_after_first=True, chunksize=100000, skip_final_dedup=False):
    """Customer: подпапки ATINN 143/604/148/151 → AUSP.
    Equipment: единый дамп (папка/файлы без split) → AUSP_EQUIPMENT.
    """
    if db_path is None:
        db_path = _resolve_db_path()
    else:
        db_path = _resolve_db_path(db_path)
    folders = []
    if ausp_folder:
        folders.append(ausp_folder)
    else:
        for resolver in (resolve_ausp_data_folder, resolve_ausp_equipment_data_folder):
            p = resolver()
            if p and p not in folders:
                folders.append(p)
    if not folders:
        # плоский AUSP_EQUIPMENT.xlsx в db/
        flat_eq = [
            f for f in list_flat_data_files(_resolve_data_path())
            if _infer_table_from_stem_fallback(os.path.splitext(os.path.basename(f))[0]) == AUSP_EQUIPMENT_TABLE_NAME
            or os.path.splitext(os.path.basename(f))[0].upper().startswith('AUSP_EQUIPMENT')
        ]
        if flat_eq:
            return merge_and_load_ausp_equipment_flat(db_path=db_path, data_files=flat_eq, skip_final_dedup=skip_final_dedup)
        print('ОШИБКА: Папка AUSP / AUSP_EQUIPMENT не найдена')
        return None
    customer_pairs = []
    equipment_flat_files = []
    for folder in folders:
        name_u = os.path.basename(folder).strip().upper()
        groups = _collect_ausp_atinn_file_groups(folder)
        if not groups:
            direct = _list_data_files(folder)
            if not direct:
                continue
            # единый дамп без подпапок ATINN
            if name_u == AUSP_EQUIPMENT_TABLE_NAME:
                equipment_flat_files.extend(direct)
            else:
                target = _classify_flat_ausp_target(direct)
                if target == AUSP_EQUIPMENT_TABLE_NAME:
                    equipment_flat_files.extend(direct)
                else:
                    for p in direct:
                        customer_pairs.append((p, None))
            continue
        for atinn, paths in groups.items():
            target = _ausp_target_table_for_atinn(atinn)
            for p in paths:
                if target == AUSP_EQUIPMENT_TABLE_NAME:
                    # даже если лежат в 24/27 — для equipment можно грузить как flat list без inject
                    equipment_flat_files.append(p)
                else:
                    customer_pairs.append((p, atinn))
    # дедуп путей equipment
    seen = set()
    eq_files = []
    for p in equipment_flat_files:
        if p not in seen:
            seen.add(p)
            eq_files.append(p)
    if not customer_pairs and not eq_files:
        print('ОШИБКА: нет файлов для AUSP / AUSP_EQUIPMENT')
        return None
    print('\n' + '=' * 80)
    print('ЗАГРУЗКА AUSP: customer (по ATINN-папкам) + equipment (единый дамп)')
    print('=' * 80)
    print(f'Customer → {AUSP_TABLE_NAME}: {len(customer_pairs)} файл(ов)')
    print(f'Equipment → {AUSP_EQUIPMENT_TABLE_NAME}: {len(eq_files)} файл(ов) (без split по папкам)')
    results = []
    if customer_pairs:
        r = merge_and_load_ausp_from_file_list(
            db_path=db_path,
            file_atinn_pairs=customer_pairs,
            skip_header_after_first=skip_header_after_first,
            chunksize=chunksize,
            skip_final_dedup=skip_final_dedup,
            target_table=AUSP_TABLE_NAME,
        )
        if r:
            results.append(r)
    if eq_files:
        r = merge_and_load_ausp_equipment_flat(
            db_path=db_path,
            data_files=eq_files,
            skip_final_dedup=skip_final_dedup,
        )
        if r:
            results.append(r)
    if not results:
        return None
    if len(results) == 1:
        return results[0]
    return {
        'tables': results,
        'table_name': ','.join(r['table_name'] for r in results),
        'db_rows': sum((r.get('db_rows') or 0) for r in results),
    }

def _read_csv_all_strings(file_path, header=0, max_rows=None):
    encodings = ('utf-8-sig', 'utf-8', 'cp1251', 'latin-1')
    last_err: Exception | None = None
    read_kw = dict(header=0 if header == 0 else None, dtype=str, keep_default_na=False, nrows=max_rows, sep=None, engine='python')
    for enc in encodings:
        try:
            try:
                df = pd.read_csv(file_path, encoding=enc, on_bad_lines='warn', **read_kw)
            except TypeError:
                df = pd.read_csv(file_path, encoding=enc, error_bad_lines=False, **read_kw)
            df = df.fillna('')
            if header == 0:
                cols, renamed = _unique_sqlite_column_names(df.columns)
                for old, new in renamed:
                    print(f'  [WARN] Дубль заголовка CSV переименован: {old!r} -> {new!r}')
                df.columns = cols
            return df.astype(str)
        except Exception as e:
            last_err = e
    raise last_err if last_err else OSError(f'Не удалось прочитать CSV: {file_path}')

def _read_data_file(file_path, header=0, max_rows=None):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.csv':
        return _read_csv_all_strings(file_path, header=header, max_rows=max_rows)
    return _read_excel_all_strings(file_path, header=header, max_rows=max_rows)

def _read_excel_all_strings(file_path, header=0, max_rows=None):
    from openpyxl import load_workbook

    def _to_rows_readonly(wb, max_rows_limit=None):
        ws = wb.active
        max_excel_row = 1048576
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=max_excel_row, values_only=True):
            try:
                rows.append([str(v) if v is not None else '' for v in row])
            except Exception:
                raise
            if max_rows_limit is not None and len(rows) >= max_rows_limit:
                break
        return rows
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        try:
            rows = _to_rows_readonly(wb, max_rows)
        finally:
            wb.close()
    except Exception:
        rows = _read_excel_all_strings_full_load(file_path, max_rows)
        if rows is None:
            raise
    if not rows:
        return pd.DataFrame()
    max_len = max((len(r) for r in rows))
    rows = [list(r) + [''] * (max_len - len(r)) for r in rows]
    if header == 0:
        columns, renamed = _unique_sqlite_column_names(rows[0])
        for old, new in renamed:
            print(f'  [WARN] Дубль заголовка Excel переименован: {old!r} -> {new!r}')
        df = pd.DataFrame(rows[1:], columns=columns)
    else:
        df = pd.DataFrame(rows)
    return df.astype(str)

def _read_excel_all_strings_full_load(file_path, max_rows=None):
    from openpyxl import load_workbook
    import openpyxl.worksheet._reader as _ox_reader
    import openpyxl.utils.datetime as _ox_dt
    _orig_cast = getattr(_ox_reader, '_cast_number', None)
    _orig_from_excel = getattr(_ox_dt, 'from_excel', None)
    _reader_has_from_excel = hasattr(_ox_reader, 'from_excel')

    def _safe_cast_number(value):
        try:
            return float(value)
        except (ValueError, TypeError):
            return value

    def _safe_from_excel(value, *args, **kwargs):
        if not isinstance(value, (int, float)):
            return value
        if _orig_from_excel is None:
            return value
        try:
            return _orig_from_excel(value, *args, **kwargs)
        except (ValueError, TypeError, ZeroDivisionError):
            return value
    try:
        _ox_reader._cast_number = _safe_cast_number
        _ox_dt.from_excel = _safe_from_excel
        if _reader_has_from_excel:
            _ox_reader.from_excel = _safe_from_excel
        wb = load_workbook(filename=file_path, read_only=False, data_only=True)
        ws = wb.active
        max_excel_row = 1048576
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=max_excel_row, values_only=True):
            rows.append([str(v) if v is not None else '' for v in row])
            if max_rows is not None and len(rows) >= max_rows:
                break
        wb.close()
    finally:
        if _orig_cast is not None:
            _ox_reader._cast_number = _orig_cast
        if _orig_from_excel is not None:
            _ox_dt.from_excel = _orig_from_excel
        if _reader_has_from_excel and _orig_from_excel is not None:
            _ox_reader.from_excel = _orig_from_excel
    return rows

def merge_and_load_xlsx_files_fast(db_path=None, data_folder=None, target_table=None, skip_header_after_first=True, chunksize=100000, skip_final_dedup=False, data_files=None):
    if db_path is None:
        db_path = _resolve_db_path()
    else:
        db_path = _resolve_db_path(db_path)
    if data_folder is None:
        data_folder = _resolve_data_path()
    else:
        data_folder = _resolve_data_path(data_folder)
    if target_table is None:
        target_table = DEFAULT_TABLE_NAME
    print('\n' + '=' * 80)
    print('ЗАГРУЗКА ТАБЛИЦЫ В БАЗУ ДАННЫХ')
    print('=' * 80)
    print(f'Папка с данными: {data_folder if data_files is None else "(список файлов)"}')
    print(f'Таблица: {target_table}')
    print(f'База данных: {db_path}')
    print('=' * 80 + '\n')
    total_steps = 4
    current_step = 1
    print_step(current_step, total_steps, 'Проверка файлов...')
    current_step += 1
    if data_files is None:
        if not os.path.exists(data_folder):
            print(f"ОШИБКА: Папка '{data_folder}' не найдена!")
            return None
        data_files = _list_data_files(data_folder)
    else:
        data_files = [f for f in data_files if os.path.isfile(f) and (not os.path.basename(f).startswith('~$'))]
        data_files = sorted(data_files)
    if not data_files:
        print(f"ОШИБКА: Нет файлов .xlsx / .xls / .csv{' в папке ' + repr(data_folder) if data_folder else ''}")
        return None
    total_files = len(data_files)
    print(f'Найдено файлов: {total_files}')
    for i, f in enumerate(data_files[:10], 1):
        print(f'  {i:2}. {os.path.basename(f)}')
    if total_files > 10:
        print(f'  ... и еще {total_files - 10} файлов')
    print_step(current_step, total_steps, 'Подготовка базы данных...')
    current_step += 1
    try:
        print('Подключение к базе данных...')
        conn = connect_sqlite(db_path)
        cursor = conn.cursor()
        print('Включение оптимизаций...')
        conn.execute('PRAGMA journal_mode = OFF')
        conn.execute('PRAGMA synchronous = OFF')
        conn.execute('PRAGMA cache_size = -20000')
        conn.execute('PRAGMA foreign_keys = OFF')
        conn.execute('PRAGMA temp_store = MEMORY')
        print('Оптимизации включены')
        print(f"Удаление старой таблицы '{target_table}'...")
        cursor.execute(f"DROP TABLE IF EXISTS '{target_table}'")
        conn.commit()
        print('Старая таблица удалена')
    except Exception as e:
        print(f'ОШИБКА при подготовке БД: {e}')
        return None
    print_step(current_step, total_steps, f'Загрузка {total_files} файлов...')
    current_step += 1
    first_file_columns = None
    table_created = False
    total_rows_loaded = 0
    start_time = time.time()
    print('\nНачало загрузки файлов:')
    print('=' * 70)
    for file_idx, file_path in enumerate(data_files, 1):
        file_name = os.path.basename(file_path)
        file_size = os.path.getsize(file_path) / (1024 * 1024)
        print(f'\nФайл {file_idx}/{total_files}: {file_name} ({file_size:.1f} MB)')
        try:
            read_start = time.time()
            if file_idx == 1:
                print(f'  Чтение с заголовками...')
                df = _normalize_df_columns(_read_data_file(file_path, header=0))
                first_file_columns = list(df.columns)
                print(f'  Столбцов: {len(first_file_columns)}')
            elif skip_header_after_first and first_file_columns:
                print(f'  Чтение без заголовков...')
                df = _read_data_file(file_path, header=None)
                if len(df) > 0:
                    rows_before = len(df)
                    df = df.iloc[1:].reset_index(drop=True)
                    rows_removed = rows_before - len(df)
                    if rows_removed > 0:
                        print(f'  Удалено заголовков: {rows_removed}')
                if len(df.columns) == len(first_file_columns):
                    df.columns = first_file_columns
                elif len(df.columns) > len(first_file_columns):
                    df.columns = first_file_columns + [f'extra_col_{j}' for j in range(len(first_file_columns), len(df.columns))]
                else:
                    df.columns = first_file_columns[:len(df.columns)]
            else:
                print(f'  Чтение с заголовками...')
                df = _normalize_df_columns(_read_data_file(file_path, header=0))
            df = _normalize_df_columns(df)
            read_time = time.time() - read_start
            print(f'  Прочитано: {len(df):,} строк за {read_time:.1f} сек')
            if len(df) > 0:
                before_dedup = len(df)
                df = df.drop_duplicates()
                if len(df) < before_dedup:
                    print(f'  Удалено дубликатов: {before_dedup - len(df):,} (осталось {len(df):,})')
            if not table_created and len(df) > 0:
                print(f'  Создание таблицы в БД...')
                create_start = time.time()
                sample_df = df.head(min(10000, len(df)))
                sample_df.to_sql(target_table, conn, if_exists='fail', index=False, chunksize=chunksize)
                table_created = True
                create_time = time.time() - create_start
                print(f'  Таблица создана за {create_time:.1f} сек')
                if len(df) > len(sample_df):
                    remaining_df = df.iloc[len(sample_df):]
                    print(f'  Запись оставшихся {len(remaining_df):,} строк...')
                    remaining_df.to_sql(target_table, conn, if_exists='append', index=False, chunksize=chunksize)
            elif table_created and len(df) > 0:
                print(f'  Добавление данных в таблицу...')
                write_start = time.time()
                df.to_sql(target_table, conn, if_exists='append', index=False, chunksize=chunksize)
                write_time = time.time() - write_start
                print(f'  Записано за {write_time:.1f} сек')
            total_rows_loaded += len(df)
            progress_percent = file_idx / total_files * 100
            ProgressBar.show(file_idx, total_files, prefix=f'Прогресс:', suffix=f'Файлов: {file_idx}/{total_files} | Строк: {total_rows_loaded:,}')
            del df
            gc.collect()
        except Exception as e:
            print(f'\n  ОШИБКА при обработке файла {file_name}: {str(e)}')
            print(f'  Пропускаем файл и продолжаем...')
            continue
    conn.commit()
    if not skip_final_dedup:
        try:
            cursor.execute(f"SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name=?", (target_table,))
            if cursor.fetchone()[0]:
                before_dedup, after_dedup, removed = _dedup_table_in_db(conn, target_table)
                if removed > 0:
                    print(f'\n  Удалено дубликатов в таблице: {removed:,} (было {before_dedup:,}, стало {after_dedup:,})')
                conn.commit()
        except Exception as e:
            print(f'\n  [WARN] Очистка дубликатов не выполнена: {e}')
    conn.execute('PRAGMA journal_mode = WAL')
    conn.execute('PRAGMA synchronous = NORMAL')
    try:
        cursor.execute(f"SELECT COUNT(*) FROM '{target_table}'")
        count = cursor.fetchone()[0]
    except sqlite3.OperationalError:
        count = 0
        print(f"\nВНИМАНИЕ: Таблица '{target_table}' не создана (все файлы завершились с ошибкой).")
    total_time = time.time() - start_time
    conn.close()
    print_step(current_step, total_steps, 'Завершение...')
    print('\n' + '=' * 80)
    print('ЗАГРУЗКА ЗАВЕРШЕНА')
    print('=' * 80)
    print(f'\nСТАТИСТИКА:')
    print(f'- Таблица: {target_table}')
    print(f'- Файлов обработано: {total_files}')
    print(f'- Строк загружено: {count:,}')
    print(f'- Общее время: {total_time:.1f} секунд')
    if count > 0 and total_time > 0:
        rows_per_second = count / total_time
        print(f'- Скорость: {rows_per_second:,.0f} строк/сек')
    print(f'- Результат: УСПЕХ')
    print('\n' + '=' * 80)
    return {'table_name': target_table, 'source_files': total_files, 'db_rows': count, 'total_time': total_time, 'rows_per_second': count / total_time if total_time > 0 else 0}

def merge_and_load_xlsx_files_ultra_fast(db_path=None, data_folder=None, target_table=None, skip_header_after_first=True, batch_size=100000, skip_final_dedup=False, data_files=None):
    if db_path is None:
        db_path = _resolve_db_path()
    else:
        db_path = _resolve_db_path(db_path)
    if data_folder is None:
        data_folder = _resolve_data_path()
    else:
        data_folder = _resolve_data_path(data_folder)
    if target_table is None:
        target_table = DEFAULT_TABLE_NAME
    print('\n' + '=' * 80)
    print('УЛЬТРА-БЫСТРАЯ ЗАГРУЗКА ТАБЛИЦЫ')
    print('=' * 80)
    try:
        if data_files is None:
            if not os.path.exists(data_folder):
                raise FileNotFoundError(f"Папка '{data_folder}' не найдена!")
            data_files = _list_data_files(data_folder)
        else:
            data_files = [f for f in data_files if os.path.isfile(f) and (not os.path.basename(f).startswith('~$'))]
            data_files = sorted(data_files)
        if not data_files:
            print(f"ОШИБКА: Нет файлов .xlsx / .xls / .csv{' в папке ' + repr(data_folder) if data_folder else ''}")
            return None
        total_files = len(data_files)
        print(f'Найдено файлов: {total_files}')
        if os.path.exists(db_path):
            import datetime

            def _run_quick_check():
                c = connect_sqlite(db_path)
                try:
                    row = c.execute('PRAGMA quick_check(1)').fetchone()
                    val = (row[0] if row else '') or ''
                    if str(val).strip().lower() != 'ok':
                        raise sqlite3.DatabaseError(f'quick_check: {val}')
                finally:
                    try:
                        c.close()
                    except Exception:
                        pass
            try:
                _run_quick_check()
            except sqlite3.OperationalError as e:
                if is_lock_error(e):
                    print('\nБД занята (database is locked) — ждём освобождения файла...')
                    ok_probe, err_probe = probe_db_writable(db_path)
                    if not ok_probe:
                        print('Не удалось получить доступ к БД для записи.')
                        print('Закройте программы с открытым файлом: DB Browser for SQLite, другой Python/загрузчик, Excel-плагины и т.п.')
                        if err_probe:
                            print(f'Детали: {err_probe}')
                        return None
                    try:
                        _run_quick_check()
                    except sqlite3.OperationalError as e2:
                        print(f'ОШИБКА: БД снова недоступна: {e2}')
                        return None
                else:
                    raise
            except sqlite3.DatabaseError as e:
                print(f'\nВНИМАНИЕ: База данных повреждена или не прошла проверку ({e}).')
                print('Если переименовать её в бэкап и создать новую — в новой БД будет только загружаемая таблица.')
                print('Остальные таблицы останутся только в файле-бэкапе.')
                try:
                    answer = input('Всё равно переименовать в бэкап и создать новую БД? (y/n): ').strip().lower()
                except Exception:
                    answer = 'n'
                if answer != 'y' and answer != 'yes':
                    print('Загрузка отменена. Восстановите БД из бэкапа (db_mrt.db.corrupted_*) или исправьте файл вручную.')
                    return None
                backup_name = f'{db_path}.corrupted_{datetime.datetime.now():%Y%m%d_%H%M%S}'
                print(f'Файл переименован в: {backup_name}')
                print('Создаётся новая пустая БД.\n')
                os.rename(db_path, backup_name)
        conn = connect_sqlite(db_path)
        conn.execute('PRAGMA journal_mode = OFF')
        conn.execute('PRAGMA synchronous = OFF')
        conn.execute('PRAGMA cache_size = -50000')
        conn.execute('PRAGMA foreign_keys = OFF')
        conn.execute('PRAGMA temp_store = MEMORY')
        conn.execute('PRAGMA locking_mode = NORMAL')
        cursor = conn.cursor()
        cursor.execute(f"DROP TABLE IF EXISTS '{target_table}'")
        first_file_columns = None
        total_rows = 0
        start_time = time.time()
        print('\nНачало ультра-быстрой загрузки...')
        print('=' * 70)
        for file_idx, file_path in enumerate(data_files, 1):
            file_name = os.path.basename(file_path)
            print(f'\nФайл {file_idx}/{total_files}: {file_name}')
            read_start = time.time()
            if file_idx == 1:
                df = _normalize_df_columns(_read_data_file(file_path, header=0))
                first_file_columns = list(df.columns)
            elif skip_header_after_first and first_file_columns:
                df = _read_data_file(file_path, header=None)
                if len(df) > 0:
                    df = df.iloc[1:].reset_index(drop=True)
                if len(df.columns) == len(first_file_columns):
                    df.columns = first_file_columns
                elif len(df.columns) > len(first_file_columns):
                    df.columns = first_file_columns + [f'extra_col_{j}' for j in range(len(first_file_columns), len(df.columns))]
                else:
                    df.columns = first_file_columns[:len(df.columns)]
            else:
                df = _normalize_df_columns(_read_data_file(file_path, header=0))
            df = _normalize_df_columns(df)
            read_time = time.time() - read_start
            print(f'  Прочитано: {len(df):,} строк за {read_time:.1f} сек')
            if len(df) > 0:
                before_dedup = len(df)
                df = df.drop_duplicates()
                if len(df) < before_dedup:
                    print(f'  Удалено дубликатов: {before_dedup - len(df):,} (осталось {len(df):,})')
            if file_idx == 1 and len(df) > 0:
                print(f'  Создание таблицы...')
                create_start = time.time()
                col_defs = []
                for col in df.columns:
                    col_defs.append(f'"{col}" TEXT')
                create_sql = f'CREATE TABLE "{target_table}" (\n'
                create_sql += ',\n'.join(col_defs)
                create_sql += '\n)'
                cursor.execute(create_sql)
                create_time = time.time() - create_start
                print(f'  Таблица создана за {create_time:.1f} сек')
            if len(df) > 0:
                print(f'  Запись данных...')
                write_start = time.time()
                num_cols = len(df.columns)
                data_tuples = [tuple(x) for x in df.values]
                SQLITE_MAX_VARS = 999
                rows_per_stmt = max(1, SQLITE_MAX_VARS // num_cols)
                one_row_ph = ','.join(['?' for _ in range(num_cols)])
                multi_ph = ','.join([f'({one_row_ph})' for _ in range(rows_per_stmt)])
                insert_sql_multi = f'INSERT INTO "{target_table}" VALUES {multi_ph}'
                insert_sql_single = f'INSERT INTO "{target_table}" VALUES ({one_row_ph})'
                for batch_start in range(0, len(data_tuples), batch_size):
                    batch = data_tuples[batch_start:batch_start + batch_size]
                    for i in range(0, len(batch), rows_per_stmt):
                        chunk = batch[i:i + rows_per_stmt]
                        if len(chunk) == rows_per_stmt:
                            flat = [v for row in chunk for v in row]
                            cursor.execute(insert_sql_multi, flat)
                        else:
                            for row in chunk:
                                cursor.execute(insert_sql_single, row)
                    current_progress = min(batch_start + batch_size, len(data_tuples))
                    percent = current_progress / len(data_tuples) * 100
                    print(f'    Записано: {current_progress:,}/{len(data_tuples):,} строк ({percent:.1f}%)', end='\r')
                print(f'    Записано: {len(data_tuples):,} строк')
                write_time = time.time() - write_start
                print(f'  Записано за {write_time:.1f} сек')
                total_rows += len(df)
                ProgressBar.show(file_idx, total_files, prefix=f'Общий прогресс:', suffix=f'Файлов: {file_idx}/{total_files} | Строк: {total_rows:,}')
            del df
            gc.collect()
        if not skip_final_dedup:
            try:
                before_dedup, after_dedup, removed = _dedup_table_in_db(conn, target_table)
                if removed > 0:
                    print(f'\n  Удалено дубликатов в таблице: {removed:,} (было {before_dedup:,}, стало {after_dedup:,})')
            except Exception as e:
                print(f'\n  [WARN] Очистка дубликатов не выполнена: {e}')
        conn.commit()
        conn.execute('PRAGMA journal_mode = WAL')
        conn.execute('PRAGMA synchronous = NORMAL')
        conn.execute('PRAGMA locking_mode = NORMAL')
        cursor.execute(f"SELECT COUNT(*) FROM '{target_table}'")
        count = cursor.fetchone()[0]
        total_time = time.time() - start_time
        conn.close()
        print('\n' + '=' * 80)
        print('УЛЬТРА-БЫСТРАЯ ЗАГРУЗКА ЗАВЕРШЕНА')
        print('=' * 80)
        print(f'\nСТАТИСТИКА:')
        print(f'- Таблица: {target_table}')
        print(f'- Файлов: {total_files}')
        print(f'- Строк: {count:,}')
        print(f'- Время: {total_time:.1f} сек')
        if total_time > 0:
            speed = count / total_time
            print(f'- Скорость: {speed:,.0f} строк/сек')
            print(f'- Метод: Ультра-быстрый')
        print('\n' + '=' * 80)
        return {'table_name': target_table, 'source_files': total_files, 'db_rows': count, 'total_time': total_time}
    except Exception as e:
        print(f'\nОШИБКА: {e}')
        import traceback
        traceback.print_exc()
        return None

def load_data_fast():
    return merge_and_load_xlsx_files_fast(db_path=DEFAULT_DB_PATH, data_folder=DEFAULT_DATA_FOLDER, target_table=DEFAULT_TABLE_NAME, skip_header_after_first=True, chunksize=100000)

def load_data_ultra_fast():
    return merge_and_load_xlsx_files_ultra_fast(db_path=DEFAULT_DB_PATH, data_folder=DEFAULT_DATA_FOLDER, target_table=DEFAULT_TABLE_NAME, skip_header_after_first=True, batch_size=50000)

def merge_and_load_xlsx_files(*args, **kwargs):
    return merge_and_load_xlsx_files_fast(*args, **kwargs)

def load_data():
    return load_data_fast()

def load_data_to_custom_table(table_name):
    print(f'\nЗагрузка данных в таблицу: {table_name}')
    result = merge_and_load_xlsx_files_fast(db_path=DEFAULT_DB_PATH, data_folder=DEFAULT_DATA_FOLDER, target_table=table_name, skip_header_after_first=True, chunksize=100000)
    if result is None:
        print(f"ОШИБКА: Не удалось загрузить данные в таблицу '{table_name}'")
    else:
        print(f"УСПЕХ: Данные загружены в таблицу '{result['table_name']}'")
    return result

def get_table_folders(base_folder=None):
    """Имена таблиц для загрузки: подпапки db/ИМЯ + плоские Excel в db/."""
    groups = collect_db_load_groups(base_folder)
    return sorted(groups.keys(), key=lambda s: str(s).upper())


# --- Excel flat-folder load matched to rules.json ---

_AUSP_DERIVED_TO_ATINN = {
    'AUSP_143': '143',
    'AUSP_604': '604',
    'AUSP_148': '148',
    'AUSP_151': '151',
    'AUSP_24': '24',
    'AUSP_27': '27',
    'AUSP_30': '30',
    'AUSP_52': '52',
}
_AUSP_EQUIPMENT_ATINN_KEYS = frozenset({'AUSP_24', 'AUSP_27', 'AUSP_30', 'AUSP_52', 'AUSP_EQUIPMENT'})
_DUMP_NUM_SUFFIX_RE = re.compile(r'^(?P<base>.+?)(?:[\s_\-\.]*\(?\d+\)?)$', re.IGNORECASE)
_AUSP_ATINN_STEM_RE = re.compile(r'^AUSP[_\-\s]*(?P<atinn>\d+)$', re.IGNORECASE)


def default_rules_path() -> str:
    return os.path.join(_PROJECT_ROOT, 'json files', 'rules.json')


def _split_table_token(raw: str) -> list[str]:
    """Разбивает 'KNB1 / KNVV', но сохраняет SAP-пути вида /LOT/GC_ADR."""
    s = str(raw or '').strip()
    if not s:
        return []
    if s.startswith('/'):
        return [s]
    if re.search(r'\s/\s', s):
        return [p.strip() for p in re.split(r'\s*/\s*', s) if p.strip()]
    return [s]


def load_rule_table_names(rules_path=None) -> list[str]:
    """Имена таблиц из ключей rules.json (+ table_name_checked)."""
    import json
    path = rules_path or default_rules_path()
    with open(path, encoding='utf-8') as f:
        data = json.load(f)
    names: set[str] = set()
    if isinstance(data, dict):
        for key in data.keys():
            for part in _split_table_token(key):
                names.add(part)
            rules = data.get(key) or []
            if isinstance(rules, list):
                for rule in rules:
                    if not isinstance(rule, dict):
                        continue
                    raw = str(rule.get('table_name_checked') or rule.get('table_name') or '').strip()
                    if not raw:
                        continue
                    for part in _split_table_token(raw):
                        names.add(part)
    return sorted(names, key=lambda s: (-len(s), s.upper()))


def table_name_aliases(table_name: str) -> list[str]:
    t = str(table_name or '').strip()
    if not t:
        return []
    aliases = [t]
    if t == '/LOT/GC_ADR' or t.upper().replace('/', '').replace('_', '') == 'LOTGCADR':
        aliases.extend(['/LOT/GC_ADR', 'LOT_GC_ADR', 'LOTGC_ADR', '_LOT_GC_ADR', 'LOT-GC-ADR', 'LOTGCADR'])
    if '/' in t:
        aliases.append(t.replace('/', '_').strip('_'))
        aliases.append(t.replace('/', ''))
        aliases.append(t.replace('/', '_'))
    # unique preserve order
    seen = set()
    out = []
    for a in aliases:
        key = a.upper()
        if key and key not in seen:
            seen.add(key)
            out.append(a)
    return out


def _build_alias_to_table(known_tables: list[str]) -> dict[str, str]:
    """alias_upper -> canonical table (longest tables registered first)."""
    alias_map: dict[str, str] = {}
    for table in sorted(known_tables, key=lambda s: (-len(s), s.upper())):
        for alias in table_name_aliases(table):
            key = alias.upper()
            if key not in alias_map:
                alias_map[key] = table
    return alias_map


def match_stem_to_rule_table(stem: str, alias_map: dict[str, str]) -> str | None:
    """Сопоставляет имя файла (без расширения) с таблицей из правил.

    Поддерживает: KNA1, KNA1_1, KNA1-2, KNA1 (3), AUSP_143, AUSP_143_1.
    Сначала точное совпадение (чтобы DFKKBPTAXNUM1 / AUSP_143 не срезались).
    """
    raw = str(stem or '').strip()
    if not raw:
        return None
    # иногда имена с путём SAP: _LOT_GC_ADR
    candidates = [raw, raw.replace(' ', '_')]
    for cand in candidates:
        hit = alias_map.get(cand.upper())
        if hit:
            return hit
    # пронумерованные выгрузки: strip trailing dump index
    for cand in candidates:
        m = _DUMP_NUM_SUFFIX_RE.match(cand)
        if not m:
            continue
        base = m.group('base').rstrip(' ._-\t')
        if not base or base.upper() == cand.upper():
            continue
        hit = alias_map.get(base.upper())
        if hit:
            return hit
        # ещё один уровень: AUSP_143_1 → после первого strip уже AUSP_143
        m2 = _DUMP_NUM_SUFFIX_RE.match(base)
        if m2:
            base2 = m2.group('base').rstrip(' ._-\t')
            hit2 = alias_map.get(base2.upper()) if base2 else None
            if hit2:
                return hit2
    return None


def list_flat_data_files(folder: str) -> list[str]:
    folder_abs = os.path.abspath(folder)
    if not os.path.isdir(folder_abs):
        return []
    return _list_data_files(folder_abs)


def _infer_table_from_stem_fallback(stem: str) -> str | None:
    """Если rules нет/не совпало: KNA1_1 → KNA1, AUSP_143 → AUSP, AUSP_24 → AUSP_EQUIPMENT."""
    raw = str(stem or '').strip()
    if not raw:
        return None
    m_ausp = _AUSP_ATINN_STEM_RE.match(raw)
    if m_ausp:
        return _ausp_target_table_for_atinn(m_ausp.group('atinn'))
    m = _DUMP_NUM_SUFFIX_RE.match(raw)
    if m:
        base = m.group('base').rstrip(' ._-\t')
        if base and base.upper() != raw.upper():
            m2 = _AUSP_ATINN_STEM_RE.match(base)
            if m2:
                return _ausp_target_table_for_atinn(m2.group('atinn'))
            return base
    return raw


def _split_ausp_file_pairs(file_atinn_pairs):
    """Делит [(path, atinn)] на customer AUSP и AUSP_EQUIPMENT. Не смешивает таблицы."""
    customer, equipment = [], []
    for path, atinn in file_atinn_pairs or []:
        target = _ausp_target_table_for_atinn(atinn)
        if target == AUSP_EQUIPMENT_TABLE_NAME:
            equipment.append((path, atinn))
        else:
            customer.append((path, atinn))
    return customer, equipment


def _load_ausp_pairs_split(db_path, file_atinn_pairs, skip_final_dedup=False, only_target=None):
    """Грузит пары в AUSP и/или AUSP_EQUIPMENT отдельно (DROP только своей таблицы)."""
    customer, equipment = _split_ausp_file_pairs(file_atinn_pairs)
    if only_target == AUSP_TABLE_NAME:
        equipment = []
    elif only_target == AUSP_EQUIPMENT_TABLE_NAME:
        customer = []
    results = []
    if customer:
        print(f'   → {AUSP_TABLE_NAME}: {len(customer)} файл(ов) (ATINN 143/604/148/151) — equipment не трогаем')
        r = merge_and_load_ausp_from_file_list(
            db_path=db_path,
            file_atinn_pairs=customer,
            skip_final_dedup=skip_final_dedup,
            target_table=AUSP_TABLE_NAME,
        )
        if r:
            results.append(r)
    if equipment:
        print(f'   → {AUSP_EQUIPMENT_TABLE_NAME}: {len(equipment)} файл(ов) (ATINN 24/27/30/52) — customer AUSP не трогаем')
        r = merge_and_load_ausp_from_file_list(
            db_path=db_path,
            file_atinn_pairs=equipment,
            skip_final_dedup=skip_final_dedup,
            target_table=AUSP_EQUIPMENT_TABLE_NAME,
        )
        if r:
            results.append(r)
    if not results:
        return None
    if len(results) == 1:
        return results[0]
    return {
        'tables': results,
        'table_name': ','.join(r['table_name'] for r in results),
        'db_rows': sum((r.get('db_rows') or 0) for r in results),
    }


def _register_ausp_atinn_folder_groups(groups: dict, folder: str, atinn_groups: dict):
    """Раскладывает подпапки ATINN по AUSP / AUSP_EQUIPMENT без смешивания."""
    for atinn, paths in atinn_groups.items():
        target = _ausp_target_table_for_atinn(atinn)
        g = groups.get(target) or {'mode': 'files', 'files': [], 'file_atinn': {}, 'folder': None}
        # folder mode только если все файлы из одной корневой AUSP-папки; для split всё равно грузим по file list
        g['mode'] = 'files'
        if not g.get('folder'):
            g['folder'] = folder
        for p in paths:
            if p not in g['files']:
                g['files'].append(p)
            g['file_atinn'][p] = str(atinn)
        groups[target] = g
        print(f'  [{target}] ATINN={atinn}: {len(paths)} файл(ов)')


def _atinn_from_stem(stem: str) -> str | None:
    s = str(stem or '').strip()
    if not s:
        return None
    if s.upper() in _AUSP_DERIVED_TO_ATINN:
        return _AUSP_DERIVED_TO_ATINN[s.upper()]
    # AUSP_143 / AUSP-143
    m = _AUSP_ATINN_STEM_RE.match(s)
    if m:
        return str(int(m.group('atinn')))
    # AUSP_143_1 → strip dump index once
    m2 = _DUMP_NUM_SUFFIX_RE.match(s)
    if m2:
        base = m2.group('base').rstrip(' ._-\t')
        if base and base.upper() != s.upper():
            if base.upper() in _AUSP_DERIVED_TO_ATINN:
                return _AUSP_DERIVED_TO_ATINN[base.upper()]
            m3 = _AUSP_ATINN_STEM_RE.match(base)
            if m3:
                return str(int(m3.group('atinn')))
    return None


def collect_db_load_groups(base_folder=None, rules_path=None) -> dict:
    """Источники загрузки в обычной папке db/: подпапки + плоские Excel.

    Возвращает:
      { table_name: { 'mode': 'files'|'ausp_atinn_folders', 'files': [...], 'file_atinn': {path: atinn}, 'folder': optional } }
    """
    base_abs = _resolve_data_path(base_folder)
    groups: dict = {}
    if not os.path.isdir(base_abs):
        return groups

    known: list[str] = []
    alias_map: dict[str, str] = {}
    try:
        rules_file = rules_path or default_rules_path()
        if os.path.isfile(rules_file):
            known = load_rule_table_names(rules_file)
            alias_map = _build_alias_to_table(known)
    except Exception as e:
        print(f'  [WARN] rules.json не прочитан ({e}) — плоские файлы по имени stem')

    def _ensure_files_group(table: str) -> dict:
        g = groups.get(table)
        if g is None:
            g = {'mode': 'files', 'files': [], 'file_atinn': {}}
            groups[table] = g
        elif g.get('mode') == 'ausp_atinn_folders':
            # уже режим папок ATINN — доп. файлы копятся в files/file_atinn
            g.setdefault('files', [])
            g.setdefault('file_atinn', {})
        else:
            g.setdefault('files', [])
            g.setdefault('file_atinn', {})
        return g

    def _resolve_table_name(stem_or_folder: str) -> str | None:
        if alias_map:
            hit = match_stem_to_rule_table(stem_or_folder, alias_map)
            if hit:
                return hit
        return _infer_table_from_stem_fallback(stem_or_folder)

    def _add_file(table: str, path: str, atinn: str | None = None):
        # AUSP_* / ATINN → AUSP (customer) или AUSP_EQUIPMENT; никогда не смешивать
        if table in _AUSP_DERIVED_TO_ATINN:
            atinn = atinn or _AUSP_DERIVED_TO_ATINN[table]
            table = _ausp_target_table_for_atinn(atinn)
        tu = str(table).strip().upper()
        if atinn:
            table = _ausp_target_table_for_atinn(atinn)
            tu = str(table).strip().upper()
        if tu == AUSP_EQUIPMENT_TABLE_NAME or tu in _AUSP_EQUIPMENT_ATINN_KEYS:
            table = AUSP_EQUIPMENT_TABLE_NAME
        elif tu == AUSP_TABLE_NAME:
            table = AUSP_TABLE_NAME
        g = _ensure_files_group(table)
        if path not in g['files']:
            g['files'].append(path)
        if atinn:
            g['file_atinn'][path] = str(atinn)

    # 1) Подпапки db/ИМЯ_ТАБЛИЦЫ (как раньше)
    for name in sorted(os.listdir(base_abs)):
        path = os.path.join(base_abs, name)
        if not os.path.isdir(path) or name.startswith('.') or name == '__pycache__':
            continue
        name_u = str(name).strip().upper()
        table = _resolve_table_name(name) or name
        if name_u in (AUSP_TABLE_NAME, AUSP_EQUIPMENT_TABLE_NAME) or str(table).strip().upper() in (AUSP_TABLE_NAME, AUSP_EQUIPMENT_TABLE_NAME):
            # Equipment: единый дамп без папок ATINN → сразу AUSP_EQUIPMENT
            if name_u == AUSP_EQUIPMENT_TABLE_NAME:
                direct = _list_data_files(path)
                atinn_groups = _collect_ausp_atinn_file_groups(path)
                if atinn_groups:
                    # если вдруг есть подпапки — всё равно всё в AUSP_EQUIPMENT одним дампом
                    for _a, paths in atinn_groups.items():
                        direct.extend(paths)
                for p in sorted(set(direct)):
                    _add_file(AUSP_EQUIPMENT_TABLE_NAME, p, None)
                if direct:
                    print(f'  [{AUSP_EQUIPMENT_TABLE_NAME}] единый дамп: {len(set(direct))} файл(ов) (ATINN внутри файла)')
                continue
            atinn_groups = _collect_ausp_atinn_file_groups(path)
            if atinn_groups:
                _register_ausp_atinn_folder_groups(groups, path, atinn_groups)
                continue
            # плоские файлы в db/AUSP/ — классифицируем по ATINN в данных
            direct = _list_data_files(path)
            if direct:
                target = _classify_flat_ausp_target(direct)
                print(f'  [AUSP flat] классификация дампа → {target}')
                for p in direct:
                    _add_file(target, p, None)
            continue
        files = _list_data_files(path)
        if not files:
            continue
        for p in files:
            _add_file(table, p, None)

    # AUSP / AUSP_EQUIPMENT вне db/ (корень проекта)
    for resolver, default_name in (
        (resolve_ausp_data_folder, AUSP_TABLE_NAME),
        (resolve_ausp_equipment_data_folder, AUSP_EQUIPMENT_TABLE_NAME),
    ):
        extra = resolver()
        if not extra:
            continue
        if os.path.abspath(extra) == os.path.abspath(os.path.join(base_abs, default_name)):
            continue
        # уже собрали из db/?
        if default_name in groups and groups[default_name].get('files'):
            continue
        atinn_groups = _collect_ausp_atinn_file_groups(extra)
        if atinn_groups:
            _register_ausp_atinn_folder_groups(groups, extra, atinn_groups)
        else:
            for p in _list_data_files(extra):
                stem = os.path.splitext(os.path.basename(p))[0]
                _add_file(default_name, p, _atinn_from_stem(stem))

    # 2) Плоские файлы прямо в db/ (KNA1.xlsx, KNA1_1.xlsx, V_EQUI.xlsx, ...)
    for path in list_flat_data_files(base_abs):
        stem = os.path.splitext(os.path.basename(path))[0]
        table = _resolve_table_name(stem)
        if not table:
            print(f'  [SKIP] не удалось определить таблицу: {os.path.basename(path)}')
            continue
        atinn = None
        if table in _AUSP_DERIVED_TO_ATINN or str(table).upper() in (AUSP_TABLE_NAME, AUSP_EQUIPMENT_TABLE_NAME) or _AUSP_ATINN_STEM_RE.match(stem):
            atinn = _atinn_from_stem(stem)
        _add_file(table, path, atinn)

    # нормализация списков
    for g in groups.values():
        g['files'] = sorted(set(g.get('files') or []))
    return groups


def discover_excel_by_rules(folder: str, rules_path=None) -> dict:
    """Сканирует плоскую папку Excel/CSV и группирует по таблицам из rules.json."""
    known = load_rule_table_names(rules_path)
    alias_map = _build_alias_to_table(known)
    files = list_flat_data_files(folder)
    matched: dict[str, list[str]] = {}
    unmatched: list[str] = []
    for path in files:
        stem = os.path.splitext(os.path.basename(path))[0]
        table = match_stem_to_rule_table(stem, alias_map)
        if table:
            matched.setdefault(table, []).append(path)
        else:
            unmatched.append(path)
    matched_keys = set(matched.keys())
    # AUSP_* считаем покрытыми, если есть AUSP или любой derived
    covered = set(matched_keys)
    if any(k == 'AUSP' or k in _AUSP_DERIVED_TO_ATINN for k in matched_keys):
        covered.add('AUSP')
        covered.update(_AUSP_DERIVED_TO_ATINN.keys())
    missing_in_folder = [t for t in known if t not in covered]
    return {
        'rules_path': rules_path or default_rules_path(),
        'folder': os.path.abspath(folder),
        'known_tables': known,
        'matched': {k: sorted(v) for k, v in sorted(matched.items())},
        'unmatched': unmatched,
        'missing_in_folder': missing_in_folder,
    }


def print_excel_rules_discovery(report: dict) -> None:
    print('\n' + '=' * 70)
    print('СОПОСТАВЛЕНИЕ EXCEL ↔ rules.json')
    print('=' * 70)
    print(f"Папка: {report['folder']}")
    print(f"Правила: {report['rules_path']}")
    print(f"Таблиц в rules: {len(report['known_tables'])}")
    print(f"Сопоставлено групп: {len(report['matched'])}")
    print(f"Файлов без совпадения: {len(report['unmatched'])}")
    print('-' * 70)
    for table, paths in report['matched'].items():
        names = ', '.join(os.path.basename(p) for p in paths)
        print(f'  {table}: {len(paths)} файл(ов) → {names}')
    if report['unmatched']:
        print('\nНе сопоставлены (не из rules или другое имя):')
        for p in report['unmatched'][:30]:
            print(f'  - {os.path.basename(p)}')
        if len(report['unmatched']) > 30:
            print(f'  ... и ещё {len(report["unmatched"]) - 30}')
    if report['missing_in_folder']:
        preview = ', '.join(report['missing_in_folder'][:20])
        more = '' if len(report['missing_in_folder']) <= 20 else f' ... (+{len(report["missing_in_folder"]) - 20})'
        print(f'\nВ rules есть, файлов нет: {preview}{more}')
    print('=' * 70)


def merge_and_load_ausp_from_file_list(db_path=None, file_atinn_pairs=None, skip_header_after_first=True, chunksize=100000, skip_final_dedup=False, target_table=None):
    """Загрузка AUSP из списка файлов; file_atinn_pairs: [(path, atinn_or_None), ...]."""
    if db_path is None:
        db_path = _resolve_db_path()
    else:
        db_path = _resolve_db_path(db_path)
    if not file_atinn_pairs:
        print('ОШИБКА: пустой список файлов AUSP')
        return None
    data_files = []
    file_atinn = {}
    for path, atinn in file_atinn_pairs:
        if not os.path.isfile(path):
            continue
        data_files.append(path)
        if atinn:
            file_atinn[path] = str(atinn)
    if not data_files:
        print('ОШИБКА: нет существующих файлов AUSP')
        return None
    if target_table is None:
        targets = {_ausp_target_table_for_atinn(a) for _, a in file_atinn_pairs if a}
        if len(targets) > 1:
            print('ОШИБКА: смешаны customer и equipment ATINN в одном вызове')
            return None
        target_table = next(iter(targets), AUSP_TABLE_NAME)
    total_files = len(data_files)
    print('\n' + '=' * 80)
    print(f'ЗАГРУЗКА {target_table} ИЗ EXCEL (ATINN из имени/папки)')
    print('=' * 80)
    print(f'Файлов: {total_files}')
    print(f'Таблица в БД: {target_table}')
    print(f'База данных: {db_path}')
    print('=' * 80 + '\n')
    try:
        conn = connect_sqlite(db_path)
        cursor = conn.cursor()
        conn.execute('PRAGMA journal_mode = OFF')
        conn.execute('PRAGMA synchronous = OFF')
        conn.execute('PRAGMA cache_size = -20000')
        conn.execute('PRAGMA foreign_keys = OFF')
        conn.execute('PRAGMA temp_store = MEMORY')
        print(f"Удаление старой таблицы '{target_table}'...")
        cursor.execute(f"DROP TABLE IF EXISTS '{target_table}'")
        conn.commit()
    except Exception as e:
        print(f'ОШИБКА при подготовке БД: {e}')
        return None
    first_file_columns = None
    table_created = False
    total_rows_loaded = 0
    start_time = time.time()
    for file_idx, file_path in enumerate(data_files, 1):
        atinn_value = file_atinn.get(file_path)
        file_name = os.path.basename(file_path)
        atinn_note = f' (ATINN={atinn_value})' if atinn_value else ''
        print(f'\nФайл {file_idx}/{total_files}: {file_name}{atinn_note}')
        try:
            if file_idx == 1:
                df = _normalize_df_columns(_read_data_file(file_path, header=0))
                first_file_columns = list(df.columns)
            elif skip_header_after_first and first_file_columns:
                df = _read_data_file(file_path, header=None)
                if len(df) > 0:
                    df = df.iloc[1:].reset_index(drop=True)
                if len(df.columns) == len(first_file_columns):
                    df.columns = first_file_columns
                elif len(df.columns) > len(first_file_columns):
                    df.columns = first_file_columns + [f'extra_col_{j}' for j in range(len(first_file_columns), len(df.columns))]
                else:
                    df.columns = first_file_columns[:len(df.columns)]
            else:
                df = _normalize_df_columns(_read_data_file(file_path, header=0))
            df = _normalize_df_columns(df)
            if atinn_value:
                df = _inject_atinn_column(df, atinn_value)
            if first_file_columns is None:
                first_file_columns = list(df.columns)
            if len(df) > 0:
                df = df.drop_duplicates()
            if not table_created and len(df) > 0:
                sample_df = df.head(min(10000, len(df)))
                sample_df.to_sql(target_table, conn, if_exists='fail', index=False, chunksize=chunksize)
                table_created = True
                if len(df) > len(sample_df):
                    df.iloc[len(sample_df):].to_sql(target_table, conn, if_exists='append', index=False, chunksize=chunksize)
            elif table_created and len(df) > 0:
                df.to_sql(target_table, conn, if_exists='append', index=False, chunksize=chunksize)
            total_rows_loaded += len(df)
            del df
            gc.collect()
        except Exception as e:
            print(f'  ОШИБКА: {e} — файл пропущен')
            continue
    conn.commit()
    if table_created and (not skip_final_dedup):
        try:
            before, after, removed = _dedup_table_in_db(conn, target_table)
            if removed:
                print(f'\nДедупликация {target_table}: {before:,} → {after:,} (−{removed:,})')
            conn.commit()
        except Exception as e:
            print(f'WARN дедуп {target_table}: {e}')
    try:
        cursor.execute(f"SELECT COUNT(*) FROM '{target_table}'")
        count = cursor.fetchone()[0]
    except Exception:
        count = total_rows_loaded
    conn.close()
    total_time = time.time() - start_time
    print('\n' + '=' * 80)
    print(f'{target_table} загружена: {count:,} строк за {total_time:.1f} сек')
    print('=' * 80)
    if count == 0:
        return None
    return {'table_name': target_table, 'source_files': total_files, 'db_rows': count, 'total_time': total_time}


def load_excel_matched_to_rules(folder, db_path=None, rules_path=None, method='fast', skip_final_dedup=False, dry_run=False, only_tables=None):
    """Загрузка Excel из плоской папки: имена файлов ↔ таблицы rules.json → SQLite."""
    report = discover_excel_by_rules(folder, rules_path=rules_path)
    print_excel_rules_discovery(report)
    if dry_run:
        print('\n[dry-run] Загрузка в БД не выполнялась.')
        return report
    matched = report['matched']
    if only_tables:
        only_set = {str(t).strip() for t in only_tables}
        matched = {k: v for k, v in matched.items() if k in only_set}
    if not matched:
        print('Нет файлов, сопоставленных с rules — нечего загружать.')
        return report
    if db_path is None:
        db_path = _resolve_db_path()
    else:
        db_path = _resolve_db_path(db_path)
    # AUSP customer vs AUSP_EQUIPMENT
    ausp_pairs = []
    equipment_pairs = []
    other = {}
    for table, paths in matched.items():
        tu = str(table).strip().upper()
        if tu == 'AUSP':
            for p in paths:
                ausp_pairs.append((p, None))
        elif tu == AUSP_EQUIPMENT_TABLE_NAME:
            for p in paths:
                equipment_pairs.append((p, None))
        elif table in _AUSP_DERIVED_TO_ATINN:
            atinn = _AUSP_DERIVED_TO_ATINN[table]
            if _ausp_target_table_for_atinn(atinn) == AUSP_EQUIPMENT_TABLE_NAME:
                for p in paths:
                    equipment_pairs.append((p, atinn))
            else:
                for p in paths:
                    ausp_pairs.append((p, atinn))
        else:
            other[table] = paths
    results = []
    tables_to_load = list(other.keys())
    if ausp_pairs:
        tables_to_load = ['AUSP'] + tables_to_load
    if equipment_pairs:
        tables_to_load = [AUSP_EQUIPMENT_TABLE_NAME] + tables_to_load
    print(f'\nК загрузке в БД: {len(tables_to_load)} таблиц(ы) → {db_path}')
    done = 0
    total = len(tables_to_load)
    if ausp_pairs:
        done += 1
        print(f'\n[{done}/{total}] AUSP ({len(ausp_pairs)} файл(ов))')
        r = merge_and_load_ausp_from_file_list(db_path=db_path, file_atinn_pairs=ausp_pairs, skip_final_dedup=skip_final_dedup, target_table=AUSP_TABLE_NAME)
        if r:
            results.append(r)
    if equipment_pairs:
        done += 1
        print(f'\n[{done}/{total}] AUSP_EQUIPMENT ({len(equipment_pairs)} файл(ов))')
        r = merge_and_load_ausp_from_file_list(db_path=db_path, file_atinn_pairs=equipment_pairs, skip_final_dedup=skip_final_dedup, target_table=AUSP_EQUIPMENT_TABLE_NAME)
        if r:
            results.append(r)
    for table, paths in other.items():
        done += 1
        print(f'\n[{done}/{total}] {table} ({len(paths)} файл(ов))')
        if method == 'ultra_fast':
            r = merge_and_load_xlsx_files_ultra_fast(db_path=db_path, data_folder=folder, target_table=table, skip_final_dedup=skip_final_dedup, data_files=paths)
        else:
            r = merge_and_load_xlsx_files_fast(db_path=db_path, data_folder=folder, target_table=table, skip_final_dedup=skip_final_dedup, data_files=paths)
        if r:
            results.append(r)
    report['load_results'] = results
    print('\n' + '=' * 70)
    print(f"ИТОГО загружено таблиц: {len(results)} из {total}")
    print('=' * 70)
    return report


def _resolve_table_data_folder(table_name, base_abs):
    tu = str(table_name or '').strip().upper()
    if tu == AUSP_TABLE_NAME:
        ausp_path = resolve_ausp_data_folder()
        if ausp_path:
            return ausp_path
    if tu == AUSP_EQUIPMENT_TABLE_NAME:
        eq_path = resolve_ausp_equipment_data_folder()
        if eq_path:
            return eq_path
    return os.path.join(base_abs, table_name)

def _load_table_folder(db_path, table_name, data_folder, method, skip_final_dedup):
    tu = str(table_name or '').strip().upper()
    if tu in (AUSP_TABLE_NAME, AUSP_EQUIPMENT_TABLE_NAME) and _collect_ausp_atinn_file_groups(data_folder):
        print(f'   [{tu}] Вложенная структура по ATINN — split customer/equipment (без перезаписи чужой таблицы)')
        return merge_and_load_ausp_from_atinn_folders(db_path=db_path, ausp_folder=data_folder, skip_final_dedup=skip_final_dedup)
    if method == 'ultra_fast':
        return merge_and_load_xlsx_files_ultra_fast(db_path=db_path, data_folder=data_folder, target_table=table_name, skip_header_after_first=True, batch_size=50000, skip_final_dedup=skip_final_dedup)
    return merge_and_load_xlsx_files_fast(db_path=db_path, data_folder=data_folder, target_table=table_name, skip_header_after_first=True, chunksize=100000, skip_final_dedup=skip_final_dedup)


def _load_table_from_group(db_path, table_name, info, method, skip_final_dedup, base_folder):
    """Загрузка одной таблицы из collect_db_load_groups."""
    files = list(info.get('files') or [])
    file_atinn = dict(info.get('file_atinn') or {})
    tu = str(table_name).strip().upper()
    if tu == AUSP_EQUIPMENT_TABLE_NAME:
        if not files and info.get('folder'):
            files = _list_data_files(info['folder'])
            atinn_groups = _collect_ausp_atinn_file_groups(info['folder']) or {}
            for _a, paths in atinn_groups.items():
                files.extend(paths)
            files = sorted(set(files))
        if not files:
            print(f'Пропуск {table_name}: нет файлов единого дампа')
            return None
        print(f'   [{tu}] единый дамп {len(files)} файл(ов); customer AUSP не трогаем')
        return merge_and_load_ausp_equipment_flat(
            db_path=db_path,
            data_files=files,
            skip_final_dedup=skip_final_dedup,
            method=method,
        )
    if tu == AUSP_TABLE_NAME:
        pairs = [(p, file_atinn.get(p)) for p in files]
        if not pairs and info.get('folder'):
            atinn_groups = _collect_ausp_atinn_file_groups(info['folder']) or {}
            for atinn, paths in atinn_groups.items():
                if _ausp_target_table_for_atinn(atinn) != AUSP_TABLE_NAME:
                    continue
                for p in paths:
                    pairs.append((p, atinn))
            if not pairs and not atinn_groups:
                # flat customer dump in AUSP/
                for p in _list_data_files(info['folder']):
                    pairs.append((p, None))
        if not pairs:
            print(f'Пропуск {table_name}: нет файлов customer AUSP')
            return None
        print(f'   [{tu}] загрузка {len(pairs)} файл(ов); AUSP_EQUIPMENT не трогаем')
        return _load_ausp_pairs_split(db_path, pairs, skip_final_dedup=skip_final_dedup, only_target=AUSP_TABLE_NAME)
    if not files:
        print(f'Пропуск {table_name}: нет файлов')
        return None
    if method == 'ultra_fast':
        return merge_and_load_xlsx_files_ultra_fast(db_path=db_path, data_folder=base_folder, target_table=table_name, skip_final_dedup=skip_final_dedup, data_files=files)
    return merge_and_load_xlsx_files_fast(db_path=db_path, data_folder=base_folder, target_table=table_name, skip_final_dedup=skip_final_dedup, data_files=files)


def load_all_tables_from_db_folders(db_path=None, base_folder=None, method='fast', skip_final_dedup=False, only_tables=None):
    if db_path is None:
        db_path = _resolve_db_path()
    else:
        db_path = _resolve_db_path(db_path)
    base_abs = _resolve_data_path(base_folder)
    if not os.path.isdir(base_abs):
        print(f'ОШИБКА: Папка не найдена: {base_abs}')
        return []
    groups = collect_db_load_groups(base_folder)
    if only_tables is not None:
        only_set = {str(t).strip() for t in only_tables}
        groups = {k: v for k, v in groups.items() if k in only_set}
    tables = sorted(groups.keys(), key=lambda s: str(s).upper())
    if not tables:
        print(f'В папке {base_abs} нет подпапок/Excel с таблицами.')
        return []
    print('\n' + '=' * 70)
    print('ЗАГРУЗКА ТАБЛИЦ ИЗ db/ (подпапки + плоские Excel)')
    print('=' * 70)
    print(f'Базовая папка: {base_abs}')
    print(f'БД: {db_path}')
    print(f'Таблиц к загрузке: {len(tables)}')
    for t in tables:
        n = len(groups[t].get('files') or [])
        print(f'  - {t}: {n} файл(ов)')
    print('=' * 70 + '\n')
    results = []
    for i, table_name in enumerate(tables, 1):
        info = groups[table_name]
        print(f'\n[{i}/{len(tables)}] Таблица: {table_name}')
        r = _load_table_from_group(db_path, table_name, info, method, skip_final_dedup, base_abs)
        if r:
            results.append(r)
    print('\n' + '=' * 70)
    print('ИТОГО ЗАГРУЖЕНО ТАБЛИЦ:', len(results), 'из', len(tables))
    print('=' * 70)
    return results

def _interactive_pick_table_name(exclude_ausp=True):
    all_tables = get_table_folders()
    ausp_names = {AUSP_TABLE_NAME, AUSP_EQUIPMENT_TABLE_NAME}
    if exclude_ausp:
        pick_list = [t for t in all_tables if str(t).strip().upper() not in ausp_names]
    else:
        pick_list = list(all_tables)
    if not pick_list:
        print('Нет доступных таблиц (подпапки или Excel в db/).')
        return None
    print('\nДоступные таблицы:')
    for i, name in enumerate(pick_list, 1):
        print(f'  {i:2}. {name}')
    if exclude_ausp and any(t in all_tables for t in (AUSP_TABLE_NAME, AUSP_EQUIPMENT_TABLE_NAME)):
        print(f'  (для AUSP / AUSP_EQUIPMENT используйте пункт меню 4 — загрузка по ATINN)')
    raw = input('\nВведите номер или имя таблицы: ').strip()
    if not raw:
        return None
    if raw.isdigit():
        idx = int(raw)
        if 1 <= idx <= len(pick_list):
            return pick_list[idx - 1]
        print(f'Неверный номер: {raw}')
        return None
    matched = next((t for t in pick_list if str(t).strip().upper() == raw.upper()), None)
    if matched:
        return matched
    print(f"Таблица '{raw}' не найдена.")
    return None

def _interactive_load_one_table(method='fast'):
    table_name = _interactive_pick_table_name(exclude_ausp=True)
    if not table_name:
        return None
    if str(table_name).strip().upper() == AUSP_TABLE_NAME:
        print('Для AUSP выберите пункт 4 меню (загрузка из подпапок ATINN).')
        return None
    groups = collect_db_load_groups()
    info = groups.get(table_name)
    if not info or not (info.get('files') or info.get('folder')):
        print(f"ОШИБКА: нет файлов для таблицы '{table_name}'")
        return None
    print(f"\nЗагрузка одной таблицы: {table_name}")
    files = info.get('files') or []
    for fp in files[:10]:
        print(f'  - {os.path.basename(fp)}')
    if len(files) > 10:
        print(f'  ... и ещё {len(files) - 10}')
    skip_dedup = input('Пропустить финальную дедупликацию? (y/n) [n]: ').strip().lower() == 'y'
    return _load_table_from_group(_resolve_db_path(), table_name, info, method, skip_dedup, _resolve_data_path())


def _interactive_load_excel_by_rules():
    default_folder = _resolve_data_path()
    print('\nРежим: что найдено в обычной папке db/ (подпапки + Excel)')
    print('Кладите файлы прямо в db/: KNA1.xlsx, KNA1_1.xlsx, V_EQUI.xlsx — или в db/KNA1/')
    print(f'Папка по умолчанию: {default_folder}')
    raw = input('Папка (Enter = db/): ').strip().strip('"')
    folder = raw or default_folder
    if not os.path.isdir(folder):
        print(f'ОШИБКА: папка не найдена: {folder}')
        return None
    groups = collect_db_load_groups(folder)
    print('\n' + '=' * 70)
    print('ИСТОЧНИКИ ЗАГРУЗКИ (как увидит пункт 3)')
    print('=' * 70)
    if not groups:
        print('(пусто)')
    for table, info in sorted(groups.items(), key=lambda x: str(x[0]).upper()):
        files = info.get('files') or []
        names = ', '.join(os.path.basename(f) for f in files[:8])
        more = '' if len(files) <= 8 else f' ... (+{len(files) - 8})'
        print(f'  {table}: {len(files)} файл(ов) → {names}{more}')
    # дополнительно: только плоские файлы ↔ rules
    load_excel_matched_to_rules(folder, dry_run=True)
    return {'matched': groups}


if __name__ == '__main__':
    print('\n' + '=' * 80)
    print('СКРИПТ ЗАГРУЗКИ ТАБЛИЦ В БАЗУ ДАННЫХ')
    print('=' * 80)
    _db, _db_src = resolve_database_path(_PROJECT_ROOT)
    print(f'Корень проекта: {_PROJECT_ROOT}')
    print(f'БД ({_db_src}): {_db}')
    print(f'Папка выгрузок: {_resolve_data_path()}')
    ausp_path = resolve_ausp_data_folder()
    if ausp_path:
        print(f'Папка AUSP (по ATINN): {ausp_path}')
    tables_in_db = get_table_folders()
    print(f"\nВ папке '{_resolve_data_path()}' найдено таблиц (подпапки + Excel): {len(tables_in_db)}")
    if tables_in_db:
        print('   ', ', '.join(tables_in_db[:15]), '...' if len(tables_in_db) > 15 else '')
    print(f'\nВыберите режим:')
    print('1. Залить одну таблицу (быстрый метод) — подпапка или Excel в db/')
    print('2. Залить одну таблицу (ультра-быстрый метод)')
    print('3. Залить ВСЕ таблицы из db/ (подпапки + плоские Excel)')
    print('4. AUSP: customer из папок ATINN (143/604/…) → AUSP; equipment единым дампом → AUSP_EQUIPMENT')
    print('5. Показать сопоставление файлов в db/ с rules.json (без загрузки)')
    try:
        choice = input('\nВведите номер (1–5): ').strip()
        if choice == '5':
            result = _interactive_load_excel_by_rules()
            if result:
                print('\nСопоставление готово.')
            else:
                print('\nОперация завершена с ошибками или отменена.')
        elif choice == '4':
            print('\nРежим: customer AUSP (папки ATINN) + equipment единым дампом (db/AUSP_EQUIPMENT/).')
            print('Equipment не делится на 24/27/30/52 — ATINN уже в колонке файла.')
            skip_dedup = input('Пропустить финальную дедупликацию? (y/n) [n]: ').strip().lower() == 'y'
            result = merge_and_load_ausp_from_atinn_folders(skip_final_dedup=skip_dedup)
            if result:
                print(f"\nГотово: {result.get('table_name')} — {result['db_rows']:,} строк")
            else:
                print('\nAUSP не загружена — проверьте структуру папок и файлы.')
        elif choice == '3':
            print('\nРежим: загрузка всех таблиц из db/ (подпапки + Excel в корне db/).')
            print('AUSP_EQUIPMENT = единый дамп (без папок ATINN); customer AUSP — отдельно.')
            skip_ausp = input('Не трогать AUSP + AUSP_EQUIPMENT? (y/n) [n]: ').strip().lower() == 'y'
            m = input('Метод: 1=быстрый, 2=ультра-быстрый [1]: ').strip() or '1'
            skip_dedup = input('Пропустить финальную дедупликацию по таблице (ускоряет загрузку)? (y/n) [n]: ').strip().lower() == 'y'
            only = None
            if skip_ausp:
                only = [
                    t for t in get_table_folders()
                    if str(t).strip().upper() not in (AUSP_TABLE_NAME, AUSP_EQUIPMENT_TABLE_NAME)
                ]
            result = load_all_tables_from_db_folders(method='ultra_fast' if m == '2' else 'fast', skip_final_dedup=skip_dedup, only_tables=only)
            if result:
                print(f'\nОперация завершена. Загружено таблиц: {len(result)}')
            else:
                print('\nНет таблиц для загрузки или ошибки.')
        elif choice == '2':
            result = _interactive_load_one_table(method='ultra_fast')
            if result:
                print(f"\nУСПЕХ: таблица '{result['table_name']}' — {result['db_rows']:,} строк")
            else:
                print(f'\nОперация завершена с ошибками или отменена.')
        elif choice == '1':
            result = _interactive_load_one_table(method='fast')
            if result:
                print(f"\nУСПЕХ: таблица '{result['table_name']}' — {result['db_rows']:,} строк")
            else:
                print(f'\nОперация завершена с ошибками или отменена.')
        else:
            print(f'\nНеверный выбор. Запуск интерактивной загрузки одной таблицы (быстрый метод).')
            result = _interactive_load_one_table(method='fast')
            if result:
                print(f"\nУСПЕХ: таблица '{result['table_name']}' — {result['db_rows']:,} строк")
    except KeyboardInterrupt:
        print(f'\n\nОперация прервана пользователем')
    except Exception as e:
        print(f'\nКРИТИЧЕСКАЯ ОШИБКА: {e}')
