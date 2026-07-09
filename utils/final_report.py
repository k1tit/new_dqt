import os
import sys
from datetime import datetime
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
import matplotlib.colors as mcolors
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import io
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)
try:
    from utils.symbols import EXCEL_SYMBOLS, TERMINAL_SYMBOLS, ts, xs
except ImportError:
    EXCEL_SYMBOLS = {'SUCCESS': '✅', 'ERROR': '❌', 'WARNING': '⚠️', 'INFO': 'ℹ️', 'ROCKET': '🚀', 'GEAR': '⚙️', 'MAGNIFYING_GLASS': '🔍', 'CHECKMARK': '✓', 'CHART_UP': '📈', 'CHART_DOWN': '📉', 'FILE_FOLDER': '📁', 'PAGE': '📄', 'TABLE': '📊', 'BOOKS': '📚', 'SAVE': '💾', 'BAR_CHART': '📊', 'CELEBRATION': '🎉', 'MEMO': '📝', 'TARGET': '🎯', 'CLOCK': '⏱️', 'SIREN': '🚨', 'TROPHY': '🏆', 'STAR': '⭐', 'FIRE': '🔥', 'SNOWFLAKE': '❄️', 'CHECKERED_FLAG': '🏁'}
    TERMINAL_SYMBOLS = {'SUCCESS': '[OK]', 'ERROR': '[ERROR]', 'WARNING': '[WARN]', 'INFO': '[INFO]', 'ROCKET': '[START]', 'GEAR': '[PROC]', 'MAGNIFYING_GLASS': '[CHECK]', 'CHECKMARK': '[DONE]', 'CELEBRATION': '[DONE]', 'MEMO': '[NOTE]', 'TARGET': '[TARGET]', 'CLOCK': '[TIME]', 'SIREN': '[ALERT]', 'TROPHY': '[BEST]', 'STAR': '[STAR]', 'FIRE': '[HOT]', 'SNOWFLAKE': '[COLD]', 'CHECKERED_FLAG': '[FINISH]'}

class FinalReportGenerator:

    def __init__(self, output_dir):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def generate_executive_summary(self, results_df):
        if results_df.empty:
            print(f'{ts("ERROR")} Нет данных для отчета')
            return None
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        report_path = os.path.join(self.output_dir, f'DQ_FINAL_REPORT_{timestamp}.xlsx')
        print(f'\n{ts("ROCKET")} СОЗДАНИЕ ИТОГОВОГО ОТЧЕТА')
        print('=' * 60)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Итоговый отчет'
        self._define_styles()
        self._create_executive_summary_sheet(ws, results_df)
        self._create_dashboard_sheet(wb, results_df)
        self._create_rules_status_sheet(wb, results_df)
        self._create_category_summary_sheet(wb, results_df)
        wb.save(report_path)
        print(f'{ts("CELEBRATION")} ИТОГОВЫЙ ОТЧЕТ СОХРАНЕН: {report_path}')
        html_path = self._create_html_dashboard(results_df)
        return report_path

    def _define_styles(self):
        self.styles = {'STATUS_PASSED': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'), 'STATUS_FAILED': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'), 'STATUS_WARNING': PatternFill(start_color='FFEB9C', end_color='FFEB9CE', fill_type='solid'), 'STATUS_CRITICAL': PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid'), 'QUALITY_0_50': PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid'), 'QUALITY_50_70': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'), 'QUALITY_70_85': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'), 'QUALITY_85_95': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'), 'QUALITY_95_100': PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'), 'HEADER_MAIN': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'), 'HEADER_SECONDARY': PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid'), 'HEADER_TERTIARY': PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid'), 'THIN_BORDER': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), 'FONT_HEADER': Font(bold=True, size=14, color='FFFFFF'), 'FONT_TITLE': Font(bold=True, size=12, color='1F4E78'), 'FONT_SUBTITLE': Font(bold=True, size=11, color='2F5496'), 'FONT_NORMAL': Font(size=10), 'FONT_BOLD': Font(bold=True, size=10), 'FONT_SMALL': Font(size=9)}

    def _create_executive_summary_sheet(self, ws, results_df):
        print(f'   {ts("GEAR")} Создаем главный отчет...')
        ws.merge_cells('A1:K1')
        ws['A1'] = f'{xs("TROPHY")} ИТОГОВЫЙ ОТЧЕТ ПО КАЧЕСТВУ ДАННЫХ'
        ws['A1'].font = Font(bold=True, size=18, color='1F4E78')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws.merge_cells('A2:K2')
        generated = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A2'] = f'Сгенерировано: {generated}'
        ws['A2'].font = Font(italic=True, size=10, color='666666')
        ws['A2'].alignment = Alignment(horizontal='center')
        total_rules = len(results_df)
        passed_rules = len(results_df[results_df['status'] == 'PASSED'])
        failed_rules = len(results_df[results_df['status'] == 'FAILED'])
        total_errors = results_df['failed'].sum()
        avg_quality = results_df['success_rate_%'].mean()
        total_records = results_df['total_records'].sum()
        if avg_quality >= 95:
            system_status = 'ОТЛИЧНО'
            status_color = '92D050'
        elif avg_quality >= 85:
            system_status = 'ХОРОШО'
            status_color = 'C6EFCE'
        elif avg_quality >= 70:
            system_status = 'УДОВЛЕТВОРИТЕЛЬНО'
            status_color = 'FFEB9C'
        else:
            system_status = 'ТРЕБУЕТ ВНИМАНИЯ'
            status_color = 'FF9999'
        ws['A4'] = f'{xs("BAR_CHART")} КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ'
        ws['A4'].font = self.styles['FONT_TITLE']
        ws['A4'].fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        ws.merge_cells('A4:C4')
        mean_exec = results_df['execution_time_sec'].mean()
        indicators = [['Общее качество данных', f'{avg_quality:.1f}%', system_status, status_color], ['Всего проверено правил', str(total_rules), f'{passed_rules} успешно', ''], ['Найдено ошибок', f'{total_errors:,}', f'{failed_rules} правил нарушено', ''], ['Обработано записей', f'{total_records:,}', '', ''], ['Среднее время проверки', f'{mean_exec:.2f}с', '', '']]
        headers = ['Показатель', 'Значение', 'Статус', '']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = self.styles['FONT_BOLD']
            cell.fill = self.styles['HEADER_TERTIARY']
            cell.border = self.styles['THIN_BORDER']
            cell.alignment = Alignment(horizontal='center')
        for idx, (label, value, status, color) in enumerate(indicators, 1):
            row = 5 + idx
            ws.cell(row=row, column=1, value=label).border = self.styles['THIN_BORDER']
            ws.cell(row=row, column=2, value=value).border = self.styles['THIN_BORDER']
            status_cell = ws.cell(row=row, column=3, value=status)
            status_cell.border = self.styles['THIN_BORDER']
            status_cell.alignment = Alignment(horizontal='center')
            if color:
                status_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                status_cell.font = Font(bold=True, color='FFFFFF' if color == 'FF9999' else '000000')
            elif 'успешно' in status.lower():
                status_cell.fill = self.styles['STATUS_PASSED']
            elif 'нарушено' in status.lower():
                status_cell.fill = self.styles['STATUS_FAILED']
        start_row = 12
        ws.cell(row=start_row, column=1, value=f'{xs("MAGNIFYING_GLASS")} ВИЗУАЛЬНАЯ СВОДКА ПРОВЕРОК')
        ws.merge_cells(f'A{start_row}:L{start_row}')
        ws.cell(row=start_row, column=1).font = self.styles['FONT_TITLE']
        ws.cell(row=start_row, column=1).fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        summary_headers = ['№', 'Код правила', 'Категория', 'Таблица', 'Колонка', 'Оценено', 'Успешно', 'Ошибок', '% успеха', 'Статус', 'Время (с)']
        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=start_row + 1, column=col, value=header)
            cell.font = self.styles['FONT_BOLD']
            cell.fill = self.styles['HEADER_SECONDARY']
            cell.border = self.styles['THIN_BORDER']
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for idx, (_, row) in enumerate(results_df.sort_values('success_rate_%').iterrows(), 1):
            data_row = start_row + 1 + idx
            row_data = [idx, row['rule_code'], row['quality_category'], row['table_name'], row.get('matched_column', row.get('column_checked', '')), row['total_records'], row['passed'], row['failed'], f"{row['success_rate_%']:.1f}%", row['status'], f"{row['execution_time_sec']:.2f}"]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=data_row, column=col, value=value)
                cell.border = self.styles['THIN_BORDER']
                if col in [6, 7, 8, 9, 11]:
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                if col == 10:
                    percent = float(str(value).replace('%', ''))
                    if percent >= 95:
                        cell.fill = self.styles['QUALITY_95_100']
                    elif percent >= 85:
                        cell.fill = self.styles['QUALITY_85_95']
                    elif percent >= 70:
                        cell.fill = self.styles['QUALITY_70_85']
                    elif percent >= 50:
                        cell.fill = self.styles['QUALITY_50_70']
                    else:
                        cell.fill = self.styles['QUALITY_0_50']
                elif col == 11:
                    if value == 'PASSED':
                        cell.fill = self.styles['STATUS_PASSED']
                        cell.font = Font(bold=True, color='006100')
                    else:
                        cell.fill = self.styles['STATUS_FAILED']
                        cell.font = Font(bold=True, color='9C0006')
                elif col == 2 and row['rule_code'] in ['RCCONF_12.2', 'RCCONF_12.3']:
                    cell.fill = self.styles['STATUS_CRITICAL']
                    cell.font = Font(bold=True, color='FFFFFF')
        category_summary_row = start_row + len(results_df) + 3
        ws.cell(row=category_summary_row, column=1, value=f'{xs("CHECKERED_FLAG")} РЕЗУЛЬТАТЫ ПО КАТЕГОРИЯМ КАЧЕСТВА')
        ws.merge_cells(f'A{category_summary_row}:L{category_summary_row}')
        ws.cell(row=category_summary_row, column=1).font = self.styles['FONT_TITLE']
        ws.cell(row=category_summary_row, column=1).fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        categories = results_df.groupby('quality_category').agg({'success_rate_%': 'mean', 'total_records': 'sum', 'failed': 'sum', 'rule_code': 'count'}).reset_index()
        categories.columns = ['Категория', 'Среднее качество', 'Оценено', 'Ошибок', 'Правил']
        cat_headers = list(categories.columns)
        for col, header in enumerate(cat_headers, 1):
            cell = ws.cell(row=category_summary_row + 1, column=col, value=header)
            cell.font = self.styles['FONT_BOLD']
            cell.fill = self.styles['HEADER_TERTIARY']
            cell.border = self.styles['THIN_BORDER']
            cell.alignment = Alignment(horizontal='center')
        for idx, (_, cat_row) in enumerate(categories.iterrows(), 1):
            data_row = category_summary_row + 1 + idx
            row_data = [cat_row['Категория'], f"{cat_row['Среднее качество']:.1f}%", cat_row['Оценено'], cat_row['Ошибок'], cat_row['Правил']]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=data_row, column=col, value=value)
                cell.border = self.styles['THIN_BORDER']
                if col > 1:
                    cell.alignment = Alignment(horizontal='center')
                if col == 2:
                    quality = cat_row['Среднее качество']
                    if quality >= 95:
                        cell.fill = self.styles['QUALITY_95_100']
                    elif quality >= 85:
                        cell.fill = self.styles['QUALITY_85_95']
                    elif quality >= 70:
                        cell.fill = self.styles['QUALITY_70_85']
                    elif quality >= 50:
                        cell.fill = self.styles['QUALITY_50_70']
                    else:
                        cell.fill = self.styles['QUALITY_0_50']
        for col in range(1, len(summary_headers) + 1):
            max_length = 0
            for row in range(start_row + 1, start_row + len(results_df) + 2):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 30)
        ws.auto_filter.ref = f'A{start_row + 1}:K{start_row + len(results_df) + 1}'
        print(f'   {ts("SUCCESS")} Главный отчет создан')

    def _create_dashboard_sheet(self, wb, results_df):
        print(f'   {ts("GEAR")} Создаем дашборд...')
        ws = wb.create_sheet('Дашборд')
        ws.merge_cells('A1:L1')
        ws['A1'] = f'{xs("BAR_CHART")} ДАШБОРД КАЧЕСТВА ДАННЫХ'
        ws['A1'].font = Font(bold=True, size=16, color='1F4E78')
        ws['A1'].alignment = Alignment(horizontal='center')
        self._add_visualizations(ws, results_df)
        print(f'   {ts("SUCCESS")} Дашборд создан')

    def _create_rules_status_sheet(self, wb, results_df):
        print(f'   {ts("GEAR")} Создаем детальный статус правил...')
        ws = wb.create_sheet('Статус правил')
        ws.merge_cells('A1:M1')
        ws['A1'] = f'{xs("MAGNIFYING_GLASS")} ДЕТАЛЬНЫЙ СТАТУС ПРАВИЛ'
        ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws['A3'] = f'{xs("PALETTE")} ЛЕГЕНДА ЦВЕТОВ:'
        ws['A3'].font = Font(bold=True, size=11)
        legend_items = [('✅ УСПЕШНО', 'C6EFCE', 'Правило успешно выполнено'), ('❌ НЕУСПЕШНО', 'FFC7CE', 'Правило нарушено'), ('⚠️ ПРОБЛЕМНОЕ', 'FF9999', 'Правило содержит ошибку в логике'), ('📊 КАЧЕСТВО 95-100%', '92D050', 'Отличное качество'), ('📊 КАЧЕСТВО 85-95%', 'C6EFCE', 'Хорошее качество'), ('📊 КАЧЕСТВО 70-85%', 'FFEB9C', 'Удовлетворительное качество'), ('📊 КАЧЕСТВО <70%', 'FF9999', 'Низкое качество')]
        for idx, (label, color, description) in enumerate(legend_items):
            row = 3 + idx
            ws.cell(row=row, column=2, value=label).font = Font(bold=True)
            ws.cell(row=row, column=3, value=description)
            color_cell = ws.cell(row=row, column=1)
            color_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            color_cell.value = '   '
        start_row = len(legend_items) + 5
        headers = ['Код правила', 'Описание', 'Категория', 'Таблица', 'Статус', 'Качество', 'Ошибок', 'Оценено', 'Рекомендация', 'Приоритет', 'Ссылка на ошибки']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.border = self.styles['THIN_BORDER']
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for idx, (_, row) in enumerate(results_df.iterrows(), 1):
            data_row = start_row + idx
            recommendation, priority = self._generate_rule_recommendation(row)
            row_data = [row['rule_code'], row['rule_description'], row['quality_category'], row['table_name'], row['status'], f"{row['success_rate_%']:.1f}%", row['failed'], row['total_records'], recommendation, priority, 'Есть' if row['error_file'] else 'Нет']
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=data_row, column=col, value=value)
                cell.border = self.styles['THIN_BORDER']
                if col == 5:
                    if value == 'PASSED':
                        cell.fill = self.styles['STATUS_PASSED']
                        cell.font = Font(bold=True, color='006100')
                    else:
                        cell.fill = self.styles['STATUS_FAILED']
                        cell.font = Font(bold=True, color='9C0006')
                elif col == 6:
                    percent = float(str(value).replace('%', ''))
                    if percent >= 95:
                        cell.fill = self.styles['QUALITY_95_100']
                    elif percent >= 85:
                        cell.fill = self.styles['QUALITY_85_95']
                    elif percent >= 70:
                        cell.fill = self.styles['QUALITY_70_85']
                    else:
                        cell.fill = self.styles['QUALITY_0_50']
                elif col == 10:
                    if value == 'Высокий':
                        cell.fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF')
                    elif value == 'Средний':
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                        cell.font = Font(bold=True)
                if row['rule_code'] in ['RCCONF_12.2', 'RCCONF_12.3']:
                    cell.fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF')
        for col in range(1, len(headers) + 1):
            max_length = 0
            for row in range(start_row, start_row + len(results_df) + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 40)
        print(f'   {ts("SUCCESS")} Статус правил создан')

    def _create_category_summary_sheet(self, wb, results_df):
        print(f'   {ts("GEAR")} Создаем сводку по категориям...')
        ws = wb.create_sheet('Сводка по категориям')
        category_stats = results_df.groupby('quality_category').agg({'rule_code': 'count', 'success_rate_%': 'mean', 'total_records': 'sum', 'passed': 'sum', 'failed': 'sum'}).round(2)
        ws['A1'] = f'{xs("CHECKERED_FLAG")} СВОДКА ПО КАТЕГОРИЯМ КАЧЕСТВА'
        ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
        ws.merge_cells('A1:F1')
        headers = ['Категория качества', 'Кол-во правил', 'Среднее качество', 'Оценено', 'Успешно', 'Ошибок']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        for idx, (category, stats) in enumerate(category_stats.iterrows(), 1):
            row = 3 + idx
            row_data = [category, int(stats['rule_code']), f"{stats['success_rate_%']:.1f}%", int(stats['total_records']), int(stats['passed']), int(stats['failed'])]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if col == 3:
                    quality = stats['success_rate_%']
                    if quality >= 95:
                        cell.fill = self.styles['QUALITY_95_100']
                    elif quality >= 85:
                        cell.fill = self.styles['QUALITY_85_95']
                    elif quality >= 70:
                        cell.fill = self.styles['QUALITY_70_85']
                    else:
                        cell.fill = self.styles['QUALITY_0_50']
        print(f'   {ts("SUCCESS")} Сводка по категориям создана')

    def _generate_rule_recommendation(self, rule_row):
        success_rate = rule_row['success_rate_%']
        failed_count = rule_row['failed']
        rule_code = rule_row['rule_code']
        if rule_code in ['RCCONF_12.2', 'RCCONF_12.3']:
            return ('Исправить логику правила в конфигурации', 'КРИТИЧЕСКИЙ')
        if success_rate >= 95:
            return ('Поддерживать текущий уровень', 'Низкий')
        elif success_rate >= 85:
            if failed_count > 100:
                return ('Провести выборочную очистку данных', 'Средний')
            else:
                return ('Мониторить качество', 'Низкий')
        elif success_rate >= 70:
            return (f'Требуется очистка {failed_count:,} записей', 'Высокий')
        else:
            return (f'СРОЧНО исправить {failed_count:,} ошибок', 'КРИТИЧЕСКИЙ')

    def _add_visualizations(self, ws, results_df):
        ws['A3'] = '📊 ВИЗУАЛИЗАЦИЯ РЕЗУЛЬТАТОВ:'
        ws['A3'].font = Font(bold=True, size=12)
        ws['A5'] = 'Распределение качества правил:'
        for idx, (_, row) in enumerate(results_df.sort_values('success_rate_%').iterrows(), 1):
            quality_bar = '█' * int(row['success_rate_%'] / 5)
            ws.cell(row=5 + idx, column=1, value=row['rule_code'])
            ws.cell(row=5 + idx, column=2, value=quality_bar)
            ws.cell(row=5 + idx, column=3, value=f"{row['success_rate_%']:.1f}%")

    def _create_html_dashboard(self, results_df):
        html_path = os.path.join(self.output_dir, 'dq_dashboard.html')
        generated = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        avg_quality = results_df['success_rate_%'].mean()
        total_failed = results_df['failed'].sum()
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Дашборд качества данных</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; background: #f5f5f5; }}
                .dashboard {{ max-width: 1200px; margin: 0 auto; }}
                .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; border-radius: 10px; margin-bottom: 30px; }}
                .kpi-card {{ background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); margin-bottom: 20px; }}
                .status-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); gap: 20px; margin: 30px 0; }}
                table {{ width: 100%; background: white; border-collapse: collapse; }}
                th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
            </style>
        </head>
        <body>
            <div class="dashboard">
                <div class="header">
                    <h1>Дашборд качества данных</h1>
                    <p>Сгенерировано: {generated}</p>
                </div>
                <div class="status-grid">
                    <div class="kpi-card"><h3>Общее качество</h3><h2>{avg_quality:.1f}%</h2></div>
                    <div class="kpi-card"><h3>Проверено правил</h3><h2>{len(results_df)}</h2></div>
                    <div class="kpi-card"><h3>Найдено ошибок</h3><h2>{total_failed:,}</h2></div>
                </div>
                <h2>Результаты проверок</h2>
                <table>
                    <tr><th>Код правила</th><th>Описание</th><th>Качество</th><th>Статус</th><th>Ошибок</th></tr>
        """
        for _, row in results_df.sort_values('success_rate_%').iterrows():
            quality_class = 'quality-low'
            rate = row['success_rate_%']
            if rate >= 95:
                quality_class = 'quality-95'
            elif rate >= 85:
                quality_class = 'quality-85'
            elif rate >= 70:
                quality_class = 'quality-70'
            status_class = 'status-passed' if row['status'] == 'PASSED' else 'status-failed'
            html_content += f"""
                    <tr>
                        <td>{row['rule_code']}</td>
                        <td>{row['rule_description'][:50]}...</td>
                        <td>{rate:.1f}%</td>
                        <td>{row['status']}</td>
                        <td>{row['failed']:,}</td>
                    </tr>
            """
        html_content += '\n                </table>\n            </div>\n        </body>\n        </html>\n'
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        print(f'   {ts("SUCCESS")} HTML дашборд создан: {html_path}')
        return html_path