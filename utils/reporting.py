import os
import sys
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)
try:
    from utils.symbols import EXCEL_SYMBOLS, TERMINAL_SYMBOLS, ts, xs
except ImportError:
    EXCEL_SYMBOLS = {'SUCCESS': '[OK]', 'ERROR': '[X]', 'WARNING': '[!]', 'INFO': '[i]', 'ROCKET': '[>]', 'GEAR': '[*]', 'MAGNIFYING_GLASS': '[?]', 'CHECKMARK': '[OK]', 'CHART_UP': '[+]', 'CHART_DOWN': '[-]', 'FILE_FOLDER': '[DIR]', 'PAGE': '[FILE]', 'TABLE': '[TBL]', 'BOOKS': '[DATA]', 'SAVE': '[SAVE]', 'BAR_CHART': '[CHART]'}
    TERMINAL_SYMBOLS = {'SUCCESS': '[OK]', 'ERROR': '[ERROR]', 'WARNING': '[WARN]', 'INFO': '[INFO]', 'SKIP': '[SKIP]', 'ROCKET': '[START]', 'GEAR': '[PROC]', 'MAGNIFYING_GLASS': '[CHECK]', 'CHECKMARK': '[DONE]', 'CHART_UP': '[STAT+]', 'CHART_DOWN': '[STAT-]', 'CLIPBOARD': '[CLIP]', 'FILE_FOLDER': '[DIR]', 'PAGE': '[FILE]', 'TABLE': '[TABLE]', 'COLUMN': '[COL]', 'BOOKS': '[DATA]', 'SAVE': '[SAVE]', 'TARGET': '[TARGET]', 'PALETTE': '[STYLE]', 'CELEBRATION': '[DONE]', 'BAR_CHART': '[CHART]', 'MEMO': '[NOTE]'}

class ExcelReportGenerator:

    def __init__(self, output_dir, error_manager=None):
        self.output_dir = output_dir
        self.error_manager = error_manager

    def create_comprehensive_report(self, results_df, include_errors=True):
        if results_df.empty:
            return None
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        excel_path = os.path.join(self.output_dir, f'DQ_Full_Report_{timestamp}.xlsx')
        print(f'{ts("PALETTE")} Создаем комплексный Excel отчет...')
        wb = Workbook()
        wb.remove(wb.active)
        self._create_color_summary_sheet(wb, results_df)
        self._create_simple_stats_sheet(wb, results_df)
        if self.error_manager and include_errors:
            if hasattr(self.error_manager, 'save_all_errors_to_excel'):
                errors_file = self.error_manager.save_all_errors_to_excel()
                if errors_file and os.path.exists(errors_file):
                    self._merge_error_sheets(wb, errors_file)
                    print(f'   {ts("SUCCESS")} Ошибки добавлены в отчет')
                else:
                    print(f'   {ts("WARNING")} Файл с ошибками не найден')
        wb.save(excel_path)
        print(f'{ts("CELEBRATION")} ПОЛНЫЙ ОТЧЕТ СОХРАНЕН: {excel_path}')
        print(f'   {xs("PAGE")} Листов: {len(wb.sheetnames)}')
        return excel_path

    def _create_color_summary_sheet(self, wb, results_df):
        ws = wb.create_sheet('Цветная сводка')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        bright_red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        dark_red_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws.merge_cells('A1:Q1')
        ws['A1'] = f'{xs("BAR_CHART")} ЦВЕТНАЯ СВОДКА ПРОВЕРОК'
        ws['A1'].font = Font(bold=True, size=16, color='1F4E78')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        total_errors = results_df['failed'].sum()
        avg_quality = results_df['success_rate_%'].mean()
        passed_rules = len(results_df[results_df['failed'] == 0])
        failed_rules = len(results_df[results_df['failed'] > 0])
        problematic_rules = ['RCCONF_12.2', 'RCCONF_12.3']
        problematic_count = len(results_df[results_df['rule_code'].isin(problematic_rules)])
        ws.merge_cells('A2:Q2')
        ws['A2'] = f'ЗЕЛЕНЫЕ = нет ошибок | КРАСНЫЕ = есть ошибки | ЯРКО-КРАСНЫЕ = ошибка в логике правила'
        ws['A2'].font = Font(bold=True, size=10, color='2F5496')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A3:Q3')
        ws['A3'] = f'[OK] Успешно: {passed_rules} | [!] С ошибками: {failed_rules} | Проблемные: {problematic_count} | Всего ошибок: {total_errors:,}'
        ws['A3'].font = Font(bold=True, size=10, color='000000')
        ws['A3'].alignment = Alignment(horizontal='center')
        headers = ['Код правила', 'Описание', 'Категория', 'Таблица', 'Колонка', 'Всего записей', 'Успешно', 'Ошибок', '% успеха', 'Статус', 'Время (сек)', 'Файл ошибок']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        PROBLEMATIC_RULES = ['RCCONF_12.2', 'RCCONF_12.3']
        for row_idx, (_, row) in enumerate(results_df.iterrows(), 1):
            data_row = 5 + row_idx
            rule_code = str(row['rule_code'])
            if rule_code in PROBLEMATIC_RULES:
                row_fill = bright_red_fill
                font_color = 'FFFFFF'
                status = '[!] ОШИБКА ПРАВИЛА'
            elif row['failed'] == 0:
                row_fill = green_fill
                font_color = '000000'
                status = '[OK] УСПЕШНО'
            else:
                row_fill = red_fill
                font_color = '000000'
                status = '[!] ОШИБКИ'
            row_data = [row['rule_code'], row['rule_description'][:60] + '...' if len(row['rule_description']) > 60 else row['rule_description'], row['quality_category'], row['table_name'], row.get('matched_column', row.get('column_checked', '')), row['total_records'], row['passed'], row['failed'], f"{row['success_rate_%']:.1f}%", status, f"{row['execution_time_sec']:.2f}", 'Есть' if row.get('error_file') else 'Нет']
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=data_row, column=col_idx, value=value)
                cell.border = thin_border
                cell.font = Font(size=9, color=font_color, bold=rule_code in PROBLEMATIC_RULES)
                cell.fill = row_fill
                if col_idx in [6, 7, 8, 9, 10, 12]:
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        column_widths = [15, 40, 12, 12, 15, 12, 12, 10, 10, 10, 12, 10, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        note_row = len(results_df) + 7
        ws.merge_cells(f'A{note_row}:M{note_row}')
        ws[f'A{note_row}'] = '[!] ВНИМАНИЕ: Правила RCCONF_12.2 и RCCONF_12.3 выделены ЯРКО-КРАСНЫМ - они содержат ошибки в логике и требуют исправления!'
        ws[f'A{note_row}'].font = Font(bold=True, size=11, color='FFFFFF')
        ws[f'A{note_row}'].fill = bright_red_fill
        ws[f'A{note_row}'].alignment = Alignment(horizontal='center')
        ws.freeze_panes = 'A6'

    def _create_simple_stats_sheet(self, wb, results_df):
        ws = wb.create_sheet('Статистика')
        header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        ws.merge_cells('A1:F1')
        ws['A1'] = f'{xs("CHART_UP")} СТАТИСТИКА ПО КАТЕГОРИЯМ'
        ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
        ws['A1'].alignment = Alignment(horizontal='center')
        category_stats = results_df.groupby('quality_category').agg({'rule_code': 'count', 'success_rate_%': 'mean', 'total_records': 'sum', 'passed': 'sum', 'failed': 'sum'}).round(2)
        headers = ['Категория', 'Правил', 'Среднее качество', 'Всего записей', 'Ошибок', 'Статус']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for idx, (category, stats) in enumerate(category_stats.iterrows(), 1):
            row = 3 + idx
            avg_quality = stats['success_rate_%']
            if avg_quality >= 95:
                row_fill = green_fill
                status = 'ОТЛИЧНО'
            elif avg_quality >= 80:
                row_fill = green_fill
                status = 'ХОРОШО'
            elif avg_quality >= 60:
                row_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                status = 'СРЕДНЕ'
            else:
                row_fill = red_fill
                status = 'ПЛОХО'
            row_data = [category, int(stats['rule_code']), f"{stats['success_rate_%']:.1f}%", int(stats['total_records']), int(stats['failed']), status]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center')
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def create_simple_color_report(self, results_df):
        if results_df.empty:
            return None
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        excel_path = os.path.join(self.output_dir, f'DQ_Color_Simple_{timestamp}.xlsx')
        print(f'{ts("PALETTE")} Создаем ПРОСТОЙ цветной отчет...')
        wb = Workbook()
        ws = wb.active
        ws.title = 'Результаты проверок'
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        ws['A1'] = '🎯 РЕЗУЛЬТАТЫ ПРОВЕРОК КАЧЕСТВА ДАННЫХ'
        ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
        ws.merge_cells('A1:F1')
        total_errors = results_df['failed'].sum()
        passed_rules = len(results_df[results_df['failed'] == 0])
        failed_rules = len(results_df[results_df['failed'] > 0])
        ws['A2'] = f'[OK] Успешно: {passed_rules} правил | [!] С ошибками: {failed_rules} правил | Всего ошибок: {total_errors:,}'
        ws.merge_cells('A2:F2')
        ws['A2'].font = Font(bold=True, size=10)
        headers = ['Код правила', 'Описание', 'Таблица', 'Ошибок', '% успеха', 'Статус']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for idx, (_, row) in enumerate(results_df.iterrows(), 1):
            row_idx = 4 + idx
            rule_code = str(row['rule_code'])
            if rule_code in ['RCCONF_12.2', 'RCCONF_12.3']:
                row_fill = red_fill
                status = '[!] ОШИБКА ПРАВИЛА'
            elif row['failed'] == 0:
                row_fill = green_fill
                status = '[OK] УСПЕШНО'
            else:
                row_fill = red_fill
                status = '[!] ОШИБКИ'
            row_data = [row['rule_code'], row['rule_description'][:40], row['table_name'], row['failed'], f"{row['success_rate_%']:.1f}%", status]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.freeze_panes = 'A5'
        wb.save(excel_path)
        print(f'{ts("SUCCESS")} ПРОСТОЙ цветной отчет сохранен: {excel_path}')
        return excel_path

    def _merge_error_sheets(self, main_wb, errors_file_path):
        try:
            errors_wb = load_workbook(errors_file_path)
            for sheet_name in errors_wb.sheetnames:
                if sheet_name in ['Сводка ошибок']:
                    continue
                source_ws = errors_wb[sheet_name]
                new_ws = main_wb.create_sheet(sheet_name)
                for row in source_ws.iter_rows(values_only=True):
                    new_ws.append(row)
                print(f'   {ts("SUCCESS")} Добавлен лист: {sheet_name}')
        except Exception as e:
            print(f'   {ts("ERROR")} Ошибка при объединении ошибок: {e}')